"""Miner: read Origin/Destination Code+Description pairs out of the real
OPUS filings in reference/2_OPUS and load them into the Location Bank as
source="sample_mined". Safe to re-run whenever new reference filings land -
it only ever adds or refreshes sample_mined rows, and
LocationBankStore.upsert_location's precedence means a hand-curated
manual_override entry is never overwritten by one mined here.

Run as: python -m mrg2opus.location_bank.bootstrap_from_samples

Rewritten 2026-08-30 because it could no longer run at all: it mined a
"Sample MRGs with OPUS FORMATS/" folder deleted long ago in favour of
reference/1_MRGs + reference/2_OPUS, so it pointed at nothing. Repointing
alone wasn't enough either - it read only '*PORT-PORT*' sheets, and NONE
of the 54 real filings carries one (that sheet is absent from every real
lane's output), so it would have mined zero rows and looked like it
worked. It now reads the RATES sheets those filings actually have.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from mrg2opus.location_bank.known_aliases import KNOWN_ALIASES
from mrg2opus.location_bank.models import LocationRecord
from mrg2opus.location_bank.store import LocationBankStore

OPUS_DIR = Path(__file__).resolve().parents[2] / "reference" / "2_OPUS"

# Origin Code/Description and Destination Code/Description columns are fixed
# by the shared OPUS RATES(-*) header layout (see schema/opus_columns.py).
ORIGIN_CODE_COL, ORIGIN_DESC_COL = 8, 9
DEST_CODE_COL, DEST_DESC_COL = 14, 15


def _has_rates_layout(ws) -> bool:
    """Whether this sheet really is the standard RATES layout, checked
    against its own header rather than its name.

    Name matching isn't safe here: real filings carry RATES-shaped sheets
    under many names ("RATES", "RATES-TZDAR", "AEW RATES") but ALSO
    vertical-rates sheets under names that look just as RATES-y - one is
    literally "V RATES" - and those are shifted a column, so columns
    8/9/14/15 land on descriptions instead of codes. Mining that put
    "ALEXANDRIA, EGYPT" into the bank as if it were a port code. Row 2 of
    a genuine RATES sheet labels those four columns Code/Description
    twice, which is a cheap and exact test.
    """
    header = [ws.cell(row=2, column=c).value for c in (ORIGIN_CODE_COL, ORIGIN_DESC_COL, DEST_CODE_COL, DEST_DESC_COL)]
    return [str(h or "").strip().lower() for h in header] == ["code", "description", "code", "description"]


def _country_from_code(code: str) -> str | None:
    return code[:2] if len(code) >= 2 and code[:2].isalpha() else None


def _scan_workbook(path: Path) -> tuple[dict[str, str], dict[str, set[str]]]:
    """Returns (clean, grouped):
    - clean: code -> name, mined from rows carrying a single port on each
      side (i.e. this port wasn't grouped with others in its raw row).
    - grouped: description_text -> set of codes seen sharing that exact
      combined description (i.e. siblings from the same raw multi-port
      group), for the elimination pass in bootstrap().
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    clean: dict[str, str] = {}
    grouped: dict[str, set[str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not _has_rates_layout(ws):
            continue
        for row in ws.iter_rows(min_row=3):
            # A short row means this sheet isn't the standard RATES layout
            # (or the row is padding) - skip rather than index off the end.
            if len(row) < DEST_DESC_COL:
                continue
            for code_col, desc_col in ((ORIGIN_CODE_COL, ORIGIN_DESC_COL), (DEST_CODE_COL, DEST_DESC_COL)):
                code = row[code_col - 1].value
                desc = row[desc_col - 1].value
                if not code or not desc:
                    continue
                codes = [c.strip() for c in str(code).split(";") if c.strip()]
                names = [n.strip() for n in str(desc).split(";") if n.strip()]
                if not codes or not names:
                    continue
                if any("  " in n for n in names):
                    # Some exports replace every comma with two spaces, so
                    # "DALIAN, LIAONING" arrives as "DALIAN  LIAONING"
                    # (confirmed for the TIER 1 filings and WEW's own RATES
                    # sheet; see test_parsers_lawc.py). Mining those would
                    # overwrite 16 correct comma'd names with the artifact,
                    # and no real port name has a double space - so skip
                    # them and take the name from a clean filing instead.
                    continue
                if len(codes) == 1 and len(names) == 1:
                    clean.setdefault(codes[0], names[0])
                elif len(codes) == len(names):
                    # A multi-port row: both lists are ";"-joined but each is
                    # sorted independently, so position N of one does NOT
                    # correspond to position N of the other. Collect the
                    # sibling set for the elimination pass rather than
                    # pairing them up and mining a wrong name.
                    grouped.setdefault("; ".join(names), set()).update(codes)
                # Mismatched lengths (a combined description repeated across
                # differently-sized groups) can't be paired at all - skipped.
    wb.close()
    return clean, grouped


def _infer_by_elimination(known: dict[str, str], grouped: dict[str, set[str]]) -> int:
    """A group's combined description is N ';'-joined name segments for its N
    codes, but neither list is in matching order (both are independently
    sorted - see saf.py comments). If N-1 of the N codes already have a known
    name, the last code and the last unmatched name segment must pair up -
    that's a logical deduction, not a guess, so it's safe to mine. Runs to a
    fixed point since resolving one group can unlock another. Returns the
    number of newly inferred entries.
    """
    total_new = 0
    while True:
        new_this_pass = 0
        for desc, codes in grouped.items():
            parts = [p.strip() for p in desc.split(";") if p.strip()]
            if len(parts) != len(codes):
                continue  # can't reliably pair up - skip rather than guess
            known_names = {known[c] for c in codes if c in known}
            remaining_codes = [c for c in codes if c not in known]
            remaining_names = [p for p in parts if p not in known_names]
            if len(remaining_codes) == 1 and len(remaining_names) == 1:
                known[remaining_codes[0]] = remaining_names[0]
                new_this_pass += 1
        total_new += new_this_pass
        if new_this_pass == 0:
            break
    return total_new


def bootstrap(store: LocationBankStore | None = None) -> int:
    store = store or LocationBankStore()

    known: dict[str, str] = {}
    grouped: dict[str, set[str]] = {}
    for path in sorted(OPUS_DIR.rglob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        clean, file_grouped = _scan_workbook(path)
        for code, name in clean.items():
            known.setdefault(code, name)
        for desc, codes in file_grouped.items():
            grouped.setdefault(desc, set()).update(codes)

    inferred = _infer_by_elimination(known, grouped)
    if inferred:
        print(f"Inferred {inferred} additional location(s) by elimination from grouped sightings")

    for code, name in known.items():
        store.upsert_location(
            LocationRecord(code=code, primary_name=name, country=_country_from_code(code), source="sample_mined")
        )

    for alias, code in KNOWN_ALIASES.items():
        if code in known:
            store.add_alias(alias, code, source="manual_override")

    return len(known)


if __name__ == "__main__":
    n = bootstrap()
    print(f"Mined and upserted {n} location records from {OPUS_DIR}")
