"""Compare a parsed MRG's OpusRowSet against a reference OPUS-format Excel
file, sheet by sheet. This is the production home of the sheet-reading
and diff logic tests/golden.py has always used for golden-file testing
against the 5 bundled samples - both the tests and the UI's Compare mode
call this module, so there is one implementation instead of two that can
drift apart.

See docs/superpowers/specs/2026-08-23-mrg-opus-comparison-design.md.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Callable

from openpyxl.workbook import Workbook

from mrg2opus.schema import opus_columns as cols


def find_sheet(wb: Workbook, sheet_name: str) -> str:
    """Sheet naming isn't fully standardized across lanes (e.g. SAF uses
    'OPUS RATES PORT - PORT' with spaces, other lanes use 'OPUS RATES
    PORT-PORT') - match loosely on whitespace/hyphen instead of exact name."""
    if sheet_name in wb.sheetnames:
        return sheet_name
    normalized_target = re.sub(r"[\s-]+", "", sheet_name).upper()
    for name in wb.sheetnames:
        if re.sub(r"[\s-]+", "", name).upper() == normalized_target:
            return name
    raise KeyError(f"No sheet matching {sheet_name!r} in {wb.sheetnames}")


def read_rates_sheet(wb: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[find_sheet(wb, sheet_name)]
    rows = []
    for row in ws.iter_rows(min_row=3):
        values = [c.value for c in row[: len(cols.RATES_ROW_FIELDS)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.RATES_ROW_FIELDS, values)))
    return rows


def read_arbs_sheet(wb: Workbook, sheet_name: str = cols.SHEET_NAME_ARBS) -> list[dict[str, Any]]:
    ws = wb[find_sheet(wb, sheet_name)]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[: len(cols.ARBS_ROW_FIELDS)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.ARBS_ROW_FIELDS, values)))
    return rows


def read_cmdt_note_sheet(wb: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[find_sheet(wb, sheet_name)]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[: len(cols.CMDT_NOTE_ROW_FIELDS)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.CMDT_NOTE_ROW_FIELDS, values)))
    return rows


def read_special_note_sheet(wb: Workbook, sheet_name: str = cols.SHEET_NAME_SPECIAL_NOTE) -> list[dict[str, Any]]:
    ws = wb[find_sheet(wb, sheet_name)]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[: len(cols.SPECIAL_NOTE_ROW_FIELDS)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.SPECIAL_NOTE_ROW_FIELDS, values)))
    return rows


def read_route_note_sheet(wb: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[find_sheet(wb, sheet_name)]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[: len(cols.RN_ROW_FIELDS)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.RN_ROW_FIELDS, values)))
    return rows


def _normalize(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        return int(value) if value.isdigit() else value
    return value


def rates_row_key(row: dict[str, Any]) -> tuple:
    # o_via_code disambiguates e.g. "Ganzhou via Shekou" vs "Ganzhou via
    # Yantian" - both resolve to the same origin_code (CNGAN) with
    # different routing and different rates, so it must be part of the key.
    return (
        row.get("origin_code"), row.get("destination_code"), row.get("cgo_type"), row.get("prefix"),
        row.get("o_via_code"), row.get("d_via_code"),
    )


def arbs_row_key(row: dict[str, Any]) -> tuple:
    return (row.get("point"), row.get("over"), row.get("per"))


# Deliberate, user-directed or externally-assigned deviations that make
# certain columns permanently non-matching between a fresh parse and a
# written OPUS file - promoted from each lane's own golden test file
# (tests/test_parsers_*.py), which documents WHY each is a known,
# accepted gap rather than a parsing bug (e.g. `type` defaults to "C" on
# every generated row, but several lanes' own ground truth leaves it
# blank; `header_seq`/`note_seq` are writer-assigned running numbers,
# never derivable from source data). Keyed by lane_id, used by the
# Compare UI so it doesn't drown real discrepancies in known noise.
RATES_IGNORE_FIELDS_BY_LANE = {
    "SAF": frozenset(),
    "EAF": frozenset({"type"}),
    "CSE": frozenset({"type", "commodity_group_description"}),
    "LAEC": frozenset({"cmdt_seq", "route_seq", "type", "commodity_group_description"}),
    "LAWC": frozenset({"cmdt_seq", "route_seq", "commodity_note", "type", "commodity_group_description"}),
}

RATES_PORT_PORT_IGNORE_FIELDS_BY_LANE = {
    **RATES_IGNORE_FIELDS_BY_LANE,
    "EAF": frozenset({"type", "route_seq", "cmdt_seq", "commodity_note"}),
}

CMDT_NOTE_IGNORE_FIELDS_BY_LANE = {
    "SAF": frozenset({"header_seq", "note_seq"}),
    "EAF": frozenset({"header_seq", "note_seq"}),
    "CSE": frozenset({"header_seq", "note_seq", "pol"}),
    "LAEC": frozenset({"header_seq", "note_seq", "pol"}),
    "LAWC": frozenset({"header_seq", "note_seq", "pol"}),
}

SPECIAL_NOTE_IGNORE_FIELDS_BY_LANE = {
    "CSE": frozenset({"header_seq", "note_seq"}),
}


@dataclass
class KeyedDiffResult:
    matched: int
    missing: set[tuple]
    extra: set[tuple]
    field_mismatches: list[tuple[tuple, str, Any, Any]] = field(default_factory=list)


def diff_by_key(
    generated: list[dict[str, Any]],
    expected: list[dict[str, Any]],
    key_fn: Callable[[dict[str, Any]], tuple],
    fields: list[str],
    ignore_fields: set[str] = frozenset(),
) -> KeyedDiffResult:
    """Row-level (keyed) diff, generalized from the matching logic every
    lane's golden test already used via rates_row_key specifically."""
    gen_by_key = {key_fn(r): r for r in generated}
    exp_by_key = {key_fn(r): r for r in expected}

    gen_keys, exp_keys = set(gen_by_key), set(exp_by_key)
    missing = exp_keys - gen_keys
    extra = gen_keys - exp_keys
    matched_keys = gen_keys & exp_keys

    field_mismatches: list[tuple[tuple, str, Any, Any]] = []
    matched = 0
    for key in matched_keys:
        g, e = gen_by_key[key], exp_by_key[key]
        row_ok = True
        for field_name in fields:
            if field_name in ignore_fields:
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            if gv != ev:
                field_mismatches.append((key, field_name, gv, ev))
                row_ok = False
        if row_ok:
            matched += 1

    return KeyedDiffResult(matched=matched, missing=missing, extra=extra, field_mismatches=field_mismatches)


@dataclass
class CmdtBlock:
    key: str
    parent: dict[str, Any]
    children: list[dict[str, Any]] = field(default_factory=list)


def reconstruct_blocks(rows: list[dict[str, Any]]) -> list[CmdtBlock]:
    """A row with non-blank `contents` starts a new block (the block's
    parent/header row - children leave `contents` blank via fill-down,
    the same invariant already used to key each block's identity below);
    every following blank-contents row belongs to it.

    Detecting boundaries via `contents` rather than `header_seq`/
    `note_seq` is required: those sequence numbers are assigned by the
    writer at Excel-export time, not present on a freshly-parsed
    OpusRowSet before it's ever been written - header_seq/note_seq are
    None on EVERY row (including parents) straight out of a parser, so
    detecting via them only works when comparing two already-written
    files, not the Compare feature's actual use case (a fresh parse vs.
    a reference file)."""
    blocks: list[CmdtBlock] = []
    current: CmdtBlock | None = None
    for row in rows:
        contents = str(row.get("contents") or "").strip()
        if contents:
            current = CmdtBlock(key=contents, parent=row, children=[])
            blocks.append(current)
        elif current is not None:
            current.children.append(row)
    return blocks


@dataclass
class BlockDiffResult:
    missing_blocks: list[str] = field(default_factory=list)
    extra_blocks: list[str] = field(default_factory=list)
    field_mismatches: list[tuple[str, int, str, Any, Any]] = field(default_factory=list)


def diff_cmdt_blocks(generated: list[dict[str, Any]], expected: list[dict[str, Any]], fields: list[str], ignore_fields: set[str] = frozenset()) -> BlockDiffResult:
    """CMDT NOTE / SPECIAL NOTE have no reliable per-row key (children
    share their parent's blank header_seq/note_seq) and a reference
    file's row order isn't guaranteed to match the generator's - unlike
    golden tests, which only ever compare against one exact known sample
    and can safely assume matching order. Blocks are matched by
    contents-text key; for matched blocks, child rows are compared
    positionally - the fragile ordering assumption is scoped to one
    block's internal order, not the whole sheet."""
    gen_blocks = {b.key: b for b in reconstruct_blocks(generated)}
    exp_blocks = {b.key: b for b in reconstruct_blocks(expected)}

    missing_blocks = sorted(set(exp_blocks) - set(gen_blocks))
    extra_blocks = sorted(set(gen_blocks) - set(exp_blocks))

    field_mismatches: list[tuple[str, int, str, Any, Any]] = []
    for key in set(gen_blocks) & set(exp_blocks):
        g_block, e_block = gen_blocks[key], exp_blocks[key]
        g_rows = [g_block.parent, *g_block.children]
        e_rows = [e_block.parent, *e_block.children]
        for idx, (g_row, e_row) in enumerate(zip(g_rows, e_rows)):
            for field_name in fields:
                if field_name in ignore_fields:
                    continue
                gv, ev = _normalize(g_row.get(field_name)), _normalize(e_row.get(field_name))
                if gv != ev:
                    field_mismatches.append((key, idx, field_name, gv, ev))
        if len(g_rows) != len(e_rows):
            field_mismatches.append((key, -1, "_row_count", len(g_rows), len(e_rows)))

    return BlockDiffResult(missing_blocks=missing_blocks, extra_blocks=extra_blocks, field_mismatches=field_mismatches)
