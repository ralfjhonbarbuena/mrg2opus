"""Merge multiple uploaded MRG workbooks into one combined workbook before
classification/parsing.

Real-world lanes sometimes arrive as more than one file - e.g. CSE ships
as a main "Tier 1" file plus a separate "...for VELAG and VEPBL" file
covering Venezuela (VELAG/VEPBL are Venezuelan port codes - La Guaira and
Puerto Cabello). The paired sample workbooks already bundle multiple
real-world files' sheets into one file for convenience (see cse.py's "CSE
VE" sheet docstring note, and eaf.py's TZDAR/KEMBA note) - merging
uploads the same way lets the existing single-workbook parser pipeline
handle them completely unchanged, no parser-side code needed.

CSE's real 2-file upload is a special case, confirmed against
reference/1_MRGs/1_CSE FAK, CSE FAK FOR VELAG AND VEPBL/ (2026-08-27):
the "...for VELAG and VEPBL" file's own "CSE" sheet is genuinely
DIFFERENT data (Venezuela) that the parser expects under the literal
name "CSE VE" (see cse.py::COMMODITY_VE) - not an accidental duplicate
upload. That same file also re-bundles 3 lane-wide support sheets ("DG
surcharges", "Yangtze ARB Add-on", "Free Time") that are either
identical to or a formatting variant of the main file's own copies (not
Venezuela-specific), so the main file's versions win and the duplicates
are silently dropped. Both behaviors are gated on the SECOND file's own
name mentioning "VELAG"/"VEPBL" (see _is_cse_venezuela_supplement) -
content alone can't safely tell "this is the known VE-file pattern"
apart from "someone accidentally uploaded a genuine duplicate", so an
unrelated same-named collision still raises DuplicateSheetError exactly
as before.

TAD FILING lanes have a different, NOT filename-gated multi-file case:
the team can issue a rate correction mid-period as a whole second raw
file with the SAME sheet names ("AEW"/"AMW"/etc.) - confirmed against
reference/1_MRGs/23_TAD FILING AEW AMW's own 2 raw files (see
parsers/common/tad_snapshots.py for how the parser side merges them).
Unlike CSE's Venezuela case, any collision on one of TAD's own known raw
sheet names IS the legitimate multi-snapshot pattern - there's no
"accidental duplicate" scenario to guard against the way there is for a
generic name like "CSE", so this doesn't need a filename marker: a
collision on _TAD_SNAPSHOT_SHEETS auto-renames "{name} (2)", "{name} (3)",
etc. instead of raising.
"""
from __future__ import annotations

from copy import copy

import openpyxl
from openpyxl.workbook import Workbook

# See the module docstring's CSE 2-file paragraph. Matched case-
# insensitively against the file's own name, not its contents.
_CSE_VENEZUELA_SUPPLEMENT_MARKERS = ("velag", "vepbl")
_CSE_MAIN_SHEET = "CSE"
_CSE_VENEZUELA_SHEET = "CSE VE"
# Lane-wide support sheets safe to dedupe (keep the first-seen file's
# copy, drop the rest) ONLY when the colliding file is confirmed to be
# the Venezuela supplement above - compared stripped, since real files
# have been seen with inconsistent trailing whitespace in sheet names
# (e.g. "Free Time ").
_CSE_DEDUPE_SAFE_SHEETS = {"DG surcharges", "Yangtze ARB Add-on", "Free Time"}

# See the module docstring's TAD paragraph. Base names only - a renamed
# occurrence ("AEW (2)") is generated here, never uploaded directly.
_TAD_SNAPSHOT_SHEETS = {
    "OEW", "OMW", "WEW", "WMW", "AEW", "AMW",
    "EX.JP AEW", "EX.JP AMW",
    "Origin ARBS", "Origin ARBS - EX JAPAN",
}


class DuplicateSheetError(Exception):
    """Two uploaded files both contain a sheet with the same name - can't
    tell which one should win, so this is surfaced rather than silently
    picking one (or worse, silently overwriting)."""


def _is_cse_venezuela_supplement(name: str | None) -> bool:
    if not name:
        return False
    lowered = name.lower()
    return any(marker in lowered for marker in _CSE_VENEZUELA_SUPPLEMENT_MARKERS)


def merge_workbooks(workbooks: list[Workbook], names: list[str | None] | None = None) -> Workbook:
    """names, when given, must be the same length as workbooks (one
    original filename per upload, in the same order) - used only for the
    CSE Venezuela-supplement special case above. Omit it (or pass None
    entries) to always raise on any collision, exactly as before."""
    if not workbooks:
        raise ValueError("merge_workbooks() needs at least one workbook")
    if len(workbooks) == 1:
        return workbooks[0]
    if names is None:
        names = [None] * len(workbooks)

    merged = openpyxl.Workbook()
    merged.remove(merged.active)
    seen: set[str] = set()
    tad_occurrence_count: dict[str, int] = {}
    for wb, name in zip(workbooks, names):
        is_venezuela_supplement = _is_cse_venezuela_supplement(name)
        for sheet_name in wb.sheetnames:
            target_name = sheet_name
            if sheet_name in seen:
                if is_venezuela_supplement and sheet_name == _CSE_MAIN_SHEET:
                    target_name = _CSE_VENEZUELA_SHEET
                elif is_venezuela_supplement and sheet_name.strip() in _CSE_DEDUPE_SAFE_SHEETS:
                    continue  # keep the main file's copy, drop this duplicate
                elif sheet_name in _TAD_SNAPSHOT_SHEETS:
                    tad_occurrence_count[sheet_name] = tad_occurrence_count.get(sheet_name, 1) + 1
                    target_name = f"{sheet_name} ({tad_occurrence_count[sheet_name]})"
                else:
                    raise DuplicateSheetError(
                        f"Sheet {sheet_name!r} appears in more than one uploaded file - "
                        "can't tell which one should win. Remove the duplicate and try again."
                    )
            seen.add(target_name)
            _copy_sheet(wb[sheet_name], merged, target_name)
    return merged


def _copy_sheet(src_ws, dst_wb: Workbook, sheet_name: str) -> None:
    dst_ws = dst_wb.create_sheet(sheet_name)
    # Touch every cell in the source's used range (not just non-blank
    # ones) so the destination sheet's max_row/max_column match the
    # source exactly - some parsers (e.g. lawc.py's single-column sheet
    # parser) iterate up to ws.max_row/max_column directly, and a
    # shrunken destination range would silently truncate them.
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
