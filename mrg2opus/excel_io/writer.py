"""Write an OpusRowSet out to an OPUS-format Excel workbook, reproducing the
exact 2-row header on RATES sheets and the fill-down parent/child semantics
on CMDT NOTE / SPECIAL NOTE sheets.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.schema import opus_columns as cols
from mrg2opus.schema.opus_rows import (
    ArbsRow,
    CmdtNoteRow,
    FreetimeRow,
    OpusRowSet,
    RatesRow,
    RouteNoteRow,
    SpecialNoteRow,
    VerticalRatesRow,
)


_HEADER_FILL = PatternFill("solid", start_color=cols.HEADER_FILL_RGB, end_color=cols.HEADER_FILL_RGB)
_HEADER_FONT = Font(name=cols.HEADER_FONT_NAME, size=cols.HEADER_FONT_SIZE, color=cols.HEADER_FONT_RGB)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin")
_HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(
    ws: Worksheet,
    header_rows: int,
    widths: dict[str, float] | None = None,
    row_height: float = cols.HEADER_ROW_HEIGHT,
) -> None:
    """Paint the header the way OPUS does - navy ground, white centred
    Arial, wrapped, thin-boxed.

    Not cosmetic: without it a MERGED header renders its label
    left-aligned at the top-left of the span against a white ground, which
    reads as a broken merge even though the ranges are right (that is what
    "the merging is incorrect" turned out to mean). Every header cell on
    every sheet of the reference carries this one style.

    Applied across the merged cells too, not just their anchors: openpyxl
    keeps the VALUE only in the top-left cell, but each covered cell still
    draws its own fill and borders, so skipping them would leave white
    gaps and missing box lines inside a span.
    """
    for row in range(1, header_rows + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _HEADER_BORDER
        ws.row_dimensions[row].height = row_height
    for column_letter, width in (widths or {}).items():
        ws.column_dimensions[column_letter].width = width


def _write_two_row_header(
    ws: Worksheet,
    group: list[str],
    field: list[str | None],
    merges: list[str],
    widths: dict[str, float] | None = None,
) -> None:
    """Write a 2-row header with OPUS's own merged cells.

    OPUS merges each group label across its span instead of repeating it,
    and the user asked for that exact format (their OPUS HEADERS.xlsx).
    The ranges come from that file - see opus_columns' *_HEADER_MERGES for
    why they are copied rather than derived from the two lists here.
    Writing the repeated label and then merging is what lines the two up:
    openpyxl keeps only the top-left cell of a merged range, so the
    duplicates collapse into the single label the reference shows.
    """
    ws.append(group)
    ws.append([f or "" for f in field])
    for merge_range in merges:
        ws.merge_cells(merge_range)
    _style_header(ws, header_rows=2, widths=widths)


def _write_rates_sheet(wb: Workbook, sheet_name: str, rows: list[RatesRow]) -> None:
    ws: Worksheet = wb.create_sheet(sheet_name)
    _write_two_row_header(ws, cols.RATES_HEADER_GROUP, cols.RATES_HEADER_FIELD, cols.RATES_HEADER_MERGES, cols.RATES_COLUMN_WIDTHS)
    for row in rows:
        data = row.model_dump()
        ws.append([data[field_name] for field_name in cols.RATES_ROW_FIELDS])


def _write_vertical_rates_sheet(wb: Workbook, sheet_name: str, rows: list[VerticalRatesRow]) -> None:
    ws: Worksheet = wb.create_sheet(sheet_name)
    _write_two_row_header(ws, cols.VERTICAL_RATES_HEADER_GROUP, cols.VERTICAL_RATES_HEADER_FIELD, cols.VERTICAL_RATES_HEADER_MERGES, cols.VERTICAL_RATES_COLUMN_WIDTHS)
    for row in rows:
        data = row.model_dump()
        ws.append([data[field_name] for field_name in cols.VERTICAL_RATES_ROW_FIELDS])


def _write_arbs_sheet(wb: Workbook, sheet_name: str, rows: list[ArbsRow]) -> None:
    ws: Worksheet = wb.create_sheet(sheet_name)
    ws.append(cols.ARBS_HEADER)
    _style_header(ws, header_rows=1, widths=cols.ARBS_COLUMN_WIDTHS, row_height=cols.ARBS_HEADER_ROW_HEIGHT)
    for row in rows:
        data = row.model_dump()
        ws.append([data[field_name] for field_name in cols.ARBS_ROW_FIELDS])


def _write_cmdt_note_sheet(wb: Workbook, sheet_name: str, rows: list[CmdtNoteRow]) -> None:
    ws: Worksheet = wb.create_sheet(sheet_name)
    ws.append(cols.CMDT_NOTE_HEADER)
    _style_header(ws, header_rows=1, widths=cols.CMDT_NOTE_COLUMN_WIDTHS)
    for row in rows:
        data = row.model_dump()
        ws.append([data[field_name] for field_name in cols.CMDT_NOTE_ROW_FIELDS])


def _write_special_note_sheet(wb: Workbook, sheet_name: str, rows: list[SpecialNoteRow]) -> None:
    # Same field set as CMDT NOTE, different column order - see
    # schema/opus_columns.py's SPECIAL_NOTE_HEADER comment.
    ws: Worksheet = wb.create_sheet(sheet_name)
    ws.append(cols.SPECIAL_NOTE_HEADER)
    _style_header(ws, header_rows=1, widths=cols.SPECIAL_NOTE_COLUMN_WIDTHS)
    for row in rows:
        data = row.model_dump()
        ws.append([data[field_name] for field_name in cols.SPECIAL_NOTE_ROW_FIELDS])


def _write_route_note_sheet(wb: Workbook, sheet_name: str, rows: list[RouteNoteRow]) -> None:
    ws: Worksheet = wb.create_sheet(sheet_name)
    ws.append(cols.RN_HEADER)
    _style_header(ws, header_rows=1, widths=cols.RN_COLUMN_WIDTHS)
    for row in rows:
        data = row.model_dump()
        ws.append([data[field_name] for field_name in cols.RN_ROW_FIELDS])


def _write_freetime_sheet(wb: Workbook, sheet_name: str, rows: list[FreetimeRow]) -> None:
    ws: Worksheet = wb.create_sheet(sheet_name)
    _write_two_row_header(ws, cols.FREETIME_HEADER_GROUP, cols.FREETIME_HEADER_FIELD, cols.FREETIME_HEADER_MERGES, cols.FREETIME_COLUMN_WIDTHS)
    for row in rows:
        data = row.model_dump()
        # Blank, not absent: the header keeps its full width and each
        # remaining value stays under its own column - see
        # opus_columns.FREETIME_UNFILED_FIELDS.
        ws.append([
            None if field_name in cols.FREETIME_UNFILED_FIELDS else data[field_name]
            for field_name in cols.FREETIME_ROW_FIELDS
        ])


def _sheet_names_for_suffix(suffix: str) -> dict[str, str]:
    """Real OPUS filing sheet names (see project-opus-note-sheet-taxonomy
    memory) - NOT the schema/opus_columns.py SHEET_NAME_* constants, which
    match the older bundled sample fixtures' own naming instead. Naming
    convention for sub-lanes confirmed against EAF.xlsx's sub-lane sheets:
    the suffix is appended directly to the base sheet name with a hyphen
    (e.g. 'RATES-TZDAR', 'RATES-TZDAR PORT-PORT', 'CMDT NOTE-TZDAR'), or the
    plain base name when there's no sub-lane (suffix "") - unconfirmed
    against real ground truth for a real multi-sub-lane filing, since no
    EAF reference file exists yet; kept as the least-surprise default."""
    tag = f"-{suffix}" if suffix else ""
    return {
        "rates": f"RATES{tag}",
        "rates_port_port": f"RATES{tag} PORT-PORT",
        "arbs": f"ORIGIN ARBS{tag}",
        "cmdt_notes": f"CMDT NOTE{tag}",
        "special_notes": f"SPECIAL NOTE{tag}",
        "route_notes": f"RN{tag}",
        "vertical_rates": f"VERTICAL RATES{tag}",
        "freetime": f"FREETIME{tag}",
    }


def resolve_sheet_names(
    suffix: str,
    sheet_name_overrides: dict[str, str] | None = None,
    scoped_sheet_name_overrides: dict[str, dict[str, str]] | None = None,
) -> dict[str, str]:
    """{OpusRowSet field name: the sheet name it will actually be written
    under} for one scope, applying both override layers.

    Public because the UI needs the SAME answer the writer will act on -
    the wizard used to keep its own hardcoded list of sheet labels, which
    drifted: it named only 5 of the 8 sheet types (so ROUTE NOTE, VERTICAL
    RATES and FREETIME were invisible and unskippable) and prefixed them
    "OPUS RATES"/"OPUS ARBS", names that appear in no output workbook.
    """
    names = _sheet_names_for_suffix(suffix)
    if sheet_name_overrides:
        tag = f"-{suffix}" if suffix else ""
        names.update({key: f"{base}{tag}" for key, base in sheet_name_overrides.items()})
    if scoped_sheet_name_overrides and suffix in scoped_sheet_name_overrides:
        names.update(scoped_sheet_name_overrides[suffix])
    return names


def _write_row_set(wb: Workbook, row_set: OpusRowSet, names: dict[str, str]) -> None:
    if row_set.rates:
        _write_rates_sheet(wb, names["rates"], row_set.rates)
    if row_set.rates_port_port:
        _write_rates_sheet(wb, names["rates_port_port"], row_set.rates_port_port)
    if row_set.arbs:
        _write_arbs_sheet(wb, names["arbs"], row_set.arbs)
    if row_set.cmdt_notes:
        _write_cmdt_note_sheet(wb, names["cmdt_notes"], row_set.cmdt_notes)
    if row_set.special_notes:
        _write_special_note_sheet(wb, names["special_notes"], row_set.special_notes)
    if row_set.route_notes:
        _write_route_note_sheet(wb, names["route_notes"], row_set.route_notes)
    if row_set.vertical_rates:
        # One sheet, however many rows - matching real ground truth, where
        # every commodity group shares a single sheet and the CMDT Seq
        # block marks each boundary. A sheet past the OPUS row limit is
        # still written; see pipeline.vertical_rates_over_cap, which
        # reports it so the filer can trim it in Excel.
        _write_vertical_rates_sheet(wb, names["vertical_rates"], row_set.vertical_rates)
    if row_set.freetime:
        _write_freetime_sheet(wb, names["freetime"], row_set.freetime)


def write_opus_workbook(
    row_set: OpusRowSet,
    out_path: Path | str,
    sheet_names: dict[str, str] | None = None,
) -> None:
    """sheet_names lets a lane override default sheet names without changing
    writer logic. For lanes with sub-lanes sharing one workbook (e.g. EAF's
    TZDAR/KEMBA), use write_opus_workbook_multi instead."""
    names = _sheet_names_for_suffix("")
    if sheet_names:
        names.update(sheet_names)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet
    _write_row_set(wb, row_set, names)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_opus_workbook_multi(
    row_sets: dict[str, OpusRowSet],
    out_path: Path | str,
    sheet_name_overrides: dict[str, str] | None = None,
    scoped_sheet_name_overrides: dict[str, dict[str, str]] | None = None,
) -> None:
    """Write several sub-lane OpusRowSets into ONE workbook, each under its
    own suffixed sheet names (e.g. {"TZDAR": ..., "KEMBA": ...} ->
    'OPUS RATES-TZDAR', 'OPUS RATES-KEMBA', ...). Pass {"": row_set} for a
    single-output lane to reuse this same path.

    sheet_name_overrides lets a lane replace one or more base sheet names
    (keyed by OpusRowSet field name, e.g. {"route_notes": "ROUTE NOTE"})
    when its own real filing convention differs from the default - see
    BaseMRGParser.SHEET_NAME_OVERRIDES. scoped_sheet_name_overrides is for
    the rarer case where even that uniform tagging is wrong for one scope -
    {scope: {field: full_name}}, applied verbatim (no tag appended) for
    whichever fields that scope lists, layered OVER sheet_name_overrides -
    see BaseMRGParser.SCOPED_SHEET_NAME_OVERRIDES."""
    wb = Workbook()
    wb.remove(wb.active)
    for suffix, row_set in row_sets.items():
        names = resolve_sheet_names(suffix, sheet_name_overrides, scoped_sheet_name_overrides)
        _write_row_set(wb, row_set, names)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
