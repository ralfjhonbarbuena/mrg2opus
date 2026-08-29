"""LAEC lane parser.

Structurally similar to CSE (direct origin/destination codes, FEBP/WPRD
group-code shorthand, D2/D4/D5 grids + a single-column "R5 NOR" reefer
sheet + an in-gauge OT/FR sheet + a Yangtze inland-origin ARBS sheet), but
with one big difference: everything is split into two parallel sections,
"Non-ISC" (Far East origins, commodity groups G0015/G0010) and "ISC"
(Indian Subcontinent origins, G0016/G0017), each with its own POD header
block on the SAME raw sheet (DRY, Ingauge FAK) and its own CMDT NOTE
charge-code set. R5 NOR (reefer) only has a Non-ISC section - ISC origins
get no reefer rate at all (verified: 0 Prefix "R" rows under G0016 in
ground truth).

Also new here: the in-gauge sheet's destination header cells are already
"/"-joined multi-code groups (e.g. "BRSSZ/BRNVT/.../UYMVD" for "ECSA BASE
PORTS"), not a single code - handled the same way as an exploded/grouped
origin (see explode_rates_row), just building destination_code/description
from multiple resolved codes instead of one.

"ECSA Add-On" sheet: a small table of extra destinations (Ushuaia, Zarate,
La Plata - "No service" rows like Fortaleza/Vitoria/Rosario are skipped)
priced as an existing destination's rate ("T/S Port") plus a fixed add-on
per container size. Verified against ground truth to apply to both G0015
and G0016 (Non-ISC and ISC main grids) but NOT to duplicate into a DG
variant (confirmed: only Prefix D/DR rows exist for these 3 destinations,
no D/DG). Not verified for the in-gauge groups (G0010/G0017) - not applied
there.

OPUS ARBS: the ground truth sample is MISSING this sheet even though the
raw "Yangtze ARB Add-on" sheet is present and structurally identical to
CSE's (confirmed by the user - an omission in how the sample was built,
not evidence this lane doesn't file ARBS). Generated here by reusing the
same, CSE-verified logic (see parsers/common/yangtze_arbs.py), but this
specific lane's ARBS output has no ground truth to check it against.

NOT implemented: OPUS SPECIAL NOTE. The "IMO charge" sheet has the same
PSA transshipment-via-Singapore surcharge table as CSE's "DG surcharges"
sheet, but LAEC's ground truth sample has no OPUS SPECIAL NOTE sheet at
all (unlike ARBS, the user hasn't flagged this as an omission), so it's
left out rather than assumed.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.store import LocationBankStore
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.cmdt_notes import build_cmdt_notes
from mrg2opus.parsers.common.commodity import (
    CommodityNoteSpec,
    build_notes_by_description,
    resolve_commodity_code,
    resolve_commodity_description,
)
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded, location_is_excluded
from mrg2opus.parsers.common.freetime import build_laec_freetime
from mrg2opus.parsers.common.group_codes import load_group_codes
from mrg2opus.parsers.common.header_grid import flatten_pod_header
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.common.yangtze_arbs import build_arbs, parse_yangtze_sheet
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import ArbsRow, OpusRowSet, RatesRow, explode_rates_row

RAW_SHEET_DRY = "DRY"
RAW_SHEET_NOR = "R5 NOR"
RAW_SHEET_INGAUGE = "Ingauge FAK"
RAW_SHEET_YANGTZE = "Yangtze ARB Add-on"
RAW_SHEET_ECSA = "ECSA Add-On"

VALIDITY_ROW, VALIDITY_FROM_COL, VALIDITY_TO_COL = 4, 3, 4

# Commodity groups, verified against LAEC.xlsx's own OPUS RATES sheet -
# just the starting defaults, user-overridable (see cse.py note).
#
# G0015 used to carry ONE description spanning 2 different raw sheets (the
# Non-ISC section of "DRY" + "R5 NOR"), hence the old "...NOR_NON-ISC"
# suffix. R5 NOR now defaults to its OWN description (its own raw sheet
# name, see NOR_DEFAULT_DESCRIPTION below) - so COMMODITY_NON_ISC_MAIN's
# description below no longer references NOR. Every override dict (code,
# description, cmdt_seq) is keyed by a group's DEFAULT description, not
# its code - see parsers/common/commodity.py's module docstring for why.
# They only end up sharing one CMDT NOTE block again if the user overrides
# them to the exact same description - see
# parsers/common/commodity.py::build_notes_by_description(), used in
# to_opus_rows() below.
COMMODITY_NON_ISC_MAIN = ("G0015", "FAK & DG_NON-ISC", None)
COMMODITY_ISC_MAIN = ("G0016", "FAK_ISC", None)
COMMODITY_NON_ISC_INGAUGE = ("G0010", "INGAUGE FAK_NON-ISC", None)
COMMODITY_ISC_INGAUGE = ("G0017", "INGAUGE FAK_ISC", None)

NOR_DEFAULT_DESCRIPTION = RAW_SHEET_NOR.strip()
# cmdt_seq for this lane (403-406 in the ground truth) doesn't follow any
# derivable pattern (unlike CSE where it happened to equal the commodity
# code's number) - it's an externally-assigned value like header_seq, so
# it's left unset (None) here rather than guessed.

# Verified per-section charge-code order for CMDT NOTE - see cse.py's
# CMDT_NOTE_CHARGE_CODES comment for why this is hardcoded rather than
# alphabetized or parsed from the raw "incl." text: the order genuinely
# isn't alphabetical in the ground truth and isn't consistent between the
# two sections either.
NON_ISC_CHARGE_CODES = ["HEA", "OBS", "EFS", "MBS", "PSS"]
ISC_CHARGE_CODES = ["OBS", "HEA", "THL", "THL", "CSS", "SLF", "EFS", "MBS", "PSS"]

# LAEC's ground truth says "HEAVY SURCHARGE(HEA)" - EAF's says "HEAVY
# WEIGHT SURCHARGE(HEA)" for the same code. Confirmed the shared
# CHARGE_CODE_NAMES can't hold both, so this lane supplies its own.
CHARGE_CODE_NAMES_OVERRIDE = {"HEA": "HEAVY SURCHARGE"}

# ECSA Add-On sheet: extra destinations priced as an existing destination's
# rate + a fixed add-on. "No service" origins are simply absent from this
# table. Verified against ground truth for all 3 (KRPUS origin spot-check).
ECSA_ADD_ON_ROW_MIN, ECSA_ADD_ON_ROW_MAX = 4, 9


@dataclass(frozen=True)
class DrySectionConfig:
    pod_code_row: int
    container_label_row: int
    data_min_row: int
    data_max_row: int
    min_col: int
    max_col: int
    commodity: tuple[str, str, int | None]


DRY_SECTIONS = [
    DrySectionConfig(9, 10, 11, 66, 3, 38, COMMODITY_NON_ISC_MAIN),
    DrySectionConfig(72, 73, 74, 87, 3, 38, COMMODITY_ISC_MAIN),
]

NOR_SECTION = DrySectionConfig(9, 10, 11, 27, 3, 10, COMMODITY_NON_ISC_MAIN)

INGAUGE_SECTIONS = [
    DrySectionConfig(10, 11, 12, 71, 3, 6, COMMODITY_NON_ISC_INGAUGE),
    DrySectionConfig(77, 78, 79, 91, 3, 6, COMMODITY_ISC_INGAUGE),
]


@dataclass
class GridRow:
    origin_code_raw: str
    dest_code_raw: str  # may be "/"-joined (in-gauge destination groups)
    sizes: dict[str, float]


@dataclass
class NorRow:
    origin_code_raw: str
    dest_code: str
    value: float


@dataclass
class EcsaAddOn:
    dest_name: str
    ts_port_code: str
    add_on_20: Decimal
    add_on_40: Decimal


@dataclass
class LAECRawData:
    validity_start: date | None
    validity_end: date | None
    grid_rows: dict[tuple[str, str, int | None], list[GridRow]]
    nor_rows: list[NorRow]
    ingauge_rows: dict[tuple[str, str, int | None], list[GridRow]]
    ecsa_add_ons: list[EcsaAddOn]
    yangtze_rows: list
    yangtze_eff_date: date | None
    yangtze_exp_date: date | None
    freetime_variant: str


class LAECParser(BaseMRGParser):
    lane_id: ClassVar[str] = "LAEC"

    def __init__(self, container_map: ContainerMap | None = None, location_store: LocationBankStore | None = None):
        self.container_map = container_map or load_container_map("laec")
        self.location_store = location_store or LocationBankStore()
        self.group_codes = load_group_codes("laec")

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        if RAW_SHEET_DRY not in wb.sheetnames:
            return 0.0
        ws = wb[RAW_SHEET_DRY]
        title = str(ws.cell(row=1, column=1).value or "")
        if "LUX" in title.upper():
            # LAEC LUX variant - a structurally different raw shape (single
            # flat POD grid, no "POL: NON-ISC" split) - see LAECLuxParser.
            return 0.0
        score = 0.5
        if "LAEC" in title.upper():
            score += 0.3
        if any(str(ws.cell(row=r, column=1).value or "").strip().upper() == "POL: NON-ISC" for r in range(1, 10)):
            score += 0.2
        return min(score, 1.0)

    def _resolve_codes(self, raw_code: str) -> list[str]:
        raw_code = raw_code.strip()
        if raw_code in self.group_codes:
            return self.group_codes[raw_code]
        return [raw_code]

    def _split_dest_codes(self, raw_text: str) -> list[str]:
        return [c.strip() for c in raw_text.replace("\n", "").split("/") if c.strip()]

    def _lookup_description(self, code: str) -> str | None:
        rec = self.location_store.get_by_code(code)
        return rec.primary_name if rec else None

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        validity_start, validity_end = _parse_validity(wb[RAW_SHEET_DRY])

        grid_rows: dict[tuple[str, str, int | None], list[GridRow]] = {}
        if RAW_SHEET_DRY in wb.sheetnames:
            ws = wb[RAW_SHEET_DRY]
            for cfg in DRY_SECTIONS:
                rows = self._parse_grid_section(ws, cfg)
                grid_rows.setdefault(cfg.commodity, []).extend(rows)

        nor_rows: list[NorRow] = []
        if RAW_SHEET_NOR in wb.sheetnames:
            nor_rows = self._parse_nor_section(wb[RAW_SHEET_NOR], NOR_SECTION)

        ingauge_rows: dict[tuple[str, str, int | None], list[GridRow]] = {}
        if RAW_SHEET_INGAUGE in wb.sheetnames:
            ws = wb[RAW_SHEET_INGAUGE]
            for cfg in INGAUGE_SECTIONS:
                rows = self._parse_grid_section(ws, cfg)
                ingauge_rows.setdefault(cfg.commodity, []).extend(rows)

        ecsa_add_ons: list[EcsaAddOn] = []
        if RAW_SHEET_ECSA in wb.sheetnames:
            ecsa_add_ons = self._parse_ecsa_sheet(wb[RAW_SHEET_ECSA])

        yangtze_rows, yangtze_eff, yangtze_exp = [], None, None
        if RAW_SHEET_YANGTZE in wb.sheetnames:
            yangtze_rows, yangtze_eff, yangtze_exp = parse_yangtze_sheet(
                wb[RAW_SHEET_YANGTZE], self._lookup_description
            )

        # FAK and TIER 1's raw workbooks are structurally and textually
        # identical (confirmed: no "Tier" mention anywhere in either, same
        # sheet set) - there is no reliable signal in the raw MRG itself to
        # pick TIER 1's own FREETIME variant (1 extra Argentina/Zarate row
        # vs FAK's), so this always defaults to "fak" here. A real TIER 1
        # filing is missing that one row as a result - a small, honestly-
        # scoped gap in the same category as every other "not derivable
        # from this file alone" gap already accepted in this project,
        # rather than guessing from an unreliable heuristic (an earlier
        # attempt keyed off a "Free Time " sheet present in one TIER 1
        # sample but absent from another).
        freetime_variant = "fak"

        return RawExtraction(
            tables={
                "laec": LAECRawData(
                    validity_start=validity_start,
                    validity_end=validity_end,
                    grid_rows=grid_rows,
                    nor_rows=nor_rows,
                    ingauge_rows=ingauge_rows,
                    ecsa_add_ons=ecsa_add_ons,
                    yangtze_rows=yangtze_rows,
                    yangtze_eff_date=yangtze_eff,
                    yangtze_exp_date=yangtze_exp,
                    freetime_variant=freetime_variant,
                )
            }
        )

    def _parse_grid_section(self, ws: Worksheet, cfg: DrySectionConfig) -> list[GridRow]:
        header_cols = flatten_pod_header(ws, cfg.pod_code_row, cfg.container_label_row, cfg.min_col, cfg.max_col)
        out: list[GridRow] = []
        for row_idx in range(cfg.data_min_row, cfg.data_max_row + 1):
            origin_cell = ws.cell(row=row_idx, column=2)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            by_dest: dict[str, dict[str, float]] = {}
            for hc in header_cols:
                suffix = self.container_map.suffix_for(hc.container_label) if "'" not in hc.container_label else (
                    "20" if "20" in hc.container_label else "40"
                )
                if suffix is None:
                    continue
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                by_dest.setdefault(hc.pod_label.strip(), {})[suffix] = cell.value
            for dest_raw, sizes in by_dest.items():
                out.append(GridRow(origin_code_raw=origin_raw, dest_code_raw=dest_raw, sizes=sizes))
        return out

    def _parse_nor_section(self, ws: Worksheet, cfg: DrySectionConfig) -> list[NorRow]:
        dest_codes = {
            col: str(ws.cell(row=cfg.pod_code_row, column=col).value or "").strip()
            for col in range(cfg.min_col, cfg.max_col + 1)
            if not is_excluded(ws.cell(row=cfg.pod_code_row, column=col))
        }
        out: list[NorRow] = []
        for row_idx in range(cfg.data_min_row, cfg.data_max_row + 1):
            origin_cell = ws.cell(row=row_idx, column=2)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            for col, dest_code in dest_codes.items():
                if not dest_code:
                    continue
                cell = ws.cell(row=row_idx, column=col)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                out.append(NorRow(origin_code_raw=origin_raw, dest_code=dest_code, value=cell.value))
        return out

    def _parse_ecsa_sheet(self, ws: Worksheet) -> list[EcsaAddOn]:
        out: list[EcsaAddOn] = []
        for row_idx in range(ECSA_ADD_ON_ROW_MIN, ECSA_ADD_ON_ROW_MAX + 1):
            dest_name = ws.cell(row=row_idx, column=1).value
            ts_port = ws.cell(row=row_idx, column=3).value
            add_20 = ws.cell(row=row_idx, column=4).value
            add_40 = ws.cell(row=row_idx, column=5).value
            if not dest_name or not ts_port or not isinstance(add_20, (int, float)):
                continue  # "No service" or a "-" placeholder T/S port
            out.append(
                EcsaAddOn(
                    dest_name=str(dest_name).strip(),
                    ts_port_code=str(ts_port).strip(),
                    add_on_20=Decimal(str(add_20)),
                    add_on_40=Decimal(str(add_40)) if isinstance(add_40, (int, float)) else Decimal(str(add_20)),
                )
            )
        return out

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: LAECRawData = raw.tables["laec"]

        rates: list[RatesRow] = []
        rates_port_port: list[RatesRow] = []
        # (commodity code) -> {(origin_code, destination_code): RatesRow}
        # for the base D/DR rows, used by the ECSA Add-On step below to
        # find each origin's T/S-port rate to build on top of.
        base_rows_by_commodity: dict[str, dict[tuple[str, str], RatesRow]] = {}
        # Descriptions now default per-sheet (see COMMODITY_NON_ISC_MAIN's
        # comment), so notes can't be built per-group as rows are built -
        # two groups might share a description (default, override, or the
        # user merging them back together) and need to share exactly one
        # CMDT NOTE block. Collect a spec per group that actually produced
        # rows, resolve all of them into notes ONCE at the end, then stamp
        # commodity_note onto every row by its own (already-final)
        # commodity_group_description.
        note_specs: list[CommodityNoteSpec] = []

        for commodity, grid_rows in data.grid_rows.items():
            code, default_description, _ = commodity
            description = resolve_commodity_description(default_description, config)
            cmdt_seq = config.commodity_sequence_overrides.get(default_description)
            output_code = resolve_commodity_code(default_description, code, config)
            charge_codes = ISC_CHARGE_CODES if code == COMMODITY_ISC_MAIN[0] else NON_ISC_CHARGE_CODES

            dr_rows, dr_pp, dg_rows, dg_pp = [], [], [], []
            base_by_key: dict[tuple[str, str], RatesRow] = {}

            for gr in grid_rows:
                row = self._build_rates_row(gr, output_code, description, cmdt_seq)
                if row is None:
                    continue
                dr_pp.extend(explode_rates_row(row))
                base_by_key[(row.origin_code, row.destination_code)] = row
                dr_rows.append(row)

                if self.container_map.cgo_type == "DR" and not config.skip_dg_generation.get(default_description, False):
                    dg_row = row.model_copy(update={"cgo_type": "DG"})
                    dg_pp.extend(explode_rates_row(dg_row))
                    dg_rows.append(dg_row)

            base_rows_by_commodity[code] = base_by_key
            rates.extend(group_by_destination(dr_rows))
            rates.extend(group_by_destination(dg_rows))
            rates_port_port.extend(group_by_destination(dr_pp))
            rates_port_port.extend(group_by_destination(dg_pp))

            if dr_rows:
                note_specs.append(CommodityNoteSpec(description, data.validity_start, data.validity_end, charge_codes))

        # R5 NOR: Prefix "R", rate in the 40HC slot, Non-ISC main group only.
        # Defaults to its OWN description now (see COMMODITY_NON_ISC_MAIN's
        # comment above) - reuses Non-ISC main's charge codes/validity since
        # there's no independent ground truth to derive its own from.
        reefer_rows, reefer_pp = [], []
        nor_main_code = COMMODITY_NON_ISC_MAIN[0]
        nor_description = resolve_commodity_description(NOR_DEFAULT_DESCRIPTION, config)
        nor_cmdt_seq = config.commodity_sequence_overrides.get(NOR_DEFAULT_DESCRIPTION)
        nor_output_code = resolve_commodity_code(NOR_DEFAULT_DESCRIPTION, nor_main_code, config)
        for nr in data.nor_rows:
            origin_codes = self._resolve_codes(nr.origin_code_raw)
            origin_names = [self._lookup_description(c) for c in origin_codes]
            if any(n is None for n in origin_names):
                continue
            dest_name = self._lookup_description(nr.dest_code)
            if dest_name is None:
                continue
            row = RatesRow(
                type="C",
                cmdt_seq=nor_cmdt_seq,
                commodity_group_code=nor_output_code,
                commodity_group_description=nor_description,
                origin_code=";".join(sorted(set(origin_codes))),
                origin_description=";".join(sorted(set(origin_names))),
                origin_term="CY",
                destination_code=nr.dest_code,
                destination_description=dest_name,
                destination_term="CY",
                prefix="R",
                cgo_type="DR",
                cur_40hc="USD",
                rate_40hc=_to_decimal(nr.value),
            )
            reefer_pp.extend(explode_rates_row(row))
            reefer_rows.append(row)
        rates.extend(group_by_destination(reefer_rows))
        rates_port_port.extend(group_by_destination(reefer_pp))
        if reefer_rows:
            note_specs.append(
                CommodityNoteSpec(nor_description, data.validity_start, data.validity_end, NON_ISC_CHARGE_CODES)
            )

        # In-gauge: Prefix O + F duplicate pair, destination itself may be
        # a "/"-joined group (e.g. "ECSA BASE PORTS").
        for commodity, grid_rows in data.ingauge_rows.items():
            code, default_description, _ = commodity
            description = resolve_commodity_description(default_description, config)
            cmdt_seq = config.commodity_sequence_overrides.get(default_description)
            output_code = resolve_commodity_code(default_description, code, config)
            charge_codes = ISC_CHARGE_CODES if code == COMMODITY_ISC_INGAUGE[0] else NON_ISC_CHARGE_CODES

            o_rows, o_pp, f_rows, f_pp = [], [], [], []
            for gr in grid_rows:
                base = self._build_rates_row(gr, output_code, description, cmdt_seq, prefix="O")
                if base is None:
                    continue
                o_pp.extend(explode_rates_row(base))
                o_rows.append(base)
                f_row = base.model_copy(update={"prefix": "F"})
                f_pp.extend(explode_rates_row(f_row))
                f_rows.append(f_row)
            rates.extend(group_by_destination(o_rows))
            rates.extend(group_by_destination(f_rows))
            rates_port_port.extend(group_by_destination(o_pp))
            rates_port_port.extend(group_by_destination(f_pp))

            if o_rows:
                note_specs.append(CommodityNoteSpec(description, data.validity_start, data.validity_end, charge_codes))

        # ECSA Add-On: extra destinations built from an existing
        # destination's ("T/S Port") rate plus a fixed add-on - verified
        # for G0015/G0016 only, no DG duplicate.
        for code, base_by_key in base_rows_by_commodity.items():
            extra_rows, extra_pp = [], []
            origin_codes_seen = {origin for origin, _dest in base_by_key}
            for origin_code in origin_codes_seen:
                for addon in data.ecsa_add_ons:
                    ts_row = base_by_key.get((origin_code, addon.ts_port_code))
                    if ts_row is None:
                        continue
                    dest_code_resolved, dest_desc_resolved = self._resolve_add_on_destination(addon.dest_name)
                    if dest_code_resolved is None:
                        continue
                    new_row = ts_row.model_copy(
                        update={
                            "destination_code": dest_code_resolved,
                            "destination_description": dest_desc_resolved,
                            "rate_20": (ts_row.rate_20 + addon.add_on_20) if ts_row.rate_20 is not None else None,
                            "rate_40": (ts_row.rate_40 + addon.add_on_40) if ts_row.rate_40 is not None else None,
                            "rate_40hc": (
                                (ts_row.rate_40hc + addon.add_on_40) if ts_row.rate_40hc is not None else None
                            ),
                        }
                    )
                    extra_rows.append(new_row)
                    extra_pp.extend(explode_rates_row(new_row))
            rates.extend(group_by_destination(extra_rows))
            rates_port_port.extend(group_by_destination(extra_pp))

        arbs = build_arbs(data.yangtze_rows, data.yangtze_eff_date, data.yangtze_exp_date)

        cmdt_notes, note_text_by_description = build_notes_by_description(
            note_specs,
            sequential_charge_seq=True,
            charge_code_names_override=CHARGE_CODE_NAMES_OVERRIDE,
            excluded_codes=frozenset(config.excluded_charge_codes),
            rfa_effective=config.rfa_effective_date,
            rfa_expiry=config.rfa_expiry_date,
        )
        for row in rates:
            row.commodity_note = note_text_by_description.get(row.commodity_group_description)

        freetime = build_laec_freetime(data.freetime_variant, data.validity_start, data.validity_end)

        return OpusRowSet(
            rates=rates, rates_port_port=rates_port_port, cmdt_notes=cmdt_notes, arbs=arbs, freetime=freetime
        )

    def _build_rates_row(
        self, gr: GridRow, code: str, description: str, cmdt_seq: int | None, prefix: str | None = None
    ) -> RatesRow | None:
        origin_codes = self._resolve_codes(gr.origin_code_raw)
        origin_names = [self._lookup_description(c) for c in origin_codes]
        if any(n is None for n in origin_names):
            return None

        dest_codes_raw = self._split_dest_codes(gr.dest_code_raw)
        dest_names = [self._lookup_description(c) for c in dest_codes_raw]
        if any(n is None for n in dest_names):
            return None

        return RatesRow(
            type="C",
            cmdt_seq=cmdt_seq,
            commodity_group_code=code,
            commodity_group_description=description,
            origin_code=";".join(sorted(set(origin_codes))),
            origin_description=";".join(sorted(set(origin_names))),
            origin_term="CY",
            destination_code=";".join(sorted(set(dest_codes_raw))),
            destination_description=";".join(sorted(set(dest_names))),
            destination_term="CY",
            prefix=prefix or self.container_map.prefix,
            cgo_type=self.container_map.cgo_type,
            cur_20="USD" if "20" in gr.sizes else None,
            rate_20=_to_decimal(gr.sizes.get("20")),
            cur_40="USD" if "40" in gr.sizes else None,
            rate_40=_to_decimal(gr.sizes.get("40")),
            cur_40hc="USD" if "40hc" in gr.sizes else None,
            rate_40hc=_to_decimal(gr.sizes.get("40hc")),
        )

    def _resolve_add_on_destination(self, dest_name: str) -> tuple[str | None, str | None]:
        # ECSA Add-On destination names (Ushuaia, Zarate, La Plata) are
        # plain city names; resolve via a direct name match against the
        # Location Bank rather than fuzzy matching (small, fixed list).
        for rec in self.location_store.all_locations():
            if rec.primary_name.upper() == dest_name.upper() or rec.primary_name.upper().startswith(
                dest_name.upper() + ","
            ):
                return rec.code, rec.primary_name
        return None, None


def _to_decimal(value: float | None) -> Decimal | None:
    return None if value is None else Decimal(str(value))


def _parse_validity(
    ws: Worksheet, row: int = VALIDITY_ROW, from_col: int = VALIDITY_FROM_COL, to_col: int = VALIDITY_TO_COL
) -> tuple[date | None, date | None]:
    start = ws.cell(row=row, column=from_col).value
    end = ws.cell(row=row, column=to_col).value
    if isinstance(start, str):
        start = date.fromisoformat(start)
    elif hasattr(start, "date"):
        start = start.date()
    if isinstance(end, str):
        end = date.fromisoformat(end)
    elif hasattr(end, "date"):
        end = end.date()
    return start, end


register(
    LayoutProfile(
        lane_id=LAECParser.lane_id,
        parser_cls=LAECParser,
        sheet_name_patterns=[r"^DRY$"],
        title_keywords=["LAEC"],
        # "POL: Non-ISC" (verbatim, confirmed on this lane's own DRY sheet)
        # replaces the old ["D2","D4","D5"] fingerprint - those tokens are
        # NOT discriminating: LAEC LUX's own DRY sheet has them too (same
        # D2/D4/D5 container labels), which without this change classified
        # a LUX raw file as this lane at confidence 1.0 (see
        # LAECLuxParser's own profile below for the LUX-specific fix).
        header_fingerprint=["POL: Non-ISC"],
    )
)


# --- LAEC LUX ("...for via LUX Service") -----------------------------------
#
# A distinct, reduced-scope LAEC variant - RATES + SUR (CMDT NOTE) +
# FREETIME only, no ORIGIN ARBS, no OPUS SPECIAL NOTE (confirmed against
# real ground truth: the OPUS workbook has exactly 3 sheets, ['RATES',
# 'SUR', 'freetime']) - filed as a raw workbook with a completely different
# shape from the main LAEC lane above:
#
# - Sheet names differ: ['DRY', 'tiger', 'ECSA Add-on', 'IMO charge'].
#   "tiger" is stale cruft - a broken VLOOKUP scratch sheet where every
#   "Freight" cell is a #REF! error and the POR codes are unrelated CN/HK/
#   TW ports, not the real Indian/Pakistani origins in the actual DRY grid
#   - never read here. "ECSA Add-on" (lowercase "on") is this variant's own
#   casing, distinct from the main lane's "ECSA Add-On" - a separate
#   RAW_SHEET_ECSA_LUX constant below avoids the case-sensitive sheet-name
#   lookup silently missing it.
# - DRY's own layout is ONE flat POD header block (no Non-ISC/ISC split):
#   POD codes at row 13, container labels (D2/D4/D5/NOR) at row 14, data
#   from row 15 (6 origins in both real samples seen, rows 15-20).
#   Validity dates live at row 8 cols C/D (not row 4 like the main lane).
# - Each POD block is 4 columns (D2/D4/D5/NOR) EXCEPT the last two (Rio De
#   Janeiro, Asuncion), which are only 3 (D2/D4/D5, no NOR) - confirmed
#   real, not a data-entry gap. flatten_pod_header() handles this natively
#   since it reads each column's own container label rather than assuming
#   a fixed block width.
# - "NOR" reads as a reefer rate slot, but every NOR cell in both real
#   samples is "-" (no rate offered). laec.yaml's container map doesn't
#   map "NOR" to a rate suffix, so flatten_pod_header/_parse_grid_section
#   silently skip it exactly like any other unmapped container label - no
#   reefer-specific wiring added since there's no confirmed ground-truth
#   reefer row to verify one against.
# - Commodity: one single group (default "G0007" / "FAK -  for via LUX
#   service only" - not present anywhere in the raw file, same
#   externally-assigned-default category as every other lane's commodity
#   identity, user-overridable). Charge-code order for CMDT NOTE, verified
#   against this variant's own SUR sheet: HEA, PSS, OBS, EFS, MBS - a
#   genuinely different order from the main lane's NON_ISC_CHARGE_CODES.
# - CMDT NOTE gets an extra line, "Rates are applicable for Vessel Service
#   Lane: LUX", and the parent row's own Lane column is set to "LUX" -
#   unique to this variant (see build_cmdt_notes()'s service_lane param).
#   The child rows' own Application Effective/Expires dates do NOT track
#   this filing's own rate validity window (unlike the parent APP row,
#   which does) - confirmed different between both real samples (2026-05-
#   14/2026-09-30 for the 20260815-20260831 filing vs 2026-08-23/2026-12-
#   31 for the 20260901-20260914 filing) and not present anywhere else in
#   the raw file - an externally-assigned RFA window, the same accepted
#   gap as the main LAEC lane's own rfa_effective_date/rfa_expiry_date
#   override (MappingProfile), reused here as-is.
# - ECSA Add-on: same shape/columns as the main lane's own sheet (just a
#   shorter row range, 4-8 vs 4-9 - reuses _parse_ecsa_sheet() unchanged
#   since row 9 is simply blank here and already skipped).
# - FREETIME uses build_laec_freetime("lux", ...) - its own, much smaller
#   table (BR/AR/UY/PY only), with a populated Seq column (1..8), unlike
#   FAK/TIER1's blank one.

RAW_SHEET_ECSA_LUX = "ECSA Add-on"

LUX_VALIDITY_ROW, LUX_VALIDITY_FROM_COL, LUX_VALIDITY_TO_COL = 8, 3, 4

COMMODITY_LUX = ("G0007", "FAK -  for via LUX service only", None)

# Verified against this variant's own SUR sheet - child-row order, not the
# alphabetized "inclusive of" text order (see build_cmdt_notes' own
# sort_text_names).
LUX_CHARGE_CODES = ["HEA", "PSS", "OBS", "EFS", "MBS"]

LUX_SECTION = DrySectionConfig(
    pod_code_row=13, container_label_row=14, data_min_row=15, data_max_row=20, min_col=3, max_col=29,
    commodity=COMMODITY_LUX,
)


@dataclass
class LAECLuxRawData:
    validity_start: date | None
    validity_end: date | None
    grid_rows: list[GridRow]
    ecsa_add_ons: list[EcsaAddOn]


class LAECLuxParser(LAECParser):
    """LAEC LUX ("...for via LUX Service") - see the module-level comment
    block above for the full, verified divergence from the main LAECParser
    above, which this subclasses only to reuse its generic instance helpers
    (_parse_grid_section, _build_rates_row, _parse_ecsa_sheet,
    _resolve_codes, _lookup_description, _resolve_add_on_destination) and
    its __init__ (same laec.yaml container map / group codes)."""

    lane_id: ClassVar[str] = "LAEC-LUX"

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        if RAW_SHEET_DRY not in wb.sheetnames:
            return 0.0
        ws = wb[RAW_SHEET_DRY]
        title = str(ws.cell(row=1, column=1).value or "")
        score = 0.5 if "LAEC" in title.upper() else 0.0
        if "LUX" in title.upper():
            score += 0.3
        nor_tokens = {
            str(ws.cell(row=LUX_SECTION.container_label_row, column=c).value or "").strip().upper()
            for c in range(LUX_SECTION.min_col, LUX_SECTION.max_col + 1)
        }
        if "NOR" in nor_tokens:
            score += 0.2
        return min(score, 1.0)

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        ws = wb[RAW_SHEET_DRY]
        validity_start, validity_end = _parse_validity(ws, LUX_VALIDITY_ROW, LUX_VALIDITY_FROM_COL, LUX_VALIDITY_TO_COL)
        grid_rows = self._parse_grid_section(ws, LUX_SECTION)

        ecsa_add_ons: list[EcsaAddOn] = []
        if RAW_SHEET_ECSA_LUX in wb.sheetnames:
            ecsa_add_ons = self._parse_ecsa_sheet(wb[RAW_SHEET_ECSA_LUX])

        return RawExtraction(
            tables={
                "laec_lux": LAECLuxRawData(
                    validity_start=validity_start,
                    validity_end=validity_end,
                    grid_rows=grid_rows,
                    ecsa_add_ons=ecsa_add_ons,
                )
            }
        )

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: LAECLuxRawData = raw.tables["laec_lux"]

        code, default_description, _ = COMMODITY_LUX
        description = resolve_commodity_description(default_description, config)
        cmdt_seq = config.commodity_sequence_overrides.get(default_description)
        output_code = resolve_commodity_code(default_description, code, config)

        dr_rows, dr_pp, dg_rows, dg_pp = [], [], [], []
        base_by_key: dict[tuple[str, str], RatesRow] = {}
        for gr in data.grid_rows:
            row = self._build_rates_row(gr, output_code, description, cmdt_seq)
            if row is None:
                continue
            dr_pp.extend(explode_rates_row(row))
            base_by_key[(row.origin_code, row.destination_code)] = row
            dr_rows.append(row)

            if self.container_map.cgo_type == "DR" and not config.skip_dg_generation.get(default_description, False):
                dg_row = row.model_copy(update={"cgo_type": "DG"})
                dg_pp.extend(explode_rates_row(dg_row))
                dg_rows.append(dg_row)

        rates = group_by_destination(dr_rows) + group_by_destination(dg_rows)
        rates_port_port = group_by_destination(dr_pp) + group_by_destination(dg_pp)

        # ECSA Add-On: extra destinations built from an existing
        # destination's ("T/S Port") rate plus a fixed add-on - no DG
        # duplicate (same as the main lane's own verified behavior).
        extra_rows, extra_pp = [], []
        origin_codes_seen = {origin for origin, _dest in base_by_key}
        for origin_code in origin_codes_seen:
            for addon in data.ecsa_add_ons:
                ts_row = base_by_key.get((origin_code, addon.ts_port_code))
                if ts_row is None:
                    continue
                dest_code_resolved, dest_desc_resolved = self._resolve_add_on_destination(addon.dest_name)
                if dest_code_resolved is None:
                    continue
                new_row = ts_row.model_copy(
                    update={
                        "destination_code": dest_code_resolved,
                        "destination_description": dest_desc_resolved,
                        "rate_20": (ts_row.rate_20 + addon.add_on_20) if ts_row.rate_20 is not None else None,
                        "rate_40": (ts_row.rate_40 + addon.add_on_40) if ts_row.rate_40 is not None else None,
                        "rate_40hc": (
                            (ts_row.rate_40hc + addon.add_on_40) if ts_row.rate_40hc is not None else None
                        ),
                    }
                )
                extra_rows.append(new_row)
                extra_pp.extend(explode_rates_row(new_row))
        rates.extend(group_by_destination(extra_rows))
        rates_port_port.extend(group_by_destination(extra_pp))

        notes = build_cmdt_notes(
            data.validity_start,
            data.validity_end,
            LUX_CHARGE_CODES,
            sequential_charge_seq=True,
            charge_code_names_override=CHARGE_CODE_NAMES_OVERRIDE,
            excluded_codes=frozenset(config.excluded_charge_codes),
            rfa_effective=config.rfa_effective_date,
            rfa_expiry=config.rfa_expiry_date,
            service_lane="LUX",
        )
        note_text = notes[0].contents if notes else None
        for row in rates:
            row.commodity_note = note_text

        freetime = build_laec_freetime("lux", data.validity_start, data.validity_end)

        return OpusRowSet(rates=rates, rates_port_port=rates_port_port, cmdt_notes=notes, freetime=freetime)


register(
    LayoutProfile(
        lane_id=LAECLuxParser.lane_id,
        parser_cls=LAECLuxParser,
        sheet_name_patterns=[r"^DRY$"],
        title_keywords=["LAEC", "LUX"],
        # "NOR" (exact cell value, the 4th rate-slot header on most POD
        # blocks in this variant's own DRY sheet) - confirmed absent as an
        # exact cell value anywhere in the main LAEC lane's own raw
        # workbook, so it discriminates the two without relying on the
        # shared, non-discriminating D2/D4/D5 tokens (see LAECParser's own
        # profile above).
        header_fingerprint=["NOR"],
    )
)
