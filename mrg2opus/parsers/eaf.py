"""EAF lane parser - two sub-lanes (TZDAR/KEMBA) sharing one raw layout:
single destination per sheet (no merged multi-POD header like SAF), rate
cells are formulas (needs a data_only=True workbook), no reefer container
column. Each sub-lane's OPUS output lives in its own suffixed sheets within
one workbook (see excel_io.writer.write_opus_workbook_multi).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationResolver
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.cmdt_notes import build_cmdt_notes, parse_included_charge_codes
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import OpusRowSet, RatesRow, explode_rates_row

SUBLANE_SHEET_RE = re.compile(r"^EAF\s+(\w+)$")
VIA_EXTRACT_RE = re.compile(r"\bvia\s+(.+)$", re.IGNORECASE)
VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s+to\s+(\d{1,2})\s+(\w+)\s+(\d{4})", re.IGNORECASE)

DEFAULT_COMMODITY_CODE = "G0001"
DEFAULT_COMMODITY_DESCRIPTION = "FAK"

DEST_LABEL_ROW = 7
DEST_VALUE_ROW = 8
CONTAINER_LABEL_ROW = 9
DATA_MIN_ROW = 10
DATA_MAX_ROW = 53
MIN_COL, MAX_COL = 4, 6  # D..F (D2, D4, D5 - no reefer column on this lane)
ORIGIN_TEXT_COL = 2  # B

PREPAID_LINE = "Ocean Freight to be Prepaid, payable at -1 by -1."


@dataclass
class RawRateCell:
    origin_text: str
    value: float | str | None
    col_idx: int


@dataclass
class EAFSubLaneData:
    destination_text: str
    validity_start: date | None
    validity_end: date | None
    included_charge_codes: list[str]
    prepaid_at_origin: bool
    rate_cells: list[RawRateCell]


class EAFParser(BaseMRGParser):
    lane_id: ClassVar[str] = "EAF"

    def __init__(self, container_map: ContainerMap | None = None, location_resolver: LocationResolver | None = None):
        self.container_map = container_map or load_container_map("eaf")
        self.location_resolver = location_resolver or LocationResolver()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        sublane_sheets = [n for n in wb.sheetnames if SUBLANE_SHEET_RE.match(n)]
        if not sublane_sheets:
            return 0.0
        score = 0.5
        ws = wb[sublane_sheets[0]]
        title = str(ws.cell(row=1, column=1).value or "")
        if "EAF" in title.upper():
            score += 0.3
        header_tokens = {ws.cell(row=CONTAINER_LABEL_ROW, column=c).value for c in range(MIN_COL, MAX_COL + 1)}
        if {"D2", "D4", "D5"} <= {str(t).strip() for t in header_tokens if t}:
            score += 0.2
        return min(score, 1.0)

    def _sublane_sheets(self, wb: Workbook) -> dict[str, Worksheet]:
        return {m.group(1): wb[m.group(0)] for name in wb.sheetnames if (m := SUBLANE_SHEET_RE.match(name))}

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        tables: dict[str, EAFSubLaneData] = {}
        for suffix, ws in self._sublane_sheets(wb).items():
            tables[suffix] = self._parse_sublane(ws)
        return RawExtraction(tables=tables)

    def _parse_sublane(self, ws: Worksheet) -> EAFSubLaneData:
        validity_start, validity_end = _parse_validity(ws)
        included_codes = parse_included_charge_codes(str(ws.cell(row=3, column=4).value or ""))
        prepaid_at_origin = "PREPAID" in str(ws.cell(row=5, column=4).value or "").upper()
        dest_cell = ws.cell(row=DEST_VALUE_ROW, column=4)
        # A struck-through/blacked-out destination cell marks the whole
        # sub-lane sheet's destination as withdrawn - no origin rate on this
        # sheet is meaningful without it, so skip the entire sheet.
        destination_text = "" if is_excluded(dest_cell) else str(dest_cell.value or "").strip()

        rate_cells: list[RawRateCell] = []
        if destination_text:
            for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
                origin_cell = ws.cell(row=row_idx, column=ORIGIN_TEXT_COL)
                if origin_cell.value in (None, ""):
                    continue
                if is_excluded(origin_cell):
                    continue
                origin_text = str(origin_cell.value).strip()
                for col in range(MIN_COL, MAX_COL + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value in (None, ""):
                        continue
                    if not isinstance(cell.value, (int, float)):
                        continue  # e.g. '-' meaning "not offered" for this container size
                    if is_excluded(cell):
                        continue
                    rate_cells.append(RawRateCell(origin_text=origin_text, value=cell.value, col_idx=col))

        return EAFSubLaneData(
            destination_text=destination_text,
            validity_start=validity_start,
            validity_end=validity_end,
            included_charge_codes=included_codes,
            prepaid_at_origin=prepaid_at_origin,
            rate_cells=rate_cells,
        )

    def _to_opus_rows_for_sublane(self, data: EAFSubLaneData, config: MappingProfile) -> OpusRowSet:
        commodity_description = resolve_commodity_description(DEFAULT_COMMODITY_DESCRIPTION, config)
        cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_COMMODITY_DESCRIPTION)
        output_commodity_code = resolve_commodity_code(DEFAULT_COMMODITY_DESCRIPTION, DEFAULT_COMMODITY_CODE, config)

        dest_matches = self.location_resolver.match_text(data.destination_text)
        dest_confident = bool(dest_matches) and not any(m.needs_review for m in dest_matches)
        dest_codes = sorted({m.code for m in dest_matches}) if dest_confident else []
        dest_names = sorted({m.primary_name for m in dest_matches}) if dest_confident else []

        # group raw cells by origin_text so one OPUS row covers all
        # container sizes for a given origin (single destination per sheet).
        groups: dict[str, dict[str, float]] = {}
        col_to_label = {MIN_COL: "D2", MIN_COL + 1: "D4", MIN_COL + 2: "D5"}
        for rc in data.rate_cells:
            suffix = self.container_map.suffix_for(col_to_label[rc.col_idx])
            if suffix is None:
                continue
            groups.setdefault(rc.origin_text, {})[suffix] = rc.value

        # CMDT NOTE is built first because the PORT-PORT sheet (but NOT the
        # grouped RATES sheet - confirmed asymmetry against ground truth)
        # copies its Contents text into every exploded row's commodity_note.
        extra_lines = [PREPAID_LINE] if data.prepaid_at_origin else []
        cmdt_notes = build_cmdt_notes(
            data.validity_start,
            data.validity_end,
            data.included_charge_codes,
            extra_content_lines=extra_lines,
            sequential_charge_seq=True,
            trailing_oft_row=data.prepaid_at_origin,
        )
        dr_rows: list[RatesRow] = []
        dg_rows: list[RatesRow] = []
        dr_pp: list = []
        dg_pp: list = []

        if dest_confident:
            for origin_text, sizes in groups.items():
                origin_matches = self.location_resolver.match_text(origin_text)
                if not origin_matches or any(m.needs_review for m in origin_matches):
                    continue

                origin_codes = sorted({m.code for m in origin_matches})
                origin_names = sorted({m.primary_name for m in origin_matches})

                o_via_code = None
                via_clause = VIA_EXTRACT_RE.search(origin_text)
                if via_clause:
                    via_match = self.location_resolver.match_token(via_clause.group(1).strip())
                    if via_match and not via_match.needs_review:
                        o_via_code = via_match.code

                row = RatesRow(
                    type="C",
                    cmdt_seq=cmdt_seq,
                    commodity_group_code=output_commodity_code,
                    commodity_group_description=commodity_description,
                    origin_code=";".join(origin_codes),
                    origin_description=";".join(origin_names),
                    origin_term="CY",
                    o_via_code=o_via_code,
                    destination_code=";".join(dest_codes),
                    destination_description=";".join(dest_names),
                    destination_term="CY",
                    prefix=self.container_map.prefix,
                    cgo_type=self.container_map.cgo_type,
                    cur_20="USD" if "20" in sizes else None,
                    rate_20=_to_decimal(sizes.get("20")),
                    cur_40="USD" if "40" in sizes else None,
                    rate_40=_to_decimal(sizes.get("40")),
                    cur_40hc="USD" if "40hc" in sizes else None,
                    rate_40hc=_to_decimal(sizes.get("40hc")),
                )
                dr_rows.append(row)
                dr_pp.extend(explode_rates_row(row))

                # Every base Dry (D/DR) row also files an identical DG (D/DG)
                # variant at the same rate - same standing filing convention
                # confirmed for SAF, holds across both EAF sub-lanes too
                # (84/84 ground-truth rows split exactly 42 DR / 42 DG).
                if self.container_map.cgo_type == "DR" and not config.skip_dg_generation.get(DEFAULT_COMMODITY_DESCRIPTION, False):
                    dg_row = row.model_copy(update={"cgo_type": "DG"})
                    dg_rows.append(dg_row)
                    dg_pp.extend(explode_rates_row(dg_row))

        rates = [*dr_rows, *dg_rows]
        rates_port_port = [*dr_pp, *dg_pp]

        # NOTE: TZDAR's ground-truth PORT-PORT sheet fills route_seq/cmdt_seq/
        # commodity_note (cross-referencing the CMDT NOTE) on every row, but
        # KEMBA's PORT-PORT sheet - in the SAME workbook, same lane - leaves
        # them blank, same as the grouped RATES sheet does for both. Since
        # the one lane that would confirm this convention contradicts itself
        # between its own two sub-lanes, it's not generated here rather than
        # guessing which sub-lane's example is the "real" rule.

        return OpusRowSet(rates=rates, rates_port_port=rates_port_port, cmdt_notes=cmdt_notes)

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        """Merged view across all sub-lanes - mainly useful for a quick
        'everything EAF produced' check. run_multi() is the real entrypoint,
        since ground truth keeps each sub-lane in its own suffixed sheets."""
        combined = OpusRowSet()
        for data in raw.tables.values():
            sub = self._to_opus_rows_for_sublane(data, config)
            combined.rates.extend(sub.rates)
            combined.rates_port_port.extend(sub.rates_port_port)
            combined.cmdt_notes.extend(sub.cmdt_notes)
        return combined

    def run_multi(self, wb: Workbook, config: MappingProfile | None = None) -> dict[str, OpusRowSet]:
        config = config or MappingProfile()
        raw = self.parse_raw(wb)
        return {suffix: self._to_opus_rows_for_sublane(data, config) for suffix, data in raw.tables.items()}


def _to_decimal(value: float | None) -> Decimal | None:
    return None if value is None else Decimal(str(value))


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    # E1: "From 19 Aug to 25 Aug 2026" - unlike SAF, start/end months can
    # differ, and the year only appears once (applies to both ends).
    text = str(ws.cell(row=1, column=5).value or "")
    m = VALIDITY_RE.search(text)
    if not m:
        return None, None
    start_day, start_month, end_day, end_month, year = m.groups()
    try:
        start = date(int(year), _month_number(start_month), int(start_day))
        end = date(int(year), _month_number(end_month), int(end_day))
    except ValueError:
        return None, None
    return start, end


def _month_number(name: str) -> int:
    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    return months.index(name.strip().lower()[:3]) + 1


register(
    LayoutProfile(
        lane_id=EAFParser.lane_id,
        parser_cls=EAFParser,
        sheet_name_patterns=[r"^EAF\s+\w+$"],
        title_keywords=["EAF"],
        header_fingerprint=["D2", "D4", "D5"],
    )
)
