"""LAWC ("Latin America West Coast") lane parser.

Four raw "Dry"-shaped grids feeding the same OPUS RATES/CMDT NOTE sheets,
each its own commodity group, direct origin/destination codes (no fuzzy
matching, same as CSE/LAEC):
  - "China_TWN_SIN_HKG_KR Dry" (G0001, main) - has a "Via" column (raw
    text like "Barge via Sha") for inland Chinese origins reaching the
    coast through a named port; also gets the "Reefer" and "LAWC NOR"
    sheets' output under the SAME commodity group (Prefix "R", cgo_type
    "RF"/"DR" respectively - both single-column, rate in the 40HC slot).
  - "S.E.A_JPN_SA_AU_NZ Dry " (G0004) and "ISC_LK_BD_AE_PK" (G0003) - no
    Via column, otherwise the same grid shape.
  - "OOG" (G0002) - the richest piece: 3 destination-groups (each a
    "/"-joined multi-code group, same handling as LAEC's ECSA groups),
    each with 4 equipment-type column-pairs (in-gauge, "OH", "OWOH",
    "OW"). Verified against ground truth: in-gauge -> Prefix O AND F
    (twin, cgo_type DR); "OH" -> Prefix O AND F (twin, cgo_type AK);
    "OWOH" and "OW" -> Prefix F only (cgo_type AK, no O counterpart).

Origin codes are sometimes a "/"-joined 2-code group directly in the raw
cell (e.g. "CNFOC/ CNFUG") - already-expanded, just split and resolve by
Location Bank lookup, no group_codes.yaml needed (unlike CSE/LAEC's FEBP/
WPRD shorthand, which doesn't appear anywhere in this lane's raw sheets).

Two destinations need Door term / D.Via (verified against ground truth,
same pattern as CSE's PACFZ): Panama City ("PAPTY", Door via Rodman/
"PAROD") and San Lorenzo ("HNSLO", CY term but via NICIO).

NOT implemented: nothing known missing at commodity-group level, but see
README/tests for the same kind of narrow, verified gaps every other lane
has turned up (child-row ordering that isn't derivable from raw text,
etc).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.store import LocationBankStore
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import (
    CommodityNoteSpec,
    build_notes_by_description,
    resolve_commodity_code,
    resolve_commodity_description,
)
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded, location_is_excluded
from mrg2opus.parsers.common.header_grid import flatten_pod_header
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import OpusRowSet, RatesPortPortRow, RatesRow, RouteNoteRow

RAW_SHEET_MAIN = "China_TWN_SIN_HKG_KR Dry"
RAW_SHEET_SEA = "S.E.A_JPN_SA_AU_NZ Dry "
RAW_SHEET_ISC = "ISC_LK_BD_AE_PK"
RAW_SHEET_REEFER = "Reefer"
RAW_SHEET_NOR = "LAWC NOR "
RAW_SHEET_OOG = "OOG"

# Commodity groups, verified against LAWC.xlsx's own OPUS RATES sheet -
# just the starting defaults, user-overridable (see cse.py note). Used
# throughout this file as internal structural keys (PP_COMMODITY,
# _charge_codes_for) - never replace these constants themselves; only the
# OUTPUT row's commodity_group_code goes through resolve_commodity_code().
#
# G0001 used to carry ONE description spanning 3 different raw sheets (the
# main dry grid + "Reefer" + "LAWC NOR"). Each now defaults to its OWN
# description - its own raw sheet name - so COMMODITY_MAIN's description
# below now only describes the main dry grid itself. Reefer/NOR's default
# descriptions are REEFER_DEFAULT_DESCRIPTION/NOR_DEFAULT_DESCRIPTION
# below. Every override dict (code, description, cmdt_seq) is keyed by a
# group's DEFAULT description, not its code - see
# parsers/common/commodity.py's module docstring for why. They only end up
# sharing one CMDT NOTE block again if the user overrides two (or all
# three) of these to the exact same description - see
# parsers/common/commodity.py::build_notes_by_description(), used in
# to_opus_rows() below.
COMMODITY_MAIN = ("G0001", RAW_SHEET_MAIN.strip(), None)
COMMODITY_OOG = ("G0002", "OOG", None)
COMMODITY_ISC = ("G0003", "ISC_LK_BD_AE_PK Dry & DG", None)
COMMODITY_SEA = ("G0004", "S.E.A_JPN_SA_AU_NZ DRY AND DG", None)

REEFER_DEFAULT_DESCRIPTION = RAW_SHEET_REEFER.strip()
NOR_DEFAULT_DESCRIPTION = RAW_SHEET_NOR.strip()

# Verified exact per-group charge-code lists (children order) - see
# cse.py/laec.py's CMDT_NOTE_CHARGE_CODES comment for why this is
# hardcoded rather than derived: the raw "incl." text doesn't reliably
# predict either the order or, for some groups, even the full set (e.g.
# G0004's actual list includes CSS/THL/DOC/CDD that no "incl." line on
# that sheet mentions). The boilerplate TEXT is confirmed alphabetized
# (default sort_text_names=True), independent of this child order.
#
# BAF added to every group (2026-08-26): confirmed present in all 5 of
# reference/2_OPUS/15_LAWC FAK's real SRCHG blocks (the real CMDT NOTE
# equivalent for this lane - see project-opus-note-sheet-taxonomy
# memory). This isn't a lane-specific quirk - see
# project-tool-mirrors-mrg-not-human-sop memory: the user's filing SOP
# tells human agents to skip BAF, but this tool should reproduce it as
# the MRG states regardless. Placed first in each list to match the two
# groups (MAIN-shaped and OOG) directly observed with BAF as the first
# child row; SEA's real block showed BAF second (after PSS) instead -
# child ORDER is cosmetic (display sequence only, not a correctness
# issue), not chased to an exact per-group match here.
MAIN_CHARGE_CODES = ["BAF", "PSS", "OBS", "MBS", "EFS"]
OOG_CHARGE_CODES = ["BAF", "PSS", "OBS", "MBS", "EFS", "HEA"]
ISC_CHARGE_CODES = ["BAF", "EFS", "MBS", "OBS", "PSS"]
SEA_CHARGE_CODES = ["BAF", "OBS", "CSS", "THL", "DOC", "CDD", "MBS", "EFS", "PSS"]

# LAWC's ground truth says "HEAVY SURCHARGE(HEA)" (same as LAEC, not EAF's
# "HEAVY WEIGHT SURCHARGE"); DOC/CDD are new codes not seen in any other
# lane - full names verified directly from this lane's own CMDT NOTE text.
CHARGE_CODE_NAMES_OVERRIDE = {
    "HEA": "HEAVY SURCHARGE",
    "DOC": "DOC FEE (ORIGIN)",
    "CDD": "CARGO DECLARATION CHARGE",
}

# "Via" column abbreviations (main/OOG sheets) -> O.Via code. Only 6
# distinct abbreviations appear across the whole lane (verified against
# ground truth's full o_via_code set) - a curated map, not a fuzzy match,
# since these are terse 3-letter codes ("SHA", "NGB", ...) fuzzy matching
# would likely mis-resolve.
VIA_ABBREV_CODES = {
    "SHA": "CNSHA",
    "NGB": "CNNGB",
    "TAO": "CNTAO",
    "YTN": "CNYTN",
    "SHK": "CNSHK",
    "XMN": "CNXMN",
}
_VIA_ABBREV_RE = re.compile(r"\b(" + "|".join(VIA_ABBREV_CODES) + r")\b", re.IGNORECASE)
# Some rows spell the via-port out as a full code directly (e.g. "CNSHA",
# or "via CNYTN RAIL") instead of the 3-letter abbreviation - check for
# this first, it's unambiguous where the abbreviation map needs a lookup.
_VIA_FULL_CODE_RE = re.compile(r"\b([A-Z]{2}[A-Z0-9]{3})\b")

# Destinations needing Door term / D.Via - verified against ground truth
# (only 2 in the whole lane, same pattern as CSE's PACFZ).
DOOR_TERM_DESTINATIONS = frozenset({"PAPTY"})
DESTINATION_VIA_CODES = {"PAPTY": "PAROD", "HNSLO": "NICIO"}

# HNSLO's raw header cell (the merged POD-name row directly above the POD-
# code row, on every Dry-section sheet) spells out "San Lorenzo CY/CY, HN
# (Via NICIO using MAR/Marex service - Trucking)" - verified per sheet:
# MAIN says "MAR service", SEA/ISC both say "Marex service". Ground truth
# uses "Truck" transmode on every HNSLO row and a per-commodity-group
# Vessel Service Lane route_note ("MAR" for MAIN, "MX2" for SEA/ISC) - the
# "Marex" -> "MX2" abbreviation isn't derivable from the text alone, so
# this is a confirmed lookup, not a parse.
HNSLO_TRANSMODE = "Truck"
HNSLO_ROUTE_NOTE_BY_COMMODITY = {
    "G0001": "Rates are applicable for Vessel Service Lane: MAR",
    "G0003": "Rates are applicable for Vessel Service Lane: MX2",
    "G0004": "Rates are applicable for Vessel Service Lane: MX2",
}

# The OPUS RATES PORT-PORT sheet uses a COMPLETELY different commodity
# code/description namespace than OPUS RATES for the exact same data,
# verified directly against ground truth (G0037-G0041 vs RATES' G0001-
# G0004) - and it splits G0001 (Dry+Reefer+NOR combined in RATES) into two
# separate codes here (G0039 dry-only, G0040 reefer+NOR). Not a derivable
# pattern, just a confirmed 1:1 (or 1:2) remap applied only to the exploded
# PORT-PORT rows, never to the base RATES rows.
PP_COMMODITY = {
    "G0001": ("G0039", "FAK - China_TWN_SIN_HKG Dry"),
    "G0002": ("G0041", "SPECIAL - INGAUGE & OUT OF GAUGE"),
    "G0003": ("G0038", "FAK - ISC LK"),
    "G0004": ("G0037", "FAK - S.E.A_JPN_AE_SA_PK DRY"),
}
PP_COMMODITY_NOR_REEFER = ("G0040", "NOR & REEFER")


_UNSET = object()


def _explode_lawc(
    row: RatesRow,
    origin_names: dict[str, str],
    dest_names: dict[str, str],
    route_note_override=_UNSET,
) -> list[RatesPortPortRow]:
    """Same key-splitting as the shared explode_rates_row(), but resolves
    each exploded row's origin/destination description from a per-code
    map instead of keeping the grouped view's combined ";"-joined string.
    Verified against LAWC's own ground truth - unlike CSE/LAEC/SAF (which
    DO keep the combined description verbatim on every exploded row, see
    schema/opus_rows.py's _explode_group docstring), LAWC's OPUS RATES
    PORT-PORT re-splits to the individual port's own name."""
    origin_codes = [c for c in row.origin_code.split(";") if c] or [row.origin_code]
    dest_codes = [c for c in row.destination_code.split(";") if c] or [row.destination_code]
    data = row.model_dump()
    data["cmdt_seq"] = None
    data["commodity_note"] = None
    if route_note_override is not _UNSET:
        data["route_note"] = route_note_override
    out = []
    for o_code in origin_codes:
        for d_code in dest_codes:
            row_data = dict(data)
            row_data["origin_code"] = o_code
            row_data["origin_description"] = origin_names.get(o_code, row.origin_description)
            row_data["destination_code"] = d_code
            row_data["destination_description"] = dest_names.get(d_code, row.destination_description)
            out.append(RatesPortPortRow(**row_data))
    return out


def _remap_pp_commodity(rows: list, code: str, description: str) -> list:
    return [
        r.model_copy(update={"commodity_group_code": code, "commodity_group_description": description})
        for r in rows
    ]


# Raw-sheet code typos: two origin codes are letter-transpositions of a
# real code already in the Location Bank, confirmed by matching the
# resolved code the ground truth actually files under - "CNXGG" (LAWC NOR
# sheet, row 21) should be "CNTXG" (Xingang, Tianjin), "INXIE" (ISC sheet,
# row 12) should be "INIXE" (Mangalore). Not missing locations - corrected
# before Location Bank lookup.
RAW_CODE_TYPOS = {"CNXGG": "CNTXG", "INXIE": "INIXE"}

VALIDITY_ROW_MAIN, VALIDITY_ROW_SEA, VALIDITY_ROW_ISC = 4, 4, 2


@dataclass(frozen=True)
class DrySectionConfig:
    sheet_name: str
    origin_code_col: int
    via_col: int | None
    pod_code_row: int
    container_label_row: int
    data_min_row: int
    data_max_row: int
    min_col: int
    max_col: int
    validity_row: int
    validity_from_col: int
    validity_to_col: int
    commodity: tuple[str, str, int | None]
    charge_codes: list[str]


# Order matters: ground truth's OPUS CMDT NOTE sheet emits blocks in
# ISC, SEA, MAIN, OOG order (verified by matching each block's charge-code
# sequence against MAIN/SEA/ISC/OOG_CHARGE_CODES) - not raw sheet-tab
# order. cmdt_notes are built per-commodity in this list's iteration
# order, so DRY_SECTIONS must follow it too.
# max_col=68, not 66: HNSLO is the last destination block on both ISC and
# SEA sheets and needs its full 3-column (20'/40'/HCD) width - a max_col of
# 66 silently truncated it to just the 20' column, discarding 40'/40HC
# rates for that one destination (confirmed by comparing raw cells against
# ground truth's HNSLO rows).
DRY_SECTIONS = [
    DrySectionConfig(
        RAW_SHEET_ISC, 2, None, 6, 7, 8, 23, 3, 68, 2, 3, 4, COMMODITY_ISC, ISC_CHARGE_CODES
    ),
    DrySectionConfig(
        RAW_SHEET_SEA, 2, None, 8, 9, 10, 37, 3, 68, 4, 3, 4, COMMODITY_SEA, SEA_CHARGE_CODES
    ),
    DrySectionConfig(
        RAW_SHEET_MAIN, 3, 2, 8, 9, 10, 107, 4, 84, 4, 4, 5, COMMODITY_MAIN, MAIN_CHARGE_CODES
    ),
]

REEFER_CONFIG = ("R", "RF")
NOR_CONFIG = ("R", "DR")
OOG_POD_CODE_ROW, OOG_CONTAINER_LABEL_ROW = 7, 8
# Rows 31-46 are a verbatim repeat of rows 9-24 (confirmed identical rate
# values cell-for-cell), preceded by a mid-sheet header-row repeat at row
# 30 - a copy-paste duplication in the raw sheet, not additional data.
# Rows 47+ are a small unrelated CM/inguage add-on note table. Bounding to
# just the first block avoids double-counting every OOG origin.
OOG_DATA_MIN_ROW, OOG_DATA_MAX_ROW = 9, 24
OOG_MIN_COL, OOG_MAX_COL = 4, 27
OOG_VALIDITY_ROW, OOG_VALIDITY_FROM_COL, OOG_VALIDITY_TO_COL = 2, 4, 5


@dataclass
class GridRow:
    origin_code_raw: str
    via_text: str | None
    dest_code: str
    sizes: dict[str, float]


@dataclass
class SingleColRow:
    origin_code_raw: str
    dest_code: str
    value: float
    vessel_service_lane: str | None = None


@dataclass
class OogRow:
    origin_code_raw: str
    dest_code_raw: str  # may be "/"-joined
    equipment: str  # "ig" | "oh" | "owoh" | "ow"
    rate_20: float | None
    rate_40: float | None
    kci: bool = False


@dataclass
class LAWCRawData:
    grid_rows: dict[tuple[str, str, int | None], list[GridRow]]
    validity: dict[str, tuple[date | None, date | None]]  # commodity_code -> (start, end)
    reefer_rows: list[SingleColRow]
    nor_rows: list[SingleColRow]
    oog_rows: list[OogRow]
    oog_validity: tuple[date | None, date | None]


class LAWCParser(BaseMRGParser):
    lane_id: ClassVar[str] = "LAWC"

    def __init__(self, container_map: ContainerMap | None = None, location_store: LocationBankStore | None = None):
        self.container_map = container_map or load_container_map("lawc")
        self.location_store = location_store or LocationBankStore()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        if RAW_SHEET_MAIN not in wb.sheetnames:
            return 0.0
        score = 0.5
        ws = wb[RAW_SHEET_MAIN]
        title = str(ws.cell(row=1, column=1).value or "") + str(ws.cell(row=1, column=4).value or "")
        if "LAWC" in title.upper():
            score += 0.3
        if RAW_SHEET_OOG in wb.sheetnames and RAW_SHEET_ISC in wb.sheetnames:
            score += 0.2
        return min(score, 1.0)

    def _resolve_codes(self, raw_text: str) -> list[str]:
        # Grouped-code cells use "/" (e.g. "CNFOC/ CNFUG") most of the
        # time, but at least one raw row uses "," instead ("VNSGN, VNCMP,
        # VNDIA, VNBHA") - split on both. Footnote markers ("CLIQQ**") get
        # stripped too.
        codes = [c.strip().rstrip("*").strip() for c in re.split(r"[/,]", raw_text) if c.strip()]
        return [RAW_CODE_TYPOS.get(c, c) for c in codes]

    def _lookup_description(self, code: str) -> str | None:
        rec = self.location_store.get_by_code(code)
        return rec.primary_name if rec else None

    def _find_via_code(self, text: str | None) -> str | None:
        if not text:
            return None
        full = _VIA_FULL_CODE_RE.search(text)
        if full and self.location_store.get_by_code(full.group(1).upper()) is not None:
            return full.group(1).upper()
        m = _VIA_ABBREV_RE.search(text)
        if not m:
            return None
        return VIA_ABBREV_CODES[m.group(1).upper()]

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        grid_rows: dict[tuple[str, str, int | None], list[GridRow]] = {}
        validity: dict[str, tuple[date | None, date | None]] = {}
        for cfg in DRY_SECTIONS:
            if cfg.sheet_name not in wb.sheetnames:
                continue
            ws = wb[cfg.sheet_name]
            rows = self._parse_dry_section(ws, cfg)
            grid_rows.setdefault(cfg.commodity, []).extend(rows)
            validity[cfg.commodity[0]] = _parse_validity(ws, cfg.validity_row, cfg.validity_from_col, cfg.validity_to_col)

        reefer_rows: list[SingleColRow] = []
        if RAW_SHEET_REEFER in wb.sheetnames:
            reefer_rows = self._parse_single_col_sheet(wb[RAW_SHEET_REEFER])

        nor_rows: list[SingleColRow] = []
        if RAW_SHEET_NOR in wb.sheetnames:
            nor_rows = self._parse_single_col_sheet(wb[RAW_SHEET_NOR])

        oog_rows: list[OogRow] = []
        oog_validity = (None, None)
        if RAW_SHEET_OOG in wb.sheetnames:
            ws = wb[RAW_SHEET_OOG]
            oog_rows = self._parse_oog_sheet(ws)
            oog_validity = _parse_validity(ws, OOG_VALIDITY_ROW, OOG_VALIDITY_FROM_COL, OOG_VALIDITY_TO_COL)

        return RawExtraction(
            tables={
                "lawc": LAWCRawData(
                    grid_rows=grid_rows,
                    validity=validity,
                    reefer_rows=reefer_rows,
                    nor_rows=nor_rows,
                    oog_rows=oog_rows,
                    oog_validity=oog_validity,
                )
            }
        )

    def _parse_dry_section(self, ws: Worksheet, cfg: DrySectionConfig) -> list[GridRow]:
        # fallback_container_cycle: Posorja's (ECPSJ) block on the main
        # sheet has all 3 container labels blank in the raw sheet (a
        # data-entry gap - the rate data itself is present, and ground
        # truth clearly files it) - see header_grid.py's docstring.
        header_cols = flatten_pod_header(
            ws, cfg.pod_code_row, cfg.container_label_row, cfg.min_col, cfg.max_col,
            fallback_container_cycle=["20'", "40'", "HCD"],
        )
        out: list[GridRow] = []
        for row_idx in range(cfg.data_min_row, cfg.data_max_row + 1):
            origin_cell = ws.cell(row=row_idx, column=cfg.origin_code_col)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            via_text = None
            if cfg.via_col is not None:
                via_val = ws.cell(row=row_idx, column=cfg.via_col).value
                via_text = str(via_val).strip() if via_val else None
            by_dest: dict[str, dict[str, float]] = {}
            for hc in header_cols:
                suffix = self.container_map.suffix_for(hc.container_label)
                if suffix is None:
                    continue
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                by_dest.setdefault(hc.pod_label.strip(), {})[suffix] = cell.value
            for dest_code, sizes in by_dest.items():
                out.append(GridRow(origin_code_raw=origin_raw, via_text=via_text, dest_code=dest_code, sizes=sizes))
        return out

    def _parse_single_col_sheet(self, ws: Worksheet) -> list[SingleColRow]:
        pod_code_row, data_min_row = 6, 8
        dest_codes = {}
        dest_service_lanes = {}
        for col in range(3, ws.max_column + 1):
            code_cell = ws.cell(row=pod_code_row, column=col)
            raw = str(code_cell.value or "").strip()
            if not raw or is_excluded(code_cell):
                continue
            # One destination cell has an annotation appended after the
            # code (e.g. "COBUN                        (on AX3 only)") -
            # take just the leading code token, and the "(on X only)" part
            # names a specific Vessel Service Lane the ground truth carries
            # as that row's route_note.
            m = _VIA_FULL_CODE_RE.match(raw)
            dest_codes[col] = m.group(1) if m else raw
            lane_m = re.search(r"\(on\s+(\S+?)\s+only\)", raw, re.IGNORECASE)
            if lane_m:
                dest_service_lanes[col] = lane_m.group(1)
        out: list[SingleColRow] = []
        for row_idx in range(data_min_row, ws.max_row + 1):
            origin_cell = ws.cell(row=row_idx, column=2)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            for col, dest_code in dest_codes.items():
                cell = ws.cell(row=row_idx, column=col)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                out.append(
                    SingleColRow(
                        origin_code_raw=origin_raw,
                        dest_code=dest_code,
                        value=cell.value,
                        vessel_service_lane=dest_service_lanes.get(col),
                    )
                )
        return out

    def _parse_oog_sheet(self, ws: Worksheet) -> list[OogRow]:
        header_cols = flatten_pod_header(ws, OOG_POD_CODE_ROW, OOG_CONTAINER_LABEL_ROW, OOG_MIN_COL, OOG_MAX_COL)
        out: list[OogRow] = []
        for row_idx in range(OOG_DATA_MIN_ROW, OOG_DATA_MAX_ROW + 1):
            origin_cell = ws.cell(row=row_idx, column=3)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            # A "KCI service only" annotation in column B (verified: only
            # IDJKT/IDSUB rows carry it) marks the origin as restricted to a
            # named vessel service lane - the ground truth's route_note text
            # differs for these rows (see COMMODITY_OOG's cmdt_notes note).
            note_cell = ws.cell(row=row_idx, column=2).value
            kci = bool(note_cell) and "KCI" in str(note_cell).upper()
            by_dest_equip: dict[tuple[str, str], dict[str, float]] = {}
            for hc in header_cols:
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                equipment = _classify_oog_equipment(hc.container_label)
                size = "20" if "20" in hc.container_label else "40"
                key = (hc.pod_label.strip(), equipment)
                by_dest_equip.setdefault(key, {})[size] = cell.value
            for (dest_raw, equipment), sizes in by_dest_equip.items():
                out.append(
                    OogRow(
                        origin_code_raw=origin_raw,
                        dest_code_raw=dest_raw,
                        equipment=equipment,
                        rate_20=sizes.get("20"),
                        rate_40=sizes.get("40"),
                        kci=kci,
                    )
                )
        return out

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: LAWCRawData = raw.tables["lawc"]

        rates: list[RatesRow] = []
        rates_port_port: list[RatesRow] = []
        # Descriptions default to per-sheet (see COMMODITY_MAIN's comment),
        # so notes can't be built per-group as each group's rows are built
        # anymore - two groups might end up sharing a description (default,
        # via override, or by the user merging them back together) and need
        # to share exactly one CMDT NOTE block. Collect a spec per group
        # that actually produced rows, resolve all of them into notes ONCE
        # at the end (build_notes_by_description merges same-description
        # specs), then stamp commodity_note onto every row by its own
        # (already-final) commodity_group_description.
        note_specs: list[CommodityNoteSpec] = []

        for commodity, grid_rows in data.grid_rows.items():
            code, default_description, _ = commodity
            description = resolve_commodity_description(default_description, config)
            cmdt_seq = config.commodity_sequence_overrides.get(default_description)
            output_code = resolve_commodity_code(default_description, code, config)
            charge_codes = _charge_codes_for(code)
            validity_start, validity_end = data.validity.get(code, (None, None))

            dr_rows, dr_pp, dg_rows, dg_pp = [], [], [], []
            for gr in grid_rows:
                built = self._build_rates_row(
                    gr.origin_code_raw, gr.dest_code, gr.sizes, code, output_code, description, cmdt_seq
                )
                if built is None:
                    continue
                row, origin_name_map = built
                o_via = self._find_via_code(gr.via_text)
                if o_via:
                    transmode = _find_transmode(gr.via_text)
                    row = row.model_copy(update={"o_via_code": o_via, "origin_transmode": transmode})
                dr_rows.append(row)
                dest_name_map = {row.destination_code: row.destination_description}
                pp_code, pp_description = PP_COMMODITY[code]
                dr_pp.extend(_remap_pp_commodity(_explode_lawc(row, origin_name_map, dest_name_map), pp_code, pp_description))
                if self.container_map.cgo_type == "DR" and not config.skip_dg_generation.get(default_description, False):
                    dg_row = row.model_copy(update={"cgo_type": "DG"})
                    dg_rows.append(dg_row)
                    dg_pp.extend(
                        _remap_pp_commodity(_explode_lawc(dg_row, origin_name_map, dest_name_map), pp_code, pp_description)
                    )

            rates.extend(group_by_destination(dr_rows))
            rates.extend(group_by_destination(dg_rows))
            rates_port_port.extend(group_by_destination(dr_pp))
            rates_port_port.extend(group_by_destination(dg_pp))

            if dr_rows:
                note_specs.append(CommodityNoteSpec(description, validity_start, validity_end, charge_codes))

        # Reefer + NOR: both Prefix "R", rate in the 40HC slot, both share
        # G0001 (main)'s code by default, no DG duplicate. Each defaults to
        # its OWN description now (see COMMODITY_MAIN's comment above), and
        # each is independently overridable (code/description/cmdt_seq)
        # since each has its own default description to key by.
        main_code = COMMODITY_MAIN[0]
        main_validity_start, main_validity_end = data.validity.get(main_code, (None, None))
        sea_output_code = resolve_commodity_code(COMMODITY_SEA[1], COMMODITY_SEA[0], config)
        sea_description = resolve_commodity_description(COMMODITY_SEA[1], config)
        for rows, (prefix, cgo_type), default_description in (
            (data.reefer_rows, REEFER_CONFIG, REEFER_DEFAULT_DESCRIPTION),
            (data.nor_rows, NOR_CONFIG, NOR_DEFAULT_DESCRIPTION),
        ):
            is_nor = default_description == NOR_DEFAULT_DESCRIPTION
            description = resolve_commodity_description(default_description, config)
            variant_cmdt_seq = config.commodity_sequence_overrides.get(default_description)
            variant_output_code = resolve_commodity_code(default_description, main_code, config)
            variant_rows, variant_pp = [], []
            nor_dg_rows, nor_dg_pp = [], []
            for sr in rows:
                origin_codes = self._resolve_codes(sr.origin_code_raw)
                origin_names = [self._lookup_description(c) for c in origin_codes]
                if any(n is None for n in origin_names):
                    continue
                dest_name = self._lookup_description(sr.dest_code)
                if dest_name is None:
                    continue
                row = RatesRow(
                    type="C",
                    cmdt_seq=variant_cmdt_seq,
                    commodity_group_code=variant_output_code,
                    commodity_group_description=description,
                    origin_code=";".join(sorted(set(origin_codes))),
                    origin_description=";".join(sorted(set(origin_names))),
                    origin_term="CY",
                    destination_code=sr.dest_code,
                    destination_description=dest_name,
                    destination_term="Door" if sr.dest_code in DOOR_TERM_DESTINATIONS else "CY",
                    destination_transmode=HNSLO_TRANSMODE if sr.dest_code == "HNSLO" else None,
                    d_via_code=DESTINATION_VIA_CODES.get(sr.dest_code),
                    prefix=prefix,
                    cgo_type=cgo_type,
                    cur_40hc="USD",
                    rate_40hc=_to_decimal(sr.value),
                    route_note=(
                        HNSLO_ROUTE_NOTE_BY_COMMODITY.get(main_code)
                        if sr.dest_code == "HNSLO"
                        else (
                            f"Rates are applicable for Vessel Service Lane: {sr.vessel_service_lane}"
                            if sr.vessel_service_lane
                            else None
                        )
                    ),
                )
                variant_rows.append(row)
                origin_name_map = dict(zip(origin_codes, origin_names))
                dest_name_map = {sr.dest_code: dest_name}
                nor_reefer_code, nor_reefer_description = PP_COMMODITY_NOR_REEFER
                variant_pp.extend(
                    _remap_pp_commodity(
                        _explode_lawc(row, origin_name_map, dest_name_map), nor_reefer_code, nor_reefer_description
                    )
                )
                if is_nor and not config.skip_dg_generation.get(default_description, False):
                    dg_row = row.model_copy(
                        update={
                            "prefix": "D",
                            "cgo_type": "DG",
                            "commodity_group_code": sea_output_code,
                            "commodity_group_description": sea_description,
                            "route_note": _sea_dg_route_note(row.route_note),
                        }
                    )
                    nor_dg_rows.append(dg_row)
                    sea_pp_code, sea_pp_description = PP_COMMODITY[COMMODITY_SEA[0]]
                    nor_dg_pp.extend(
                        _remap_pp_commodity(
                            _explode_lawc(dg_row, origin_name_map, dest_name_map), sea_pp_code, sea_pp_description
                        )
                    )
            rates.extend(group_by_destination(variant_rows))
            rates.extend(group_by_destination(nor_dg_rows))
            rates_port_port.extend(group_by_destination(variant_pp))
            rates_port_port.extend(group_by_destination(nor_dg_pp))
            if variant_rows:
                note_specs.append(
                    CommodityNoteSpec(description, main_validity_start, main_validity_end, MAIN_CHARGE_CODES)
                )

        # OOG: in-gauge -> O+F twin (DR); "OH" -> O+F twin (AK); "OWOH"/"OW"
        # -> F only (AK). Destination itself may be a "/"-joined group.
        oog_code, oog_default_description, _ = COMMODITY_OOG
        oog_description = resolve_commodity_description(oog_default_description, config)
        oog_cmdt_seq = config.commodity_sequence_overrides.get(oog_default_description)
        oog_output_code = resolve_commodity_code(oog_default_description, oog_code, config)
        oog_rows_out, oog_pp = [], []
        for gr in data.oog_rows:
            origin_codes = self._resolve_codes(gr.origin_code_raw)
            origin_names = [self._lookup_description(c) for c in origin_codes]
            if any(n is None for n in origin_names):
                continue
            dest_codes = self._resolve_codes(gr.dest_code_raw)
            dest_names = [self._lookup_description(c) for c in dest_codes]
            if any(n is None for n in dest_names):
                continue
            origin_name_map = dict(zip(origin_codes, origin_names))
            dest_name_map = dict(zip(dest_codes, dest_names))

            base = RatesRow(
                type="C",
                cmdt_seq=oog_cmdt_seq,
                commodity_group_code=oog_output_code,
                commodity_group_description=oog_description,
                origin_code=";".join(sorted(set(origin_codes))),
                origin_description=";".join(sorted(set(origin_names))),
                origin_term="CY",
                destination_code=";".join(sorted(set(dest_codes))),
                destination_description=";".join(sorted(set(dest_names))),
                destination_term="CY",
                prefix="O",
                cgo_type="DR" if gr.equipment == "ig" else "AK",
                cur_20="USD" if gr.rate_20 is not None else None,
                rate_20=_to_decimal(gr.rate_20),
                cur_40="USD" if gr.rate_40 is not None else None,
                rate_40=_to_decimal(gr.rate_40),
                route_note=_oog_route_note(gr.equipment, gr.kci),
            )
            oog_pp_code, oog_pp_description = PP_COMMODITY[oog_code]
            f_row = base.model_copy(update={"prefix": "F"})
            oog_rows_out.append(f_row)
            oog_pp.extend(
                _remap_pp_commodity(
                    _explode_lawc(f_row, origin_name_map, dest_name_map), oog_pp_code, oog_pp_description
                )
            )
            if gr.equipment in ("ig", "oh"):
                oog_rows_out.append(base)
                oog_pp.extend(
                    _remap_pp_commodity(
                        _explode_lawc(base, origin_name_map, dest_name_map), oog_pp_code, oog_pp_description
                    )
                )

        rates.extend(group_by_destination(oog_rows_out))
        rates_port_port.extend(group_by_destination(oog_pp))
        if oog_rows_out:
            note_specs.append(
                CommodityNoteSpec(oog_description, data.oog_validity[0], data.oog_validity[1], OOG_CHARGE_CODES)
            )

        cmdt_notes, note_text_by_description = build_notes_by_description(
            note_specs,
            sequential_charge_seq=True,
            charge_code_names_override=CHARGE_CODE_NAMES_OVERRIDE,
            excluded_codes=frozenset(config.excluded_charge_codes),
        )
        for row in rates:
            row.commodity_note = note_text_by_description.get(row.commodity_group_description)

        route_notes = _derive_route_notes(rates, main_validity_start, main_validity_end)

        return OpusRowSet(
            rates=rates, rates_port_port=rates_port_port, cmdt_notes=cmdt_notes, route_notes=route_notes
        )

    def _build_rates_row(
        self,
        origin_raw: str,
        dest_code: str,
        sizes: dict[str, float],
        structural_code: str,
        output_code: str,
        description: str,
        cmdt_seq: int | None,
    ) -> tuple[RatesRow, dict[str, str]] | None:
        origin_codes = self._resolve_codes(origin_raw)
        origin_names = [self._lookup_description(c) for c in origin_codes]
        if any(n is None for n in origin_names):
            return None
        dest_name = self._lookup_description(dest_code)
        if dest_name is None:
            return None
        origin_name_map = dict(zip(origin_codes, origin_names))

        row = RatesRow(
            type="C",
            cmdt_seq=cmdt_seq,
            commodity_group_code=output_code,
            commodity_group_description=description,
            origin_code=";".join(sorted(set(origin_codes))),
            origin_description=";".join(sorted(set(origin_names))),
            origin_term="CY",
            destination_code=dest_code,
            destination_description=dest_name,
            destination_term="Door" if dest_code in DOOR_TERM_DESTINATIONS else "CY",
            destination_transmode=HNSLO_TRANSMODE if dest_code == "HNSLO" else None,
            d_via_code=DESTINATION_VIA_CODES.get(dest_code),
            prefix=self.container_map.prefix,
            cgo_type=self.container_map.cgo_type,
            cur_20="USD" if "20" in sizes else None,
            rate_20=_to_decimal(sizes.get("20")),
            cur_40="USD" if "40" in sizes else None,
            rate_40=_to_decimal(sizes.get("40")),
            cur_40hc="USD" if "40hc" in sizes else None,
            rate_40hc=_to_decimal(sizes.get("40hc")),
            route_note=HNSLO_ROUTE_NOTE_BY_COMMODITY.get(structural_code) if dest_code == "HNSLO" else None,
        )
        return row, origin_name_map


def _derive_route_notes(
    rates: list[RatesRow], validity_start: date | None, validity_end: date | None
) -> list[RouteNoteRow]:
    """Every RatesRow with a non-null route_note needs a matching entry on
    the real RN sheet (see project-opus-note-sheet-taxonomy memory) - real
    RN rows are header-only (charge_seq/code always 1/"APP", no child
    charge-code rows, unlike CMDT NOTE). header_seq/route_seq/note_seq are
    placeholder running numbers, not reproductions of any real OPUS-assigned
    number - confirmed acceptable since OPUS renumbers these on import,
    same treatment already given to CMDT NOTE's header_seq/note_seq
    elsewhere in this codebase. header_seq groups by distinct route_note
    text (verified: every real RN row sharing one route_note text shares
    one header_seq); route_seq is a running counter across all qualifying
    rows, also stamped back onto the RatesRow it came from so the two
    sheets stay linkable."""
    route_notes: list[RouteNoteRow] = []
    header_seq_by_text: dict[str, int] = {}
    next_route_seq = 1
    for row in rates:
        if not row.route_note:
            continue
        header_seq = header_seq_by_text.setdefault(row.route_note, len(header_seq_by_text) + 1)
        row.route_seq = next_route_seq
        route_notes.append(
            RouteNoteRow(
                header_seq=header_seq,
                route_seq=next_route_seq,
                note_seq=1,
                contents=row.route_note,
                charge_seq=1,
                code="APP",
                application="S",
                application_effective=validity_start,
                application_expires=validity_end,
            )
        )
        next_route_seq += 1
    return route_notes


def _charge_codes_for(commodity_code: str) -> list[str]:
    return {
        COMMODITY_MAIN[0]: MAIN_CHARGE_CODES,
        COMMODITY_SEA[0]: SEA_CHARGE_CODES,
        COMMODITY_ISC[0]: ISC_CHARGE_CODES,
    }[commodity_code]


def _find_transmode(via_text: str | None) -> str | None:
    # The raw via-column text carries the inland transmode directly (e.g.
    # "Barge via Sha", "via CNYTN RAIL") for SOME rows, but not all - many
    # via rows just name the port ("Via SHA", "Via TAO") with no mode word,
    # and ground truth leaves origin_transmode blank for those.
    if not via_text:
        return None
    lower = via_text.lower()
    if "rail" in lower:
        return "Rail"
    if "barge" in lower:
        return "Barge"
    return None


def _oog_route_note(equipment: str, kci: bool) -> str | None:
    # In-gauge ("ig") is treated as the "default" equipment: no route note
    # at all when there's no KCI restriction, and the bare "...KCI" text
    # (no "(IG)" parenthetical) when there is - verified directly against
    # reference/2_OPUS/17_LAWC TIER 1's real RATES sheet (every prefix O/F,
    # cgo_type DR row has Route Note blank unless KCI-flagged, and the
    # KCI-flagged ones read exactly "...Vessel Service Lane: KCI", no
    # "(IG)"). OH/OWOH/OW always get a route note, KCI or not.
    if equipment == "ig":
        return "Rates are applicable for Vessel Service Lane: KCI" if kci else None
    label = equipment.upper()
    if kci:
        return f"Rates are applicable for Vessel Service Lane: KCI ({label})"
    return label


# "REEFER DRY AS DANGEROUS" - user-confirmed (2026-08-26, not derivable
# from raw MRG text): Non-Operating Reefer ("LAWC NOR" sheet) cargo that's
# dangerous gets filed as D/DG (folded into G0004/S.E.A_JPN_SA_AU_NZ's
# regular dry-and-dangerous bucket) instead of R/DG, with this route note
# explaining why. Applies ONLY to NOR (never Reefer - REEFER_CONFIG's
# cgo_type "RF" never matches the DR-only DG-duplication rule anyway, so
# Reefer never had a DG variant to begin with) - see the nor_rows branch
# below. Combines with any route_note the NOR row already carries (e.g.
# COBUN's AX3 vessel-lane note) via " | ", confirmed against
# reference/2_OPUS/15_LAWC FAK and 17_LAWC TIER 1's real RN/RATES sheets.
SEA_DG_ROUTE_NOTE = "REEFER DRY AS DANGEROUS"


def _sea_dg_route_note(existing: str | None) -> str:
    return f"{SEA_DG_ROUTE_NOTE} | {existing}" if existing else SEA_DG_ROUTE_NOTE


def _classify_oog_equipment(container_label: str) -> str:
    label = container_label.upper()
    if "OWOH" in label:
        return "owoh"
    if "OH" in label:
        return "oh"
    if "OW" in label:
        return "ow"
    return "ig"


def _to_decimal(value: float | None) -> Decimal | None:
    return None if value is None else Decimal(str(value))


def _parse_validity(ws: Worksheet, row: int, from_col: int, to_col: int) -> tuple[date | None, date | None]:
    start = ws.cell(row=row, column=from_col).value
    end = ws.cell(row=row, column=to_col).value
    if isinstance(start, str):
        start = date.fromisoformat(start)
    elif hasattr(start, "date"):
        start = start.date()
    if isinstance(end, str):
        end = date.fromisoformat(end)
    elif hasattr(end, "date"):
        end = end.date()
    return start, end


register(
    LayoutProfile(
        lane_id=LAWCParser.lane_id,
        parser_cls=LAWCParser,
        sheet_name_patterns=[r"^China_TWN_SIN_HKG_KR Dry$"],
        title_keywords=["LAWC"],
        header_fingerprint=["20'", "40'", "HCD"],
    )
)
