"""West Asia to West Africa lane parser.

Single raw sheet: an 8-POD x 3-container-size (D2/D4/D5) rate grid, ex a
fixed list of ~14 South Asia (India/Bangladesh/Pakistan/Sri Lanka) origins.
Structurally close to the West Africa WAF lane (waf.py) - same D2/D4/D5
grid shape, same "Rate structure: incl. X, Y, Z ;" comma-separated
charge-code text, same automatic D/DG duplicate at the same rate - but
genuinely different in one important way: BOTH the origin ("Port Code"
column) and destination (embedded in each POD's own label, e.g.
"POD: Tema (GHTEM)") codes are given directly in the raw sheet, so this
lane needs NO fuzzy Location Bank matching for the MATCH step - codes are
parsed straight off the sheet, then resolved to their OPUS description via
a direct LocationBankStore.get_by_code() lookup (confirmed: every raw
label's own display text, e.g. "Calcutta"/"Apapa/Tincan", sometimes
disagrees with the Location Bank's current primary_name, e.g. "KOLKATA" -
ground truth always matches the Location Bank, not the raw label).

Also new here: an origin-scoped charge-code override ("incl. THL for cargo
ex LKCMB/PKKHI/BDCGP") layered on top of the shared "incl. BAF, HEA, ..."
list - filed as 3 extra POL-scoped THL child rows on the one CMDT NOTE
block (same "extra child row per origin" pattern as AUEC/AUWC's own
POL-scoped surcharges), not a second commodity group.

Real filings seen use TWO different raw sheet names for the same shape
("WAF" and "West Asia - West Africa") - detect() matches on title text
rather than a fixed sheet name for this reason.

Known, accepted gap: the "inclusive of" child rows are filed alphabetized
(confirmed against the first week's ground truth), but the SECOND week's
raw file adds a brand-new code (EFS, previously only in the "subj to"
list) - its ground truth appends that new code's child row at the very
END of the block (after even the THL origin-scoped rows) rather than
re-alphabetizing, a human filing-history artifact ("we don't reorder
existing entries when adding one mid-filing-period") that isn't
derivable from a single week's raw file in isolation. This parser keeps
the simpler, fully-alphabetized order (matching week 1 exactly).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.store import LocationBankStore
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.header_grid import flatten_pod_header
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.common.sequencing import assign_cmdt_seq_numbers
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import CHARGE_CODE_NAMES, is_known_charge_code
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow

TITLE_KEYWORD = "WEST ASIA WAF"

POD_LABEL_ROW = 9
CONTAINER_LABEL_ROW = 10
DATA_MIN_ROW = 11
DATA_MAX_ROW = 24
ORIGIN_CODE_COL = 3
MIN_COL, MAX_COL = 5, 35

VALIDITY_ROW, VALIDITY_TEXT_COL = 1, 7
INCLUDES_ROW, INCLUDES_COL = 2, 6
THL_OVERRIDE_ROW, THL_OVERRIDE_COL = 4, 7

DEFAULT_DR_DESCRIPTION = "West Asia West Africa MRG"
DEFAULT_DR_CODE = "G0002"
DEFAULT_DG_DESCRIPTION = "West Asia West Africa MRG - DG"
DEFAULT_DG_CODE = "G0004"

_POD_LABEL_RE = re.compile(r"POD:\s*.+?\(([A-Z/;]+)\)", re.IGNORECASE)
_VALIDITY_RE = re.compile(r"(\d{1,2})\w{0,2}\s+(\w+)\s*-\s*\s*(\d{1,2})\w{0,2}\s+(\w+)\s+(\d{4})")
_INCLUDES_RE = re.compile(r"incl\.?\s*([A-Za-z,\s]+?)\s*;", re.IGNORECASE)
_THL_ORIGINS_RE = re.compile(r"incl\.\s*THL\s+for\s+cargo\s+ex\s+([A-Z/]+)", re.IGNORECASE)

_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=VALIDITY_ROW, column=VALIDITY_TEXT_COL).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, d2, mon2, year = m.groups()
    try:
        end = date(int(year), _month_number(mon2), int(d2))
        start_month = _month_number(mon1)
        return date(int(year), start_month, int(d1)), end
    except ValueError:
        return None, None


def _parse_included_charge_codes(ws: Worksheet) -> list[str]:
    """Raw text: 'Rate structure: incl. BAF, HEA, AMS, EPH, ERS, CGD, LSF,
    OBS, MBS ; ' - comma separated, alphabetized in the ground-truth note
    text (confirmed both weeks) - same convention as WAF's own
    _parse_included_charge_codes. ERS isn't a real OPUS charge code
    (absent from CHARGE_CODE_NAMES, so is_known_charge_code drops it) and
    is correctly dropped by ground truth too."""
    text = str(ws.cell(row=INCLUDES_ROW, column=INCLUDES_COL).value or "")
    m = _INCLUDES_RE.search(text)
    if not m:
        return []
    codes = [c.strip().upper() for c in m.group(1).split(",") if c.strip()]
    codes = [c for c in codes if is_known_charge_code(c)]
    return sorted(dict.fromkeys(codes))


def _parse_thl_override_origins(ws: Worksheet) -> list[str]:
    """Raw text: 'incl. THL for cargo ex LKCMB/PKKHI/BDCGP.' - confirmed
    against ground truth: 3 extra POL-scoped THL child rows, one per
    origin named here, appended after the shared charge-code list."""
    text = str(ws.cell(row=THL_OVERRIDE_ROW, column=THL_OVERRIDE_COL).value or "")
    m = _THL_ORIGINS_RE.search(text)
    if not m:
        return []
    return [c.strip() for c in m.group(1).split("/") if c.strip()]


@dataclass
class RawRateCell:
    origin_code: str
    dest_codes: tuple[str, ...]
    container_label: str
    value: float


@dataclass
class RawData:
    validity_start: date | None
    validity_end: date | None
    included_charge_codes: list[str]
    thl_override_origins: list[str]
    rate_cells: list[RawRateCell] = field(default_factory=list)


def _find_sheet(wb: Workbook) -> Worksheet | None:
    for name in wb.sheetnames:
        ws = wb[name]
        title = str(ws.cell(row=1, column=1).value or "")
        if TITLE_KEYWORD in title.upper():
            return ws
    return None


class WestAsiaWAFParser(BaseMRGParser):
    lane_id: ClassVar[str] = "WEST-ASIA-WAF"

    def __init__(self, container_map: ContainerMap | None = None, location_store: LocationBankStore | None = None):
        self.container_map = container_map or load_container_map("west_asia_waf")
        self.location_store = location_store or LocationBankStore()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        ws = _find_sheet(wb)
        if ws is None:
            return 0.0
        score = 0.5
        header_tokens = {str(ws.cell(row=CONTAINER_LABEL_ROW, column=c).value or "").strip() for c in range(MIN_COL, MAX_COL + 1)}
        if {"D2", "D4", "D5"} <= header_tokens:
            score += 0.3
        if str(ws.cell(row=POD_LABEL_ROW, column=ORIGIN_CODE_COL + 2).value or "").upper().startswith("POD:"):
            score += 0.2
        return min(score, 1.0)

    def _lookup_description(self, code: str) -> str | None:
        rec = self.location_store.get_by_code(code)
        return rec.primary_name if rec else None

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        ws = _find_sheet(wb)
        if ws is None:
            return RawExtraction(tables={"data": RawData(None, None, [], [])})

        validity_start, validity_end = _parse_validity(ws)
        included_codes = _parse_included_charge_codes(ws)
        thl_origins = _parse_thl_override_origins(ws)

        header_cols = flatten_pod_header(ws, POD_LABEL_ROW, CONTAINER_LABEL_ROW, MIN_COL, MAX_COL)

        rate_cells: list[RawRateCell] = []
        for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
            origin_cell = ws.cell(row=row_idx, column=ORIGIN_CODE_COL)
            if origin_cell.value in (None, ""):
                continue
            if is_excluded(origin_cell):
                continue
            origin_code = str(origin_cell.value).strip()
            for hc in header_cols:
                m = _POD_LABEL_RE.search(hc.pod_label)
                if not m:
                    continue
                dest_codes = tuple(sorted(c.strip() for c in m.group(1).split("/") if c.strip()))
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                rate_cells.append(RawRateCell(origin_code, dest_codes, hc.container_label, cell.value))

        return RawExtraction(
            tables={"data": RawData(validity_start, validity_end, included_codes, thl_origins, rate_cells)}
        )

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: RawData = raw.tables["data"]

        dr_description = resolve_commodity_description(DEFAULT_DR_DESCRIPTION, config)
        dr_code = resolve_commodity_code(DEFAULT_DR_DESCRIPTION, DEFAULT_DR_CODE, config)
        dg_description = resolve_commodity_description(DEFAULT_DG_DESCRIPTION, config)
        dg_code = resolve_commodity_code(DEFAULT_DG_DESCRIPTION, DEFAULT_DG_CODE, config)
        block_seq = assign_cmdt_seq_numbers([DEFAULT_DR_DESCRIPTION, DEFAULT_DG_DESCRIPTION], config.commodity_sequence_overrides)
        dr_cmdt_seq = block_seq[DEFAULT_DR_DESCRIPTION]
        dg_cmdt_seq = block_seq[DEFAULT_DG_DESCRIPTION]

        groups: dict[tuple[str, tuple[str, ...]], dict] = {}
        for rc in data.rate_cells:
            key = (rc.origin_code, rc.dest_codes)
            bucket = groups.setdefault(key, {})
            suffix = self.container_map.suffix_for(rc.container_label)
            if suffix is None:
                continue
            bucket[suffix] = rc.value

        dr_rows: list[RatesRow] = []
        for (origin_code, dest_codes), sizes in groups.items():
            origin_name = self._lookup_description(origin_code)
            dest_names = [self._lookup_description(c) for c in dest_codes]
            if origin_name is None or any(n is None for n in dest_names):
                continue
            dr_rows.append(
                RatesRow(
                    cmdt_seq=dr_cmdt_seq,
                    commodity_group_code=dr_code,
                    commodity_group_description=dr_description,
                    origin_code=origin_code,
                    origin_description=origin_name,
                    origin_term="CY",
                    destination_code=";".join(dest_codes),
                    destination_description=";".join(dest_names),
                    destination_term="CY",
                    prefix=self.container_map.prefix,
                    cgo_type=self.container_map.cgo_type,
                    cur_20="USD" if "20" in sizes else None,
                    rate_20=sizes.get("20"),
                    cur_40="USD" if "40" in sizes else None,
                    rate_40=sizes.get("40"),
                    cur_40hc="USD" if "40hc" in sizes else None,
                    rate_40hc=sizes.get("40hc"),
                )
            )

        dg_rows: list[RatesRow] = []
        if not config.skip_dg_generation.get(DEFAULT_DR_DESCRIPTION, False):
            dg_rows = [
                row.model_copy(update={"cgo_type": "DG", "commodity_group_code": dg_code, "commodity_group_description": dg_description, "cmdt_seq": dg_cmdt_seq})
                for row in dr_rows
            ]

        dr_rows = group_by_destination(dr_rows)
        dg_rows = group_by_destination(dg_rows)
        for i, row in enumerate(dr_rows, start=1):
            row.route_seq = i
        for i, row in enumerate(dg_rows, start=1):
            row.route_seq = i
        rates = [*dr_rows, *dg_rows]

        cmdt_notes: list[CmdtNoteRow] = []
        note_text_by_description: dict[str, str | None] = {}
        for description, description_key in ((dr_description, DEFAULT_DR_DESCRIPTION), (dg_description, DEFAULT_DG_DESCRIPTION)):
            if not any(r.commodity_group_description == description for r in rates):
                continue
            notes = self._build_cmdt_notes(data, config)
            for note in notes:
                note.header_seq = block_seq[description_key]
            cmdt_notes.extend(notes)
            note_text_by_description[description] = notes[0].contents if notes else None
        for row in rates:
            row.commodity_note = note_text_by_description.get(row.commodity_group_description)

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)

    def _build_cmdt_notes(self, data: RawData, config: MappingProfile) -> list[CmdtNoteRow]:
        excluded = frozenset(config.excluded_charge_codes)
        codes = [c for c in data.included_charge_codes if c not in excluded]
        thl_origins = list(data.thl_override_origins) if "THL" not in excluded else []
        if not codes or data.validity_start is None or data.validity_end is None:
            return []

        # THL appears in the main "inclusive of" sentence (always sorts
        # last among these codes anyway) but gets NO blank-origin child
        # row of its own - only the 3 origin-scoped ones below, confirmed
        # against ground truth's exact row count (8 base codes + 3 THL = 11
        # children, not 12).
        names = {**CHARGE_CODE_NAMES, "HEA": "HEAVY SURCHARGE"}
        text_codes = sorted([*codes, "THL"]) if thl_origins else codes
        names_line = " and the ".join(f"{names.get(c, c)}({c})" for c in text_codes)
        contents = "\n".join(
            [
                f"Rates are valid from {data.validity_start:%Y%m%d} to {data.validity_end:%Y%m%d}",
                f"Rates are inclusive of the {names_line}",
                "Rates are subject to all other surcharges, including those, if any, "
                "specified in the contract and those published in the Governing Tariff(s) at the time of shipment.",
            ]
        )
        parent = CmdtNoteRow(
            contents=contents, charge_seq=1, code="APP",
            application_effective=data.validity_start, application_expires=data.validity_end, application="S",
        )
        child_effective = config.rfa_effective_date if config.rfa_effective_date is not None else data.validity_start
        child_expiry = config.rfa_expiry_date if config.rfa_expiry_date is not None else data.validity_end
        children = [
            CmdtNoteRow(charge_seq=i + 2, code=c, application_effective=child_effective, application_expires=child_expiry, application="I")
            for i, c in enumerate(codes)
        ]
        children.extend(
            CmdtNoteRow(
                charge_seq=len(codes) + 2 + i, code="THL", pol=origin,
                application_effective=child_effective, application_expires=child_expiry, application="I",
            )
            for i, origin in enumerate(thl_origins)
        )
        return [parent, *children]


register(
    LayoutProfile(
        lane_id=WestAsiaWAFParser.lane_id,
        parser_cls=WestAsiaWAFParser,
        sheet_name_patterns=[r"^WAF$", r"^West Asia"],
        title_keywords=[TITLE_KEYWORD],
        header_fingerprint=["D2", "D4", "D5"],
    )
)
