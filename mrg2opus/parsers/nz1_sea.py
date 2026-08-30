"""NZ1 SEA to NZBP FAK (Southeast/South Asia origins to New Zealand
destinations) lane parser.

Single raw sheet ("ex SEA to NZ"): a flat row-per-origin table, one shared
POD group for the whole file ("Auckland (Metro Port) / Lyttelton / Napier /
Tauranga" - the "(Metro Port)" annotation drops before Location Bank
matching). Columns: D2/D4/D5 = Dry 20'/40'/40'HC; R2/R5 = Reefer 20'/40'HC
(only 2 sizes, no plain 40' reefer); R2-RAD/R5-RAD = "RAD" (Reefer
container filed as Dry cargo - Prefix R, cgo_type DR - same concept as
AUEC/AUWC's own RAD/NOR columns, same 20'/40'HC-only shape).

Origins run well past Southeast Asia proper - the raw sheet continues
through India, Pakistan, Bangladesh, Sri Lanka and a block of Middle
East/Gulf origins (UAE, Saudi Arabia, Oman, Kuwait, Jordan, Bahrain, Egypt,
Qatar) all priced "NA" - naturally excluded by the same "value must be
numeric" filter every lane's rate-cell scan already applies, so no special
handling is needed for them.

Six origin rows are grouped under an "India ICD" country label (ICD Delhi,
Ahmedabad ICD, Ludhiana Cluster, Indore Cluster, Faridabad Cluster,
Moradabad Cluster) whose own names don't exist in the Location Bank -
confirmed against both reference weeks' ground truth, these resolve
instead via their own Remark cell's "Via <port> [& <port>]" text (e.g.
"Via Nhava Sheva & Mundra" -> INMUN;INNSA), not via the cluster name
itself. Ground truth does NOT merge/dedupe these 6 rows even though several
share the identical resolved origin_code (5 of the 6 resolve to
"INMUN;INNSA") - each raw row becomes its own OPUS row at its own rate,
same "one row per raw row" behavior as every other origin here.

CMDT NOTE surcharge overrides (DOC/ISL/THL scoped to specific origins) are
hardcoded rather than parsed from each origin's own "(incl X)" annotation -
confirmed byte-identical across both reference weeks: Colombo's own
"(incl ISL/THL/DOC)" annotation maps to POL="LKCMB" (its own port code),
but Chittagong's "(Incl THL)" annotation maps to POL="BD" (Bangladesh's
country code, not "BDCGP") - a genuinely asymmetric, non-derivable
convention, same category as AUEC's own hardcoded TW/HK/KR POL scoping.

Known, deliberate scope limit: ground truth's own ROUTE NOTE sheet (rnt/
RNT - Yangon's "(Freight Collect Only)" annotation and the 6 India ICD
rows' own footnote code lists) is NOT reproduced by this parser - out of
scope for this pass; only RATES + CMDT NOTE are generated.

TIER 1 ground truth ("NZ1 SEA to NZBP TIER 1", reference folders 43/44)
looked at first glance like it needed real structural changes: its RATES
sheet has 3006 rows across 48 distinct CMDT Seq. groups for the same raw
file that a plain MappingProfile() run produces only 159 rows for, and
a handful of origins (e.g. KHPNH) carry 8 different rate values across
those groups. Both turned out to be fully explained without any parser
change:

1. The 48 groups are NOT a customer-tier schedule. Each group's own CMDT
   NOTE text ("Rates are valid from <date> to <date>") pins it to one
   specific half-month validity window - the ground truth workbook is a
   running historical+future log of every filing OPUS has ever held for
   this carrier's Oceania Tier-1 program, not just the one matching this
   MRG. Week 1's file additionally bundles in the *other* Tier-1 lane's
   own groups (NZJ NEA-to-NZ's G0005/G0006/G0007, alongside this lane's
   own G0008/G0009/G0010) - confirmed by cross-checking: the NZJ TIER 1
   ground truth (reference folders 47/48) contains exactly one of those
   NZJ blocks, byte-identical to the matching block embedded here. Week
   2's file has no such bundling and only the single current-period
   block - the accumulation isn't a fixed shape, just whatever OPUS had
   on file at export time. The correct single block to compare against
   is the one whose own CMDT NOTE validity dates equal this MRG's own
   Validity row (row 2) - see tests/test_parsers_nz1_sea.py's TIER1_PAIRS
   for exactly this filter. Once filtered down, that one block matches
   this parser's plain-MappingProfile() output field-for-field.
2. Like AUEC/AUWC's own TIER 1 ground truth, the CMDT NOTE child rows
   use a separate, longer-lived RFA effective/expiry window (e.g.
   20260510-20261231) instead of the rate validity window - supplied via
   config.rfa_effective_date/rfa_expiry_date, same mechanism as every
   other TIER 1 lane.
3. TIER 1 filings scope the Chittagong THL charge to POL="BDCGP" (its
   own port code), not FAK's own POL="BD" (Bangladesh's country code) -
   confirmed identical across both TIER 1 weeks, and genuinely different
   from FAK's own confirmed "BD". Detected via the raw title cell's own
   "- Tier 1" suffix (TIER1_TITLE_KEYWORD/_is_tier1_filing) - present on
   all 4 real reference files, absent/present exactly where expected -
   and used to pick between INCLUDED_CHARGES and its TIER 1 counterpart
   INCLUDED_CHARGES_TIER1.

One separate, genuinely unresolved gap surfaced during this check: TIER
1's 6 India ICD cluster rows (ICD Delhi, Ahmedabad ICD, Ludhiana
Cluster, Indore Cluster, Faridabad Cluster, Moradabad Cluster) resolve
to a DIFFERENT origin identity than in FAK's own ground truth. FAK
resolves them via the Remark cell's "Via <port>" text (this parser's
current, confirmed behavior - see _resolve_origin and the ICD test
below). TIER 1 instead resolves each cluster to its own distinct set of
specific inland ICD facility codes/descriptions (e.g. "Ludhiana
Cluster" -> IN1LU;IN2LU;INCWL;INLUH;INSWA, described as "LUDHIANA -
CHAWAPAIL;LUDHIANA - CONCOR ICD;..."), confirmed identical across both
TIER 1 reference weeks. The raw sheet's own bottom-of-page definition
table (rows ~82-87, "Ludhiana Cluster : IN1LU,IN2LU,...") does list
these same codes for both FAK's and TIER 1's raw files alike - but this
project's Location Bank does not contain entries for these specific
inland-facility codes (confirmed: LocationResolver.match_token on each
one fuzzy-matches to an unrelated port with needs_review=True), so
their real-world descriptions are not derivable from data this project
has access to. Rather than guess at descriptions, these 6 rows' TIER 1
identity is left unreproduced (rate value itself is correct and
consistent - 1675 in both TIER 1 weeks, matching the raw MRG exactly;
only the origin code/description differ) - see the TIER1 rates test's
own exclusion and comment for the precise, verified scope of this gap.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationResolver
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import CHARGE_CODE_NAMES
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow

SHEET_NAME = "ex SEA to NZ"
TITLE_KEYWORD = "SEA TO NZ"

ORIGIN_COUNTRY_COL, POL_COL = 1, 2
COL_D2, COL_D4, COL_D5 = 3, 4, 5
COL_R2, COL_R5 = 6, 7
COL_RAD2, COL_RAD5 = 8, 9
REMARK_COL = 10

VALIDITY_ROW, VALIDITY_COL = 2, 2
POD_ROW, POD_COL = 5, 3
HEADER_ROW = 6
DATA_MIN_ROW, DATA_MAX_ROW = 8, 80

ICD_COUNTRY_LABEL = "INDIA ICD"

DEFAULT_DESCRIPTION = "EX SEA TO NZ"
DEFAULT_CODE = "G0001"

# (code, POL scope) - see module docstring on why this is hardcoded rather
# than parsed per-origin. Order matches ground truth's own charge_seq order
# exactly (confirmed identical to alphabetical here, unlike AUEC's own
# non-alphabetical order).
INCLUDED_CHARGES: list[tuple[str, str | None]] = [
    ("DOC", "LKCMB"),
    ("EFS", None),
    ("ISL", "LKCMB"),
    ("OBS", None),
    ("THL", "LKCMB"),
    ("THL", "BD"),
]

# TIER 1 filings scope the Chittagong THL charge to the port code
# "BDCGP" instead of FAK's own country code "BD" - confirmed identical
# across both TIER 1 reference weeks (folders 43/44), byte-different from
# FAK's own confirmed "BD" (folders 41/42). Same non-derivable, per-filing
# convention category as the FAK/TIER1 India ICD origin difference (see
# module docstring); the raw MRG gives no signal for which POL grain a
# given filing type expects, so it's selected by filing type instead.
INCLUDED_CHARGES_TIER1: list[tuple[str, str | None]] = [
    ("DOC", "LKCMB"),
    ("EFS", None),
    ("ISL", "LKCMB"),
    ("OBS", None),
    ("THL", "LKCMB"),
    ("THL", "BDCGP"),
]

# "ONE Minimum Rate Guideline SEA to NZ - Tier 1" (TIER 1 filings) vs
# "ONE Minimum Rate Guideline SEA to NZ" (FAK filings) - confirmed
# present/absent identically across all 4 real reference files.
TIER1_TITLE_KEYWORD = "TIER 1"

_VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s+(\d{4})\s+to\s+(\d{1,2})\s+(\w+)\s+(\d{4})", re.IGNORECASE)
_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
_TRAILING_PAREN_RE = re.compile(r"\s*\([^)]*\)\s*$")
_METRO_PORT_RE = re.compile(r"\s*\(Metro Port\)\s*", re.IGNORECASE)
_ICD_VIA_RE = re.compile(r"via\s+(.+)", re.IGNORECASE)
# "Batam (Door/CY)" - confirmed against ground truth: the only origin_term
# override in either reference week (every other origin, including every
# other parenthetical-annotated one, stays the "CY" default).
_ORIGIN_TERM_RE = re.compile(r"\(([A-Za-z]+)/CY\)\s*$", re.IGNORECASE)


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=VALIDITY_ROW, column=VALIDITY_COL).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, y1, d2, mon2, y2 = m.groups()
    try:
        return date(int(y1), _month_number(mon1), int(d1)), date(int(y2), _month_number(mon2), int(d2))
    except ValueError:
        return None, None


def _find_sheet(wb: Workbook) -> Worksheet | None:
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    for name in wb.sheetnames:
        ws = wb[name]
        title = str(ws.cell(row=1, column=1).value or "")
        if TITLE_KEYWORD in title.upper():
            return ws
    return None


def _is_tier1_filing(ws: Worksheet) -> bool:
    title = str(ws.cell(row=1, column=1).value or "")
    return TIER1_TITLE_KEYWORD in title.upper()


@dataclass
class OriginRow:
    origin_text: str
    is_icd: bool
    remark_text: str
    values: dict[int, float] = field(default_factory=dict)


@dataclass
class RawData:
    validity_start: date | None
    validity_end: date | None
    destination_text: str
    is_tier1: bool = False
    origins: list[OriginRow] = field(default_factory=list)


class Nz1SeaParser(BaseMRGParser):
    lane_id: ClassVar[str] = "NZ1-SEA"

    def __init__(self, location_resolver: LocationResolver | None = None):
        self.location_resolver = location_resolver or LocationResolver()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        ws = _find_sheet(wb)
        if ws is None:
            return 0.0
        score = 0.5
        header_tokens = {str(ws.cell(row=HEADER_ROW, column=c).value or "").strip().upper() for c in range(1, 11)}
        if {"D2", "D4", "D5", "R2", "R5"} <= header_tokens:
            score += 0.3
        if any("RAD" in t for t in header_tokens):
            score += 0.2
        return min(score, 1.0)

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        ws = _find_sheet(wb)
        if ws is None:
            return RawExtraction(tables={"data": RawData(None, None, "")})

        validity_start, validity_end = _parse_validity(ws)
        destination_text = str(ws.cell(row=POD_ROW, column=POD_COL).value or "")
        is_tier1 = _is_tier1_filing(ws)

        origins: list[OriginRow] = []
        current_country = ""
        for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
            country_cell = ws.cell(row=row_idx, column=ORIGIN_COUNTRY_COL)
            if country_cell.value not in (None, ""):
                current_country = str(country_cell.value).strip()
            pol_cell = ws.cell(row=row_idx, column=POL_COL)
            if pol_cell.value in (None, ""):
                continue
            if is_excluded(pol_cell):
                continue

            values: dict[int, float] = {}
            for col in (COL_D2, COL_D4, COL_D5, COL_R2, COL_R5, COL_RAD2, COL_RAD5):
                cell = ws.cell(row=row_idx, column=col)
                if isinstance(cell.value, (int, float)) and not is_excluded(cell):
                    values[col] = cell.value
            if not values:
                continue

            remark_text = str(ws.cell(row=row_idx, column=REMARK_COL).value or "")
            origins.append(
                OriginRow(
                    origin_text=str(pol_cell.value).strip(),
                    is_icd=(current_country.upper() == ICD_COUNTRY_LABEL),
                    remark_text=remark_text,
                    values=values,
                )
            )

        return RawExtraction(
            tables={"data": RawData(validity_start, validity_end, destination_text, is_tier1, origins)}
        )

    def _resolve_origin(self, origin: OriginRow) -> tuple[str, str, str] | None:
        """Returns (origin_code, origin_description, origin_term) or None if
        unresolved."""
        origin_term = "CY"
        if origin.is_icd:
            m = _ICD_VIA_RE.search(origin.remark_text)
            if not m:
                return None
            tokens = re.split(r"\s*&\s*|\s*,\s*", m.group(1).strip())
            matches = [self.location_resolver.match_token(t) for t in tokens if t.strip()]
            matches = [mm for mm in matches if mm is not None]
            if not matches or any(mm.needs_review for mm in matches):
                return None
        else:
            term_match = _ORIGIN_TERM_RE.search(origin.origin_text)
            if term_match:
                origin_term = term_match.group(1).title()
            clean = _TRAILING_PAREN_RE.sub("", origin.origin_text).strip()
            matches = self.location_resolver.match_text(clean)
            if not matches or any(mm.needs_review for mm in matches):
                return None

        code = ";".join(sorted({mm.code for mm in matches}))
        desc = ";".join(sorted({mm.primary_name for mm in matches}))
        return code, desc, origin_term

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: RawData = raw.tables["data"]

        description = resolve_commodity_description(DEFAULT_DESCRIPTION, config)
        code = resolve_commodity_code(DEFAULT_DESCRIPTION, DEFAULT_CODE, config)
        # One CMDT NOTE block covers DR/RF/RAD/DG all together here (a
        # constant, not something to auto-number across multiple blocks).
        cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_DESCRIPTION, 1)

        dest_matches = self.location_resolver.match_text(_METRO_PORT_RE.sub(" ", data.destination_text))
        if not dest_matches or any(m.needs_review for m in dest_matches):
            return OpusRowSet(rates=[], cmdt_notes=[])
        dest_code = ";".join(sorted({m.code for m in dest_matches}))
        dest_desc = ";".join(sorted({m.primary_name for m in dest_matches}))

        dr_rows: list[RatesRow] = []
        rf_rows: list[RatesRow] = []
        rad_rows: list[RatesRow] = []

        for origin in data.origins:
            resolved = self._resolve_origin(origin)
            if resolved is None:
                continue
            origin_code, origin_desc, origin_term = resolved
            v = origin.values

            base = dict(
                commodity_group_code=code,
                commodity_group_description=description,
                cmdt_seq=cmdt_seq,
                origin_code=origin_code,
                origin_description=origin_desc,
                origin_term=origin_term,
                destination_code=dest_code,
                destination_description=dest_desc,
                destination_term="CY",
            )

            if COL_D2 in v or COL_D4 in v or COL_D5 in v:
                dr_rows.append(
                    RatesRow(
                        **base, prefix="D", cgo_type="DR",
                        cur_20="USD" if COL_D2 in v else None, rate_20=v.get(COL_D2),
                        cur_40="USD" if COL_D4 in v else None, rate_40=v.get(COL_D4),
                        cur_40hc="USD" if COL_D5 in v else None, rate_40hc=v.get(COL_D5),
                    )
                )
            if COL_R2 in v or COL_R5 in v:
                rf_rows.append(
                    RatesRow(
                        **base, prefix="R", cgo_type="RF",
                        cur_20="USD" if COL_R2 in v else None, rate_20=v.get(COL_R2),
                        cur_40hc="USD" if COL_R5 in v else None, rate_40hc=v.get(COL_R5),
                    )
                )
            if COL_RAD2 in v or COL_RAD5 in v:
                rad_rows.append(
                    RatesRow(
                        **base, prefix="R", cgo_type="DR",
                        cur_20="USD" if COL_RAD2 in v else None, rate_20=v.get(COL_RAD2),
                        cur_40hc="USD" if COL_RAD5 in v else None, rate_40hc=v.get(COL_RAD5),
                    )
                )

        non_dg = group_by_destination([*dr_rows, *rf_rows, *rad_rows])
        dg_rows: list[RatesRow] = []
        if not config.skip_dg_generation.get(DEFAULT_DESCRIPTION, False):
            dg_rows = group_by_destination(
                [row.model_copy(update={"cgo_type": "DG"}) for row in dr_rows]
            )

        rates = [*non_dg, *dg_rows]
        for i, row in enumerate(rates, start=1):
            row.route_seq = i

        cmdt_notes = self._build_cmdt_notes(data, config)
        for note in cmdt_notes:
            note.header_seq = cmdt_seq
        note_text = cmdt_notes[0].contents if cmdt_notes else None
        for row in rates:
            row.commodity_note = note_text

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)

    def _build_cmdt_notes(self, data: RawData, config: MappingProfile) -> list[CmdtNoteRow]:
        if data.validity_start is None or data.validity_end is None:
            return []
        excluded = frozenset(config.excluded_charge_codes)
        included_charges = INCLUDED_CHARGES_TIER1 if data.is_tier1 else INCLUDED_CHARGES
        charges = [(c, pol) for c, pol in included_charges if c not in excluded]
        if not charges:
            return []

        unique_codes = sorted(dict.fromkeys(c for c, _ in charges))
        names_line = " and the ".join(f"{CHARGE_CODE_NAMES.get(c, c)}({c})" for c in unique_codes)
        contents = "\n".join(
            [
                f"Rates are valid from {data.validity_start:%Y%m%d} to {data.validity_end:%Y%m%d}",
                f"Rates are inclusive of the {names_line}",
                "Rates are subject to all other surcharges, including those, if any, "
                "specified in the contract and those published in the Governing Tariff(s) at the time of shipment.",
            ]
        )
        parent = CmdtNoteRow(
            contents=contents, charge_seq=1, code="APP",
            application_effective=data.validity_start, application_expires=data.validity_end, application="S",
        )
        child_effective = config.rfa_effective_date if config.rfa_effective_date is not None else data.validity_start
        child_expiry = config.rfa_expiry_date if config.rfa_expiry_date is not None else data.validity_end
        children = [
            CmdtNoteRow(
                charge_seq=i + 2, code=c, pol=pol,
                application_effective=child_effective, application_expires=child_expiry, application="I",
            )
            for i, (c, pol) in enumerate(charges)
        ]
        return [parent, *children]


register(
    LayoutProfile(
        lane_id=Nz1SeaParser.lane_id,
        parser_cls=Nz1SeaParser,
        sheet_name_patterns=[re.escape(SHEET_NAME)],
        title_keywords=[TITLE_KEYWORD],
        header_fingerprint=["D2", "D4", "D5", "RAD"],
    )
)
