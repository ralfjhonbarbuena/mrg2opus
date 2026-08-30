"""West Africa WAF lane parser.

Single raw sheet: a 9-POD x 3-container-size (D2/D4/D5) rate grid, ex a
fixed list of ~41 Asia/SEA origins. No ARBS/SPECIAL NOTE/ROUTE NOTE scope
(confirmed against reference/2_OPUS/9_West Africa WAF and 10_West Africa
WAF - both real ground-truth files carry only RATES + SRCHG, SRCHG being
the CMDT-NOTE naming drift, see project-opus-note-sheet-taxonomy memory).

Every DR row also files an identical D/DG duplicate at the same rate,
under its own commodity group ("<desc> - DG") - the raw sheet's separate
HAZ/PSA per-IMO-class add-on tables (rows 57+) are NOT reflected anywhere
in either ground-truth file (no ARBS, no extra charge codes for them in
SRCHG), so they're out of this filing's scope entirely and deliberately
not parsed.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationResolver
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import (
    CommodityNoteSpec,
    build_notes_by_description,
    resolve_commodity_code,
    resolve_commodity_description,
)
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.header_grid import flatten_pod_header
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.common.sequencing import assign_cmdt_seq_numbers
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import is_known_charge_code
from mrg2opus.schema.opus_rows import OpusRowSet, RatesRow

SHEET_NAME_RE = re.compile(r"^Asia WAF MRG", re.IGNORECASE)
TITLE_KEYWORD = "ASIA WAF MRG"

POD_LABEL_ROW = 5
CONTAINER_LABEL_ROW = 6
DATA_MIN_ROW = 7
DATA_MAX_ROW = 50
MIN_COL, MAX_COL = 4, 38
ORIGIN_TEXT_COL = 2

DEFAULT_DR_DESCRIPTION = "West Africa MRG"
DEFAULT_DR_CODE = "G0001"
DEFAULT_DG_DESCRIPTION = "West Africa MRG - DG"
DEFAULT_DG_CODE = "G0002"

_VIA_PAREN_RE = re.compile(r"\s*\(via\s+([^)]+)\)", re.IGNORECASE)
_POD_LABEL_RE = re.compile(r"POD:\s*(.+?)\s+(\S+)$", re.IGNORECASE)
_VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s+(\d{4})\s*-\s*(\d{1,2})\s+(\w+)\s+(\d{4})")
_INCLUDES_RE = re.compile(r"Incl\.\s*([A-Za-z,\s]+);")

_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _split_via(origin_text: str) -> tuple[str, str | None]:
    """'Ganzhou (via Shekou)' -> ('Ganzhou', 'Shekou'). Confirmed against
    ground truth: the via-port becomes O.Via, the paren clause is dropped
    from the origin name itself rather than resolved as a 2nd location."""
    m = _VIA_PAREN_RE.search(origin_text)
    via = m.group(1).strip() if m else None
    clean = _VIA_PAREN_RE.sub("", origin_text).strip()
    return clean, via


@dataclass
class RawRateCell:
    origin_text: str
    pod_name: str
    pod_term: str
    container_label: str
    value: float


@dataclass
class WAFRawData:
    validity_start: date | None
    validity_end: date | None
    included_charge_codes: list[str]
    rate_cells: list[RawRateCell] = field(default_factory=list)


def _find_sheet(wb: Workbook) -> Worksheet | None:
    for name in wb.sheetnames:
        if SHEET_NAME_RE.search(name):
            return wb[name]
    return None


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=1, column=5).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, y1, d2, mon2, y2 = m.groups()
    try:
        start = date(int(y1), _month_number(mon1), int(d1))
        end = date(int(y2), _month_number(mon2), int(d2))
    except ValueError:
        return None, None
    return start, end


def _parse_included_charge_codes(ws: Worksheet) -> list[str]:
    """Raw F2: 'Incl. BAF, HEA, EPH, BRS,LSF, CGD, OBS, MBS, EFS;' - comma
    separated (not the '/'-separated convention parse_included_charge_codes
    handles), and the ground truth lists the survivors alphabetically for
    both the note text AND each child row's order - confirmed against both
    reference/2_OPUS/9 and 10's SRCHG sheets (charge_seq 2..9 = BAF, CGD,
    EFS, EPH, HEA, LSF, MBS, OBS), so sort here rather than relying on
    build_cmdt_notes' text-only sort_text_names."""
    text = str(ws.cell(row=2, column=6).value or "")
    m = _INCLUDES_RE.search(text)
    if not m:
        return []
    codes = [c.strip().upper() for c in m.group(1).split(",") if c.strip()]
    codes = [c for c in codes if is_known_charge_code(c)]
    return sorted(dict.fromkeys(codes))


def _parse_pod_label(label: str) -> tuple[str, str]:
    m = _POD_LABEL_RE.match(label)
    if not m:
        return label, "CY"
    return m.group(1).strip(), m.group(2).strip()


def _assign_route_seq(rows: list[RatesRow]) -> None:
    """Route Seq. is a single running counter across the WHOLE group (1..369
    covering all 9 PODs x 41 origins, NOT resetting per destination),
    restarting at 1 only for the next commodity group (DR vs DG) -
    confirmed against both reference/2_OPUS/9 and 10 (e.g. NGTIN's block
    continues 42, 43, ... rather than restarting at 1)."""
    for i, row in enumerate(rows, start=1):
        row.route_seq = i


def _clean_description(name: str) -> str:
    """The Location Bank's primary_name uses 'CITY, SUBDIVISION' (comma),
    mined from other lanes' ground truth - but West Africa WAF's own
    ground truth spells every multi-part name 'CITY  SUBDIVISION' (double
    space, no comma) instead, confirmed with zero exceptions across all 37
    distinct origin/destination descriptions in both reference weeks."""
    return name.replace(", ", "  ")


class WAFParser(BaseMRGParser):
    lane_id: ClassVar[str] = "WAF"

    def __init__(self, container_map: ContainerMap | None = None, location_resolver: LocationResolver | None = None):
        self.container_map = container_map or load_container_map("waf")
        self.location_resolver = location_resolver or LocationResolver()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        ws = _find_sheet(wb)
        if ws is None:
            return 0.0
        score = 0.5
        header_tokens = {ws.cell(row=CONTAINER_LABEL_ROW, column=c).value for c in range(MIN_COL, MAX_COL + 1)}
        if {"D2", "D4", "D5"} <= {str(t).strip() for t in header_tokens if t}:
            score += 0.3
        title = str(ws.cell(row=1, column=1).value or "")
        if "WAF" in title.upper():
            score += 0.2
        return min(score, 1.0)

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        ws = _find_sheet(wb)
        if ws is None:
            return RawExtraction(tables={"waf": WAFRawData(None, None, [])})

        validity_start, validity_end = _parse_validity(ws)
        included_codes = _parse_included_charge_codes(ws)

        header_cols = flatten_pod_header(ws, POD_LABEL_ROW, CONTAINER_LABEL_ROW, MIN_COL, MAX_COL)

        rate_cells: list[RawRateCell] = []
        for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
            origin_cell = ws.cell(row=row_idx, column=ORIGIN_TEXT_COL)
            if origin_cell.value in (None, ""):
                continue
            if is_excluded(origin_cell):
                continue
            origin_text = str(origin_cell.value).strip()
            for hc in header_cols:
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, ""):
                    continue
                if not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                pod_name, pod_term = _parse_pod_label(hc.pod_label)
                rate_cells.append(RawRateCell(origin_text, pod_name, pod_term, hc.container_label, cell.value))

        return RawExtraction(
            tables={"waf": WAFRawData(validity_start, validity_end, included_codes, rate_cells)}
        )

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: WAFRawData = raw.tables["waf"]

        dr_description = resolve_commodity_description(DEFAULT_DR_DESCRIPTION, config)
        dr_code = resolve_commodity_code(DEFAULT_DR_DESCRIPTION, DEFAULT_DR_CODE, config)

        dg_description = resolve_commodity_description(DEFAULT_DG_DESCRIPTION, config)
        dg_code = resolve_commodity_code(DEFAULT_DG_DESCRIPTION, DEFAULT_DG_CODE, config)

        # DR and DG are two distinct CMDT NOTE blocks under the same lane -
        # auto-numbered 1, 2 (or an explicit override), never left blank.
        block_seq = assign_cmdt_seq_numbers([DEFAULT_DR_DESCRIPTION, DEFAULT_DG_DESCRIPTION], config.commodity_sequence_overrides)
        dr_cmdt_seq = block_seq[DEFAULT_DR_DESCRIPTION]
        dg_cmdt_seq = block_seq[DEFAULT_DG_DESCRIPTION]

        # Group raw cells by (origin_text, pod_name) so one OPUS row covers
        # all 3 container sizes for a given origin/destination pair.
        groups: dict[tuple[str, str], dict] = {}
        for rc in data.rate_cells:
            key = (rc.origin_text, rc.pod_name)
            bucket = groups.setdefault(key, {"sizes": {}, "term": rc.pod_term})
            suffix = self.container_map.suffix_for(rc.container_label)
            if suffix is None:
                continue
            bucket["sizes"][suffix] = rc.value

        dr_rows: list[RatesRow] = []
        dg_rows: list[RatesRow] = []

        for (origin_text, pod_name), bucket in groups.items():
            sizes = bucket["sizes"]
            term = bucket["term"]

            clean_origin, via_text = _split_via(origin_text)
            origin_matches = self.location_resolver.match_text(clean_origin)
            dest_matches = self.location_resolver.match_text(pod_name)
            if not origin_matches or not dest_matches:
                continue
            # A low-confidence fuzzy hit is worse than no match at all (see
            # saf.py's identical rationale) - skip the whole row rather
            # than file a partial/wrong-port group.
            if any(m.needs_review for m in origin_matches) or any(m.needs_review for m in dest_matches):
                continue

            origin_codes = sorted({m.code for m in origin_matches})
            origin_names = sorted({_clean_description(m.primary_name) for m in origin_matches})
            dest_codes = sorted({m.code for m in dest_matches})
            dest_names = sorted({_clean_description(m.primary_name) for m in dest_matches})

            o_via_code = None
            if via_text:
                via_match = self.location_resolver.match_token(via_text)
                if via_match and not via_match.needs_review:
                    o_via_code = via_match.code

            dr_row = RatesRow(
                cmdt_seq=dr_cmdt_seq,
                commodity_group_code=dr_code,
                commodity_group_description=dr_description,
                origin_code=";".join(origin_codes),
                origin_description=";".join(origin_names),
                origin_term="CY",
                o_via_code=o_via_code,
                destination_code=";".join(dest_codes),
                destination_description=";".join(dest_names),
                destination_term=term,
                prefix=self.container_map.prefix,
                cgo_type=self.container_map.cgo_type,
                cur_20="USD" if "20" in sizes else None,
                rate_20=sizes.get("20"),
                cur_40="USD" if "40" in sizes else None,
                rate_40=sizes.get("40"),
                cur_40hc="USD" if "40hc" in sizes else None,
                rate_40hc=sizes.get("40hc"),
            )
            dr_rows.append(dr_row)

            # Every base Dry (D/DR) row also files an identical D/DG variant
            # at the same rate, under its own commodity group - confirmed
            # against both reference/2_OPUS/9 and 10 (100% consistent: same
            # rate, only CGO TYPE + commodity group flip DR->DG).
            if not config.skip_dg_generation.get(DEFAULT_DR_DESCRIPTION, False):
                dg_row = dr_row.model_copy(
                    update={
                        "cgo_type": "DG",
                        "commodity_group_code": dg_code,
                        "commodity_group_description": dg_description,
                        "cmdt_seq": dg_cmdt_seq,
                    }
                )
                dg_rows.append(dg_row)

        dr_rows = group_by_destination(dr_rows)
        dg_rows = group_by_destination(dg_rows)
        _assign_route_seq(dr_rows)
        _assign_route_seq(dg_rows)
        rates = [*dr_rows, *dg_rows]

        note_specs: list[CommodityNoteSpec] = []
        if dr_rows:
            note_specs.append(
                CommodityNoteSpec(dr_description, data.validity_start, data.validity_end, data.included_charge_codes)
            )
        if dg_rows:
            note_specs.append(
                CommodityNoteSpec(dg_description, data.validity_start, data.validity_end, data.included_charge_codes)
            )

        cmdt_notes, note_text_by_description = build_notes_by_description(
            note_specs,
            sequential_charge_seq=True,
            excluded_codes=frozenset(config.excluded_charge_codes),
            rfa_effective=config.rfa_effective_date,
            rfa_expiry=config.rfa_expiry_date,
        )
        for row in rates:
            row.commodity_note = note_text_by_description.get(row.commodity_group_description)
        for note in cmdt_notes:
            note.header_seq = block_seq.get(note.group_description)

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)


register(
    LayoutProfile(
        lane_id=WAFParser.lane_id,
        parser_cls=WAFParser,
        sheet_name_patterns=[r"^Asia WAF MRG"],
        title_keywords=[TITLE_KEYWORD],
        header_fingerprint=["D2", "D4", "D5"],
    )
)
