"""AUS NEA to AUWC (ex North/Central-East Asia to Australia West Coast)
FAK lane parser.

Two raw sheets, "ex NEA to AUFRE" (Fremantle) and "ex NEA to AUADL "
(Adelaide), sharing the exact same 49-origin grid layout (rows 6-54,
footnote city-group definitions at rows 56-57 - byte-identical PRDA/PRDB
regional groupings to AUEC's, reused from there rather than
re-transcribed) and feeding the SAME main commodity group (confirmed:
both destinations share one CMDT NOTE block in ground truth, same
pattern as AUEC's AUBNE_AUMEL+AUSYD split).

Structurally different from AUEC in two ways that matter:
  - Reefer (RF) and NOR ("Non-Operating Reefer", filed at Prefix R / CGO
    TYPE DR - same concept as AUEC's "RAD") each get their OWN SEPARATE
    commodity group here ("FAK - NEA to AUWC (RF)" / "(NOR)"), NOT folded
    into the main group the way AUEC's RF/RAD rows are. Confirmed via
    ground truth: 3 distinct CMDT Seq/commodity-group blocks, not 2.
  - CMDT NOTE child rows carry NO per-code POL scope here (every POL cell
    is blank) - same hardcoded charge/duplicate-EFS list as AUEC
    (INCLUDED_CHARGE_CODES below), but plain, so this lane uses the
    shared cmdt_notes.py::build_notes_by_description() helper directly
    rather than needing AUEC's lane-specific POL-aware builder.

Output scope is RATES + CMDT NOTE only (ground truth's own sheet is
named "SRCHG" - the known naming drift), confirmed against the real
reference week (33; also present for 34/35/36 but not separately
verified here).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationMatch, LocationResolver
from mrg2opus.parsers.auec import PRDA_CODE, PRDA_DESCRIPTION, PRDB_CODE, PRDB_DESCRIPTION
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import CHARGE_CODE_NAMES
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow

SHEET_AUFRE = "ex NEA to AUFRE"
SHEET_AUADL = "ex NEA to AUADL"

DATA_MIN_ROW, DATA_MAX_ROW = 6, 54
ORIGIN_COL, DEST_COL = 2, 3
VALIDITY_ROW, VALIDITY_COL = 3, 2

# Column -> RatesRow rate slot. RF populates 20'/40'HC (no plain 40');
# NOR only ever populates 20' - its own "40'HC NOR" column is blank in
# every real row seen (confirmed both origins' sheets) - not a bug, just
# this container type's own data shape here.
COL_DRY_20, COL_DRY_40, COL_DRY_40HC = 5, 6, 7
COL_RF_20, COL_RF_40HC = 8, 9
COL_NOR_20, COL_NOR_40HC = 10, 11

DEFAULT_MAIN_DESCRIPTION = "FAK - NEA to AUWC"
DEFAULT_MAIN_CODE = "G0001"
DEFAULT_RF_DESCRIPTION = "FAK - NEA to AUWC (RF)"
DEFAULT_RF_CODE = "G0002"
DEFAULT_NOR_DESCRIPTION = "FAK - NEA to AUWC (NOR)"
DEFAULT_NOR_CODE = "G0003"

# Hardcoded rather than parsed from the raw sheet's own "Remarks:" text -
# same reasoning as auec.py::INCLUDED_CHARGES (this lane's remarks text
# is near-identical: "Rate incl OBS, EFS (ex KR & HK), PSS, subject to
# ISL/..."). Confirmed against reference/2_OPUS/33's SRCHG sheet: this
# lane scopes the SAME way AUEC does (ISL->Taiwan, EFS split into Korea
# and Hong Kong) but stamps it on the **POR** column, not POL like AUEC -
# same business meaning, different OPUS field per lane. Every POL cell
# here is blank.
INCLUDED_CHARGES: list[tuple[str, str | None]] = [
    ("OBS", None),
    ("ISL", "TW"),
    ("EFS", "KR"),
    ("EFS", "HK"),
    ("PSS", None),
]

_PRD_GROUPS = {"PRDA": (PRDA_CODE, PRDA_DESCRIPTION), "PRDB": (PRDB_CODE, PRDB_DESCRIPTION)}

_VIA_PAREN_RE = re.compile(r"\s*\(via\s+([^)]+?)(\s+by\s+rail)?\)", re.IGNORECASE)
_VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s+to\s+(\d{1,2})\s+(\w+)\s+(\d{4})", re.IGNORECASE)
_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=VALIDITY_ROW, column=VALIDITY_COL).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, d2, mon2, year = m.groups()
    try:
        y = int(year)
        start = date(y, _month_number(mon1), int(d1))
        end = date(y, _month_number(mon2), int(d2))
    except ValueError:
        return None, None
    return start, end


def _split_via(origin_text: str) -> tuple[str, str | None, bool]:
    """See auec.py::_split_via - identical raw-text convention in this
    lane's own origin column ('Chengdu (via CNYTN by rail)', etc.)."""
    m = _VIA_PAREN_RE.search(origin_text)
    if not m:
        return origin_text.strip(), None, False
    via = m.group(1).strip()
    by_rail = bool(m.group(2))
    clean = _VIA_PAREN_RE.sub("", origin_text).strip()
    return clean, via, by_rail


def _build_group_cmdt_notes(
    validity_start: date | None,
    validity_end: date | None,
    excluded_codes: frozenset[str],
    rfa_effective: date | None,
    rfa_expiry: date | None,
) -> list[CmdtNoteRow]:
    """Lane-specific CMDT NOTE builder (not the shared cmdt_notes.py::
    build_cmdt_notes - see INCLUDED_CHARGES above for why: this lane's
    ISL/EFS rows need a per-child-row POR scope the shared helper has no
    concept of). Mirrors auec.py::_build_group_cmdt_notes exactly except
    for stamping `por` instead of `pol`."""
    if validity_start is None or validity_end is None:
        return []
    charges = [(code, por) for code, por in INCLUDED_CHARGES if code not in excluded_codes]
    if not charges:
        return []

    unique_codes = sorted(dict.fromkeys(code for code, _ in charges))
    names_line = " and the ".join(f"{CHARGE_CODE_NAMES.get(code, code)}({code})" for code in unique_codes)
    contents = "\n".join(
        [
            f"Rates are valid from {validity_start:%Y%m%d} to {validity_end:%Y%m%d}",
            f"Rates are inclusive of the {names_line}",
            "Rates are subject to all other surcharges, including those, if any, specified in "
            "the contract and those published in the Governing Tariff(s) at the time of shipment.",
        ]
    )
    parent = CmdtNoteRow(
        contents=contents, charge_seq=1, code="APP",
        application_effective=validity_start, application_expires=validity_end, application="S",
    )
    child_effective = rfa_effective if rfa_effective is not None else validity_start
    child_expires = rfa_expiry if rfa_expiry is not None else validity_end
    children = [
        CmdtNoteRow(
            charge_seq=i + 2, code=code, por=por,
            application_effective=child_effective, application_expires=child_expires, application="I",
        )
        for i, (code, por) in enumerate(charges)
    ]
    return [parent, *children]


@dataclass
class OriginRow:
    origin_text: str
    destination_text: str
    values: dict[int, float]


@dataclass
class AUWCRawData:
    validity_start: date | None
    validity_end: date | None
    origins: list[OriginRow] = field(default_factory=list)


def _read_sheet(ws: Worksheet) -> list[OriginRow]:
    rows: list[OriginRow] = []
    for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
        origin_cell = ws.cell(row=row_idx, column=ORIGIN_COL)
        if origin_cell.value in (None, ""):
            continue
        if is_excluded(origin_cell):
            continue
        dest_cell = ws.cell(row=row_idx, column=DEST_COL)
        if dest_cell.value in (None, ""):
            continue
        values: dict[int, float] = {}
        for col in (COL_DRY_20, COL_DRY_40, COL_DRY_40HC, COL_RF_20, COL_RF_40HC, COL_NOR_20, COL_NOR_40HC):
            v = ws.cell(row=row_idx, column=col).value
            if isinstance(v, (int, float)):
                values[col] = v
        if not values:
            continue
        rows.append(
            OriginRow(origin_text=str(origin_cell.value).strip(), destination_text=str(dest_cell.value).strip(), values=values)
        )
    return rows


class AUWCParser(BaseMRGParser):
    lane_id: ClassVar[str] = "AUWC"

    def __init__(self, location_resolver: LocationResolver | None = None):
        self.location_resolver = location_resolver or LocationResolver()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        names = set(wb.sheetnames)
        score = 0.0
        for expected in (SHEET_AUFRE, SHEET_AUADL):
            if any(expected in n for n in names):
                score += 0.5
        if score == 0.0:
            return 0.0
        first_sheet = wb.sheetnames[0]
        ws = wb[first_sheet]
        header_tokens = {str(ws.cell(row=4, column=c).value or "").strip() for c in range(4, 12)}
        if any("NOR" in t for t in header_tokens):
            score = min(score + 0.2, 1.0)
        return score

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        validity_start, validity_end = None, None
        origins: list[OriginRow] = []
        for sheet_name in wb.sheetnames:
            if SHEET_AUFRE not in sheet_name and SHEET_AUADL not in sheet_name:
                continue
            ws = wb[sheet_name]
            if validity_start is None:
                validity_start, validity_end = _parse_validity(ws)
            origins.extend(_read_sheet(ws))

        return RawExtraction(
            tables={"auwc": AUWCRawData(validity_start=validity_start, validity_end=validity_end, origins=origins)}
        )

    def _resolve_origin(self, origin_text: str) -> tuple[str, str, str | None, bool] | None:
        """See auec.py::AUECParser._resolve_origin - identical logic
        (whole-text-first, then PRDA/PRDB, then comma-split fallback)."""
        clean, via, by_rail = _split_via(origin_text)

        prd_key = clean.rstrip("*").strip().upper()
        if prd_key in _PRD_GROUPS:
            code, desc = _PRD_GROUPS[prd_key]
        else:
            whole = self.location_resolver.match_token(clean)
            matches: list[LocationMatch]
            if whole is not None and not whole.needs_review:
                matches = [whole]
            else:
                matches = self.location_resolver.match_text(clean)
            if not matches or any(m.needs_review for m in matches):
                return None
            code = ";".join(sorted({m.code for m in matches}))
            desc = ";".join(sorted({m.primary_name for m in matches}))

        via_code = None
        if via:
            via_match = self.location_resolver.match_token(via)
            if via_match and not via_match.needs_review:
                via_code = via_match.code

        return code, desc, via_code, by_rail

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: AUWCRawData = raw.tables["auwc"]

        main_description = resolve_commodity_description(DEFAULT_MAIN_DESCRIPTION, config)
        main_code = resolve_commodity_code(DEFAULT_MAIN_DESCRIPTION, DEFAULT_MAIN_CODE, config)
        main_cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_MAIN_DESCRIPTION)

        rf_description = resolve_commodity_description(DEFAULT_RF_DESCRIPTION, config)
        rf_code = resolve_commodity_code(DEFAULT_RF_DESCRIPTION, DEFAULT_RF_CODE, config)
        rf_cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_RF_DESCRIPTION)

        nor_description = resolve_commodity_description(DEFAULT_NOR_DESCRIPTION, config)
        nor_code = resolve_commodity_code(DEFAULT_NOR_DESCRIPTION, DEFAULT_NOR_CODE, config)
        nor_cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_NOR_DESCRIPTION)

        dr_rows: list[RatesRow] = []
        rf_rows: list[RatesRow] = []
        nor_rows: list[RatesRow] = []

        dest_cache: dict[str, tuple[str, str] | None] = {}

        for origin in data.origins:
            resolved = self._resolve_origin(origin.origin_text)
            if resolved is None:
                continue
            origin_code, origin_desc, via_code, by_rail = resolved

            if origin.destination_text not in dest_cache:
                dest_matches = self.location_resolver.match_text(origin.destination_text)
                if not dest_matches or any(m.needs_review for m in dest_matches):
                    dest_cache[origin.destination_text] = None
                else:
                    dest_cache[origin.destination_text] = (
                        ";".join(sorted({m.code for m in dest_matches})),
                        ";".join(sorted({m.primary_name for m in dest_matches})),
                    )
            dest = dest_cache[origin.destination_text]
            if dest is None:
                continue
            dest_code, dest_desc = dest

            v = origin.values
            base = dict(
                commodity_group_code="",
                commodity_group_description="",
                origin_code=origin_code,
                origin_description=origin_desc,
                origin_term="CY",
                origin_transmode="Rail" if by_rail else None,
                o_via_code=via_code,
                destination_code=dest_code,
                destination_description=dest_desc,
                destination_term="CY",
            )

            if COL_DRY_20 in v or COL_DRY_40 in v or COL_DRY_40HC in v:
                dr_rows.append(
                    RatesRow(
                        **base,
                        prefix="D",
                        cgo_type="DR",
                        cur_20="USD" if COL_DRY_20 in v else None,
                        rate_20=v.get(COL_DRY_20),
                        cur_40="USD" if COL_DRY_40 in v else None,
                        rate_40=v.get(COL_DRY_40),
                        cur_40hc="USD" if COL_DRY_40HC in v else None,
                        rate_40hc=v.get(COL_DRY_40HC),
                    )
                )
            if COL_RF_20 in v or COL_RF_40HC in v:
                rf_rows.append(
                    RatesRow(
                        **base,
                        prefix="R",
                        cgo_type="RF",
                        cur_20="USD" if COL_RF_20 in v else None,
                        rate_20=v.get(COL_RF_20),
                        cur_40hc="USD" if COL_RF_40HC in v else None,
                        rate_40hc=v.get(COL_RF_40HC),
                    )
                )
            # NOR ("Non-Operating Reefer" - same concept as AUEC's "RAD"):
            # physically a reefer container filed as dry cargo, Prefix R /
            # CGO TYPE DR. Never gets a DG duplicate (raw sheet's own
            # remark: "not accepting DG in Reefer").
            if COL_NOR_20 in v or COL_NOR_40HC in v:
                nor_rows.append(
                    RatesRow(
                        **base,
                        prefix="R",
                        cgo_type="DR",
                        cur_20="USD" if COL_NOR_20 in v else None,
                        rate_20=v.get(COL_NOR_20),
                        cur_40hc="USD" if COL_NOR_40HC in v else None,
                        rate_40hc=v.get(COL_NOR_40HC),
                    )
                )

        dr_rows = group_by_destination(dr_rows)
        rf_rows = group_by_destination(rf_rows)
        nor_rows = group_by_destination(nor_rows)

        dg_rows: list[RatesRow] = []
        if dr_rows and not config.skip_dg_generation.get(DEFAULT_MAIN_DESCRIPTION, False):
            dg_rows = [row.model_copy(update={"cgo_type": "DG"}) for row in dr_rows]

        def stamp(rows: list[RatesRow], code: str, description: str, cmdt_seq: int | None) -> None:
            for row in rows:
                row.commodity_group_code = code
                row.commodity_group_description = description
                row.cmdt_seq = cmdt_seq

        stamp(dr_rows, main_code, main_description, main_cmdt_seq)
        stamp(dg_rows, main_code, main_description, main_cmdt_seq)
        stamp(rf_rows, rf_code, rf_description, rf_cmdt_seq)
        stamp(nor_rows, nor_code, nor_description, nor_cmdt_seq)

        # Route Seq. is a single running counter per COMMODITY GROUP, not
        # per cgo_type/prefix block - confirmed against ground truth: DG
        # continues the same counter as its DR parent (same group), but
        # RF and NOR each restart at 1 (separate groups). Same scope
        # rule as auec.py's own finding, just at a different boundary
        # here since RF/NOR aren't folded into the main group.
        main_group_rows = [*dr_rows, *dg_rows]
        for i, row in enumerate(main_group_rows, start=1):
            row.route_seq = i
        for i, row in enumerate(rf_rows, start=1):
            row.route_seq = i
        for i, row in enumerate(nor_rows, start=1):
            row.route_seq = i

        rates = [*main_group_rows, *rf_rows, *nor_rows]

        excluded_codes = frozenset(config.excluded_charge_codes)
        cmdt_notes: list[CmdtNoteRow] = []
        for group_rows, description in (
            (main_group_rows, main_description),
            (rf_rows, rf_description),
            (nor_rows, nor_description),
        ):
            if not group_rows:
                continue
            notes = _build_group_cmdt_notes(
                data.validity_start, data.validity_end, excluded_codes,
                config.rfa_effective_date, config.rfa_expiry_date,
            )
            cmdt_notes.extend(row.model_copy(update={"group_description": description}) for row in notes)
            note_text = notes[0].contents if notes else None
            for row in group_rows:
                row.commodity_note = note_text

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)


register(
    LayoutProfile(
        lane_id=AUWCParser.lane_id,
        parser_cls=AUWCParser,
        sheet_name_patterns=[re.escape(SHEET_AUFRE), re.escape(SHEET_AUADL)],
        title_keywords=["ONE MINIMUM RATE GUIDELINE"],
        header_fingerprint=["20'NOR", "40'HC NOR"],
    )
)
