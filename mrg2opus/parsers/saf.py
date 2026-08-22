"""SAF lane parser - the simplest layout (single raw sheet, no ARBS/SPECIAL
NOTE) and the first Phase 1 milestone target.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationResolver, split_location_text
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.cmdt_notes import build_cmdt_notes, parse_included_charge_codes
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.header_grid import flatten_pod_header
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import OpusRowSet, RatesRow, explode_rates_row

VIA_EXTRACT_RE = re.compile(r"\bvia\s+(.+)$", re.IGNORECASE)

RAW_SHEET_NAME = "SAF"
DEFAULT_COMMODITY_CODE = "G0001"
DEFAULT_COMMODITY_DESCRIPTION = "FAK"

POD_LABEL_ROW = 5
CONTAINER_LABEL_ROW = 6
DATA_MIN_ROW = 7
DATA_MAX_ROW = 47
MIN_COL, MAX_COL = 4, 11  # D..K
ORIGIN_TEXT_COL = 2  # B


@dataclass
class RawRateCell:
    origin_text: str
    pod_label: str
    container_label: str
    value: float | str | None


@dataclass
class SAFRawData:
    validity_start: date | None
    validity_end: date | None
    included_charge_codes: list[str]
    rate_cells: list[RawRateCell]


class SAFParser(BaseMRGParser):
    lane_id: ClassVar[str] = "SAF"

    def __init__(self, container_map: ContainerMap | None = None, location_resolver: LocationResolver | None = None):
        self.container_map = container_map or load_container_map("saf")
        self.location_resolver = location_resolver or LocationResolver()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        if RAW_SHEET_NAME not in wb.sheetnames:
            return 0.0
        score = 0.5
        ws = wb[RAW_SHEET_NAME]
        title = str(ws.cell(row=1, column=1).value or "")
        if "SAF" in title.upper():
            score += 0.3
        header_tokens = {ws.cell(row=CONTAINER_LABEL_ROW, column=c).value for c in range(MIN_COL, MAX_COL + 1)}
        if {"D2", "D4", "D5", "RD5"} <= {str(t).strip() for t in header_tokens if t}:
            score += 0.2
        return min(score, 1.0)

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        ws: Worksheet = wb[RAW_SHEET_NAME]

        validity_start, validity_end = _parse_validity(ws)
        included_codes = _parse_included_charge_codes(ws)

        header_cols = flatten_pod_header(ws, POD_LABEL_ROW, CONTAINER_LABEL_ROW, MIN_COL, MAX_COL)

        rate_cells: list[RawRateCell] = []
        for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
            origin_cell = ws.cell(row=row_idx, column=ORIGIN_TEXT_COL)
            if origin_cell.value in (None, ""):
                continue
            if is_excluded(origin_cell):
                continue
            origin_text = str(origin_cell.value).strip()
            for hc in header_cols:
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, ""):
                    continue
                if not isinstance(cell.value, (int, float)):
                    continue  # e.g. '-' meaning "not offered" for this container size
                if is_excluded(cell):
                    continue
                rate_cells.append(
                    RawRateCell(
                        origin_text=origin_text,
                        pod_label=hc.pod_label,
                        container_label=hc.container_label,
                        value=cell.value,
                    )
                )

        return RawExtraction(
            tables={
                "saf": SAFRawData(
                    validity_start=validity_start,
                    validity_end=validity_end,
                    included_charge_codes=included_codes,
                    rate_cells=rate_cells,
                )
            }
        )

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: SAFRawData = raw.tables["saf"]

        commodity_description = resolve_commodity_description(DEFAULT_COMMODITY_DESCRIPTION, config)
        cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_COMMODITY_DESCRIPTION)
        output_commodity_code = resolve_commodity_code(DEFAULT_COMMODITY_DESCRIPTION, DEFAULT_COMMODITY_CODE, config)

        # Collected into separate per-type blocks and concatenated at the end
        # (dr_rows, then dg_rows, then reefer_rows) so the output reads as
        # OPUS RATES normally does: all D/DR rows together, then all D/DG,
        # then all R/DR - rather than interleaved per origin/destination.
        dr_rows: list[RatesRow] = []
        dg_rows: list[RatesRow] = []
        reefer_rows: list[RatesRow] = []
        dr_pp: list = []
        dg_pp: list = []
        reefer_pp: list = []

        # group raw cells by (origin_text, pod_label) so one OPUS row covers
        # all container sizes for a given origin/destination pair. The
        # "reefer" container label (e.g. RD5) is tracked separately - it
        # does NOT become a 4th rate slot on this row, it becomes a whole
        # separate Prefix="R" row (see below).
        groups: dict[tuple[str, str], dict] = {}
        for rc in data.rate_cells:
            key = (rc.origin_text, rc.pod_label)
            bucket = groups.setdefault(key, {"sizes": {}, "reefer": None})
            if rc.container_label == self.container_map.reefer_container_label:
                bucket["reefer"] = rc.value
                continue
            suffix = self.container_map.suffix_for(rc.container_label)
            if suffix is None:
                continue
            bucket["sizes"][suffix] = rc.value

        for (origin_text, pod_label), bucket in groups.items():
            sizes = bucket["sizes"]
            reefer_value = bucket["reefer"]

            origin_matches = self.location_resolver.match_text(origin_text)
            dest_matches = self.location_resolver.match_text(pod_label)
            if not origin_matches or not dest_matches:
                continue

            # A low-confidence fuzzy hit is worse than no match at all - it
            # would silently file freight rates under the wrong port, or
            # (for a multi-city origin) a partial/incomplete code group. If
            # ANY token in the origin/destination text is unresolved, skip
            # the whole row rather than guess or file a partial group; the
            # gap is a Location Bank coverage issue for Step 3 / the Audit
            # Gate to surface, not something to paper over here.
            if any(m.needs_review for m in origin_matches) or any(m.needs_review for m in dest_matches):
                continue

            origin_codes = sorted({m.code for m in origin_matches})
            origin_names = sorted({m.primary_name for m in origin_matches})
            dest_codes = sorted({m.code for m in dest_matches})
            dest_names = sorted({m.primary_name for m in dest_matches})

            # A "<city> via <transship port>" origin (e.g. "Ganzhou via
            # Yantian") names routing, not an additional origin location
            # (split_location_text already strips it before location
            # matching) - but the via-port DOES belong in O.Via, confirmed
            # against ground truth (e.g. Ganzhou via Yantian -> O.Via=CNYTN).
            o_via_code = None
            via_clause = VIA_EXTRACT_RE.search(origin_text)
            if via_clause:
                via_match = self.location_resolver.match_token(via_clause.group(1).strip())
                if via_match and not via_match.needs_review:
                    o_via_code = via_match.code

            row = RatesRow(
                type="C",
                cmdt_seq=cmdt_seq,
                commodity_group_code=output_commodity_code,
                commodity_group_description=commodity_description,
                origin_code=";".join(origin_codes),
                origin_description=";".join(origin_names),
                origin_term="CY",
                o_via_code=o_via_code,
                destination_code=";".join(dest_codes),
                destination_description=";".join(dest_names),
                destination_term="CY",
                prefix=self.container_map.prefix,
                cgo_type=self.container_map.cgo_type,
                cur_20="USD" if "20" in sizes else None,
                rate_20=sizes.get("20"),
                cur_40="USD" if "40" in sizes else None,
                rate_40=sizes.get("40"),
                cur_40hc="USD" if "40hc" in sizes else None,
                rate_40hc=sizes.get("40hc"),
            )
            dr_rows.append(row)
            dr_pp.extend(explode_rates_row(row))

            # Every base Dry (D/DR) row also files an identical DG (D/DG)
            # variant at the same rate - confirmed against all 39 ground-truth
            # origin rows in SAF.xlsx (100% consistent: same commodity group,
            # same rate, only CGO TYPE flips DR->DG). This is NOT derivable
            # from anything in the raw SAF sheet itself; it's a standing
            # filing convention.
            if self.container_map.cgo_type == "DR":
                dg_row = row.model_copy(update={"cgo_type": "DG"})
                dg_rows.append(dg_row)
                dg_pp.extend(explode_rates_row(dg_row))

            # The "RD5" raw column ("Reefer D5") is NOT a 4th container-size
            # slot on the base row - it produces a wholly separate row with
            # Prefix="R" whose value sits in the 40HC rate slot (confirmed:
            # ground truth's Prefix="R" rows carry cur_40hc/rate_40hc only,
            # all other rate slots blank). It only appears when the raw
            # sheet actually has an RD5 value for this origin/POD, which is
            # why it's present for some origins and absent for others -
            # directly derived from the source data, not guessed.
            if reefer_value is not None and self.container_map.reefer_prefix:
                reefer_row = row.model_copy(update={
                    "prefix": self.container_map.reefer_prefix,
                    "cur_20": None, "rate_20": None,
                    "cur_40": None, "rate_40": None,
                    "cur_40hc": "USD", "rate_40hc": Decimal(str(reefer_value)),
                })
                reefer_rows.append(reefer_row)
                reefer_pp.extend(explode_rates_row(reefer_row))

        # Within each D/DR, D/DG, R/DR block, group same-destination rows
        # together (all Durban rows, then all Cape Town rows, ...) instead
        # of interleaving by origin - destinations keep the order they
        # first appear in (Durban before Cape Town, matching the raw
        # header layout), origins keep their relative order within each.
        rates = [
            *group_by_destination(dr_rows),
            *group_by_destination(dg_rows),
            *group_by_destination(reefer_rows),
        ]
        rates_port_port = [
            *group_by_destination(dr_pp),
            *group_by_destination(dg_pp),
            *group_by_destination(reefer_pp),
        ]

        cmdt_notes = build_cmdt_notes(data.validity_start, data.validity_end, data.included_charge_codes)

        return OpusRowSet(rates=rates, rates_port_port=rates_port_port, cmdt_notes=cmdt_notes)


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=1, column=5).value or "")  # E1: "from 21 - 27 Aug 2026"
    m = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})\s+(\w+)\s+(\d{4})", text)
    if not m:
        return None, None
    start_day, end_day, month_name, year = m.groups()
    try:
        start = date(int(year), _month_number(month_name), int(start_day))
        end = date(int(year), _month_number(month_name), int(end_day))
    except ValueError:
        return None, None
    return start, end


def _month_number(name: str) -> int:
    months = [
        "jan", "feb", "mar", "apr", "may", "jun",
        "jul", "aug", "sep", "oct", "nov", "dec",
    ]
    return months.index(name.strip().lower()[:3]) + 1


def _parse_included_charge_codes(ws: Worksheet) -> list[str]:
    """Raw D3: 'Rate structure: incl. OBS/EFS/MBS, subj to THL/THD, ...'."""
    text = str(ws.cell(row=3, column=4).value or "")
    return parse_included_charge_codes(text)


register(
    LayoutProfile(
        lane_id=SAFParser.lane_id,
        parser_cls=SAFParser,
        sheet_name_patterns=[r"^SAF$"],
        title_keywords=["SAF"],
        header_fingerprint=["D2", "D4", "D5", "RD5"],
    )
)
