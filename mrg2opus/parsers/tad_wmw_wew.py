"""TAD FILING WMW WEW lane parser.

Two raw sheets, "WEW" and "WMW", each its own Service Scope and own OPUS
output file. Same "Tool for TAD.xlsm" DATA-sheet-shaped raw layout as
TAD-OEW-OMW (see tad_oew_omw.py and the project_tad_vba_tool_analysis
memory) - no Location Bank fuzzy matching needed, every origin/destination
code + description is given directly.

Output scope is RATES + CMDT NOTE + ROUTE NOTE (confirmed against
reference/2_OPUS/27_TAD FILING WMW WEW). Unlike OEW/OMW's raw file (where
POL/POD are always blank), this raw file populates them for a subset of
rows and they map straight through to RATES' O.Via/D.Via Code columns
(POL->o_via_code, POD->d_via_code; confirmed against ground truth: INCCU
rows carry POL=INNSA -> o_via_code=INNSA, and a BDCGP->PTLEI row carries
POD=NLRTM -> d_via_code=NLRTM). One real difference from OEW/OMW:
a raw row can populate BOTH T/S Port and Service Lane at once (ground truth
WMW POLLY.xlsx route_seq 24, origin INCOK/dest EGALY), and when it does the
two note lines are combined into a single ROUTE NOTE cell joined by " | " -
never split into two note_seq rows. No SPECIAL_NODE_ROUTE_NOTES-style
hardcoded commodity-group-name table is needed here: every row seen in
ground truth uses plain "FAK" as its Commodity Group Name, so that lookup
path (used by OEW/OMW for HAYDARPASA/MARPORT/EGALY special nodes) is
unconfirmed for this lane and deliberately not copied over.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.tad_snapshots import find_snapshot_sheets, merge_dated_snapshots
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import CHARGE_CODE_NAMES
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow, RouteNoteRow

SHEET_WEW = "WEW"
SHEET_WMW = "WMW"

DATA_MIN_ROW = 2
COL_EFF_DATE, COL_EXP_DATE, COL_SCOPE, COL_CMDT_NAME = 1, 2, 3, 4
COL_POR, COL_POR_DESC, COL_ORIGIN_TERM, COL_ORIGIN_TRANSMODE = 5, 6, 7, 8
COL_POL, COL_POD, COL_DEL, COL_DEL_DESC, COL_DEST_TERM = 9, 10, 11, 12, 13
COL_CARGO_TYPE, COL_TS_PORT, COL_SERVICE_LANE, COL_CUR = 14, 15, 16, 17
COL_OFT20, COL_OFT40, COL_OFTHC, COL_OFT45, COL_INCLUDE_SURCHARGE = 18, 19, 20, 21, 22

# Same CARGO sheet equivalent as TAD-OEW-OMW - see that module's comment.
CARGO_TYPE_MAP: dict[str, tuple[str, str]] = {
    "Dry General": ("D", "DR"),
    "Reefer": ("R", "RF"),
    "Reefer Dry": ("R", "DR"),
    "Dry Dangerous": ("D", "DG"),
}

DEFAULT_COMMODITY_DESCRIPTION = "FAK"
DEFAULT_COMMODITY_CODE = "G0001"

# TAD's own CMDT NOTE boilerplate ("surcharges including", one comma short
# of the shared cmdt_notes.py template) and the HEA->"HEAVY SURCHARGE"
# override are the same confirmed quirks as TAD-OEW-OMW (see that module),
# reconfirmed here against reference/2_OPUS/27_TAD FILING WMW WEW.
_HEA_OVERRIDE = "HEAVY SURCHARGE"


def _route_note_for(ts_port: str | None, service_lane: str | None) -> str | None:
    """T/S Port and Service Lane are independent flags on the same row and
    can both be populated at once (confirmed: WMW POLLY.xlsx route_seq 24,
    INCOK->EGALY, T/S Port=INMUN + Service Lane=IOM). When both fire, their
    two note lines join into ONE cell separated by " | ", never split
    across two note_seq rows.
    """
    parts = []
    if ts_port:
        parts.append(f"Rates are Subject to Transhipment Port: {ts_port}")
    if service_lane:
        parts.append(f"Rates are applicable for Vessel Service Lane: {service_lane}")
    return " | ".join(parts) if parts else None


def _parse_decimal(text: str | None) -> Decimal | None:
    if not text:
        return None
    try:
        return Decimal(str(text).replace(",", "").strip())
    except InvalidOperation:
        return None


def _parse_date(text: str | None) -> date | None:
    if not text:
        return None
    try:
        return date.fromisoformat(str(text).strip())
    except ValueError:
        return None


def _parse_included_codes(text: str | None) -> list[str]:
    """Same authoritative literal comma-separated list as TAD-OEW-OMW - see
    that module's comment for why this bypasses is_known_charge_code()."""
    if not text:
        return []
    return [c.strip().upper() for c in str(text).split(",") if c.strip()]


@dataclass
class RawRow:
    validity_start: date | None
    validity_end: date | None
    origin_code: str
    origin_description: str
    origin_term: str | None
    origin_transmode: str | None
    pol: str | None
    pod: str | None
    dest_code: str
    dest_description: str
    dest_term: str | None
    cargo_type: str
    ts_port: str | None
    service_lane: str | None
    included_codes: list[str]
    rate_20: Decimal | None
    rate_40: Decimal | None
    rate_40hc: Decimal | None
    rate_45: Decimal | None


@dataclass
class ScopeData:
    scope: str
    rows: list[RawRow] = field(default_factory=list)


def _read_sheet(ws: Worksheet, scope: str) -> ScopeData:
    data = ScopeData(scope=scope)
    for row_idx in range(DATA_MIN_ROW, ws.max_row + 1):
        origin = ws.cell(row=row_idx, column=COL_POR).value
        if origin in (None, ""):
            continue
        origin_cell = ws.cell(row=row_idx, column=COL_POR)
        if is_excluded(origin_cell):
            continue
        data.rows.append(
            RawRow(
                validity_start=_parse_date(ws.cell(row=row_idx, column=COL_EFF_DATE).value),
                validity_end=_parse_date(ws.cell(row=row_idx, column=COL_EXP_DATE).value),
                origin_code=str(origin).strip(),
                origin_description=str(ws.cell(row=row_idx, column=COL_POR_DESC).value or "").strip(),
                origin_term=str(ws.cell(row=row_idx, column=COL_ORIGIN_TERM).value or "").strip() or None,
                origin_transmode=str(ws.cell(row=row_idx, column=COL_ORIGIN_TRANSMODE).value or "").strip() or None,
                pol=str(ws.cell(row=row_idx, column=COL_POL).value or "").strip() or None,
                pod=str(ws.cell(row=row_idx, column=COL_POD).value or "").strip() or None,
                dest_code=str(ws.cell(row=row_idx, column=COL_DEL).value or "").strip(),
                dest_description=str(ws.cell(row=row_idx, column=COL_DEL_DESC).value or "").strip(),
                dest_term=str(ws.cell(row=row_idx, column=COL_DEST_TERM).value or "").strip() or None,
                cargo_type=str(ws.cell(row=row_idx, column=COL_CARGO_TYPE).value or "").strip(),
                ts_port=str(ws.cell(row=row_idx, column=COL_TS_PORT).value or "").strip() or None,
                service_lane=str(ws.cell(row=row_idx, column=COL_SERVICE_LANE).value or "").strip() or None,
                included_codes=_parse_included_codes(ws.cell(row=row_idx, column=COL_INCLUDE_SURCHARGE).value),
                rate_20=_parse_decimal(ws.cell(row=row_idx, column=COL_OFT20).value),
                rate_40=_parse_decimal(ws.cell(row=row_idx, column=COL_OFT40).value),
                rate_40hc=_parse_decimal(ws.cell(row=row_idx, column=COL_OFTHC).value),
                rate_45=_parse_decimal(ws.cell(row=row_idx, column=COL_OFT45).value),
            )
        )
    return data


class TADWmwWewParser(BaseMRGParser):
    lane_id: ClassVar[str] = "TAD-WMW-WEW"
    SHEET_NAME_OVERRIDES: ClassVar[dict[str, str]] = {"route_notes": "ROUTE NOTE"}

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        names = set(wb.sheetnames)
        if not ({SHEET_WEW, SHEET_WMW} & names):
            return 0.0
        score = 0.5 * len({SHEET_WEW, SHEET_WMW} & names)
        ws = wb[SHEET_WEW] if SHEET_WEW in names else wb[SHEET_WMW]
        header_tokens = {str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, 23)}
        if "Include Surcharge" in header_tokens and "DEL Description" in header_tokens:
            score = min(score + 0.3, 1.0)
        return score

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        scopes: dict[str, ScopeData] = {}
        for scope_name in (SHEET_WEW, SHEET_WMW):
            sheet_names = find_snapshot_sheets(wb, scope_name)
            if not sheet_names:
                continue
            occurrences = [_read_sheet(wb[name], scope_name).rows for name in sheet_names]
            scopes[scope_name] = ScopeData(scope=scope_name, rows=merge_dated_snapshots(occurrences))
        return RawExtraction(tables={"scopes": scopes})

    def _build_one_scope(self, data: ScopeData, config: MappingProfile) -> OpusRowSet:
        description = resolve_commodity_description(DEFAULT_COMMODITY_DESCRIPTION, config)
        code = resolve_commodity_code(DEFAULT_COMMODITY_DESCRIPTION, DEFAULT_COMMODITY_CODE, config)
        excluded_codes = frozenset(config.excluded_charge_codes)

        # CMDT NOTE grouping is data-driven (DATA's own per-row validity
        # dates + Include Surcharge list) - see TAD-OEW-OMW's comment.
        group_order: list[tuple] = []
        group_rows: dict[tuple, list[RawRow]] = defaultdict(list)
        for row in data.rows:
            key = (row.validity_start, row.validity_end, tuple(row.included_codes))
            if key not in group_rows:
                group_order.append(key)
            group_rows[key].append(row)
        cmdt_seq_by_key = {key: i + 1 for i, key in enumerate(group_order)}

        rates: list[RatesRow] = []
        for row in data.rows:
            mapped = CARGO_TYPE_MAP.get(row.cargo_type)
            if mapped is None:
                continue
            prefix, cgo_type = mapped
            key = (row.validity_start, row.validity_end, tuple(row.included_codes))
            rates.append(
                RatesRow(
                    cmdt_seq=cmdt_seq_by_key[key],
                    commodity_group_code=code,
                    commodity_group_description=description,
                    origin_code=row.origin_code,
                    origin_description=row.origin_description,
                    origin_term=row.origin_term,
                    origin_transmode=row.origin_transmode,
                    o_via_code=row.pol,
                    d_via_code=row.pod,
                    destination_code=row.dest_code,
                    destination_description=row.dest_description,
                    destination_term=row.dest_term,
                    prefix=prefix,
                    cgo_type=cgo_type,
                    cur_20="USD" if row.rate_20 is not None else None,
                    rate_20=row.rate_20,
                    cur_40="USD" if row.rate_40 is not None else None,
                    rate_40=row.rate_40,
                    cur_40hc="USD" if row.rate_40hc is not None else None,
                    rate_40hc=row.rate_40hc,
                    cur_45="USD" if row.rate_45 is not None else None,
                    rate_45=row.rate_45,
                    route_note=_route_note_for(row.ts_port, row.service_lane),
                )
            )

        # Opt-in only (config.generate_tad_dg_duplicate, default off) -
        # see TAD-OEW-OMW's identical comment.
        if config.generate_tad_dg_duplicate:
            rates.extend(
                row.model_copy(update={"cgo_type": "DG"}) for row in list(rates) if row.prefix == "D" and row.cgo_type == "DR"
            )

        # Route Seq.: one continuous counter per commodity group (same
        # scope rule as every other lane).
        for key in group_order:
            group = [r for r in rates if r.cmdt_seq == cmdt_seq_by_key[key]]
            for i, r in enumerate(group, start=1):
                r.route_seq = i

        cmdt_notes: list[CmdtNoteRow] = []
        note_text_by_seq: dict[int, str | None] = {}
        for key in group_order:
            validity_start, validity_end, codes = key
            seq = cmdt_seq_by_key[key]
            notes = self._build_cmdt_notes(validity_start, validity_end, list(codes), excluded_codes, config)
            for note in notes:
                note.header_seq = seq
            cmdt_notes.extend(notes)
            note_text_by_seq[seq] = notes[0].contents if notes else None

        for row in rates:
            row.commodity_note = note_text_by_seq.get(row.cmdt_seq)

        route_notes: list[RouteNoteRow] = []
        validity_by_seq = {cmdt_seq_by_key[key]: (key[0], key[1]) for key in group_order}
        for row in rates:
            if not row.route_note:
                continue
            validity_start, validity_end = validity_by_seq.get(row.cmdt_seq, (None, None))
            route_notes.append(
                RouteNoteRow(
                    header_seq=row.cmdt_seq,
                    route_seq=row.route_seq,
                    contents=row.route_note,
                    charge_seq=1,
                    code="APP",
                    application_effective=validity_start,
                    application_expires=validity_end,
                    application="S",
                )
            )

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes, route_notes=route_notes)

    def _build_cmdt_notes(
        self,
        validity_start: date | None,
        validity_end: date | None,
        included_codes: list[str],
        excluded_codes: frozenset[str],
        config: MappingProfile,
    ) -> list[CmdtNoteRow]:
        codes = [c for c in included_codes if c not in excluded_codes]
        if not codes or validity_start is None or validity_end is None:
            return []
        unique_codes = sorted(dict.fromkeys(codes))
        names = {**CHARGE_CODE_NAMES, "HEA": _HEA_OVERRIDE}
        names_line = " and the ".join(f"{names.get(c, c)}({c})" for c in unique_codes)
        contents = "\n".join(
            [
                f"Rates are valid from {validity_start:%Y%m%d} to {validity_end:%Y%m%d}",
                f"Rates are inclusive of the {names_line}",
                "Rates are subject to all other surcharges including those, if any, "
                "specified in the contract and those published in the Governing Tariff(s) at the time of shipment.",
            ]
        )
        parent = CmdtNoteRow(
            contents=contents, charge_seq=1, code="APP",
            application_effective=validity_start, application_expires=validity_end, application="S",
        )
        child_effective = config.rfa_effective_date if config.rfa_effective_date is not None else validity_start
        child_expires = config.rfa_expiry_date if config.rfa_expiry_date is not None else validity_end
        children = [
            CmdtNoteRow(
                charge_seq=i + 2, code=c,
                application_effective=child_effective, application_expires=child_expires, application="I",
            )
            for i, c in enumerate(codes)
        ]
        return [parent, *children]

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        # Single-scope entry point (BaseMRGParser.run()) - used when only
        # one of WEW/WMW is present in the workbook.
        scopes: dict[str, ScopeData] = raw.tables["scopes"]
        if not scopes:
            return OpusRowSet()
        return self._build_one_scope(next(iter(scopes.values())), config)

    def run_multi(self, wb: Workbook, config: MappingProfile | None = None) -> dict[str, OpusRowSet]:
        config = config or MappingProfile()
        raw = self.parse_raw(wb)
        scopes: dict[str, ScopeData] = raw.tables["scopes"]
        return {scope: self._build_one_scope(data, config) for scope, data in scopes.items()}


register(
    LayoutProfile(
        lane_id=TADWmwWewParser.lane_id,
        parser_cls=TADWmwWewParser,
        sheet_name_patterns=[r"^WEW$", r"^WMW$"],
        title_keywords=["EFFECTIVE DATE", "COMMODITY GROUP NAME", "INCLUDE SURCHARGE"],
        header_fingerprint=["DEL Description", "Include Surcharge"],
    )
)
