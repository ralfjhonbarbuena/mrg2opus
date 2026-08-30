"""NZJ NEA to NZ (ex North East Asia to New Zealand, service NZJ) FAK lane
parser.

Single raw sheet ("Ex. NEA to NZ"): a 49-origin grid with THREE parallel
rate blocks sharing one commodity group, distinguished by which New
Zealand ports they price to:
  - "Auckland (only)" - its own D2/D4/D5 (20'/40'/40'HC dry) columns,
    destination NZAKL alone.
  - "Lyttelton / Napier / Tauranga" - its own separate D2/D4/D5 columns,
    destination the 3-code combined group NZLYT;NZNPE;NZTRG.
  - "Auckland / Lyttelton / Napier / Tauranga" - ONE combined R2/R5
    (reefer) and R2-RAD/R5-RAD (Reefer As Dry, see auec.py) column set,
    destination the full 4-code combined group.

Confirmed against both reference weeks (45, 46): every origin row feeds
BOTH dry destination blocks (Auckland alone, and the other 3 combined)
with its own D/DR rate row each - not a fan-out choice, the raw sheet
genuinely files the same rate twice under two different POD labels. RF/
RAD rows go to the 4-way combined destination only, and never get a DG
duplicate (same "not accepting DG in Reefer" rule as AUEC/AUWC). D/DG
duplicates ARE generated for both dry destination blocks.

Route Seq. is one continuous counter across the whole commodity group, in
generation order: Auckland-DR, combined-DR(LNT), RF, RAD, Auckland-DG,
combined-DG(LNT) - confirmed via the exact route_seq transition points in
both reference weeks' RATES sheets.

The 'Country'/'Region' columns are fill-down grouping labels only (e.g.
"China" / "North China") - never appear in ground truth CMDT NOTE POL
scoping (unlike AUEC/AUWC, every child row here is unscoped) and aren't
otherwise used in the parser.

PRDA/PRDB footnote groups are byte-identical to AUEC's (verified against
both reference weeks' own ground truth code/description strings) - reused
from auec.py rather than re-transcribed, same precedent as auwc.py.

Genuinely new here vs. AUEC/AUWC: an origin's "via" clause isn't always
parenthetical - "Keelung \\n via Kaohsiung" is a plain trailing "via
<city name>" suffix (no parens, no rail), not "(via <code> by rail)" -
needed a more general via-extraction than auec.py's paren-only regex.
Also: "Taoyuen / Hsinchu" needed one new Location Bank entry (TWHSZ /
HSINCHU) - Hsinchu was completely absent from the bank, not just a
low-confidence fuzzy match.

CMDT NOTE: hardcoded INCLUDED_CHARGE_CODES = EFS, OBS, ISL (both weeks'
ground truth includes ISL despite the raw surcharge table itself marking
ISL as "ONE Tariff", not "included" - the same raw-remarks-are-unreliable
finding as AUEC's own ISL/EFS handling). Child row order (EFS, OBS, ISL)
is identical across both weeks; the "inclusive of" text line's own word
order is NOT (week 1 says "...EFS...OBS...ISL", week 2 says
"...EFS...ISL...OBS") - a human filing inconsistency, not a parsing gap.
This parser keeps one fixed order (matching week 1's text exactly, and
both weeks' child rows), same accepted-gap category as
west_asia_waf.py's own documented text-ordering inconsistency.

TIER 1 (folders 47/48): unlike LAEC's own TIER 1 (raw file textually
identical to FAK, no signal to detect it by), this lane's TIER 1 raw
workbook carries an extra "T-1 Customer list" sheet (a customer roster,
confirmed present in both TIER 1 weeks, absent from both FAK weeks) - a
reliable detection signal, see TIER1_SHEET_NAME/NZJRawData.is_tier1.
TIER 1's own ground truth CMDT NOTE differs from FAK's in 3 confirmed
ways: (1) an extra "Rates are applicable for Vessel Service Lane: NZJ"
text line (see common/cmdt_notes.py's service_lane parameter), (2) the
"inclusive of" text alphabetizes (EFS, ISL, OBS) instead of FAK's fixed
order, (3) ISL is scoped via POR instead of POL for the identical
business rule - the same FAK-uses-POL/TIER1-switches-to-POR convention
already confirmed for AUEC/AUWC/AUBP. TIER 1's RATES sheet also has an
unexplained, non-derivable duplication in one of the two weeks (6
header_seq groups repeating the same 159-row content) - same accepted-gap
category as AUEC TIER 1's own note-block duplication, not chased further.

Output scope is RATES + CMDT NOTE only (ground truth's own sheet names
drift between weeks - 'rates'/'SUR' in week 1, 'RATES'/'SURCHARGE' in
week 2 - no Route Note, no ARBS, no Special Note in either week).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationMatch, LocationResolver
from mrg2opus.parsers.auec import PRDA_CODE, PRDA_DESCRIPTION, PRDB_CODE, PRDB_DESCRIPTION
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.cmdt_notes import build_cmdt_notes
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow

SHEET_NAME = "Ex. NEA to NZ"
# Present only in real TIER 1 raw files (a customer roster, absent from
# FAK's own) - confirmed a reliable, consistent signal across both TIER 1
# reference weeks (unlike some other lanes, where FAK/TIER1 raw files are
# textually indistinguishable). TIER 1's own CMDT NOTE ground truth adds a
# "Vessel Service Lane: NZJ" line and alphabetizes the "inclusive of" text
# (FAK's own stays un-alphabetized, matching INCLUDED_CHARGE_CODES' own
# order) - confirmed in both TIER 1 weeks.
TIER1_SHEET_NAME = "T-1 Customer list"

DATA_MIN_ROW, DATA_MAX_ROW = 8, 56
ORIGIN_COL = 3
VALIDITY_ROW, VALIDITY_COL = 3, 4
POD_LABEL_ROW = 5

# Column -> RatesRow rate slot, for each of the 3 parallel rate blocks.
COL_AKL_20, COL_AKL_40, COL_AKL_40HC = 4, 5, 6
COL_LNT_20, COL_LNT_40, COL_LNT_40HC = 7, 8, 9
COL_RF_20, COL_RF_40HC = 10, 11
COL_RAD_20, COL_RAD_40HC = 12, 13

# Top-left cell of each POD label's merged range on POD_LABEL_ROW - fixed
# positions (D5/G5/J5), confirmed identical in both reference weeks.
AKL_POD_COL, LNT_POD_COL, COMBINED_POD_COL = COL_AKL_20, COL_LNT_20, COL_RF_20

DEFAULT_DESCRIPTION = "EX. NEA TO NZ"
DEFAULT_CODE = "G0003"

# Hardcoded rather than parsed from the raw sheet's own surcharge table
# (rows 64-71 there mark ISL as "ONE Tariff", not "included") - see module
# docstring. Order matches ground truth's own charge_seq order in BOTH
# weeks (not alphabetical: alphabetical would be EFS, ISL, OBS).
INCLUDED_CHARGE_CODES = ["EFS", "OBS", "ISL"]

# ISL is POL-scoped to Taiwan only (every other included code is
# unscoped) - confirmed in BOTH weeks' ground truth (POL="TW" on the ISL
# child row, every other POL cell blank). Same business rule as AUEC's
# own ISL->TW scoping (auec.py's INCLUDED_CHARGES), just on a lane with
# no other scoped codes, so the shared build_cmdt_notes() helper is used
# as-is and this one row is patched afterward rather than needing a
# lane-specific builder.
ISL_SCOPE_POL = "TW"

_PRD_GROUPS = {"PRDA": (PRDA_CODE, PRDA_DESCRIPTION), "PRDB": (PRDB_CODE, PRDB_DESCRIPTION)}

# General "via" extraction: handles both AUEC-style "(via CNYTN by rail)"
# and this lane's own plain, non-parenthetical "via Kaohsiung" suffix
# (raw text "Keelung \n    via Kaohsiung") - see module docstring.
_VIA_RE = re.compile(r"\bvia\b\s*", re.IGNORECASE)
_BY_RAIL_RE = re.compile(r"by\s+rail", re.IGNORECASE)
_VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s*(?:\d{4}\s+)?to\s+(\d{1,2})\s+(\w+)\s+(\d{4})", re.IGNORECASE)
_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=VALIDITY_ROW, column=VALIDITY_COL).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, d2, mon2, year = m.groups()
    try:
        y = int(year)
        start = date(y, _month_number(mon1), int(d1))
        end = date(y, _month_number(mon2), int(d2))
    except ValueError:
        return None, None
    return start, end


def _split_via(origin_text: str) -> tuple[str, str | None, bool]:
    """'Chengdu** (via CNYTN by rail)' -> ('Chengdu**', 'CNYTN', True).
    'Keelung \\n    via Kaohsiung' -> ('Keelung', 'Kaohsiung', False) - a
    plain trailing via-clause with no parens, resolved as a location name
    rather than a bare code (see caller). Only the literal 'by rail' text
    sets by_rail; origin_term stays 'CY' either way."""
    parts = _VIA_RE.split(origin_text, maxsplit=1)
    if len(parts) == 1:
        return origin_text.strip(), None, False
    before, after = parts
    by_rail = bool(_BY_RAIL_RE.search(after))
    after = _BY_RAIL_RE.sub("", after)
    via = after.strip().rstrip(")").strip()
    clean = before.strip().rstrip("(").strip()
    return clean, (via or None), by_rail


@dataclass
class OriginRow:
    origin_text: str
    values: dict[int, float]


@dataclass
class NZJRawData:
    validity_start: date | None
    validity_end: date | None
    akl_pod_text: str
    lnt_pod_text: str
    combined_pod_text: str
    origins: list[OriginRow] = field(default_factory=list)
    is_tier1: bool = False


def _read_sheet(ws: Worksheet, is_tier1: bool = False) -> NZJRawData:
    validity_start, validity_end = _parse_validity(ws)
    akl_pod_text = str(ws.cell(row=POD_LABEL_ROW, column=AKL_POD_COL).value or "").strip()
    lnt_pod_text = str(ws.cell(row=POD_LABEL_ROW, column=LNT_POD_COL).value or "").strip()
    combined_pod_text = str(ws.cell(row=POD_LABEL_ROW, column=COMBINED_POD_COL).value or "").strip()

    origins: list[OriginRow] = []
    for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
        origin_cell = ws.cell(row=row_idx, column=ORIGIN_COL)
        if origin_cell.value in (None, ""):
            continue
        if is_excluded(origin_cell):
            continue
        values: dict[int, float] = {}
        for col in (
            COL_AKL_20, COL_AKL_40, COL_AKL_40HC,
            COL_LNT_20, COL_LNT_40, COL_LNT_40HC,
            COL_RF_20, COL_RF_40HC, COL_RAD_20, COL_RAD_40HC,
        ):
            v = ws.cell(row=row_idx, column=col).value
            if isinstance(v, (int, float)):
                values[col] = v
        if not values:
            continue
        origins.append(OriginRow(origin_text=str(origin_cell.value).strip(), values=values))

    return NZJRawData(
        validity_start=validity_start,
        validity_end=validity_end,
        akl_pod_text=akl_pod_text,
        lnt_pod_text=lnt_pod_text,
        combined_pod_text=combined_pod_text,
        origins=origins,
        is_tier1=is_tier1,
    )


class NZJParser(BaseMRGParser):
    lane_id: ClassVar[str] = "NZJ"

    def __init__(self, location_resolver: LocationResolver | None = None):
        self.location_resolver = location_resolver or LocationResolver()

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        if SHEET_NAME not in wb.sheetnames:
            return RawExtraction(tables={})
        data = _read_sheet(wb[SHEET_NAME], is_tier1=TIER1_SHEET_NAME in wb.sheetnames)
        return RawExtraction(tables={"nzj": data})

    def _resolve_destination(self, label_text: str) -> tuple[str, str] | None:
        matches = self.location_resolver.match_text(label_text)
        if not matches or any(m.needs_review for m in matches):
            return None
        code = ";".join(sorted({m.code for m in matches}))
        desc = ";".join(sorted({m.primary_name for m in matches}))
        return code, desc

    def _resolve_origin(self, origin_text: str) -> tuple[str, str, str | None, bool] | None:
        """Returns (origin_code, origin_description, o_via_code, is_rail) or
        None if unresolved (skip the row rather than guess)."""
        clean, via, by_rail = _split_via(origin_text)

        prd_key = clean.rstrip("*").strip().upper()
        if prd_key in _PRD_GROUPS:
            code, desc = _PRD_GROUPS[prd_key]
        else:
            matches: list[LocationMatch]
            if "/" in clean:
                # A slash always means "these are N distinct locations
                # sharing one rate" in this raw sheet (same convention as
                # the POD labels' own "/"-joined destination groups) -
                # never try the whole string as one location here.
                # Confirmed necessary: "Taoyuen / Hsinchu" whole-text
                # fuzzy-matches HSINCHU alone (score 90, not needs_review)
                # once TWHSZ exists in the Location Bank, which would
                # silently drop TWTYN/Taoyuan.
                matches = self.location_resolver.match_text(clean)
            else:
                # Try the whole (unsplit) text first - handles single
                # locations written as "City, Province" (e.g. "Taizhou,
                # Jiangsu") that would otherwise be wrongly split into two
                # independent, incorrectly-matched tokens by the comma.
                whole = self.location_resolver.match_token(clean)
                if whole is not None and not whole.needs_review:
                    matches = [whole]
                else:
                    matches = self.location_resolver.match_text(clean)
            if not matches or any(m.needs_review for m in matches):
                return None
            code = ";".join(sorted({m.code for m in matches}))
            desc = ";".join(sorted({m.primary_name for m in matches}))

        via_code = None
        if via:
            via_match = self.location_resolver.match_token(via)
            if via_match and not via_match.needs_review:
                via_code = via_match.code

        return code, desc, via_code, by_rail

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: NZJRawData | None = raw.tables.get("nzj")
        if data is None:
            return OpusRowSet()

        description = resolve_commodity_description(DEFAULT_DESCRIPTION, config)
        code = resolve_commodity_code(DEFAULT_DESCRIPTION, DEFAULT_CODE, config)
        # One CMDT NOTE block covers every prefix/cargo-type combo here (a
        # constant, not something to auto-number across multiple blocks).
        cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_DESCRIPTION, 1)

        akl_dest = self._resolve_destination(data.akl_pod_text)
        lnt_dest = self._resolve_destination(data.lnt_pod_text)
        combined_dest = self._resolve_destination(data.combined_pod_text)

        dr_akl_rows: list[RatesRow] = []
        dr_lnt_rows: list[RatesRow] = []
        rf_rows: list[RatesRow] = []
        rad_rows: list[RatesRow] = []

        for origin in data.origins:
            resolved = self._resolve_origin(origin.origin_text)
            if resolved is None:
                continue
            origin_code, origin_desc, via_code, by_rail = resolved
            v = origin.values

            base = dict(
                commodity_group_code="",
                commodity_group_description="",
                origin_code=origin_code,
                origin_description=origin_desc,
                origin_term="CY",
                origin_transmode="Rail" if by_rail else None,
                o_via_code=via_code,
                destination_term="CY",
            )

            if akl_dest is not None and (COL_AKL_20 in v or COL_AKL_40 in v or COL_AKL_40HC in v):
                dest_code, dest_desc = akl_dest
                dr_akl_rows.append(
                    RatesRow(
                        **base,
                        destination_code=dest_code,
                        destination_description=dest_desc,
                        prefix="D",
                        cgo_type="DR",
                        cur_20="USD" if COL_AKL_20 in v else None,
                        rate_20=v.get(COL_AKL_20),
                        cur_40="USD" if COL_AKL_40 in v else None,
                        rate_40=v.get(COL_AKL_40),
                        cur_40hc="USD" if COL_AKL_40HC in v else None,
                        rate_40hc=v.get(COL_AKL_40HC),
                    )
                )
            if lnt_dest is not None and (COL_LNT_20 in v or COL_LNT_40 in v or COL_LNT_40HC in v):
                dest_code, dest_desc = lnt_dest
                dr_lnt_rows.append(
                    RatesRow(
                        **base,
                        destination_code=dest_code,
                        destination_description=dest_desc,
                        prefix="D",
                        cgo_type="DR",
                        cur_20="USD" if COL_LNT_20 in v else None,
                        rate_20=v.get(COL_LNT_20),
                        cur_40="USD" if COL_LNT_40 in v else None,
                        rate_40=v.get(COL_LNT_40),
                        cur_40hc="USD" if COL_LNT_40HC in v else None,
                        rate_40hc=v.get(COL_LNT_40HC),
                    )
                )
            if combined_dest is not None and (COL_RF_20 in v or COL_RF_40HC in v):
                dest_code, dest_desc = combined_dest
                rf_rows.append(
                    RatesRow(
                        **base,
                        destination_code=dest_code,
                        destination_description=dest_desc,
                        prefix="R",
                        cgo_type="RF",
                        cur_20="USD" if COL_RF_20 in v else None,
                        rate_20=v.get(COL_RF_20),
                        cur_40hc="USD" if COL_RF_40HC in v else None,
                        rate_40hc=v.get(COL_RF_40HC),
                    )
                )
            # RAD ("Reefer As Dry") - see auec.py module docstring. Never
            # gets a DG duplicate.
            if combined_dest is not None and (COL_RAD_20 in v or COL_RAD_40HC in v):
                dest_code, dest_desc = combined_dest
                rad_rows.append(
                    RatesRow(
                        **base,
                        destination_code=dest_code,
                        destination_description=dest_desc,
                        prefix="R",
                        cgo_type="DR",
                        cur_20="USD" if COL_RAD_20 in v else None,
                        rate_20=v.get(COL_RAD_20),
                        cur_40hc="USD" if COL_RAD_40HC in v else None,
                        rate_40hc=v.get(COL_RAD_40HC),
                    )
                )

        dg_akl_rows: list[RatesRow] = []
        dg_lnt_rows: list[RatesRow] = []
        if not config.skip_dg_generation.get(DEFAULT_DESCRIPTION, False):
            dg_akl_rows = [row.model_copy(update={"cgo_type": "DG"}) for row in dr_akl_rows]
            dg_lnt_rows = [row.model_copy(update={"cgo_type": "DG"}) for row in dr_lnt_rows]

        rates = [*dr_akl_rows, *dr_lnt_rows, *rf_rows, *rad_rows, *dg_akl_rows, *dg_lnt_rows]
        for row in rates:
            row.commodity_group_code = code
            row.commodity_group_description = description
            row.cmdt_seq = cmdt_seq
        for i, row in enumerate(rates, start=1):
            row.route_seq = i

        excluded_codes = frozenset(config.excluded_charge_codes)
        included_codes = [c for c in INCLUDED_CHARGE_CODES if c not in excluded_codes]
        cmdt_notes: list[CmdtNoteRow] = build_cmdt_notes(
            data.validity_start, data.validity_end, included_codes,
            sequential_charge_seq=True, sort_text_names=data.is_tier1,
            rfa_effective=config.rfa_effective_date, rfa_expiry=config.rfa_expiry_date,
            service_lane="NZJ" if data.is_tier1 else None,
        )
        # TIER 1's own ground truth scopes ISL via POR instead of POL for
        # the identical business rule (confirmed both TIER 1 weeks) - the
        # same FAK-uses-POL/TIER1-switches-to-POR convention already seen
        # in AUEC/AUWC/AUBP.
        isl_scope_field = "por" if data.is_tier1 else "pol"
        cmdt_notes = [
            row.model_copy(update={"group_description": description, "header_seq": cmdt_seq, isl_scope_field: ISL_SCOPE_POL})
            if row.code == "ISL"
            else row.model_copy(update={"group_description": description, "header_seq": cmdt_seq})
            for row in cmdt_notes
        ]
        note_text = cmdt_notes[0].contents if cmdt_notes else None
        for row in rates:
            row.commodity_note = note_text

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)


register(
    LayoutProfile(
        lane_id=NZJParser.lane_id,
        parser_cls=NZJParser,
        sheet_name_patterns=[re.escape(SHEET_NAME)],
        title_keywords=["AUCKLAND (ONLY)"],
        header_fingerprint=["R2 - RAD", "R5 - RAD"],
    )
)
