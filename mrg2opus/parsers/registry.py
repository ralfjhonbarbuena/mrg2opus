"""Rules-based MRG layout classifier (Step 1: "identify what kind of MRG it is").

Deliberately rules-based rather than ML: only a handful of known lanes exist,
and a wrong classification silently produces wrong freight rates, so the
score must be explainable and shown to the user rather than a black box.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import ClassVar

from openpyxl.workbook import Workbook

from mrg2opus.parsers.base import BaseMRGParser, UnclassifiedMRGError


@dataclass(frozen=True)
class LayoutProfile:
    lane_id: str
    parser_cls: type[BaseMRGParser]
    sheet_name_patterns: list[str] = field(default_factory=list)
    title_keywords: list[str] = field(default_factory=list)
    header_fingerprint: list[str] = field(default_factory=list)
    min_confidence: float = 0.7


@dataclass(frozen=True)
class ClassificationResult:
    profile: LayoutProfile
    confidence: float
    breakdown: dict[str, float]


_REGISTRY: list[LayoutProfile] = []


def register(profile: LayoutProfile) -> None:
    _REGISTRY.append(profile)


def _sheet_name_score(wb: Workbook, patterns: list[str]) -> float:
    if not patterns:
        return 0.0
    hits = 0
    for pattern in patterns:
        if any(re.search(pattern, name, re.IGNORECASE) for name in wb.sheetnames):
            hits += 1
    return hits / len(patterns)


def _title_keyword_score(wb: Workbook, keywords: list[str], scan_rows: int = 5) -> float:
    if not keywords:
        return 0.0
    text_blobs: list[str] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.upper().startswith("OPUS"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=scan_rows):
            for cell in row:
                if isinstance(cell.value, str):
                    text_blobs.append(cell.value.upper())
    blob = " ".join(text_blobs)
    hits = sum(1 for kw in keywords if kw.upper() in blob)
    return hits / len(keywords)


def _header_fingerprint_score(wb: Workbook, tokens: list[str], scan_rows: int = 20) -> float:
    if not tokens:
        return 0.0
    found: set[str] = set()
    for sheet_name in wb.sheetnames:
        if sheet_name.upper().startswith("OPUS"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=scan_rows):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in tokens:
                    found.add(cell.value.strip())
    return len(found) / len(tokens)


def classify_all(wb: Workbook) -> list[ClassificationResult]:
    """Score every registered lane against wb, best first. Used directly by
    classify() and by the UI's manual lane picker (Step 1) so a
    below-threshold upload still shows the full ranked breakdown instead of
    just an error."""
    results = []
    for profile in _REGISTRY:
        sheet_score = _sheet_name_score(wb, profile.sheet_name_patterns)
        title_score = _title_keyword_score(wb, profile.title_keywords)
        header_score = _header_fingerprint_score(wb, profile.header_fingerprint)
        confidence = 0.4 * sheet_score + 0.3 * title_score + 0.3 * header_score
        breakdown = {"sheet_name": sheet_score, "title_keywords": title_score, "header_fingerprint": header_score}
        results.append(ClassificationResult(profile=profile, confidence=confidence, breakdown=breakdown))
    return sorted(results, key=lambda r: r.confidence, reverse=True)


def all_profiles() -> list[LayoutProfile]:
    return list(_REGISTRY)


def get_profile(lane_id: str) -> LayoutProfile:
    for profile in _REGISTRY:
        if profile.lane_id == lane_id:
            return profile
    raise KeyError(f"No registered layout profile for lane_id={lane_id!r}")


def classify(wb: Workbook) -> ClassificationResult:
    results = classify_all(wb)
    best = results[0] if results else None

    if best is None or best.confidence < best.profile.min_confidence:
        raise UnclassifiedMRGError(
            f"No layout profile matched with sufficient confidence "
            f"(best={best.confidence if best else 0:.2f} for {best.profile.lane_id if best else 'none'})"
        )
    return best
