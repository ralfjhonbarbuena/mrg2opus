"""AUS NEA to AUEC (ex North/Central-East Asia to Australia East Coast) FAK
lane parser.

Three raw sheets, all sharing the exact same 48-origin grid layout (rows
7-54, footnote city-group definitions at rows 56-57) but different
destination(s):
  - "ex NEA to AUBNE_AUMEL": destination Brisbane/Melbourne (one combined
    POD) - feeds the main "EX NEA TO AUEC" commodity group.
  - "ex NEA to AUSYD": destination Sydney - same commodity group as above
    (confirmed: both share one CMDT NOTE block/cmdt_seq in ground truth).
  - "ex NEA to AUBNE on NZJ": destination Brisbane via a different vessel
    operator (NZJ) - its own SEPARATE commodity group, "EX NEA TO AUBNE
    ON NZJ".

Output scope is RATES + CMDT NOTE only (ground truth's own sheet is named
"SUR" - the already-known CMDT-NOTE naming drift), confirmed against both
real reference weeks (37, 38).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationMatch, LocationResolver
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.common.sequencing import assign_cmdt_seq_numbers
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import CHARGE_CODE_NAMES
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow

SHEET_MAIN = "ex NEA to AUBNE_AUMEL"
SHEET_AUSYD = "ex NEA to AUSYD"
SHEET_NZJ = "ex NEA to AUBNE on NZJ"

DATA_MIN_ROW, DATA_MAX_ROW = 7, 54
ORIGIN_COL = 2
POD_LABEL_ROW, POD_LABEL_COL = 4, 4
VALIDITY_ROW, VALIDITY_COL = 4, 2
FOOTNOTE_ROWS = (56, 57)

# Column -> RatesRow rate slot. RF/RAD only ever populate 20'/40'HC (no
# plain 40' column exists for either in the raw sheet) - confirmed across
# all 3 sheets, both reference weeks.
COL_DRY_20, COL_DRY_40, COL_DRY_40HC = 4, 5, 6
COL_RF_20, COL_RF_40HC = 7, 8
COL_RAD_20, COL_RAD_40HC = 9, 10

DEFAULT_MAIN_DESCRIPTION = "EX NEA TO AUEC"
DEFAULT_MAIN_CODE = "G0001"
DEFAULT_NZJ_DESCRIPTION = "EX NEA TO AUBNE ON NZJ"
DEFAULT_NZJ_CODE = "G0002"

# Hardcoded rather than parsed from the raw sheet's own "Remarks:" text -
# confirmed against both reference weeks' SUR sheets, both commodity
# groups. NOT recoverable from the raw remarks ("Rate incl OBS, EFS (ex KR
# & HK), PSS, subject to ISL/THL/THD/LND/DOC/DOF...", "Ex TW only: Rate
# incl OBS/ISL, PSS..."): each (code, pol) pair below is its own CMDT NOTE
# child row with that POL scope - ISL only for Taiwan-origin shipments,
# EFS as two separate rows scoped to Hong Kong and Korea respectively
# (both genuinely present in ground truth, not a duplicate-row bug, same
# category as CSE's real duplicate THL child row). Order matches ground
# truth's charge_seq order exactly (not alphabetical - only the
# "inclusive of" text line is alphabetized, same convention as
# cmdt_notes.py::build_cmdt_notes).
INCLUDED_CHARGES: list[tuple[str, str | None]] = [
    ("OBS", None),
    ("ISL", "TW"),
    ("EFS", "HK"),
    ("EFS", "KR"),
    ("PSS", None),
]

# Verified verbatim against reference/2_OPUS/37 and 38's RATES sheet (both
# identical) - see module docstring on why this is hardcoded rather than
# resolved per-city through the Location Bank: PRDA/PRDB group several
# genuinely new, ambiguous-without-context Chinese place names (e.g.
# "Longhua", "Sihui (Mafang)") that a fuzzy match could easily mis-resolve
# to an unrelated existing code - wrong freight rates are worse than a
# hardcoded lookup for a fixed, standing regional grouping confirmed
# byte-identical across both reference weeks.
PRDA_CODE = (
    "CNCAN;CNGGY;CNGOM;CNHDU;CNJMN;CNLGH;CNLIH;CNLUD;CNLUU;CNNSA;"
    "CNROQ;CNSHG;CNSJQ;CNSWA;CNXIL;CNXIN;CNYQS;CNZSN;CNZUH"
)
PRDA_DESCRIPTION = (
    "BEIJIAO, GUANGDONG;GAOMING, GUANGDONG;GONGYI, GUANGDONG;GUANGZHOU, GUANGDONG;"
    "HUADU, GUANGDONG;JIANGMEN, GUANGDONG;LELIU, GUANGDONG;LIANHUASHAN, GUANGDONG;"
    "LIUDU, GUANGDONG;LONGHUA, GUANGDONG;NANSHA, GUANGDONG;RONGQI, GUANGDONG;"
    "SANSHAN, GUANGDONG;SANSHUI, GUANGDONG;SHANTOU, GUANGDONG;XIAOLAN, GUANGDONG;"
    "XINHUI, GUANGDONG;ZHONGSHAN, GUANGDONG;ZHUHAI, GUANGDONG"
)
PRDB_CODE = (
    "CNBHY;CNDGG;CNDOU;CNFAN;CNGON;CNHAK;CNHSN;CNJJG;CNKPN;CNMFG;"
    "CNNAH;CNNGG;CNQYN;CNQZH;CNSBU;CNSIH;CNWUZ;CNYNF;CNZHA;CNZQG"
)
PRDB_DESCRIPTION = (
    "BEIHAI, GUANGXI;DONGGUAN, GUANGDONG;DOUMEN, GUANGDONG;FANGCHENG, GUANGXI;"
    "GAOLAN, GUANGDONG;HAIKOU, HAINAN;HESHAN, GUANGDONG;JIUJIANG, GUANGDONG;"
    "KAIPING, GUANGDONG;MAFANG, GUANGDONG;NANGANG, GUANGDONG;NANHAI, GUANGDONG;"
    "QINGYUAN, GUANGDONG;QINZHOU, GUANGXI;SANBU, GUANGDONG;SIHUI, GUANGDONG;"
    "WUZHOU, GUANGXI;YUNFU, GUANGDONG;ZHANJIANG, GUANGDONG;ZHAOQING, GUANGDONG"
)
_PRD_GROUPS = {"PRDA": (PRDA_CODE, PRDA_DESCRIPTION), "PRDB": (PRDB_CODE, PRDB_DESCRIPTION)}

_VIA_PAREN_RE = re.compile(r"\s*\(via\s+([^)]+?)(\s+by\s+rail)?\)", re.IGNORECASE)
_VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s+to\s+(\d{1,2})\s+(\w+)\s+(\d{4})", re.IGNORECASE)
_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=VALIDITY_ROW, column=VALIDITY_COL).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, d2, mon2, year = m.groups()
    try:
        y = int(year)
        start = date(y, _month_number(mon1), int(d1))
        end = date(y, _month_number(mon2), int(d2))
    except ValueError:
        return None, None
    return start, end


def _split_via(origin_text: str) -> tuple[str, str | None, bool]:
    """'Chengdu (via CNYTN by rail)' -> ('Chengdu', 'CNYTN', True).
    '(via CNSHA)' without 'by rail' sets O.Via but leaves origin_transmode
    blank - only the literal 'by rail' text sets origin_transmode='Rail'
    (origin_term stays 'CY' either way) - confirmed against ground truth
    (Chongqing (via CNSHA) has term=CY/transmode=None, Chongqing (via
    CNYTN by rail) has term=CY/transmode=Rail - both keep O.Via set)."""
    m = _VIA_PAREN_RE.search(origin_text)
    if not m:
        return origin_text.strip(), None, False
    via = m.group(1).strip()
    by_rail = bool(m.group(2))
    clean = _VIA_PAREN_RE.sub("", origin_text).strip()
    return clean, via, by_rail


def _build_group_cmdt_notes(
    validity_start: date | None,
    validity_end: date | None,
    excluded_codes: frozenset[str],
    rfa_effective: date | None,
    rfa_expiry: date | None,
) -> list[CmdtNoteRow]:
    """Lane-specific CMDT NOTE builder (not the shared cmdt_notes.py::
    build_cmdt_notes - that helper has no notion of a per-child-row POL
    scope, which this lane's ISL/EFS rows genuinely need, see
    INCLUDED_CHARGES above)."""
    if validity_start is None or validity_end is None:
        return []
    charges = [(code, pol) for code, pol in INCLUDED_CHARGES if code not in excluded_codes]
    if not charges:
        return []

    unique_codes = sorted(dict.fromkeys(code for code, _ in charges))
    names_line = " and the ".join(f"{CHARGE_CODE_NAMES.get(code, code)}({code})" for code in unique_codes)
    contents = "\n".join(
        [
            f"Rates are valid from {validity_start:%Y%m%d} to {validity_end:%Y%m%d}",
            f"Rates are inclusive of the {names_line}",
            "Rates are subject to all other surcharges, including those, if any, specified in "
            "the contract and those published in the Governing Tariff(s) at the time of shipment.",
        ]
    )
    parent = CmdtNoteRow(
        contents=contents, charge_seq=1, code="APP",
        application_effective=validity_start, application_expires=validity_end, application="S",
    )
    child_effective = rfa_effective if rfa_effective is not None else validity_start
    child_expires = rfa_expiry if rfa_expiry is not None else validity_end
    children = [
        CmdtNoteRow(
            charge_seq=i + 2, code=code, pol=pol,
            application_effective=child_effective, application_expires=child_expires, application="I",
        )
        for i, (code, pol) in enumerate(charges)
    ]
    return [parent, *children]


@dataclass
class OriginRow:
    origin_text: str
    values: dict[int, float]  # column index -> raw numeric value


@dataclass
class SheetData:
    destination_text: str
    origins: list[OriginRow] = field(default_factory=list)


@dataclass
class AUECRawData:
    validity_start: date | None
    validity_end: date | None
    main_sheets: list[SheetData] = field(default_factory=list)
    nzj_sheet: SheetData | None = None


def _read_pod_text(ws: Worksheet) -> str:
    """'POD: Melbourne/Brisbane' -> 'Melbourne/Brisbane'. Location Bank's
    match_text already splits on '/' (see fuzzy_match.split_location_text),
    so the combined POD resolves to both cities automatically."""
    raw = str(ws.cell(row=POD_LABEL_ROW, column=POD_LABEL_COL).value or "")
    return raw.split(":", 1)[1].strip() if ":" in raw else raw.strip()


def _read_sheet(ws: Worksheet) -> SheetData:
    data = SheetData(destination_text=_read_pod_text(ws))
    for row_idx in range(DATA_MIN_ROW, DATA_MAX_ROW + 1):
        origin_cell = ws.cell(row=row_idx, column=ORIGIN_COL)
        if origin_cell.value in (None, ""):
            continue
        if is_excluded(origin_cell):
            continue
        values: dict[int, float] = {}
        for col in (COL_DRY_20, COL_DRY_40, COL_DRY_40HC, COL_RF_20, COL_RF_40HC, COL_RAD_20, COL_RAD_40HC):
            v = ws.cell(row=row_idx, column=col).value
            if isinstance(v, (int, float)):
                values[col] = v
        if not values:
            continue
        data.origins.append(OriginRow(origin_text=str(origin_cell.value).strip(), values=values))
    return data


class AUECParser(BaseMRGParser):
    lane_id: ClassVar[str] = "AUEC"

    def __init__(self, location_resolver: LocationResolver | None = None):
        self.location_resolver = location_resolver or LocationResolver()

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        validity_start, validity_end = (None, None)
        main_sheets: list[SheetData] = []
        nzj_sheet: SheetData | None = None

        if SHEET_MAIN in wb.sheetnames:
            ws = wb[SHEET_MAIN]
            validity_start, validity_end = _parse_validity(ws)
            main_sheets.append(_read_sheet(ws))
        if SHEET_AUSYD in wb.sheetnames:
            ws = wb[SHEET_AUSYD]
            if validity_start is None:
                validity_start, validity_end = _parse_validity(ws)
            main_sheets.append(_read_sheet(ws))
        if SHEET_NZJ in wb.sheetnames:
            ws = wb[SHEET_NZJ]
            if validity_start is None:
                validity_start, validity_end = _parse_validity(ws)
            nzj_sheet = _read_sheet(ws)

        return RawExtraction(
            tables={
                "auec": AUECRawData(
                    validity_start=validity_start,
                    validity_end=validity_end,
                    main_sheets=main_sheets,
                    nzj_sheet=nzj_sheet,
                )
            }
        )

    def _resolve_origin(self, origin_text: str) -> tuple[str, str, str | None, bool] | None:
        """Returns (origin_code, origin_description, o_via_code, is_rail) or
        None if unresolved (skip the row rather than guess - same
        low-confidence-is-worse-than-no-match rationale as every other
        lane)."""
        clean, via, by_rail = _split_via(origin_text)

        prd_key = clean.rstrip("*").strip().upper()
        if prd_key in _PRD_GROUPS:
            code, desc = _PRD_GROUPS[prd_key]
        else:
            # Try the whole (unsplit) text first - handles single
            # locations written as "City, Province" (e.g. "Taizhou,
            # Jiangsu") that would otherwise be wrongly split into two
            # independent, incorrectly-matched tokens by the comma.
            whole = self.location_resolver.match_token(clean)
            matches: list[LocationMatch]
            if whole is not None and not whole.needs_review:
                matches = [whole]
            else:
                matches = self.location_resolver.match_text(clean)
            if not matches or any(m.needs_review for m in matches):
                return None
            code = ";".join(sorted({m.code for m in matches}))
            desc = ";".join(sorted({m.primary_name for m in matches}))

        via_code = None
        if via:
            via_match = self.location_resolver.match_token(via)
            if via_match and not via_match.needs_review:
                via_code = via_match.code

        return code, desc, via_code, by_rail

    def _build_rows_for_sheet(self, sheet: SheetData) -> tuple[list[RatesRow], list[RatesRow], list[RatesRow]]:
        """Returns (dr_rows, rf_rows, rad_dr_rows) for one raw sheet - DG
        duplicates are built by the caller from dr_rows."""
        dest_matches = self.location_resolver.match_text(sheet.destination_text)
        if not dest_matches or any(m.needs_review for m in dest_matches):
            return [], [], []
        dest_code = ";".join(sorted({m.code for m in dest_matches}))
        dest_desc = ";".join(sorted({m.primary_name for m in dest_matches}))

        dr_rows: list[RatesRow] = []
        rf_rows: list[RatesRow] = []
        rad_rows: list[RatesRow] = []

        for origin in sheet.origins:
            resolved = self._resolve_origin(origin.origin_text)
            if resolved is None:
                continue
            origin_code, origin_desc, via_code, by_rail = resolved
            v = origin.values

            base = dict(
                # Placeholder - overwritten by to_opus_rows()'s stamp()
                # once the final (post-override) group identity is known;
                # RatesRow requires these fields non-blank at construction.
                commodity_group_code="",
                commodity_group_description="",
                origin_code=origin_code,
                origin_description=origin_desc,
                origin_term="CY",
                origin_transmode="Rail" if by_rail else None,
                o_via_code=via_code,
                destination_code=dest_code,
                destination_description=dest_desc,
                destination_term="CY",
            )

            if COL_DRY_20 in v or COL_DRY_40 in v or COL_DRY_40HC in v:
                dr_rows.append(
                    RatesRow(
                        **base,
                        prefix="D",
                        cgo_type="DR",
                        cur_20="USD" if COL_DRY_20 in v else None,
                        rate_20=v.get(COL_DRY_20),
                        cur_40="USD" if COL_DRY_40 in v else None,
                        rate_40=v.get(COL_DRY_40),
                        cur_40hc="USD" if COL_DRY_40HC in v else None,
                        rate_40hc=v.get(COL_DRY_40HC),
                    )
                )
            if COL_RF_20 in v or COL_RF_40HC in v:
                rf_rows.append(
                    RatesRow(
                        **base,
                        prefix="R",
                        cgo_type="RF",
                        cur_20="USD" if COL_RF_20 in v else None,
                        rate_20=v.get(COL_RF_20),
                        cur_40hc="USD" if COL_RF_40HC in v else None,
                        rate_40hc=v.get(COL_RF_40HC),
                    )
                )
            # RAD ("Reefer As Dry"): physically a reefer container, but
            # priced/filed as dry cargo - Prefix R, CGO TYPE DR, same
            # 20'/40'HC-only rate shape as RF. Never gets a DG duplicate
            # (raw sheet's own remark: "For dry non-DG cargo only (*not
            # accepting DG in Reefer)").
            if COL_RAD_20 in v or COL_RAD_40HC in v:
                rad_rows.append(
                    RatesRow(
                        **base,
                        prefix="R",
                        cgo_type="DR",
                        cur_20="USD" if COL_RAD_20 in v else None,
                        rate_20=v.get(COL_RAD_20),
                        cur_40hc="USD" if COL_RAD_40HC in v else None,
                        rate_40hc=v.get(COL_RAD_40HC),
                    )
                )

        return dr_rows, rf_rows, rad_rows

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: AUECRawData = raw.tables["auec"]

        main_description = resolve_commodity_description(DEFAULT_MAIN_DESCRIPTION, config)
        main_code = resolve_commodity_code(DEFAULT_MAIN_DESCRIPTION, DEFAULT_MAIN_CODE, config)

        nzj_description = resolve_commodity_description(DEFAULT_NZJ_DESCRIPTION, config)
        nzj_code = resolve_commodity_code(DEFAULT_NZJ_DESCRIPTION, DEFAULT_NZJ_CODE, config)

        block_seq = assign_cmdt_seq_numbers([DEFAULT_MAIN_DESCRIPTION, DEFAULT_NZJ_DESCRIPTION], config.commodity_sequence_overrides)
        main_cmdt_seq = block_seq[DEFAULT_MAIN_DESCRIPTION]
        nzj_cmdt_seq = block_seq[DEFAULT_NZJ_DESCRIPTION]

        def stamp(rows: list[RatesRow], code: str, description: str, cmdt_seq: int | None) -> None:
            for row in rows:
                row.commodity_group_code = code
                row.commodity_group_description = description
                row.cmdt_seq = cmdt_seq

        def build_group(sheets: list[SheetData], default_description: str) -> list[RatesRow]:
            non_dg: list[RatesRow] = []
            dg: list[RatesRow] = []
            for sheet in sheets:
                dr_rows, rf_rows, rad_rows = self._build_rows_for_sheet(sheet)
                non_dg.extend(group_by_destination([*dr_rows, *rf_rows, *rad_rows]))
                if not config.skip_dg_generation.get(default_description, False):
                    dg.extend(row.model_copy(update={"cgo_type": "DG"}) for row in dr_rows)
            return [*non_dg, *group_by_destination(dg)]

        main_rows = build_group(data.main_sheets, DEFAULT_MAIN_DESCRIPTION)
        stamp(main_rows, main_code, main_description, main_cmdt_seq)

        nzj_rows: list[RatesRow] = []
        if data.nzj_sheet is not None:
            nzj_rows = build_group([data.nzj_sheet], DEFAULT_NZJ_DESCRIPTION)
            stamp(nzj_rows, nzj_code, nzj_description, nzj_cmdt_seq)

        for group_rows in (main_rows, nzj_rows):
            for i, row in enumerate(group_rows, start=1):
                row.route_seq = i

        rates = [*nzj_rows, *main_rows]

        excluded_codes = frozenset(config.excluded_charge_codes)
        cmdt_notes: list[CmdtNoteRow] = []
        if nzj_rows:
            nzj_notes = _build_group_cmdt_notes(
                data.validity_start, data.validity_end, excluded_codes,
                config.rfa_effective_date, config.rfa_expiry_date,
            )
            cmdt_notes.extend(row.model_copy(update={"group_description": nzj_description, "header_seq": nzj_cmdt_seq}) for row in nzj_notes)
            note_text = nzj_notes[0].contents if nzj_notes else None
            for row in nzj_rows:
                row.commodity_note = note_text
        if main_rows:
            main_notes = _build_group_cmdt_notes(
                data.validity_start, data.validity_end, excluded_codes,
                config.rfa_effective_date, config.rfa_expiry_date,
            )
            cmdt_notes.extend(row.model_copy(update={"group_description": main_description, "header_seq": main_cmdt_seq}) for row in main_notes)
            note_text = main_notes[0].contents if main_notes else None
            for row in main_rows:
                row.commodity_note = note_text

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)


register(
    LayoutProfile(
        lane_id=AUECParser.lane_id,
        parser_cls=AUECParser,
        sheet_name_patterns=[re.escape(SHEET_MAIN), re.escape(SHEET_AUSYD), re.escape(SHEET_NZJ)],
        title_keywords=["ONE AU MINIMUM RATE GUIDELINE"],
        header_fingerprint=["20'RAD", "40'HC RAD"],
    )
)
