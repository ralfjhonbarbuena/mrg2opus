"""TAD FILING OEW OMW lane parser.

Two raw sheets, "OEW" and "OMW", each its own Service Scope and own OPUS
output file. Unlike every other lane built so far, the raw sheet is
ALREADY pre-structured in the team's own internal "DATA sheet" shape
(same column set as "Tool for TAD.xlsm" - see project_tad_vba_tool_
analysis memory) - every origin (POR) and destination (DEL) code +
description is given directly, so this lane needs NO Location Bank
fuzzy matching at all. Every rate/date cell is stored as literal text
("1,941", "2026-09-01"), not a native Excel number/date.

Output scope is RATES + CMDT NOTE + ROUTE NOTE (confirmed against
reference/2_OPUS/25_TAD FILING OEW OMW - the real filing's own ROUTE
NOTE sheet is literally named that, not "RN" like every other lane built
so far, see SHEET_NAME_OVERRIDES below). VERTICAL RATES is a general,
user-toggled OPUS feature (see schema/opus_rows.py::build_vertical_rates),
not something this parser generates itself.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.tad_snapshots import find_snapshot_sheets, merge_dated_snapshots
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.charge_codes import CHARGE_CODE_NAMES
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow, RouteNoteRow

SHEET_OEW = "OEW"
SHEET_OMW = "OMW"

DATA_MIN_ROW = 2
COL_EFF_DATE, COL_EXP_DATE, COL_SCOPE, COL_CMDT_NAME = 1, 2, 3, 4
COL_POR, COL_POR_DESC, COL_ORIGIN_TERM, COL_ORIGIN_TRANSMODE = 5, 6, 7, 8
COL_POL, COL_POD, COL_DEL, COL_DEL_DESC, COL_DEST_TERM = 9, 10, 11, 12, 13
COL_CARGO_TYPE, COL_TS_PORT, COL_SERVICE_LANE, COL_CUR = 14, 15, 16, 17
COL_OFT20, COL_OFT40, COL_OFTHC, COL_OFT45, COL_INCLUDE_SURCHARGE = 18, 19, 20, 21, 22

# CARGO sheet equivalent (same 4-entry mapping confirmed in "Tool for
# TAD.xlsm"'s own CARGO sheet - Reefer Dry is the same "physically
# reefer, filed as dry cargo" concept as AUEC's "RAD"/AUWC's "NOR",
# a third name for the identical business rule).
CARGO_TYPE_MAP: dict[str, tuple[str, str]] = {
    "Dry General": ("D", "DR"),
    "Reefer": ("R", "RF"),
    "Reefer Dry": ("R", "DR"),
    "Dry Dangerous": ("D", "DG"),
}

DEFAULT_COMMODITY_DESCRIPTION = "FAK"
DEFAULT_COMMODITY_CODE = "G0001"

# Hardcoded (raw "Commodity Group Name" value -> RATES/ROUTE NOTE Contents
# text) - confirmed against reference/2_OPUS/25_TAD FILING OEW OMW/OMW.xlsx,
# the only real example seen so far. These 4 raw values do NOT create a
# separate commodity group (RATES.commodity_group_description stays plain
# "FAK" for these rows too) - they only trigger a route note, the same
# mechanism as a populated T/S Port column (see _route_note_for below).
# NOT derivable by formula from the raw text alone (e.g. "ACCHCO" and the
# exact wording aren't present in the raw tag) - only extend this table
# when a new special-node case is directly confirmed against real ground
# truth, same rule as every other hardcoded lookup in this project.
SPECIAL_NODE_ROUTE_NOTES: dict[str, str] = {
    "FAK Alexandria (EGALY20 DEKHEILA PORT)": "EGALY20 - DEKHEILA PORT (ACCHCO)",
    "FAK Alexandria (EGALY21 Old Port)": "EGALY21 - ALEXANDRIA OLD PORT ( ACCHCO )",
    "HAYDARPASA FAK": "TRIST21 - PORT OF HAYDARPASA (ISTANBUL)",
    "MARPORT FAK": "TRIST02 - ISTANBUL MARPORT",
}

# TAD's own CMDT NOTE boilerplate is one comma short of the shared
# cmdt_notes.py::build_cmdt_notes() template ("surcharges including",
# not "surcharges, including") - confirmed against ground truth AND the
# TAD VBA tool's own GenerateWording() sub, which hardcodes this exact
# wording. HEA is also named "HEAVY SURCHARGE" here (matching LAEC's own
# confirmed override, NOT this project's EAF/WAF/AUEC-confirmed "HEAVY
# WEIGHT SURCHARGE") - both reasons this lane needs its own builder
# rather than the shared one.
_HEA_OVERRIDE = "HEAVY SURCHARGE"


def _route_note_for(commodity_name: str, ts_port: str | None) -> str | None:
    """A row's route note comes from EITHER a populated T/S Port column OR
    a recognized special-node commodity-group-name tag - never both at
    once in any real row seen so far, so this doesn't attempt to combine
    them; T/S Port wins if a future file somehow has both (arbitrary,
    unverified - flag if that combination ever shows up in ground truth).
    """
    if ts_port:
        return f"Rates are Subject to Transhipment Port: {ts_port}"
    return SPECIAL_NODE_ROUTE_NOTES.get(commodity_name)


def _parse_decimal(text: str | None) -> Decimal | None:
    if not text:
        return None
    try:
        return Decimal(str(text).replace(",", "").strip())
    except InvalidOperation:
        return None


def _parse_date(text: str | None) -> date | None:
    if not text:
        return None
    try:
        return date.fromisoformat(str(text).strip())
    except ValueError:
        return None


def _parse_included_codes(text: str | None) -> list[str]:
    """DATA's own 'Include Surcharge' column is already a literal,
    authoritative comma-separated charge-code list (e.g.
    'CAF,CSS,EFS,HEA,MBS') - no free-text parsing or lane-wide hardcoding
    needed, unlike every other lane built so far. Deliberately NOT
    filtered through INDIVIDUAL_CHARGE_CODES - that frozenset exists to
    disambiguate OTHER lanes' free-text "Includes X/Y/Z" parsing (see
    common/cmdt_notes.py::parse_included_charge_codes) and does not
    include CAF, even though CAF is confirmed (ground truth: OEW POLLY.xlsx
    / OMW.xlsx CMDT NOTE, both list CAF as a child row) to always be a
    legitimate, individually-filed code here."""
    if not text:
        return []
    return [c.strip().upper() for c in str(text).split(",") if c.strip()]


@dataclass
class RawRow:
    validity_start: date | None
    validity_end: date | None
    commodity_name: str
    origin_code: str
    origin_description: str
    origin_term: str | None
    origin_transmode: str | None
    dest_code: str
    dest_description: str
    dest_term: str | None
    cargo_type: str
    ts_port: str | None
    included_codes: list[str]
    rate_20: Decimal | None
    rate_40: Decimal | None
    rate_40hc: Decimal | None
    rate_45: Decimal | None


@dataclass
class ScopeData:
    scope: str
    rows: list[RawRow] = field(default_factory=list)


def _read_sheet(ws: Worksheet, scope: str) -> ScopeData:
    data = ScopeData(scope=scope)
    for row_idx in range(DATA_MIN_ROW, ws.max_row + 1):
        origin = ws.cell(row=row_idx, column=COL_POR).value
        if origin in (None, ""):
            continue
        origin_cell = ws.cell(row=row_idx, column=COL_POR)
        if is_excluded(origin_cell):
            continue
        data.rows.append(
            RawRow(
                validity_start=_parse_date(ws.cell(row=row_idx, column=COL_EFF_DATE).value),
                validity_end=_parse_date(ws.cell(row=row_idx, column=COL_EXP_DATE).value),
                commodity_name=str(ws.cell(row=row_idx, column=COL_CMDT_NAME).value or "").strip(),
                origin_code=str(origin).strip(),
                origin_description=str(ws.cell(row=row_idx, column=COL_POR_DESC).value or "").strip(),
                origin_term=str(ws.cell(row=row_idx, column=COL_ORIGIN_TERM).value or "").strip() or None,
                origin_transmode=str(ws.cell(row=row_idx, column=COL_ORIGIN_TRANSMODE).value or "").strip() or None,
                dest_code=str(ws.cell(row=row_idx, column=COL_DEL).value or "").strip(),
                dest_description=str(ws.cell(row=row_idx, column=COL_DEL_DESC).value or "").strip(),
                dest_term=str(ws.cell(row=row_idx, column=COL_DEST_TERM).value or "").strip() or None,
                cargo_type=str(ws.cell(row=row_idx, column=COL_CARGO_TYPE).value or "").strip(),
                ts_port=str(ws.cell(row=row_idx, column=COL_TS_PORT).value or "").strip() or None,
                included_codes=_parse_included_codes(ws.cell(row=row_idx, column=COL_INCLUDE_SURCHARGE).value),
                rate_20=_parse_decimal(ws.cell(row=row_idx, column=COL_OFT20).value),
                rate_40=_parse_decimal(ws.cell(row=row_idx, column=COL_OFT40).value),
                rate_40hc=_parse_decimal(ws.cell(row=row_idx, column=COL_OFTHC).value),
                rate_45=_parse_decimal(ws.cell(row=row_idx, column=COL_OFT45).value),
            )
        )
    return data


class TADOewOmwParser(BaseMRGParser):
    lane_id: ClassVar[str] = "TAD-OEW-OMW"
    SHEET_NAME_OVERRIDES: ClassVar[dict[str, str]] = {"route_notes": "ROUTE NOTE"}

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        names = set(wb.sheetnames)
        if not ({SHEET_OEW, SHEET_OMW} & names):
            return 0.0
        score = 0.5 * len({SHEET_OEW, SHEET_OMW} & names)
        ws = wb[SHEET_OEW] if SHEET_OEW in names else wb[SHEET_OMW]
        header_tokens = {str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, 23)}
        if "Include Surcharge" in header_tokens and "DEL Description" in header_tokens:
            score = min(score + 0.3, 1.0)
        return score

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        scopes: dict[str, ScopeData] = {}
        for scope_name in (SHEET_OEW, SHEET_OMW):
            sheet_names = find_snapshot_sheets(wb, scope_name)
            if not sheet_names:
                continue
            occurrences = [_read_sheet(wb[name], scope_name).rows for name in sheet_names]
            scopes[scope_name] = ScopeData(scope=scope_name, rows=merge_dated_snapshots(occurrences))
        return RawExtraction(tables={"scopes": scopes})

    def _build_one_scope(self, data: ScopeData, config: MappingProfile) -> OpusRowSet:
        description = resolve_commodity_description(DEFAULT_COMMODITY_DESCRIPTION, config)
        code = resolve_commodity_code(DEFAULT_COMMODITY_DESCRIPTION, DEFAULT_COMMODITY_CODE, config)
        excluded_codes = frozenset(config.excluded_charge_codes)

        # CMDT NOTE grouping is data-driven (DATA's own per-row validity
        # dates + Include Surcharge list), not a lane-wide constant like
        # every other lane built so far - GroupedCargoNumbering in the TAD
        # VBA tool works the same way. In every real file seen, every row
        # shares one (dates, codes) combo, so this degenerates to a single
        # group, but the general case is handled rather than assumed.
        group_order: list[tuple] = []
        group_rows: dict[tuple, list[RawRow]] = defaultdict(list)
        for row in data.rows:
            key = (row.validity_start, row.validity_end, tuple(row.included_codes))
            if key not in group_rows:
                group_order.append(key)
            group_rows[key].append(row)
        cmdt_seq_by_key = {key: i + 1 for i, key in enumerate(group_order)}

        rates: list[RatesRow] = []
        for row in data.rows:
            mapped = CARGO_TYPE_MAP.get(row.cargo_type)
            if mapped is None:
                continue
            prefix, cgo_type = mapped
            key = (row.validity_start, row.validity_end, tuple(row.included_codes))
            rates.append(
                RatesRow(
                    cmdt_seq=cmdt_seq_by_key[key],
                    commodity_group_code=code,
                    commodity_group_description=description,
                    origin_code=row.origin_code,
                    origin_description=row.origin_description,
                    origin_term=row.origin_term,
                    origin_transmode=row.origin_transmode,
                    destination_code=row.dest_code,
                    destination_description=row.dest_description,
                    destination_term=row.dest_term,
                    prefix=prefix,
                    cgo_type=cgo_type,
                    cur_20="USD" if row.rate_20 is not None else None,
                    rate_20=row.rate_20,
                    cur_40="USD" if row.rate_40 is not None else None,
                    rate_40=row.rate_40,
                    cur_40hc="USD" if row.rate_40hc is not None else None,
                    rate_40hc=row.rate_40hc,
                    cur_45="USD" if row.rate_45 is not None else None,
                    rate_45=row.rate_45,
                    route_note=_route_note_for(row.commodity_name, row.ts_port),
                )
            )

        # Opt-in only (config.generate_tad_dg_duplicate, default off) -
        # mirrors the team's own VBA tool's "Include Dry Dangerous" toggle
        # (see project_tad_vba_tool_analysis memory). A duplicate shares
        # its parent's cmdt_seq (same validity/charge-code key - only
        # cgo_type changes), so it needs no separate group bookkeeping;
        # inserted before Route Seq. numbering below so it's included.
        if config.generate_tad_dg_duplicate:
            rates.extend(
                row.model_copy(update={"cgo_type": "DG"}) for row in list(rates) if row.prefix == "D" and row.cgo_type == "DR"
            )

        # Route Seq.: one continuous counter per commodity group (same
        # scope rule confirmed for every lane so far - see auec.py/
        # auwc.py's own findings).
        for key in group_order:
            group = [r for r in rates if r.cmdt_seq == cmdt_seq_by_key[key]]
            for i, r in enumerate(group, start=1):
                r.route_seq = i

        cmdt_notes: list[CmdtNoteRow] = []
        note_text_by_seq: dict[int, str | None] = {}
        for key in group_order:
            validity_start, validity_end, codes = key
            seq = cmdt_seq_by_key[key]
            notes = self._build_cmdt_notes(validity_start, validity_end, list(codes), excluded_codes, config)
            for note in notes:
                note.header_seq = seq
            cmdt_notes.extend(notes)
            note_text_by_seq[seq] = notes[0].contents if notes else None

        for row in rates:
            row.commodity_note = note_text_by_seq.get(row.cmdt_seq)

        route_notes: list[RouteNoteRow] = []
        validity_by_seq = {cmdt_seq_by_key[key]: (key[0], key[1]) for key in group_order}
        for row in rates:
            if not row.route_note:
                continue
            validity_start, validity_end = validity_by_seq.get(row.cmdt_seq, (None, None))
            route_notes.append(
                RouteNoteRow(
                    header_seq=row.cmdt_seq,
                    route_seq=row.route_seq,
                    contents=row.route_note,
                    charge_seq=1,
                    code="APP",
                    application_effective=validity_start,
                    application_expires=validity_end,
                    application="S",
                )
            )

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes, route_notes=route_notes)

    def _build_cmdt_notes(
        self,
        validity_start: date | None,
        validity_end: date | None,
        included_codes: list[str],
        excluded_codes: frozenset[str],
        config: MappingProfile,
    ) -> list[CmdtNoteRow]:
        codes = [c for c in included_codes if c not in excluded_codes]
        if not codes or validity_start is None or validity_end is None:
            return []
        unique_codes = sorted(dict.fromkeys(codes))
        names = {**CHARGE_CODE_NAMES, "HEA": _HEA_OVERRIDE}
        names_line = " and the ".join(f"{names.get(c, c)}({c})" for c in unique_codes)
        contents = "\n".join(
            [
                f"Rates are valid from {validity_start:%Y%m%d} to {validity_end:%Y%m%d}",
                f"Rates are inclusive of the {names_line}",
                "Rates are subject to all other surcharges including those, if any, "
                "specified in the contract and those published in the Governing Tariff(s) at the time of shipment.",
            ]
        )
        parent = CmdtNoteRow(
            contents=contents, charge_seq=1, code="APP",
            application_effective=validity_start, application_expires=validity_end, application="S",
        )
        child_effective = config.rfa_effective_date if config.rfa_effective_date is not None else validity_start
        child_expires = config.rfa_expiry_date if config.rfa_expiry_date is not None else validity_end
        children = [
            CmdtNoteRow(
                charge_seq=i + 2, code=c,
                application_effective=child_effective, application_expires=child_expires, application="I",
            )
            for i, c in enumerate(codes)
        ]
        return [parent, *children]

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        # Single-scope entry point (BaseMRGParser.run()) - used when only
        # one of OEW/OMW is present in the workbook.
        scopes: dict[str, ScopeData] = raw.tables["scopes"]
        if not scopes:
            return OpusRowSet()
        return self._build_one_scope(next(iter(scopes.values())), config)

    def run_multi(self, wb: Workbook, config: MappingProfile | None = None) -> dict[str, OpusRowSet]:
        config = config or MappingProfile()
        raw = self.parse_raw(wb)
        scopes: dict[str, ScopeData] = raw.tables["scopes"]
        return {scope: self._build_one_scope(data, config) for scope, data in scopes.items()}


register(
    LayoutProfile(
        lane_id=TADOewOmwParser.lane_id,
        parser_cls=TADOewOmwParser,
        sheet_name_patterns=[r"^OEW$", r"^OMW$"],
        title_keywords=["EFFECTIVE DATE", "COMMODITY GROUP NAME", "INCLUDE SURCHARGE"],
        header_fingerprint=["DEL Description", "Include Surcharge"],
    )
)
