"""AUS SEA to AUBP (ex Southeast Asia to Australia Brisbane/Perth-area) FAK
lane parser.

Three raw sheets, all flat row-per-origin tables (NOT a POD-header grid like
AUEC/AUWC) sharing the same 83-origin list (rows 6-84 on the main sheet,
5-87 on the other two - see DATA_MIN_ROW/DATA_MAX_ROW per sheet):
  - "ex SEA to MEL BNE ADL": destinations Brisbane/Melbourne/Adelaide, split
    across 5 column-groups (Dry applies to all 3; Reefer and RAD each split
    into a BNE/MEL-only pair of columns and an ADL-only pair) - row 4 gives
    the actual destination CODE(s) for each column-group directly (e.g.
    "AUBNE/AUMEL/AUADL"), so destinations need only a "/"-split, never
    fuzzy Location Bank matching.
  - "ex SEA to SYD" / "ex SEA to FRE": single destination per sheet (POD
    column reads "Sydney"/"Fremantle" on every row - resolved once via the
    Location Bank, same as AUEC's own POD-label pattern), Dry/Reefer/RAD in
    3 simple column pairs each.

Output scope is RATES + SUR (the already-known CMDT-NOTE naming drift, see
project-opus-note-sheet-taxonomy memory), confirmed against both real
reference weeks (folders 29, 30).

**Commodity groups**: 3 distinct, NOT folded together - confirmed via
ground truth's own (cmdt_seq, code, description) tuple counts, matching
AUWC's shape (RF/RAD each get their own group), not AUEC's (RAD folds into
main):
  - G0001 "FAK - SEA": Dry (D/DR + auto D/DG duplicate at the same rate).
  - G0002 "REEFER - SEA": Reefer (R/RF).
  - G0003 "NOR - SEA": RAD (R/DR - a reefer container filed as dry cargo,
    same concept as AUEC's "RAD"/AUWC's "NOR", just this lane's own default
    description is "NOR - SEA"). For BOTH Reefer and RAD, the raw sheet's
    own "40'" rate column maps into the OPUS 40HC slot, not 40' (which
    stays blank) - confirmed across every RF/RAD row in ground truth; Dry
    uses the normal 20'/40'/40'HC 3-column layout as-is.

**Destination merging**: ground truth does NOT emit one row per
origin-per-POD; it emits one row per origin per column-group, with every
POD in that group's destination joined into ONE row
(destination_code="AUADL;AUBNE;AUMEL", not 3 separate rows) - codes and
descriptions are each independently alphabetized (not a positionally
zipped pair - same convention used for the India ICD/cluster origins
below).

**Origin resolution - mostly hardcoded, not fuzzy-matched.** Location Bank
fuzzy matching was tried first and produced WRONG matches (not just
low-confidence "needs_review" ones - confidently wrong top hits) for a
concerning fraction of this lane's real origin names (e.g. "Kuantan" ->
Yantian China; "Muara" -> Mundra India; "Balikpapan" -> Apapa Nigeria;
"Cagayan" -> Taoyuan Taiwan) - this is a short, fixed, real-world list of
~64 named ports, safer to hardcode verbatim against both reference weeks'
ground truth than to trust fuzzy scoring on short ASEAN/Indian city names
that collide with unrelated global ports. ORIGIN_OVERRIDES below is the
single source of truth; the Location Bank is not consulted for plain
origin names at all in this lane (unlike AUEC/AUWC, which fuzzy-match
freely and only hardcode narrow regional groups).

Two names are genuinely ABSENT from the Location Bank outright (not just
mis-scored): "Batam" (IDBTM) and "Thilawa" (MMTLA) - both still resolve
correctly here since ORIGIN_OVERRIDES supplies the code/description
directly rather than depending on the bank.

A handful of raw origin names carry a parenthetical suffix that needs
special handling before the base name is looked up in ORIGIN_OVERRIDES:
  - "Batam (Door/CY)" -> term override: the word before the slash becomes
    origin_term ("Door"), replacing the lane's usual default of "CY".
  - "Yangon (Freight Collect only)" / "Thilawa (Freight Collect only)" ->
    drives route_note "(FRT collect only)"; see the Yangon/Thilawa merge
    note below.
  - "Tuticorin (via Colombo)" / "Cochin (via Colombo)" -> dropped entirely,
    no effect (NOT parsed as an O.Via the way AUEC/AUWC's "(via X by
    rail)" convention works - confirmed against ground truth, O.Via stays
    blank for these).
  - "Chittagong (incl THL)" / "Colombo (incl ISL/THL/DOC)" -> drives
    per-origin SUR/CMDT-NOTE charge-code scoping, see below. (The same
    "(incl THL/OCR)" annotation also appears on 3 Middle-East origins
    whose rates are all "N/A" - those origins produce zero output rows
    regardless, so the annotation is naturally never surfaced for them.)

**Yangon/Thilawa merge - a real, narrow special case, not a general
"merge identical rates" rule.** Both origins carry byte-identical rates in
every raw sheet/column-group they appear in (Dry only - neither has
Reefer/RAD data) and both share the "(Freight Collect only)" annotation.
Week 2's ground truth merges them into ONE row (origin_code
"MMRGN;MMTLA") with the FRT-collect route note; week 1's ground truth
instead drops Thilawa entirely (both weeks end up with the same row count
- week 1 by omission, week 2 by merge - which points at week 1 being a
human ground-truth defect, not a second valid design). This parser always
merges (matches week 2); week 1's own test carries a documented, narrow
exception for exactly this pair - see tests/test_parsers_aubp.py.

**CMDT NOTE / SUR per-origin charge scoping**: OBS/EFS/PSS are blanket
(no scope) for G0001 and G0002 (both share the exact same note text and
child-row set in ground truth, regardless of whether a given scoped
origin actually has data in that particular group); THL/ISL/DOC are
additionally scoped to whichever real origins carry a "(incl ...)"
annotation AND have at least one Dry rate (Chittagong/BDCGP for THL only,
Colombo/LKCMB for all three) - derived from the raw sheet's own
annotations rather than hardcoded origin codes, so a future week's
different scoped origin(s) would flow through automatically. G0003 (RAD)
gets a permanently reduced note (OBS/EFS/PSS only, no origin scoping ever)
- confirmed identical both weeks, matching AUWC's finding that a
reefer-like sub-group can carry different default charge coverage than
the main group. The scoping column is POR (not POL) - week 1's own ground
truth used POL for the exact same rule, week 2 switched to POR for
identical business logic; POR is treated as the intended convention
(matches AUWC's own POR-based scoping) and POL is not reproduced. Charge
child rows are emitted grouped by CODE (all THL rows, then ISL, then DOC),
each code's rows in raw-sheet origin order - confirmed exact match against
ground truth's own charge_seq order.

**Route Seq.**: one continuous counter per commodity group, spanning every
raw sheet and every DR/DG duplication within that group - confirmed
against ground truth (1->390 for G0001 across all 3 raw sheets' Dry data
plus its DG duplicate, 1->144 for G0002, 1->30 for G0003) - same "count
every row in this group's final output order" scope as AUEC's own finding,
not narrower.

India ICD/cluster origins ("ICD Delhi", "Ahmedabad ICD", "Ludhiana
Cluster", "Indore Cluster", "Faridabad Cluster", "Moradabad Cluster") each
expand to several individual INxxx codes via a footnote block elsewhere on
the main raw sheet (rows 86-91: "ICD Delhi: IN1MZ,INCML,...") - the codes
ARE given directly in that footnote text, but NONE of those ~28 individual
INxxx codes exist in the Location Bank at all (confirmed via
LocationBankStore.get_by_code), so their descriptions are hardcoded from
ground truth verbatim below (ICD_CLUSTERS), same "verified, stable,
narrow regional group" rationale as AUEC's PRDA/PRDB groups.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.fuzzy_match import LocationResolver
from mrg2opus.location_bank.store import LocationBankStore
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.cmdt_notes import build_cmdt_notes
from mrg2opus.parsers.common.commodity import resolve_commodity_code, resolve_commodity_description
from mrg2opus.parsers.common.exclusion import is_excluded
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import CmdtNoteRow, OpusRowSet, RatesRow

SHEET_MAIN = "ex SEA to MEL BNE ADL"
SHEET_SYD = "ex SEA to SYD"
SHEET_FRE = "ex SEA to FRE"

ORIGIN_COL = 2
POD_COL = 3
ROW4_DEST_ROW = 4

DEFAULT_MAIN_DESCRIPTION = "FAK - SEA"
DEFAULT_MAIN_CODE = "G0001"
DEFAULT_RF_DESCRIPTION = "REEFER - SEA"
DEFAULT_RF_CODE = "G0002"
DEFAULT_RAD_DESCRIPTION = "NOR - SEA"
DEFAULT_RAD_CODE = "G0003"

# Verified verbatim against both reference weeks' RATES sheets - see module
# docstring for why plain origin names are hardcoded rather than
# fuzzy-matched in this lane. Keyed by the origin's cleaned base text
# (after stripping any parenthetical suffix), matched case-insensitively.
ORIGIN_OVERRIDES: dict[str, tuple[str, str]] = {
    "tanjung pelepas": ("MYTPP", "TANJUNG PELEPAS"),
    "port klang": ("MYPKG", "PORT KLANG"),
    "penang": ("MYPEN", "PENANG"),
    "pasir gudang": ("MYPGU", "PASIR GUDANG"),
    "kuantan": ("MYKUA", "KUANTAN"),
    "bintulu": ("MYBTU", "BINTULU"),
    "kuching": ("MYKCH", "KUCHING"),
    "muara": ("BNMUA", "MUARA"),
    "singapore": ("SGSIN", "SINGAPORE"),
    "bangkok": ("THBKK", "BANGKOK"),
    "laem chabang": ("THLCH", "LAEM CHABANG"),
    "lat krabang": ("THLKR", "LAT KRABANG"),
    "songkhla": ("THSGZ", "SONGKHLA"),
    "jakarta": ("IDJKT", "JAKARTA"),
    "surabaya": ("IDSUB", "SURABAYA"),
    "semarang": ("IDSRG", "SEMARANG"),
    "belawan": ("IDBLW", "BELAWAN"),
    "kuala tanjung": ("IDKTJ", "KUALA TANJUNG"),
    "panjang": ("IDPNJ", "PANJANG"),
    "makassar": ("IDMAK", "MAKASSAR"),
    "bitung": ("IDBIT", "BITUNG"),
    "palembang": ("IDPLM", "PALEMBANG"),
    "padang": ("IDPDG", "PADANG"),
    "banjarmasin": ("IDBDJ", "BANJARMASIN"),
    "balikpapan": ("IDBPN", "BALIKPAPAN"),
    "pontianak": ("IDPNK", "PONTIANAK"),
    "batam": ("IDBTM", "BATAM"),  # absent from Location Bank entirely
    "samarinda": ("IDSRI", "SAMARINDA"),
    "ho chi minh": ("VNSGN", "HO CHI MINH"),
    "cai mep": ("VNCMP", "CAI MEP"),
    "haiphong": ("VNHPH", "HAI PHONG"),
    "bien hoa": ("VNBHA", "DONG NAI, BIEN HOA"),
    "di an": ("VNDIA", "DI AN, BINH DUONG"),
    "danang": ("VNDAD", "DA NANG"),
    "qui nhon": ("VNUIH", "QUI NHON"),
    "sihanouk ville": ("KHKOS", "SIHANOUKVILLE"),
    "phnom penh": ("KHPNH", "PHNOM PENH"),
    "yangon": ("MMRGN", "YANGON"),
    "thilawa": ("MMTLA", "THILAWA"),  # absent from Location Bank entirely
    "manila": ("PHMNL", "MANILA"),
    "subic bay": ("PHSFS", "SUBIC BAY"),
    "cebu": ("PHCEB", "CEBU"),
    "davao": ("PHDVO", "DAVAO"),
    "general santos": ("PHGES", "GENERAL SANTOS"),
    "cagayan": ("PHCGY", "CAGAYAN DE ORO"),
    "nhava sheva": ("INNSA", "NHAVA SHEVA"),
    "pipavav": ("INPAV", "PIPAVAV"),
    "mundra": ("INMUN", "MUNDRA"),
    "chennai": ("INMAA", "CHENNAI"),
    "kattupalli": ("INKTP", "KATTUPALLI PORT"),
    "hazira": ("INHZA", "HAZIRA"),
    "visakhapatnam": ("INVTZ", "VISAKHAPATNAM"),
    "haldia": ("INHAL", "HALDIA"),
    "kolkata / calcutta": ("INCCU", "KOLKATA"),
    "tuticorin": ("INTUT", "TUTICORIN"),
    "cochin": ("INCOK", "COCHIN"),
    "karachi": ("PKKHI", "KARACHI"),
    "muhammad bin qasim": ("PKBQM", "MUHAMMAD BIN QASIM"),
    "chittagong": ("BDCGP", "CHITTAGONG"),
    "mongla": ("BDMGL", "MONGLA"),  # always "N/A"-rated - never actually reaches output
    "colombo": ("LKCMB", "COLOMBO"),
}

# India ICD/cluster origins: the raw sheet's own footnote block (rows
# 86-91 on the main sheet) gives the exact member codes directly, but none
# of those individual codes exist in the Location Bank - see module
# docstring. Hardcoded verbatim from both reference weeks' ground truth
# (byte-identical set both weeks); values are the FINAL semicolon-joined,
# independently-alphabetized strings ground truth itself uses, so no
# per-code pairing is needed.
ICD_CLUSTERS: dict[str, tuple[str, str]] = {
    "icd delhi": (
        "IN1MZ;INCML;INDRI;INGGN;INGHR;INGZD;INPAL;INPAP;INREA;INSON;INTKD",
        "DADRI;DELHI - CONCOR ICD   - TUGHLAKABAD;GURUGRAM - GARHI HARSARU ICD - GATEWAY DISTRIPARKS;"
        "GURUGRAM - PATLI ICD - ADANI LOGISTICS LIMITED;KATHUWAS - CONCOR ICD - NEEMRANA;LONI, GHAZIABAD;"
        "MODINAGAR ICD - DP WORLD MULTIMODAL LOGISTICS PVT;PALI ICD - DP WORLD;"
        "PANIPAT - JHATTIPUR ICD - DP WORLD;REWARI ICD - PRISTINE LOGISTICS;SONIPAT",
    ),
    "ahmedabad icd": (
        "INAMD;INJKA;INSAA;INVDL;INVRM",
        "AHMEDABAD - CONCOR ICD   - KHODIYAR;AHMEDABAD - SACHANA;AHMEDABAD - SANAND ICD - HASTI PETRO;"
        "AHMEDABAD - VIRAMGAM ICD - GATEWAY DISTRIPARKS LTD;AHMEDABAD - VIROCHANNAGAR ICD - ADANI LOGISTICS LI",
    ),
    "ludhiana cluster": (
        "IN1LU;IN2LU;INCWL;INLUH;INSWA",
        "LUDHIANA - CHAWAPAIL;LUDHIANA - CONCOR ICD;LUDHIANA - KILA RAIPUR ICD - ADANI LOGISTICS LIMIT;"
        "LUDHIANA - KILARAIPUR DEHLON HTPL;LUDHIANA - SAHNEWAL",
    ),
    "indore cluster": ("INMNP;INTIH", "INDORE - TIHI ICD - CONCOR;MANDIDEEP - CONCOR ICD"),
    "faridabad cluster": (
        "INFBD;INPLW;INPYL",
        "FARIDABAD - ASSOCIATE CONTAINER TERMINAL PVT LTD;FARIDABAD - PIYALA ICD - GATEWAY DISTRIPARKS LTD;PALWAL",
    ),
    "moradabad cluster": (
        "INKSR;INMOR;INPNT",
        "KASHIPUR ICD - GATEWAY DISTRIPARKS LTD;MORADABAD - CONCOR ICD;PANTNAGAR",
    ),
}

# Codes confirmed (both reference weeks) to always be blanket-included, no
# origin scoping, for G0001/G0002. G0003 (RAD) uses this same blanket list
# with no scoped codes ever added (see module docstring).
BLANKET_CHARGES = ["OBS", "EFS", "PSS"]
# Codes that CAN carry per-origin POR scoping (derived dynamically from
# "(incl X)" annotations - see _parse_origin_paren), in the fixed priority
# order ground truth's own charge_seq uses (confirmed week 1: all THL rows
# before ISL before DOC, not grouped by origin).
SCOPABLE_CHARGE_ORDER = ["THL", "ISL", "DOC"]

_PAREN_RE = re.compile(r"^(.*?)\s*\(([^)]*)\)\s*$")
_TERM_OVERRIDE_RE = re.compile(r"^(\w+)\s*/\s*CY$", re.IGNORECASE)
_INCL_RE = re.compile(r"^incl\.?\s*(.+)$", re.IGNORECASE)
# "01 Sept 2026 to 14 Sept 2026" - this lane's own validity text repeats
# the year after BOTH the start and end month (unlike AUEC/AUWC's "d Mon to
# d Mon YYYY" single-year form), so the first year is an optional,
# non-captured group.
_VALIDITY_RE = re.compile(r"(\d{1,2})\s+(\w+)\s+(?:\d{4}\s+)?to\s+(\d{1,2})\s+(\w+)\s+(\d{4})", re.IGNORECASE)
_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1


def _parse_validity(ws: Worksheet, row: int, col: int) -> tuple[date | None, date | None]:
    """Returns (None, None) only when the validity text doesn't match the
    expected shape at all (a genuinely different/absent cell). Once matched,
    an unrecognized month name or out-of-range day raises rather than
    silently swallowing the error - those would otherwise produce a fully
    empty CMDT NOTE/SUR sheet with no signal anything was wrong (see
    _build_cmdt_notes' own validity_start/validity_end is None guard)."""
    text = str(ws.cell(row=row, column=col).value or "")
    m = _VALIDITY_RE.search(text)
    if not m:
        return None, None
    d1, mon1, d2, mon2, year = m.groups()
    y = int(year)
    return date(y, _month_number(mon1), int(d1)), date(y, _month_number(mon2), int(d2))


@dataclass
class ParsedOrigin:
    base_text: str
    term_override: str | None
    is_freight_collect: bool
    incl_codes: list[str]


def _parse_origin_paren(origin_text: str) -> ParsedOrigin:
    """Splits a trailing parenthetical suffix off the origin's base name and
    classifies its meaning - see module docstring for the 4 known kinds
    ("Door/CY" term override, "Freight Collect only", "incl X/Y", "via X" -
    the last dropped with no effect). An unrecognized parenthetical (none
    occur in either reference week) is treated the same as "via" - dropped,
    base name kept as-is."""
    m = _PAREN_RE.match(origin_text.strip())
    if not m:
        return ParsedOrigin(origin_text.strip(), None, False, [])
    base, paren = m.group(1).strip(), m.group(2).strip()

    term_match = _TERM_OVERRIDE_RE.match(paren)
    if term_match:
        return ParsedOrigin(base, term_match.group(1), False, [])

    if "freight collect" in paren.lower():
        return ParsedOrigin(base, None, True, [])

    incl_match = _INCL_RE.match(paren)
    if incl_match:
        codes = [c.strip().upper() for c in re.split(r"[/,]", incl_match.group(1)) if c.strip()]
        return ParsedOrigin(base, None, False, codes)

    return ParsedOrigin(base, None, False, [])


def _merge_incl(target: dict[str, list[str]], source: dict[str, list[str]]) -> None:
    """Unions `source`'s codes into `target` in place, keyed by origin code -
    NOT a first-wins `setdefault`, since the same origin's "(incl ...)"
    scoping is expected to be merged (not silently overridden) if it's ever
    seen on more than one raw sheet or column-group for the same origin."""
    for code, codes in source.items():
        existing = target.setdefault(code, [])
        for c in codes:
            if c not in existing:
                existing.append(c)


def _resolve_origin(base_text: str) -> tuple[str, str] | None:
    key = base_text.strip().lower()
    if key in ICD_CLUSTERS:
        return ICD_CLUSTERS[key]
    if key in ORIGIN_OVERRIDES:
        return ORIGIN_OVERRIDES[key]
    return None


class UnrecognizedOriginError(ValueError):
    pass


@dataclass(frozen=True)
class ColumnGroup:
    """One rate-column-group on a raw sheet: which columns feed which
    output rate slot, and which commodity group it belongs to.

    dest_col: for the MAIN sheet, the column whose row-4 cell carries this
    group's own destination code(s) (see module docstring) - carried on
    the same object as the rate columns so the two can never drift out of
    index-alignment the way two separate parallel lists could. None for
    SINGLE_DEST_COLUMN_GROUPS, which don't have a per-group destination."""

    commodity: str  # "MAIN" | "RF" | "RAD"
    prefix: str
    cgo_type: str
    col_20: int
    col_40: int | None  # None when this group's 2nd column feeds 40HC instead (RF/RAD)
    col_40hc: int
    dest_col: int | None = None


# MAIN sheet: 5 column-groups, each with its own destination read off row 4
# (see module docstring). Order here matches the raw sheet's own column
# order left-to-right.
MAIN_COLUMN_GROUPS = [
    ColumnGroup("MAIN", "D", "DR", 5, 6, 7, dest_col=5),
    ColumnGroup("RF", "R", "RF", 8, None, 9, dest_col=8),
    ColumnGroup("RF", "R", "RF", 10, None, 11, dest_col=10),
    ColumnGroup("RAD", "R", "DR", 12, None, 13, dest_col=12),
    ColumnGroup("RAD", "R", "DR", 14, None, 15, dest_col=14),
]

# SYD/FRE sheets: single destination for the whole sheet, 3 column-groups.
SINGLE_DEST_COLUMN_GROUPS = [
    ColumnGroup("MAIN", "D", "DR", 5, 6, 7),
    ColumnGroup("RF", "R", "RF", 8, None, 9),
    ColumnGroup("RAD", "R", "DR", 10, None, 11),
]

# Both Yangon and Thilawa always carry byte-identical rates and the same
# "Freight Collect only" annotation - see module docstring for why this is
# a narrow, hardcoded pair rather than a general "merge on identical
# rates" rule.
YANGON_CODE, THILAWA_CODE = "MMRGN", "MMTLA"

# SYD/FRE's own POD label is always exactly "Sydney"/"Fremantle" - each
# sheet has exactly one fixed, known destination by construction, so this
# is checked before falling back to fuzzy matching (which both currently
# resolve correctly, but a future ambiguous fuzzy score would otherwise
# silently drop the entire sheet - see module docstring's origin-resolution
# rationale for the same "hardcode a short, fixed, verified list" approach).
DEST_OVERRIDES: dict[str, tuple[str, str]] = {
    "sydney": ("AUSYD", "SYDNEY, NSW"),
    "fremantle": ("AUFRE", "FREMANTLE, WA"),
}


@dataclass
class OriginValueRow:
    origin_code: str
    origin_desc: str
    origin_term: str
    route_note: str | None
    values: dict[str, float]  # "20" | "40" | "40hc" -> rate


def _values_match(a: dict[str, float], b: dict[str, float]) -> bool:
    """Rounded comparison (2dp, these are USD rates) rather than exact float
    equality - a formula-derived rate can differ from a literal by a tiny
    floating-point epsilon while being economically identical."""
    if a.keys() != b.keys():
        return False
    return all(round(a[k], 2) == round(b[k], 2) for k in a)


def _merge_yangon_thilawa(rows: list[OriginValueRow]) -> list[OriginValueRow]:
    """See module docstring - merges Yangon+Thilawa into one row when both
    are present with matching rates AND term; otherwise leaves rows as-is
    (a term mismatch means the two origins aren't actually interchangeable
    this week, so the merge shouldn't silently pick one and drop the
    other's term)."""
    yangon = next((r for r in rows if r.origin_code == YANGON_CODE), None)
    thilawa = next((r for r in rows if r.origin_code == THILAWA_CODE), None)
    if (
        yangon is None
        or thilawa is None
        or not _values_match(yangon.values, thilawa.values)
        or yangon.origin_term != thilawa.origin_term
    ):
        return rows
    merged = OriginValueRow(
        origin_code=";".join(sorted([YANGON_CODE, THILAWA_CODE])),
        origin_desc=";".join(sorted([yangon.origin_desc, thilawa.origin_desc])),
        origin_term=yangon.origin_term,
        route_note=yangon.route_note or thilawa.route_note,
        values=yangon.values,
    )
    return [merged, *(r for r in rows if r.origin_code not in (YANGON_CODE, THILAWA_CODE))]


@dataclass
class SheetRawData:
    dest_text: str | None = None
    rows: list[tuple[ParsedOrigin, dict[int, float]]] = field(default_factory=list)


@dataclass
class AUBPRawData:
    validity_start: date | None
    validity_end: date | None
    # (ColumnGroup, raw destination-codes text from row 4) pairs, in
    # MAIN_COLUMN_GROUPS order.
    main_group_dests: list[tuple[ColumnGroup, str]] = field(default_factory=list)
    main_rows: list[tuple[ParsedOrigin, dict[int, float]]] = field(default_factory=list)
    syd: SheetRawData | None = None
    fre: SheetRawData | None = None


def _read_flat_sheet(ws: Worksheet, min_row: int, max_row: int, rate_cols: list[int]) -> list[tuple[ParsedOrigin, dict[int, float]]]:
    rows: list[tuple[ParsedOrigin, dict[int, float]]] = []
    for row_idx in range(min_row, max_row + 1):
        origin_cell = ws.cell(row=row_idx, column=ORIGIN_COL)
        if origin_cell.value in (None, ""):
            continue
        if is_excluded(origin_cell):
            continue
        values: dict[int, float] = {}
        for col in rate_cols:
            v = ws.cell(row=row_idx, column=col).value
            if isinstance(v, (int, float)):
                values[col] = v
        if not values:
            continue
        rows.append((_parse_origin_paren(str(origin_cell.value).strip()), values))
    return rows


def _build_cmdt_notes(
    validity_start: date | None,
    validity_end: date | None,
    charges: list[tuple[str, str | None]],
    rfa_effective: date | None,
    rfa_expiry: date | None,
) -> list[CmdtNoteRow]:
    """charges is an ordered list of (code, por) pairs - por is None for a
    blanket (unscoped) code. Delegates to the shared
    common/cmdt_notes.py::build_cmdt_notes builder (extended with a
    scope_values/scope_field parameter for exactly this need) instead of
    forking its own copy of the boilerplate, as auec.py/auwc.py each still
    do for their own `pol`-scoped equivalent."""
    if not charges:
        return []
    included_codes = [code for code, _ in charges]
    scope_values = [por for _, por in charges]
    return build_cmdt_notes(
        validity_start, validity_end, included_codes,
        sequential_charge_seq=True,
        rfa_effective=rfa_effective, rfa_expiry=rfa_expiry,
        scope_values=scope_values, scope_field="por",
    )


def _build_scoped_charge_list(
    incl_by_origin: dict[str, list[str]], excluded_codes: frozenset[str]
) -> list[tuple[str, str | None]]:
    """Builds the ordered (code, por) list for G0001/G0002's shared CMDT
    NOTE: blanket OBS/EFS/PSS first, then each scopable code (THL, ISL,
    DOC, in that fixed priority order) once per origin that carries it, in
    first-seen (raw sheet row) order - see module docstring."""
    charges: list[tuple[str, str | None]] = [(c, None) for c in BLANKET_CHARGES if c not in excluded_codes]
    for code in SCOPABLE_CHARGE_ORDER:
        if code in excluded_codes:
            continue
        for origin_code, codes in incl_by_origin.items():
            if code in codes:
                charges.append((code, origin_code))
    return charges


class AUBPParser(BaseMRGParser):
    lane_id: ClassVar[str] = "AUBP"

    def __init__(self, location_resolver: LocationResolver | None = None, location_store: LocationBankStore | None = None):
        self.location_resolver = location_resolver or LocationResolver()
        self.location_store = location_store or LocationBankStore()

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        names = set(wb.sheetnames)
        score = 0.0
        for expected in (SHEET_MAIN, SHEET_SYD, SHEET_FRE):
            if expected in names:
                score += 1 / 3
        if score == 0.0:
            return 0.0
        ws = wb[SHEET_MAIN] if SHEET_MAIN in names else wb[wb.sheetnames[0]]
        header_tokens = {str(ws.cell(row=3, column=c).value or "").strip() for c in range(4, 16)}
        if any("RAD" in t for t in header_tokens):
            score = min(score + 0.2, 1.0)
        return score

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        validity_start, validity_end = None, None
        main_group_dests: list[tuple[ColumnGroup, str]] = []
        main_rows: list[tuple[ParsedOrigin, dict[int, float]]] = []
        syd: SheetRawData | None = None
        fre: SheetRawData | None = None

        if SHEET_MAIN in wb.sheetnames:
            ws = wb[SHEET_MAIN]
            validity_start, validity_end = _parse_validity(ws, 2, 2)
            main_rows = _read_flat_sheet(ws, 6, 84, list(range(5, 16)))
            for group in MAIN_COLUMN_GROUPS:
                dest_cell = ws.cell(row=ROW4_DEST_ROW, column=group.dest_col)
                if is_excluded(dest_cell):
                    continue
                dest_text = str(dest_cell.value or "").strip()
                if dest_text:
                    main_group_dests.append((group, dest_text))

        for sheet_name, holder in ((SHEET_SYD, "syd"), (SHEET_FRE, "fre")):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if validity_start is None:
                validity_start, validity_end = _parse_validity(ws, 2, 2)
            pod_cell = ws.cell(row=5, column=POD_COL)
            dest_text = "" if is_excluded(pod_cell) else str(pod_cell.value or "").strip()
            rows = _read_flat_sheet(ws, 5, 83, list(range(5, 12)))
            data = SheetRawData(dest_text=dest_text, rows=rows)
            if holder == "syd":
                syd = data
            else:
                fre = data

        return RawExtraction(
            tables={
                "aubp": AUBPRawData(
                    validity_start=validity_start, validity_end=validity_end,
                    main_group_dests=main_group_dests, main_rows=main_rows, syd=syd, fre=fre,
                )
            }
        )

    def _dest_from_codes(self, codes_text: str) -> tuple[str, str] | None:
        """Resolves whichever of the "/"-joined codes the Location Bank
        actually has, rather than dropping the WHOLE destination group the
        moment a single code fails to resolve - a group's other, still-valid
        destinations shouldn't lose their entire Dry/Reefer/RAD table just
        because one code (e.g. a renamed/removed AUADL) went missing."""
        codes = sorted({c.strip() for c in codes_text.split("/") if c.strip()})
        resolved: list[tuple[str, str]] = []
        for c in codes:
            rec = self.location_store.get_by_code(c)
            if rec is not None:
                resolved.append((c, rec.primary_name))
        if not resolved:
            return None
        codes_out = [c for c, _ in resolved]
        names_out = sorted(name for _, name in resolved)
        return ";".join(codes_out), ";".join(names_out)

    def _resolve_single_dest(self, text: str) -> tuple[str, str] | None:
        override = DEST_OVERRIDES.get(text.strip().lower())
        if override is not None:
            return override
        m = self.location_resolver.match_token(text)
        if m is None or m.needs_review:
            return None
        return m.code, m.primary_name

    def _build_group_rows(
        self,
        raw_rows: list[tuple[ParsedOrigin, dict[int, float]]],
        group: ColumnGroup,
        dest: tuple[str, str],
    ) -> tuple[list[RatesRow], dict[str, list[str]]]:
        """Returns (rows, incl_codes_by_origin_code) for this column-group -
        the second element feeds the CMDT NOTE origin-scoping derivation
        (only meaningful for the MAIN/Dry commodity, see caller)."""
        dest_code, dest_desc = dest
        candidates: list[OriginValueRow] = []
        incl_by_origin: dict[str, list[str]] = {}

        for parsed, raw_values in raw_rows:
            values: dict[str, float] = {}
            if group.col_20 in raw_values:
                values["20"] = raw_values[group.col_20]
            if group.col_40 is not None and group.col_40 in raw_values:
                values["40"] = raw_values[group.col_40]
            if group.col_40hc in raw_values:
                values["40hc"] = raw_values[group.col_40hc]
            if not values:
                continue

            resolved = _resolve_origin(parsed.base_text)
            if resolved is None:
                # _read_flat_sheet only ever emits a row once it has at
                # least one real numeric rate somewhere, and `values` above
                # confirms THIS column-group has real data for it too - so
                # this is never a legitimate "N/A-rated origin" case (those
                # never reach here), always a genuine gap in ORIGIN_OVERRIDES/
                # ICD_CLUSTERS (a renamed or new port) that would otherwise
                # silently drop real rate data - see module docstring.
                raise UnrecognizedOriginError(
                    f"AUBP: origin {parsed.base_text!r} has real rate data but no entry in "
                    "ORIGIN_OVERRIDES/ICD_CLUSTERS"
                )
            origin_code, origin_desc = resolved

            if parsed.incl_codes and group.commodity == "MAIN":
                incl_by_origin.setdefault(origin_code, parsed.incl_codes)

            route_note = "(FRT collect only)" if parsed.is_freight_collect else None
            candidates.append(
                OriginValueRow(
                    origin_code=origin_code, origin_desc=origin_desc,
                    origin_term=parsed.term_override or "CY",
                    route_note=route_note, values=values,
                )
            )

        candidates = _merge_yangon_thilawa(candidates)

        rows = [
            RatesRow(
                commodity_group_code="", commodity_group_description="",
                origin_code=c.origin_code, origin_description=c.origin_desc, origin_term=c.origin_term,
                destination_code=dest_code, destination_description=dest_desc, destination_term="CY",
                prefix=group.prefix, cgo_type=group.cgo_type,
                cur_20="USD" if "20" in c.values else None, rate_20=c.values.get("20"),
                cur_40="USD" if "40" in c.values else None, rate_40=c.values.get("40"),
                cur_40hc="USD" if "40hc" in c.values else None, rate_40hc=c.values.get("40hc"),
                route_note=c.route_note,
            )
            for c in candidates
        ]
        return rows, incl_by_origin

    def _build_main_sheet_rows(
        self, data: AUBPRawData
    ) -> tuple[list[RatesRow], list[RatesRow], list[RatesRow], dict[str, list[str]]]:
        main_rows: list[RatesRow] = []
        rf_rows: list[RatesRow] = []
        rad_rows: list[RatesRow] = []
        incl_by_origin: dict[str, list[str]] = {}

        for group, dest_codes_text in data.main_group_dests:
            dest = self._dest_from_codes(dest_codes_text)
            if dest is None:
                continue
            rows, incl = self._build_group_rows(data.main_rows, group, dest)
            if group.commodity == "MAIN":
                main_rows.extend(rows)
            elif group.commodity == "RF":
                rf_rows.extend(rows)
            else:
                rad_rows.extend(rows)
            _merge_incl(incl_by_origin, incl)
        return main_rows, rf_rows, rad_rows, incl_by_origin

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: AUBPRawData = raw.tables["aubp"]

        main_description = resolve_commodity_description(DEFAULT_MAIN_DESCRIPTION, config)
        main_code = resolve_commodity_code(DEFAULT_MAIN_DESCRIPTION, DEFAULT_MAIN_CODE, config)
        main_cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_MAIN_DESCRIPTION)

        rf_description = resolve_commodity_description(DEFAULT_RF_DESCRIPTION, config)
        rf_code = resolve_commodity_code(DEFAULT_RF_DESCRIPTION, DEFAULT_RF_CODE, config)
        rf_cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_RF_DESCRIPTION)

        rad_description = resolve_commodity_description(DEFAULT_RAD_DESCRIPTION, config)
        rad_code = resolve_commodity_code(DEFAULT_RAD_DESCRIPTION, DEFAULT_RAD_CODE, config)
        rad_cmdt_seq = config.commodity_sequence_overrides.get(DEFAULT_RAD_DESCRIPTION)

        main_rows, rf_rows, rad_rows, incl_by_origin = self._build_main_sheet_rows(data)

        for sheet_data in (data.syd, data.fre):
            if sheet_data is None or not sheet_data.dest_text:
                continue
            dest = self._resolve_single_dest(sheet_data.dest_text)
            if dest is None:
                continue
            for group in SINGLE_DEST_COLUMN_GROUPS:
                rows, incl = self._build_group_rows(sheet_data.rows, group, dest)
                if group.commodity == "MAIN":
                    main_rows.extend(rows)
                elif group.commodity == "RF":
                    rf_rows.extend(rows)
                else:
                    rad_rows.extend(rows)
                _merge_incl(incl_by_origin, incl)

        dg_rows: list[RatesRow] = []
        if main_rows and not config.skip_dg_generation.get(DEFAULT_MAIN_DESCRIPTION, False):
            dg_rows = [row.model_copy(update={"cgo_type": "DG"}) for row in main_rows]

        main_rows = group_by_destination(main_rows)
        dg_rows = group_by_destination(dg_rows)
        rf_rows = group_by_destination(rf_rows)
        rad_rows = group_by_destination(rad_rows)

        def stamp(rows: list[RatesRow], code: str, description: str, cmdt_seq: int | None) -> None:
            for row in rows:
                row.commodity_group_code = code
                row.commodity_group_description = description
                row.cmdt_seq = cmdt_seq

        stamp(main_rows, main_code, main_description, main_cmdt_seq)
        stamp(dg_rows, main_code, main_description, main_cmdt_seq)
        stamp(rf_rows, rf_code, rf_description, rf_cmdt_seq)
        stamp(rad_rows, rad_code, rad_description, rad_cmdt_seq)

        # Route Seq.: one continuous counter per commodity group, spanning
        # all raw sheets and the DR/DG duplication within G0001 - see
        # module docstring.
        main_group_rows = [*main_rows, *dg_rows]
        for i, row in enumerate(main_group_rows, start=1):
            row.route_seq = i
        for i, row in enumerate(rf_rows, start=1):
            row.route_seq = i
        for i, row in enumerate(rad_rows, start=1):
            row.route_seq = i

        rates = [*main_group_rows, *rf_rows, *rad_rows]

        excluded_codes = frozenset(config.excluded_charge_codes)
        cmdt_notes: list[CmdtNoteRow] = []
        full_charges = _build_scoped_charge_list(incl_by_origin, excluded_codes)
        for group_rows, description in ((main_group_rows, main_description), (rf_rows, rf_description)):
            if not group_rows:
                continue
            notes = _build_cmdt_notes(
                data.validity_start, data.validity_end, full_charges,
                config.rfa_effective_date, config.rfa_expiry_date,
            )
            cmdt_notes.extend(row.model_copy(update={"group_description": description}) for row in notes)
            note_text = notes[0].contents if notes else None
            for row in group_rows:
                row.commodity_note = note_text

        if rad_rows:
            blanket_only = [(c, None) for c in BLANKET_CHARGES if c not in excluded_codes]
            notes = _build_cmdt_notes(
                data.validity_start, data.validity_end, blanket_only,
                config.rfa_effective_date, config.rfa_expiry_date,
            )
            cmdt_notes.extend(row.model_copy(update={"group_description": rad_description}) for row in notes)
            note_text = notes[0].contents if notes else None
            for row in rad_rows:
                row.commodity_note = note_text

        return OpusRowSet(rates=rates, cmdt_notes=cmdt_notes)


register(
    LayoutProfile(
        lane_id=AUBPParser.lane_id,
        parser_cls=AUBPParser,
        sheet_name_patterns=[re.escape(SHEET_MAIN), re.escape(SHEET_SYD), re.escape(SHEET_FRE)],
        title_keywords=["ONE MINIMUM RATE GUIDELINE"],
        header_fingerprint=["20'RAD", "40'RAD"],
    )
)
