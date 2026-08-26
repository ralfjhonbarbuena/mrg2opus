"""CSE lane parser - the richest lane so far:

- Three D2/D4/D5 rate grids on separate raw sheets, each its own service/
  commodity group: "CSE" (main Caribbean/Central America, G0001), "CSE
  (MAOVLD)" (Manaus/Vila do Conde, Brazil, G0002), "CSE VE" (Venezuela,
  reuses G0001). Unlike SAF/EAF, origin AND destination codes are given
  directly in the raw sheet (column B / a header row) - no fuzzy matching
  needed, just Location Bank description lookups by code.
- Two single-column "R5 NOR" (reefer) sheets - NOR(PA) for the main
  service's PAMIT/PACFZ destinations, NOR (MAOVLD) for BRMAO/BRVLD - each
  producing a Prefix "R" row (rate lands in the 40HC slot, no DG dup).
- "In guage guideline": a 13-destination, 2-column (20'/40' OT/FR) grid
  that produces an identical-content Prefix "O" + Prefix "F" row pair per
  origin/destination (G0003).
- "Yangtze ARB Add-on": inland China origins connecting through a main
  port, feeding OPUS ARBS (one row per container size).
- Two origin codes ("FEBP", "WPRD") are regional group shorthand that
  expand to a member-code list - see parsers/common/group_codes.py.
- "DG surcharges" sheet, rows 21-24: the PSA transshipment-via-Singapore
  DG surcharge table, feeding OPUS SPECIAL NOTE. Amounts are read from the
  sheet; the per-class row order and the (undated-in-source) validity
  start are hardcoded, verified constants - see the PSA_* constants below.

Freetime is out of scope project-wide (see saf.py/plan).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, replace
from datetime import date
from decimal import Decimal
from typing import ClassVar

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.location_bank.store import LocationBankStore
from mrg2opus.parsers.base import BaseMRGParser, RawExtraction
from mrg2opus.parsers.common.commodity import (
    CommodityNoteSpec,
    build_notes_by_description,
    resolve_commodity_code,
    resolve_commodity_description,
)
from mrg2opus.parsers.common.container_map import ContainerMap, load_container_map
from mrg2opus.parsers.common.exclusion import is_excluded, location_is_excluded
from mrg2opus.parsers.common.group_codes import load_group_codes
from mrg2opus.parsers.common.header_grid import flatten_pod_header
from mrg2opus.parsers.common.ordering import group_by_destination
from mrg2opus.parsers.common.yangtze_arbs import YangtzeRow, build_arbs, parse_yangtze_sheet
from mrg2opus.parsers.registry import LayoutProfile, register
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema.opus_rows import ArbsRow, OpusRowSet, RatesRow, SpecialNoteRow, explode_rates_row

RAW_SHEET_MAIN = "CSE"

# --- commodity groups (verified against CSE.xlsx's own OPUS RATES sheet;
# these are just the starting defaults - the actual output code/description
# are user-overridable via MappingProfile, see parsers/common/commodity.py)
# -----------------------------------------------------------------------
#
# G0001 used to carry ONE description spanning 3 different raw sheets
# ("CSE" main + "NOR(PA)" reefer + "CSE VE"), and G0002 spanned 2 ("CSE
# (MAOVLD)" main + "NOR (MAOVLD)" reefer). Each raw sheet now defaults to
# its OWN description (its own sheet name); they only end up sharing one
# CMDT NOTE block again if the user overrides them to the exact same
# description - see parsers/common/commodity.py::build_notes_by_description(),
# used in to_opus_rows() below.
#
# Each commodity tuple is (default_description, cmdt_seq,
# default_output_code): default_description doubles as the STABLE lookup
# key for every override dict (see parsers/common/commodity.py's module
# docstring - it's the one identity always unique per group) AND as the
# data.grid_rows/nor_rows dict key (so sheets that used to share one entry
# - "CSE" and "CSE VE" both feeding COMMODITY_MAIN's rows - now parse into
# separate buckets); default_output_code is the CODE shown unless
# overridden, which stays the shared parent code by default even for a
# split-out sheet like "CSE VE".
COMMODITY_MAIN = (RAW_SHEET_MAIN, 1, "G0001")
COMMODITY_VE = ("CSE VE", 1, "G0001")
COMMODITY_MAOVLD = ("CSE (MAOVLD)", 2, "G0002")
COMMODITY_NOR_PA = ("NOR(PA)", 1, "G0001")
COMMODITY_NOR_MAOVLD = ("NOR (MAOVLD)", 2, "G0002")
COMMODITY_INGAUGE = ("IN GUAGE GUIDELINE (IG)", 3, "G0003")

# Destination whose delivery term is "Door" instead of "CY", and its
# corresponding D.Via code - verified against ground truth (PACFZ is the
# only one, matching its raw label "Colon Free Zone (Door via Manzanillo,
# PA)": Door term because delivery is at the door, D.Via=PAMIT because
# cargo transits through Manzanillo to get there).
DOOR_TERM_DESTINATIONS = frozenset({"PACFZ"})
DESTINATION_VIA_CODES = {"PACFZ": "PAMIT"}

# CSE's raw "Rate structure: incl. X/Y/Z" lines (found on the CSE and CSE
# (MAOVLD) sheets) list PSS/OBS/EFS/MBS as included - but the ground truth
# CMDT NOTE also has CSS, THL (twice), and SLF as child rows despite those
# being in the "subject to" clause, not "incl.". This contradicts the
# incl./subj-to distinction that held for SAF/EAF, and both raw sheets'
# slightly different wording produce the exact same child list anyway - so
# this is a verified, hardcoded constant for this lane rather than parsed.
CMDT_NOTE_CHARGE_CODES = ["PSS", "OBS", "EFS", "MBS", "THL", "THL", "CSS", "SLF"]

VALIDITY_ROW, VALIDITY_FROM_COL, VALIDITY_TO_COL = 4, 3, 4


@dataclass(frozen=True)
class GridSheetConfig:
    sheet_name: str
    pod_code_row: int
    container_label_row: int
    data_min_row: int
    min_col: int
    max_col: int
    commodity: tuple[str, int, str]


# max_col below is the bundled sample's verified destination-column extent;
# real filings can add destinations past it (see _resolve_grid_config,
# which widens it to the actual sheet's extent). data_min_row..data_max_row
# used to be a hardcoded pair too, but the row range is now open-ended -
# _parse_grid_sheet stops at the first blank origin-code cell, which is the
# actual boundary of the data block in every sample and real file checked
# (a blank gap always separates it from trailing freetime/footnote text).
GRID_SHEETS = [
    GridSheetConfig("CSE", 8, 9, 10, 3, 56, COMMODITY_MAIN),
    GridSheetConfig("CSE (MAOVLD)", 7, 8, 9, 3, 8, COMMODITY_MAOVLD),
    GridSheetConfig("CSE VE", 8, 9, 10, 3, 8, COMMODITY_VE),
]

RAW_SHEET_NAMES = {c.sheet_name for c in GRID_SHEETS}

# Literal anchor text preceding the CSE/CSE VE header block. Real filings
# have been seen with an extra note row inserted above it (shifting every
# row below down by 1) - searching for this marker and re-deriving the
# header/data rows from its actual position handles that without needing a
# second hardcoded row set. "CSE (MAOVLD)" has no such marker in any sample
# seen, so it's left on its verified fixed rows.
SERVICE_SCOPE_MARKER = "SERVICE SCOPE = CSE"


def _find_anchor_row(ws: Worksheet, marker: str, max_search_row: int = 20, max_search_col: int = 8) -> int | None:
    marker_upper = marker.strip().upper()
    for r in range(1, max_search_row + 1):
        for c in range(1, max_search_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().upper() == marker_upper:
                return r
    return None


def _resolve_grid_config(ws: Worksheet, cfg: GridSheetConfig) -> GridSheetConfig:
    anchor = _find_anchor_row(ws, SERVICE_SCOPE_MARKER)
    if anchor is None:
        return cfg
    shift = anchor - (cfg.pod_code_row - 2)
    if shift == 0:
        return cfg
    return replace(
        cfg,
        pod_code_row=cfg.pod_code_row + shift,
        container_label_row=cfg.container_label_row + shift,
        data_min_row=cfg.data_min_row + shift,
    )


@dataclass(frozen=True)
class NorSheetConfig:
    sheet_name: str
    dest_code_row: int
    data_min_row: int
    data_max_row: int
    dest_cols: list[int]
    commodity: tuple[str, int, str]


NOR_SHEETS = [
    NorSheetConfig("NOR(PA)", 7, 9, 25, [3, 4], COMMODITY_NOR_PA),
    NorSheetConfig("NOR (MAOVLD)", 7, 9, 25, [3, 4], COMMODITY_NOR_MAOVLD),
]

INGAUGE_SHEET = "In guage guideline"
INGAUGE_POD_CODE_ROW = 7
INGAUGE_CONTAINER_LABEL_ROW = 8
INGAUGE_DATA_MIN_ROW, INGAUGE_DATA_MAX_ROW = 9, 77
INGAUGE_MIN_COL, INGAUGE_MAX_COL = 3, 28

YANGTZE_SHEET = "Yangtze ARB Add-on"  # parsing itself lives in common/yangtze_arbs.py

# OPUS SPECIAL NOTE source: "DG surcharges" sheet rows 21-24, the PSA
# transshipment-via-Singapore DG surcharge table. Row 22's 3 non-storable
# classes each get a real rate (2 rows: 20', 40'); row 23's 4 storable
# classes plus row 24's single class 3 all get a $0/BX row. The exact
# per-class ORDER in both groups is verified against ground truth, not a
# guessable rule (2B/1D/2 for the first group; 1S/2A/3/2F/2S for the
# second, interleaving row 23 and row 24's classes) - only one example
# exists, so this is hardcoded rather than derived.
DG_SURCHARGES_SHEET = "DG surcharges"
PSA_NON_STORABLE_ROW = 22
PSA_STORABLE_ROW = 23
PSA_CLASS3_ROW = 24
PSA_NON_STORABLE_CLASS_ORDER = ["2B", "1D", "2"]
PSA_STORABLE_CLASS_ORDER = ["1S", "2A", "3", "2F", "2S"]
PSA_TS_PORT = "SGSIN"  # "for any shpt tranship via Singapore"
# CORRECTION (2026-08-26): originally hardcoded as a fixed "standing
# policy" validity window (start May 22 2026, contents text "March 22,
# 2026 until August 31, 2026") verified only against one old bundled
# sample. A real reference file (reference/2_OPUS/1_CSE FAK.../..., a
# different filing week) shows the SPECIAL NOTE's validity - both start
# and end, and the Contents text's own "Valid from X until Y" line -
# simply mirrors the main filing's own validity window (data.validity_
# start/data.validity_end), not a fixed policy date at all. The old
# sample's specific week's dates were mistaken for a standing constant.
PSA_CONTENTS_TEMPLATE = (
    "Valid from {start} until {end}.\n\n"
    "For dangerous cargo under PSA Group: 1S, 2S, 2A, 2F, 3, the PSA DG SURCHARGE "
    "(APPLICABLE FOR CARGOES VIA SIN)(PSA) is fixed at USD 0.00 per Container.\n\n"
    "For dangerous cargo, under PSA Group: 1D, 2 and 2B the PSA DG SURCHARGE "
    "(APPLICABLE FOR CARGOES VIA SIN)(PSA) is fixed at USD 538.00 per 20 Foot Equivalent Unit.\n"
    "For dangerous cargo, under PSA Group: 1D, 2 and 2B the PSA DG SURCHARGE "
    "(APPLICABLE FOR CARGOES VIA SIN)(PSA) is fixed at USD 753.00 per 40 Foot Equivalent Unit."
)


def _psa_contents(validity_start: date, validity_end: date) -> str:
    def _fmt(d: date) -> str:
        return f"{d.strftime('%B')} {d.day}, {d.year}"

    return PSA_CONTENTS_TEMPLATE.format(start=_fmt(validity_start), end=_fmt(validity_end))


@dataclass
class GridRow:
    origin_code_raw: str
    dest_code: str
    sizes: dict[str, float]


@dataclass
class NorRow:
    origin_code_raw: str
    dest_code: str
    value: float


@dataclass
class InGaugeRow:
    origin_code_raw: str
    dest_code: str
    rate_20: float | None
    rate_40: float | None


@dataclass
class CSERawData:
    validity_start: date | None
    validity_end: date | None
    grid_rows: dict[tuple[str, int, str], list[GridRow]]  # keyed by commodity tuple
    nor_rows: dict[tuple[str, int, str], list[NorRow]]
    ingauge_rows: list[InGaugeRow]
    yangtze_rows: list[YangtzeRow]
    yangtze_eff_date: date | None
    yangtze_exp_date: date | None
    psa_rate_20: Decimal | None
    psa_rate_40: Decimal | None


class CSEParser(BaseMRGParser):
    lane_id: ClassVar[str] = "CSE"

    def __init__(self, container_map: ContainerMap | None = None, location_store: LocationBankStore | None = None):
        self.container_map = container_map or load_container_map("cse")
        self.location_store = location_store or LocationBankStore()
        self.group_codes = load_group_codes("cse")

    @classmethod
    def detect(cls, wb: Workbook) -> float:
        if RAW_SHEET_MAIN not in wb.sheetnames:
            return 0.0
        score = 0.5
        ws = wb[RAW_SHEET_MAIN]
        title = str(ws.cell(row=1, column=1).value or "")
        if "CARIBBEAN" in title.upper() or "CSE" in title.upper():
            score += 0.2
        if str(ws.cell(row=6, column=3).value or "").strip().upper() == "SERVICE SCOPE = CSE":
            score += 0.3
        return min(score, 1.0)

    def _resolve_codes(self, raw_code: str) -> list[str]:
        raw_code = raw_code.strip()
        if raw_code in self.group_codes:
            return self.group_codes[raw_code]
        return [raw_code]

    def _lookup_description(self, code: str) -> str | None:
        rec = self.location_store.get_by_code(code)
        return rec.primary_name if rec else None

    def parse_raw(self, wb: Workbook) -> RawExtraction:
        validity_start, validity_end = _parse_validity(wb[RAW_SHEET_MAIN])

        grid_rows: dict[tuple[str, int, str], list[GridRow]] = {}
        for cfg in GRID_SHEETS:
            if cfg.sheet_name not in wb.sheetnames:
                continue
            ws = wb[cfg.sheet_name]
            cfg = _resolve_grid_config(ws, cfg)
            rows = self._parse_grid_sheet(ws, cfg)
            grid_rows.setdefault(cfg.commodity, []).extend(rows)

        nor_rows: dict[tuple[str, int, str], list[NorRow]] = {}
        for ncfg in NOR_SHEETS:
            if ncfg.sheet_name not in wb.sheetnames:
                continue
            ws = wb[ncfg.sheet_name]
            rows = self._parse_nor_sheet(ws, ncfg)
            nor_rows.setdefault(ncfg.commodity, []).extend(rows)

        ingauge_rows: list[InGaugeRow] = []
        if INGAUGE_SHEET in wb.sheetnames:
            ingauge_rows = self._parse_ingauge_sheet(wb[INGAUGE_SHEET])

        yangtze_rows: list[YangtzeRow] = []
        yangtze_eff, yangtze_exp = None, None
        if YANGTZE_SHEET in wb.sheetnames:
            yangtze_rows, yangtze_eff, yangtze_exp = parse_yangtze_sheet(
                wb[YANGTZE_SHEET], self._lookup_description
            )

        psa_rate_20, psa_rate_40 = None, None
        if DG_SURCHARGES_SHEET in wb.sheetnames:
            psa_rate_20, psa_rate_40 = self._parse_psa_rates(wb[DG_SURCHARGES_SHEET])

        return RawExtraction(
            tables={
                "cse": CSERawData(
                    validity_start=validity_start,
                    validity_end=validity_end,
                    grid_rows=grid_rows,
                    nor_rows=nor_rows,
                    ingauge_rows=ingauge_rows,
                    yangtze_rows=yangtze_rows,
                    yangtze_eff_date=yangtze_eff,
                    yangtze_exp_date=yangtze_exp,
                    psa_rate_20=psa_rate_20,
                    psa_rate_40=psa_rate_40,
                )
            }
        )

    def _parse_psa_rates(self, ws: Worksheet) -> tuple[Decimal | None, Decimal | None]:
        rate_20 = _parse_money(ws.cell(row=PSA_NON_STORABLE_ROW, column=3).value)
        rate_40 = _parse_money(ws.cell(row=PSA_NON_STORABLE_ROW, column=4).value)
        return rate_20, rate_40

    def _parse_grid_sheet(self, ws: Worksheet, cfg: GridSheetConfig) -> list[GridRow]:
        # max_col is widened to the sheet's real extent rather than trusting
        # the sample-derived minimum - flatten_pod_header already skips any
        # column with no container label, so this only ever picks up real
        # destination columns a wider real file added past the sample.
        max_col = max(cfg.max_col, ws.max_column)
        header_cols = flatten_pod_header(ws, cfg.pod_code_row, cfg.container_label_row, cfg.min_col, max_col)
        out: list[GridRow] = []
        row_idx = cfg.data_min_row
        while True:
            origin_cell = ws.cell(row=row_idx, column=2)
            if origin_cell.value in (None, ""):
                break
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                row_idx += 1
                continue
            origin_raw = str(origin_cell.value).strip()
            by_dest: dict[str, dict[str, float]] = {}
            for hc in header_cols:
                suffix = self.container_map.suffix_for(hc.container_label)
                if suffix is None:
                    continue
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                by_dest.setdefault(hc.pod_label.strip(), {})[suffix] = cell.value
            for dest_code, sizes in by_dest.items():
                out.append(GridRow(origin_code_raw=origin_raw, dest_code=dest_code, sizes=sizes))
            row_idx += 1
        return out

    def _parse_nor_sheet(self, ws: Worksheet, cfg: NorSheetConfig) -> list[NorRow]:
        dest_codes = {
            col: str(ws.cell(row=cfg.dest_code_row, column=col).value or "").strip()
            for col in cfg.dest_cols
            if not is_excluded(ws.cell(row=cfg.dest_code_row, column=col))
        }
        out: list[NorRow] = []
        for row_idx in range(cfg.data_min_row, cfg.data_max_row + 1):
            origin_cell = ws.cell(row=row_idx, column=2)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            for col, dest_code in dest_codes.items():
                if not dest_code:
                    continue
                cell = ws.cell(row=row_idx, column=col)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                out.append(NorRow(origin_code_raw=origin_raw, dest_code=dest_code, value=cell.value))
        return out

    def _parse_ingauge_sheet(self, ws: Worksheet) -> list[InGaugeRow]:
        # 2 columns per destination (20'OT/FR, 40'OT/FR); the code sits in
        # the first of each pair (flatten_pod_header carries it forward).
        header_cols = flatten_pod_header(
            ws, INGAUGE_POD_CODE_ROW, INGAUGE_CONTAINER_LABEL_ROW, INGAUGE_MIN_COL, INGAUGE_MAX_COL
        )
        out: list[InGaugeRow] = []
        for row_idx in range(INGAUGE_DATA_MIN_ROW, INGAUGE_DATA_MAX_ROW + 1):
            origin_cell = ws.cell(row=row_idx, column=2)
            if origin_cell.value in (None, ""):
                continue
            if location_is_excluded([ws.cell(row=row_idx, column=1), origin_cell]):
                continue
            origin_raw = str(origin_cell.value).strip()
            by_dest: dict[str, dict[str, float]] = {}
            for hc in header_cols:
                cell = ws.cell(row=row_idx, column=hc.col_idx)
                if cell.value in (None, "") or not isinstance(cell.value, (int, float)):
                    continue
                if is_excluded(cell):
                    continue
                size = "20" if "20" in hc.container_label else "40"
                by_dest.setdefault(hc.pod_label.strip(), {})[size] = cell.value
            for dest_code, sizes in by_dest.items():
                out.append(
                    InGaugeRow(
                        origin_code_raw=origin_raw,
                        dest_code=dest_code,
                        rate_20=sizes.get("20"),
                        rate_40=sizes.get("40"),
                    )
                )
        return out

    def to_opus_rows(self, raw: RawExtraction, config: MappingProfile) -> OpusRowSet:
        data: CSERawData = raw.tables["cse"]

        rates: list[RatesRow] = []
        rates_port_port = []
        # Descriptions now default per-sheet (see the commodity constants'
        # comment above), so notes can't be built per-group as rows are
        # built - two groups might share a description (default, override,
        # or the user merging them back together) and need to share exactly
        # one CMDT NOTE block. Collect a spec per group that actually
        # produced rows, resolve all of them into notes ONCE at the end.
        # Unlike LAWC/LAEC, CSE's own ground truth never copies the CMDT
        # NOTE text into RATES rows' commodity_note field (verified: every
        # ground-truth row leaves it blank) - so rows are NOT stamped here.
        note_specs: list[CommodityNoteSpec] = []

        for commodity, grid_rows in data.grid_rows.items():
            default_description, cmdt_seq, default_code = commodity
            description = resolve_commodity_description(default_description, config)
            output_code = resolve_commodity_code(default_description, default_code, config)
            dr_rows, dr_pp, dg_rows, dg_pp = [], [], [], []

            for gr in grid_rows:
                origin_codes = self._resolve_codes(gr.origin_code_raw)
                origin_names = [self._lookup_description(c) for c in origin_codes]
                if any(n is None for n in origin_names):
                    continue  # unresolved code - don't fabricate a row
                dest_name = self._lookup_description(gr.dest_code)
                if dest_name is None:
                    continue

                row = RatesRow(
                    type="C",
                    cmdt_seq=cmdt_seq,
                    commodity_group_code=output_code,
                    commodity_group_description=description,
                    origin_code=";".join(sorted(set(origin_codes))),
                    origin_description=";".join(sorted(set(origin_names))),
                    origin_term="CY",
                    destination_code=gr.dest_code,
                    destination_description=dest_name,
                    destination_term="Door" if gr.dest_code in DOOR_TERM_DESTINATIONS else "CY",
                    d_via_code=DESTINATION_VIA_CODES.get(gr.dest_code),
                    prefix=self.container_map.prefix,
                    cgo_type=self.container_map.cgo_type,
                    cur_20="USD" if "20" in gr.sizes else None,
                    rate_20=_to_decimal(gr.sizes.get("20")),
                    cur_40="USD" if "40" in gr.sizes else None,
                    rate_40=_to_decimal(gr.sizes.get("40")),
                    cur_40hc="USD" if "40hc" in gr.sizes else None,
                    rate_40hc=_to_decimal(gr.sizes.get("40hc")),
                )
                dr_rows.append(row)
                dr_pp.extend(explode_rates_row(row))

                if self.container_map.cgo_type == "DR" and not config.skip_dg_generation.get(default_description, False):
                    dg_row = row.model_copy(update={"cgo_type": "DG"})
                    dg_rows.append(dg_row)
                    dg_pp.extend(explode_rates_row(dg_row))

            rates.extend(group_by_destination(dr_rows))
            rates.extend(group_by_destination(dg_rows))
            rates_port_port.extend(group_by_destination(dr_pp))
            rates_port_port.extend(group_by_destination(dg_pp))
            if dr_rows:
                note_specs.append(CommodityNoteSpec(description, data.validity_start, data.validity_end, CMDT_NOTE_CHARGE_CODES))

        for commodity, nor_rows in data.nor_rows.items():
            default_description, cmdt_seq, default_code = commodity
            description = resolve_commodity_description(default_description, config)
            output_code = resolve_commodity_code(default_description, default_code, config)
            reefer_rows, reefer_pp = [], []
            for nr in nor_rows:
                origin_codes = self._resolve_codes(nr.origin_code_raw)
                origin_names = [self._lookup_description(c) for c in origin_codes]
                if any(n is None for n in origin_names):
                    continue
                dest_name = self._lookup_description(nr.dest_code)
                if dest_name is None:
                    continue
                row = RatesRow(
                    type="C",
                    cmdt_seq=cmdt_seq,
                    commodity_group_code=output_code,
                    commodity_group_description=description,
                    origin_code=";".join(sorted(set(origin_codes))),
                    origin_description=";".join(sorted(set(origin_names))),
                    origin_term="CY",
                    destination_code=nr.dest_code,
                    destination_description=dest_name,
                    destination_term="Door" if nr.dest_code in DOOR_TERM_DESTINATIONS else "CY",
                    d_via_code=DESTINATION_VIA_CODES.get(nr.dest_code),
                    prefix="R",
                    cgo_type="DR",
                    cur_40hc="USD",
                    rate_40hc=_to_decimal(nr.value),
                )
                reefer_rows.append(row)
                reefer_pp.extend(explode_rates_row(row))
            rates.extend(group_by_destination(reefer_rows))
            rates_port_port.extend(group_by_destination(reefer_pp))
            if reefer_rows:
                note_specs.append(CommodityNoteSpec(description, data.validity_start, data.validity_end, CMDT_NOTE_CHARGE_CODES))

        ig_default_description, ig_cmdt_seq, ig_default_code = COMMODITY_INGAUGE
        ig_description = resolve_commodity_description(ig_default_description, config)
        ig_output_code = resolve_commodity_code(ig_default_description, ig_default_code, config)
        o_rows, o_pp, f_rows, f_pp = [], [], [], []
        for ir in data.ingauge_rows:
            origin_codes = self._resolve_codes(ir.origin_code_raw)
            origin_names = [self._lookup_description(c) for c in origin_codes]
            if any(n is None for n in origin_names):
                continue
            dest_name = self._lookup_description(ir.dest_code)
            if dest_name is None:
                continue
            base = RatesRow(
                type="C",
                cmdt_seq=ig_cmdt_seq,
                commodity_group_code=ig_output_code,
                commodity_group_description=ig_description,
                origin_code=";".join(sorted(set(origin_codes))),
                origin_description=";".join(sorted(set(origin_names))),
                origin_term="CY",
                destination_code=ir.dest_code,
                destination_description=dest_name,
                destination_term="Door" if ir.dest_code in DOOR_TERM_DESTINATIONS else "CY",
                d_via_code=DESTINATION_VIA_CODES.get(ir.dest_code),
                prefix="O",
                cgo_type="DR",
                cur_20="USD" if ir.rate_20 is not None else None,
                rate_20=_to_decimal(ir.rate_20),
                cur_40="USD" if ir.rate_40 is not None else None,
                rate_40=_to_decimal(ir.rate_40),
            )
            o_rows.append(base)
            o_pp.extend(explode_rates_row(base))
            f_row = base.model_copy(update={"prefix": "F"})
            f_rows.append(f_row)
            f_pp.extend(explode_rates_row(f_row))
        rates.extend(group_by_destination(o_rows))
        rates.extend(group_by_destination(f_rows))
        rates_port_port.extend(group_by_destination(o_pp))
        rates_port_port.extend(group_by_destination(f_pp))
        if o_rows:
            note_specs.append(CommodityNoteSpec(ig_description, data.validity_start, data.validity_end, CMDT_NOTE_CHARGE_CODES))

        cmdt_notes, _ = build_notes_by_description(
            note_specs,
            sequential_charge_seq=True,
            sort_text_names=False,  # verified: CSE's text preserves input order, not alphabetical
            excluded_codes=frozenset(config.excluded_charge_codes),
        )

        arbs = self._build_arbs(data)
        special_notes = self._build_special_notes(data)

        return OpusRowSet(
            rates=rates, rates_port_port=rates_port_port, cmdt_notes=cmdt_notes, arbs=arbs,
            special_notes=special_notes,
        )

    def _build_special_notes(self, data: CSERawData) -> list[SpecialNoteRow]:
        if data.psa_rate_20 is None or data.psa_rate_40 is None or data.validity_start is None or data.validity_end is None:
            return []

        rows: list[SpecialNoteRow] = []
        charge_seq = 1
        for cls in PSA_NON_STORABLE_CLASS_ORDER:
            for per, amount in (("20", data.psa_rate_20), ("40", data.psa_rate_40)):
                rows.append(
                    SpecialNoteRow(
                        header_seq=1 if charge_seq == 1 else None,
                        note_seq=3 if charge_seq == 1 else None,
                        contents=_psa_contents(data.validity_start, data.validity_end) if charge_seq == 1 else None,
                        charge_seq=charge_seq,
                        code="PSA",
                        application="Fix Amount",
                        application_effective=data.validity_start,
                        application_expires=data.validity_end,
                        cur="USD",
                        amount=amount,
                        per=per,
                        cgo_type="DG",
                        psa_grp=cls,
                        ts_port=PSA_TS_PORT,
                    )
                )
                charge_seq += 1
        for cls in PSA_STORABLE_CLASS_ORDER:
            rows.append(
                SpecialNoteRow(
                    charge_seq=charge_seq,
                    code="PSA",
                    application="Fix Amount",
                    application_effective=data.validity_start,
                    application_expires=data.validity_end,
                    cur="USD",
                    amount=Decimal(0),
                    per="BX",
                    cgo_type="DG",
                    psa_grp=cls,
                    ts_port=PSA_TS_PORT,
                )
            )
            charge_seq += 1
        return rows

    def _build_arbs(self, data: CSERawData) -> list[ArbsRow]:
        return build_arbs(data.yangtze_rows, data.yangtze_eff_date, data.yangtze_exp_date)


def _to_decimal(value: float | None) -> Decimal | None:
    return None if value is None else Decimal(str(value))


_MONEY_RE = re.compile(r"[\d.]+")


def _parse_money(value) -> Decimal | None:
    """Handles the DG surcharges sheet's inconsistent formatting: some
    amounts are literal numbers (0), others are strings like "USD 538"."""
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        m = _MONEY_RE.search(value)
        if m:
            return Decimal(m.group())
    return None


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    start = ws.cell(row=VALIDITY_ROW, column=VALIDITY_FROM_COL).value
    end = ws.cell(row=VALIDITY_ROW, column=VALIDITY_TO_COL).value
    start_date = start.date() if hasattr(start, "date") else start
    end_date = end.date() if hasattr(end, "date") else end
    return start_date, end_date


register(
    LayoutProfile(
        lane_id=CSEParser.lane_id,
        parser_cls=CSEParser,
        sheet_name_patterns=[r"^CSE$"],
        title_keywords=["CSE", "CARIBBEAN"],
        header_fingerprint=["D2", "D4", "D5"],
    )
)
