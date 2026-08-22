"""Flatten the merged-cell POD/container-type header block that every raw
MRG rate grid uses: a top row of merged destination-port (POD) labels, each
spanning N sub-columns, with a row directly below naming the container type
in each sub-column (e.g. D2/D4/D5/RD5 or 20'/40'/40HC/45').
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.excel_io.style_utils import ExclusionConfig, is_excluded

_EXCLUDED = object()  # sentinel: pod_label_for found a label cell, but it's struck/blacked


@dataclass(frozen=True)
class HeaderColumn:
    col_idx: int
    pod_label: str
    container_label: str


def flatten_pod_header(
    ws: Worksheet,
    pod_label_row: int,
    container_label_row: int,
    min_col: int,
    max_col: int,
    fallback_container_cycle: list[str] | None = None,
    exclusion_config: ExclusionConfig | None = None,
) -> list[HeaderColumn]:
    """Return one HeaderColumn per data column in [min_col, max_col].

    pod_label_row is expected to contain merged cells naming each POD, with
    the label only physically present in the top-left cell of each merged
    range (openpyxl returns None for the other cells in a merged range).
    container_label_row holds one label per physical column, no merging.

    A struck-through or blacked-out pod_label_row cell marks that whole POD
    as withdrawn/not offered (a raw-sheet convention also used for
    individual rate cells, see exclusion.py) - every column under it is
    dropped, the same as if no label had ever been found there.

    fallback_container_cycle: if the container_label_row cell is blank for
    a column (confirmed real-world case: a raw sheet with one destination
    block missing all 3 container labels, a data-entry gap, while the rate
    data itself is present and the ground truth clearly includes it), use
    this list cycling by position within min_col..max_col instead of
    skipping the column. Only pass this when the container-column pattern
    is verified fixed-width across the whole header (e.g. always
    20'/40'/HCD every 3 columns) - it's a positional guess, not a read.
    """
    merged_ranges = ws.merged_cells.ranges

    def pod_label_for(col: int) -> str | None | object:
        cell = ws.cell(row=pod_label_row, column=col)
        if cell.value not in (None, ""):
            if is_excluded(cell, exclusion_config):
                return _EXCLUDED
            return str(cell.value).strip()
        for rng in merged_ranges:
            if (
                rng.min_row <= pod_label_row <= rng.max_row
                and rng.min_col <= col <= rng.max_col
            ):
                top_left = ws.cell(row=rng.min_row, column=rng.min_col)
                if top_left.value not in (None, ""):
                    if is_excluded(top_left, exclusion_config):
                        return _EXCLUDED
                    return str(top_left.value).strip()
        return None

    columns: list[HeaderColumn] = []
    current_pod: str | None = None
    for col in range(min_col, max_col + 1):
        label = pod_label_for(col)
        if label is _EXCLUDED:
            current_pod = None
        elif label is not None:
            current_pod = label
        container = ws.cell(row=container_label_row, column=col).value
        if container in (None, ""):
            if fallback_container_cycle:
                container = fallback_container_cycle[(col - min_col) % len(fallback_container_cycle)]
            else:
                continue
        if current_pod is None:
            continue
        columns.append(HeaderColumn(col, current_pod, str(container).strip()))
    return columns
