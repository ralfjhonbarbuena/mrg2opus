"""Shared "Yangtze ARB Add-on" sheet parser -> OPUS ARBS builder.

Confirmed identical sheet structure/layout in both CSE and LAEC: inland
China origins (reached by barge/rail) connecting through a main port, with
per-container-size (D2/D4/D5) add-on rates. One ArbsRow per (origin, size)
combination. Verified against CSE's ground truth OPUS ARBS sheet (CSE has
no ground truth OPUS ARBS sheet in the samples - see cse.py/laec.py for
per-lane wiring notes); reused as-is for LAEC since the raw sheet layout
is identical and the user confirmed LAEC should get the same treatment.
"""
from __future__ import annotations

import calendar
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Callable

from openpyxl.worksheet.worksheet import Worksheet

from mrg2opus.schema.opus_rows import ArbsRow

HEADER_DATE_ROW, HEADER_DATE_COL = 1, 7
DATA_MIN_ROW = 4
SIZE_COLS = {7: "D2", 8: "D4", 9: "D5"}

_DATE_RE = re.compile(r"From\s+(\w+)\s+(\d{1,2})\s*-\s*(\w+)\s+(\d{1,2}),\s*(\d{4})", re.IGNORECASE)
_MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


@dataclass
class YangtzeRow:
    origin_code: str
    origin_description: str
    trans_mode: str
    via_code: str
    per: str
    value: Decimal


def parse_yangtze_sheet(
    ws: Worksheet, lookup_description: Callable[[str], str | None]
) -> tuple[list[YangtzeRow], date | None, date | None]:
    """lookup_description: Location Bank lookup by code - authoritative
    when available (verified against ground truth: it correctly says
    "CHENGDU, SICHUAN" where the raw sheet's own Province column mistakenly
    repeats the city name "Chengdu"); falls back to raw Origin+Province
    text for inland points that never appear as a rate origin/destination
    elsewhere and so were never mined into the bank (e.g. Anji). Apostrophes
    in raw pinyin names (e.g. "Hai'An") are dropped in the ground truth
    ("HAIAN") - a couple of inland points use a different romanization
    entirely ("Lu'An" -> "LIUAN") that isn't derivable from the raw sheet
    at all; those remain a known Location Bank gap (see README)."""
    eff_date, exp_date = _parse_validity(ws)
    out: list[YangtzeRow] = []
    for row_idx in range(DATA_MIN_ROW, ws.max_row + 1):
        origin_code = ws.cell(row=row_idx, column=4).value
        if origin_code in (None, ""):
            continue
        province = str(ws.cell(row=row_idx, column=2).value or "").strip()
        origin = str(ws.cell(row=row_idx, column=3).value or "").strip()
        via_code = str(ws.cell(row=row_idx, column=6).value or "").strip()
        trans_mode = str(ws.cell(row=row_idx, column=1).value or "").strip()
        fallback = f"{origin.upper()}, {province.upper()}".replace("'", "")
        description = lookup_description(str(origin_code).strip()) or fallback
        for col, per in SIZE_COLS.items():
            cell = ws.cell(row=row_idx, column=col)
            if not isinstance(cell.value, (int, float)):
                continue  # e.g. "Nil" / "check with SLS for details"
            out.append(
                YangtzeRow(
                    origin_code=str(origin_code).strip(),
                    origin_description=description,
                    trans_mode=trans_mode,
                    via_code=via_code,
                    per=per,
                    value=Decimal(str(cell.value)),
                )
            )
    return out, eff_date, exp_date


def build_arbs(rows: list[YangtzeRow], eff_date: date | None, exp_date: date | None) -> list[ArbsRow]:
    return [
        ArbsRow(
            point=yr.origin_code,
            description=yr.origin_description,
            trans_mode=yr.trans_mode,
            term="CY",
            over=yr.via_code,
            per=yr.per,
            cur="USD",
            proposal=yr.value,
            final=yr.value,
            eff_date=eff_date,
            exp_date=exp_date,
        )
        for yr in rows
    ]


def _parse_validity(ws: Worksheet) -> tuple[date | None, date | None]:
    text = str(ws.cell(row=HEADER_DATE_ROW, column=HEADER_DATE_COL).value or "")
    m = _DATE_RE.search(text)
    if not m:
        return None, None
    start_month, start_day, end_month, end_day, year = m.groups()
    start = _safe_date(int(year), _month_number(start_month), int(start_day))
    end = _safe_date(int(year), _month_number(end_month), int(end_day))
    return start, end


def _safe_date(year: int, month: int, day: int) -> date | None:
    """Clamps an out-of-range day to the last real day of the month - the
    raw sheet has a genuine typo ("Sept 31") the ground truth silently
    corrects to Sept 30."""
    last_day = calendar.monthrange(year, month)[1]
    try:
        return date(year, month, min(day, last_day))
    except ValueError:
        return None


def _month_number(name: str) -> int:
    return _MONTHS.index(name.strip().lower()[:3]) + 1
