"""Regression check: run the NEW production compare engine (audit/compare.py,
not tests/golden.py's own helpers) against each bundled sample's own
ground-truth RATES sheet. Every row that should exist must exist on both
sides - the same row-level guarantee test_parsers_*.py already asserts
via `assert not missing` / `assert not extra`, re-derived through the new
production code path as a consistency check on top of what those tests
already cover.

Deliberately does NOT assert zero field_mismatches - several lanes have
documented, accepted field-level gaps (see each lane's own test file
comments, e.g. test_parsers_cse.py's PAMIT rate_20 note). This test
checks row existence, not the field-level nuances the existing
test_parsers_*.py files already own and assert on directly.
"""
from __future__ import annotations

import openpyxl
import pytest

from mrg2opus.audit.compare import (
    CMDT_NOTE_IGNORE_FIELDS_BY_LANE,
    SPECIAL_NOTE_IGNORE_FIELDS_BY_LANE,
    diff_by_key,
    diff_cmdt_blocks,
    rates_row_key,
    read_cmdt_note_sheet,
    read_rates_sheet,
    read_special_note_sheet,
)
from mrg2opus.parsers.cse import CSEParser
from mrg2opus.parsers.eaf import EAFParser
from mrg2opus.parsers.laec import LAECParser
from mrg2opus.parsers.lawc import LAWCParser
from mrg2opus.parsers.saf import SAFParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols


@pytest.mark.parametrize(
    "path,parser_cls",
    [
        ("Sample MRGs with OPUS FORMATS/SAF.xlsx", SAFParser),
        ("Sample MRGs with OPUS FORMATS/EAF.xlsx", EAFParser),
        ("Sample MRGs with OPUS FORMATS/CSE.xlsx", CSEParser),
        ("Sample MRGs with OPUS FORMATS/LAEC.xlsx", LAECParser),
        ("Sample MRGs with OPUS FORMATS/LAWC.xlsx", LAWCParser),
    ],
)
def test_compare_engine_finds_no_missing_or_extra_rates_rows_against_own_ground_truth(path, parser_cls):
    wb = openpyxl.load_workbook(path, data_only=True)
    parser = parser_cls()
    row_sets = parser.run_multi(wb, MappingProfile())
    for suffix, row_set in row_sets.items():
        tag = f"-{suffix}" if suffix else ""
        generated = [r.model_dump() for r in row_set.rates]
        expected = read_rates_sheet(wb, f"OPUS RATES{tag}")

        result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS)
        assert not result.missing, f"{path} [{suffix or '(default)'}]: missing {len(result.missing)} rows"
        assert not result.extra, f"{path} [{suffix or '(default)'}]: {len(result.extra)} unexpected rows"


@pytest.mark.parametrize(
    "path,parser_cls,lane_id",
    [
        ("Sample MRGs with OPUS FORMATS/SAF.xlsx", SAFParser, "SAF"),
        ("Sample MRGs with OPUS FORMATS/EAF.xlsx", EAFParser, "EAF"),
        ("Sample MRGs with OPUS FORMATS/CSE.xlsx", CSEParser, "CSE"),
        ("Sample MRGs with OPUS FORMATS/LAEC.xlsx", LAECParser, "LAEC"),
        ("Sample MRGs with OPUS FORMATS/LAWC.xlsx", LAWCParser, "LAWC"),
    ],
)
def test_compare_engine_finds_no_missing_or_extra_cmdt_note_blocks_against_own_ground_truth(path, parser_cls, lane_id):
    wb = openpyxl.load_workbook(path, data_only=True)
    parser = parser_cls()
    row_sets = parser.run_multi(wb, MappingProfile())
    ignore = CMDT_NOTE_IGNORE_FIELDS_BY_LANE.get(lane_id, frozenset())
    for suffix, row_set in row_sets.items():
        if not row_set.cmdt_notes:
            continue
        tag = f"-{suffix}" if suffix else ""
        generated = [r.model_dump() for r in row_set.cmdt_notes]
        expected = read_cmdt_note_sheet(wb, f"OPUS CMDT NOTE{tag}")

        result = diff_cmdt_blocks(generated, expected, cols.CMDT_NOTE_ROW_FIELDS, ignore_fields=ignore)
        assert not result.missing_blocks, f"{path} [{suffix or '(default)'}]: missing blocks {result.missing_blocks}"
        assert not result.extra_blocks, f"{path} [{suffix or '(default)'}]: extra blocks {result.extra_blocks}"


def test_compare_engine_finds_no_missing_or_extra_special_note_blocks_against_own_ground_truth():
    """CSE is the only bundled sample lane that produces OPUS SPECIAL NOTE."""
    wb = openpyxl.load_workbook("Sample MRGs with OPUS FORMATS/CSE.xlsx", data_only=True)
    parser = CSEParser()
    row_set = parser.run(wb, MappingProfile())
    ignore = SPECIAL_NOTE_IGNORE_FIELDS_BY_LANE.get("CSE", frozenset())
    generated = [r.model_dump() for r in row_set.special_notes]
    expected = read_special_note_sheet(wb)

    result = diff_cmdt_blocks(generated, expected, cols.SPECIAL_NOTE_ROW_FIELDS, ignore_fields=ignore)
    assert not result.missing_blocks, f"missing blocks {result.missing_blocks}"
    assert not result.extra_blocks, f"extra blocks {result.extra_blocks}"
