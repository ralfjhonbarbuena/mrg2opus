from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.auec import AUECParser, DEFAULT_MAIN_DESCRIPTION, DEFAULT_NZJ_DESCRIPTION
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

# 2 weekly pairs, both real ground truth (tracker status "Completed").
PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "37_AUS NEA to AUEC FAK" / "ONE AU MRG 2026_0815 to 2026_0831 - ex NEA to AUEC (07 August 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "37_AUS NEA to AUEC FAK" / "ONE AU MRG 2026_0815 to 2026_0831 - ex NEA to AUEC (07 August 2026)_OPUS.xlsx",
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "38_AUS NEA to AUEC FAK" / "ONE AU MRG 2026_0901 to 2026_0914 - ex NEA to AUEC (24 August 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "38_AUS NEA to AUEC FAK" / "ONE AU MRG 2026_0901 to 2026_0914 - ex NEA to AUEC (24 August 2026)_opus.xlsx",
    ),
]

pytestmark = pytest.mark.skipif(
    any(not p.exists() for pair in PAIRS for p in pair),
    reason="reference/ ground-truth files not present in this checkout",
)

# Known, deliberate deviations from ground truth (same categories already
# documented for every other lane, see RATES_IGNORE_FIELDS_BY_LANE in
# audit/compare.py):
#   - type: forced to "C" on every row, a user-directed business rule.
#   - commodity_group_code/commodity_group_description/cmdt_seq: OPUS's
#     own global running sequence (G0005/G0003 in this ground truth) is
#     not reproducible from the raw MRG alone - user-customizable per
#     filing, same gap already documented for every other lane.
#   - route_seq: single running counter per commodity group across every
#     destination/container-type block - verified correct in shape (see
#     test_auec_route_seq_is_one_continuous_counter_per_group below) but
#     not pinned to the exact literal ground-truth numbers, same
#     "OPUS renumbers on import" placeholder status as every other lane.
#   - commodity_note: exact text verified separately via the CMDT NOTE
#     sheet comparison below.
RATES_IGNORE_FIELDS = {
    "type", "commodity_group_code", "commodity_group_description", "cmdt_seq", "commodity_note", "route_seq",
}


def _run_auec(raw_path: Path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    return AUECParser().run(wb, MappingProfile())


@pytest.mark.parametrize("raw_path,opus_path", PAIRS)
def test_auec_rates_matches_ground_truth(raw_path, opus_path):
    row_set = _run_auec(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize("raw_path,opus_path", PAIRS)
def test_auec_cmdt_note_matches_ground_truth(raw_path, opus_path):
    """Covers the lane-specific POL scoping (ISL only for Taiwan-origin
    shipments; EFS filed as two separate rows scoped to Hong Kong and
    Korea respectively) - see auec.py's INCLUDED_CHARGES module comment."""
    row_set = _run_auec(raw_path)
    generated = [r.model_dump() for r in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SUR")

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            if field_name in ("application_effective", "application_expires") and g.get("code") != "APP":
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_auec_route_seq_is_one_continuous_counter_per_group():
    """Confirmed against ground truth: Route Seq. is a single running
    counter across the WHOLE commodity group (every destination and
    container-type block), NOT resetting per destination or per
    cgo_type/prefix - restarting at 1 only for the next commodity group
    (main vs NZJ)."""
    raw_path, _ = PAIRS[0]
    row_set = _run_auec(raw_path)

    for description in (DEFAULT_MAIN_DESCRIPTION, DEFAULT_NZJ_DESCRIPTION):
        group_rows = [r for r in row_set.rates if r.commodity_group_description == description]
        assert [r.route_seq for r in group_rows] == list(range(1, len(group_rows) + 1))


def test_auec_excluded_charge_codes_drops_baf_end_to_end():
    """MappingProfile.excluded_charge_codes wired through: excluding EFS
    drops BOTH its POL-scoped rows (Hong Kong and Korea), not just one."""
    raw_path, _ = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUECParser().run(wb, MappingProfile(excluded_charge_codes=["EFS"]))

    codes = [n.code for n in row_set.cmdt_notes]
    assert "EFS" not in codes
    assert "ISL" in codes  # untouched - only EFS was excluded


def test_auec_skip_dg_generation_suppresses_dg_rows_for_main_group_only():
    """skip_dg_generation is keyed per commodity group's default
    description - suppressing the main group's DG rows must not affect
    the separate NZJ group."""
    raw_path, _ = PAIRS[0]
    default_row_set = _run_auec(raw_path)
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUECParser().run(wb, MappingProfile(skip_dg_generation={DEFAULT_MAIN_DESCRIPTION: True}))

    default_main_cgo = {r.cgo_type for r in default_row_set.rates if r.commodity_group_description == DEFAULT_MAIN_DESCRIPTION}
    assert "DG" in default_main_cgo

    main_cgo = {r.cgo_type for r in row_set.rates if r.commodity_group_description == DEFAULT_MAIN_DESCRIPTION}
    assert "DG" not in main_cgo
    assert "DR" in main_cgo

    nzj_cgo = {r.cgo_type for r in row_set.rates if r.commodity_group_description == DEFAULT_NZJ_DESCRIPTION}
    assert "DG" in nzj_cgo  # NZJ group is unaffected


# AUEC TIER 1 - same parser, no code changes needed for RATES (confirmed
# exact match, including route_seq's literal numbers). CMDT NOTE child rows
# need the RFA effective/expiry override wired through (each week uses its
# own distinct, longer-lived RFA window - 20260510/20261231 for week 39,
# 20260526/20261231 for week 40 - not derivable from the raw MRG, same as
# every other lane's RFA feature). Row COUNT still differs beyond that:
# real ground truth repeats each group's identical note block 3x
# (header_seq 841-843 for main, 844-846 for NZJ; same pattern in the other
# TIER 1 week) with no signal anywhere in the raw MRG explaining "3" - the
# raw sheets are the same 3-sheet shape as FAK's own (which produces only
# 2 blocks total), and the workbook's own "Tier 1 list" sheet is a much
# longer, unrelated customer roster, not a 3-item list. Treated as a
# non-derivable, accepted gap (same category as header_seq itself) rather
# than guessed at - verified below by CONTENT only, not row count.
TIER1_PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "39_AUS NEA to AUEC TIER 1" / "ONE AU MRG 2026_0815 to 2026_0831 - ex NEA to AUEC (07 August 2026) - Tier 1.xlsx",
        REFERENCE_DIR / "2_OPUS" / "39_AUS NEA to AUEC TIER 1" / "AUEC 15 TO 31 T1.xlsx",
        date(2026, 5, 10), date(2026, 12, 31),
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "40_AUS NEA to AUEC TIER 1" / "ONE AU MRG 2026_0901 to 2026_0914 - ex NEA to AUEC (24 August 2026) - Tier 1.xlsx",
        REFERENCE_DIR / "2_OPUS" / "40_AUS NEA to AUEC TIER 1" / "AUEC 1 TO 14.xlsx",
        date(2026, 5, 26), date(2026, 12, 31),
    ),
]


@pytest.mark.parametrize("raw_path,opus_path,rfa_eff,rfa_exp", TIER1_PAIRS)
def test_auec_tier1_rates_matches_ground_truth(raw_path, opus_path, rfa_eff, rfa_exp):
    if not raw_path.exists() or not opus_path.exists():
        pytest.skip("reference/ ground-truth files not present in this checkout")
    row_set = _run_auec(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize("raw_path,opus_path,rfa_eff,rfa_exp", TIER1_PAIRS)
def test_auec_tier1_cmdt_note_content_is_a_subset_of_ground_truth(raw_path, opus_path, rfa_eff, rfa_exp):
    """Ground truth repeats each generated block 3x under separate
    header_seq numbers (see module comment above) - so every row this
    parser generates must appear in ground truth, just not the reverse.
    charge_seq is also ignored here: TIER 1's ground truth orders each
    block's children OBS/EFS/EFS/ISL/PSS while FAK's own (confirmed
    exact-match elsewhere) orders them OBS/ISL/EFS/EFS/PSS - a genuine,
    confirmed difference between the two real filings' own conventions,
    not a parsing gap, so charge_seq shouldn't gate content correctness
    here."""
    if not raw_path.exists() or not opus_path.exists():
        pytest.skip("reference/ ground-truth files not present in this checkout")
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUECParser().run(wb, MappingProfile(rfa_effective_date=rfa_eff, rfa_expiry_date=rfa_exp))
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SRCHG")

    ignore = {"header_seq", "note_seq", "charge_seq"}

    def key(row):
        return tuple(
            _normalize(row.get(f)) for f in cols.CMDT_NOTE_ROW_FIELDS if f not in ignore
        )

    expected_keys = {key(e) for e in expected}
    missing = [g for g in generated if key(g) not in expected_keys]
    assert not missing, f"{len(missing)} generated rows have no match in ground truth, e.g. {missing[:2]}"
