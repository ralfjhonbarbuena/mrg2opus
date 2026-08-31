"""Route-note verification against reference/2_OPUS's real LAWC ground
truth (delivered 2026-08-26 - see project-opus-note-sheet-taxonomy
memory), kept separate from test_parsers_lawc.py's older bundled-sample
tests because that sample predates route notes entirely for two cases
(see that file's RATES_IGNORE_FIELDS route_note comment). One raw MRG
input file, one separately-filed real OPUS output file - matching how
these were actually delivered (unlike the bundled sample, which combines
both in one workbook).
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import diff_by_key, rates_row_key, read_rates_sheet, read_route_note_sheet
from mrg2opus.parsers.lawc import LAWCParser
from mrg2opus.presets.models import MappingProfile

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"
RAW_PATH = (
    REFERENCE_DIR / "1_MRGs" / "15_LAWC FAK"
    / "20260812_MRG guideline template China_HKG_SIN_TWN_KR (15-21 Aug) and SEA ISC (15-31 Aug)_FAK (1).xlsx"
)
OPUS_PATH = REFERENCE_DIR / "2_OPUS" / "15_LAWC FAK" / "LWE ( 20260815 - 20260821 ).xlsx"

pytestmark = pytest.mark.skipif(
    not RAW_PATH.exists() or not OPUS_PATH.exists(),
    reason="reference/ ground-truth files not present in this checkout",
)

# Reefer/NOR (prefix "R") rows are excluded from both checks below: this
# reference file's own commodity_group_code for NOR is G0004, but that
# code is a user-customizable override (confirmed - see
# project-opus-note-sheet-taxonomy memory), not a derivable default, so a
# fresh parse with no override applied can't be expected to reproduce
# whatever code this particular filing batch was actually override'd to.
# Separately, NOR's DR rows don't get a DG-duplicate at all yet (a real,
# but out-of-scope-for-this-task gap - see that same memory) - which is
# the other reason its route notes (REEFER DRY AS DANGEROUS combined with
# AX3) can't be reproduced here. HNSLO/OOG/the main SEA grid's own DG rows
# are unaffected by either gap and are exactly what this test verifies.
_NOR_PREFIX = "R"


def _run_lawc_fak():
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    return LAWCParser().run(wb, MappingProfile())


def test_lawc_route_note_text_matches_reference_rates_sheet():
    """Every non-NOR RATES row's route_note text (HNSLO's MAR/MX2, OOG's
    KCI/OH/OWOH/OW cases including the blank-for-plain-in-gauge fix, and
    the main SEA grid's own G0004 DG rows' REEFER DRY AS DANGEROUS note)
    matches the real filed RATES sheet, keyed the same way
    test_parsers_lawc.py's own RATES tests key rows. Only asserts on the
    matched intersection (not missing/extra) - this reference file has
    known, separately-tracked gaps unrelated to route notes (see the
    _NOR_PREFIX comment, and the general "LAWC Tier 1/FAK real-file
    fidelity" follow-up noted in project-opus-note-sheet-taxonomy)."""
    row_set = _run_lawc_fak()
    generated = [r.model_dump() for r in row_set.rates if r.prefix != _NOR_PREFIX]

    ref_wb = openpyxl.load_workbook(OPUS_PATH, data_only=True, read_only=True)
    expected = [r for r in read_rates_sheet(ref_wb, "RATES") if r.get("prefix") != _NOR_PREFIX]

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=["route_note"])
    assert result.matched > 0, "expected at least some matched rows to actually verify route_note against"
    assert not result.field_mismatches, (
        f"{len(result.field_mismatches)} route_note mismatches, e.g. {result.field_mismatches[:10]}"
    )


def test_lawc_route_notes_rn_sheet_content_matches_reference():
    """The derived RN sheet's note-text counts match the real RN sheet
    exactly, for every category - ignoring header_seq/route_seq/note_seq,
    which OPUS assigns itself (this file numbers its headers from 1015).

    MAR and MX2 used to be excluded here: MX2 was 0 vs 88 and MAR 190 vs
    196, blamed on an HNSLO-detection gap. It was really three truncated
    row ranges - see DRY_SECTIONS' data_max_row/max_col notes - and with
    those corrected every count agrees, so the whole sheet is asserted
    rather than a hand-picked subset."""
    row_set = _run_lawc_fak()
    generated = Counter(r.contents for r in row_set.route_notes)

    ref_wb = openpyxl.load_workbook(OPUS_PATH, data_only=True, read_only=True)
    expected = Counter(r["contents"] for r in read_route_note_sheet(ref_wb, "RN") if r["contents"])

    assert expected, "expected the reference RN sheet to have some notes to compare against"
    mismatches = {
        k: (generated.get(k, 0), expected.get(k, 0))
        for k in set(generated) | set(expected)
        if generated.get(k, 0) != expected.get(k, 0)
    }
    assert not mismatches, f"generated vs expected counts differ: {mismatches}"


def test_every_rn_row_points_at_a_rates_row_that_carries_the_note():
    """An RN row addresses its RATES row by (Header Seq, Route Seq), so
    every RN row must land on a real RATES row whose own Route Note
    (column AC) is populated - otherwise the RN sheet describes routes the
    RATES sheet says have no note (user-reported, 2026-08-31).

    Route Seq only counts 1..N within a commodity block, so this holds
    only while Header Seq names that block; grouping it by note text
    instead - as this lane used to - broke 602 of 649 links. The other
    RN-emitting lanes (TAD's three, West Asia) pass this by construction,
    having always keyed on cmdt_seq.
    """
    row_set = _run_lawc_fak()
    assert row_set.route_notes, "expected this lane to emit RN rows at all"

    rates_by_key = {(r.cmdt_seq, r.route_seq): r for r in row_set.rates}
    unmatched = [rn for rn in row_set.route_notes if (rn.header_seq, rn.route_seq) not in rates_by_key]
    assert not unmatched, f"{len(unmatched)} RN rows point at no RATES row, e.g. {unmatched[:3]}"

    blank_ac = [
        rn for rn in row_set.route_notes
        if not rates_by_key[(rn.header_seq, rn.route_seq)].route_note
    ]
    assert not blank_ac, f"{len(blank_ac)} RN rows point at a RATES row with a blank Route Note"

    # ...and the reverse: nothing on RATES is left without its RN row.
    assert len(row_set.route_notes) == sum(1 for r in row_set.rates if r.route_note)


def test_san_lorenzo_rows_are_filed_for_every_dry_sheet():
    """HNSLO (San Lorenzo, via NICIO by truck) is the LAST destination
    block on all three dry sheets, which is what made it the casualty of
    every truncated range: ISC and SEA stopped at column 68, one short of
    HNSLO's 69-71, so neither filed a single San Lorenzo rate, and all
    three stopped a row or more above their last rated origin.

    Counts are per (origin, o_via) against this file's own RATES sheet,
    so the origin ranges are pinned too - not just that HNSLO appears.
    """
    row_set = _run_lawc_fak()
    ref_wb = openpyxl.load_workbook(OPUS_PATH, data_only=True, read_only=True)
    expected_rows = [
        r for r in read_rates_sheet(ref_wb, "RATES")
        if r.get("destination_code") == "HNSLO" and r.get("cgo_type") == "DR"
    ]
    assert expected_rows, "expected the reference RATES sheet to file HNSLO at all"

    # This lane's structural codes: MAIN G0001, ISC G0003, SEA G0004. The
    # reference files MAIN/SEA/ISC as G0001/G0002/G0003 - commodity codes
    # are a user override, not a derivable default (see the _NOR_PREFIX
    # note above), so the two are mapped rather than compared.
    for ref_code, our_code in (("G0001", "G0001"), ("G0002", "G0004"), ("G0003", "G0003")):
        expected = Counter(
            (r["origin_code"], r.get("o_via_code"))
            for r in expected_rows if r.get("commodity_group_code") == ref_code
        )
        generated = Counter(
            (r.origin_code, r.o_via_code) for r in row_set.rates
            if r.destination_code == "HNSLO" and r.cgo_type == "DR" and r.commodity_group_code == our_code
        )
        assert generated == expected, (
            f"{our_code}: missing {sorted(set(expected) - set(generated))[:5]}, "
            f"extra {sorted(set(generated) - set(expected))[:5]}"
        )

    # Every one routes via Corinto on a truck, and none is a direct call.
    hnslo = [r for r in row_set.rates if r.destination_code == "HNSLO"]
    assert {r.d_via_code for r in hnslo} == {"NICIO"}
    assert {r.destination_transmode for r in hnslo} == {"Truck"}
