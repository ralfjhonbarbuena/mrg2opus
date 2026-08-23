"""Golden-file diff helpers: read the ground-truth OPUS sheets already baked
into the paired sample workbooks and compare them against a parser's output.

Sheet-reading and keyed-diff logic live in mrg2opus.audit.compare (the
production module the UI's Compare mode also calls) - this module is a
thin path-loading wrapper so every existing test call site here keeps its
current path-based signature and (matched, missing, extra, mismatches)
4-tuple return shape unchanged.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl

from mrg2opus.audit.compare import diff_by_key, rates_row_key
from mrg2opus.audit.compare import read_arbs_sheet as _read_arbs_sheet_from_wb
from mrg2opus.audit.compare import read_cmdt_note_sheet as _read_cmdt_note_sheet_from_wb
from mrg2opus.audit.compare import read_rates_sheet as _read_rates_sheet_from_wb
from mrg2opus.audit.compare import read_special_note_sheet as _read_special_note_sheet_from_wb
from mrg2opus.schema import opus_columns as cols


def read_rates_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return _read_rates_sheet_from_wb(wb, sheet_name)
    finally:
        wb.close()


def read_arbs_sheet(path: Path, sheet_name: str = cols.SHEET_NAME_ARBS) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return _read_arbs_sheet_from_wb(wb, sheet_name)
    finally:
        wb.close()


def read_cmdt_note_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return _read_cmdt_note_sheet_from_wb(wb, sheet_name)
    finally:
        wb.close()


def read_special_note_sheet(path: Path, sheet_name: str = cols.SHEET_NAME_SPECIAL_NOTE) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return _read_special_note_sheet_from_wb(wb, sheet_name)
    finally:
        wb.close()


def _normalize(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        return int(value) if value.isdigit() else value
    return value


def _normalize_cmdt_value(value: Any) -> Any:
    if hasattr(value, "date") and callable(getattr(value, "date")):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        if value.isdigit():
            return int(value)
        return value
    return value


def diff_rates(generated: list[dict[str, Any]], expected: list[dict[str, Any]], ignore_fields: set[str] = frozenset()):
    """Row-level (keyed) diff. Returns (matched, missing, extra, field_mismatches) -
    unchanged shape from before this refactor, so every existing call
    site's 4-way unpack keeps working."""
    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=ignore_fields)
    return result.matched, result.missing, result.extra, result.field_mismatches
