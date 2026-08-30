"""The wizard's view of what an export contains must equal what the
writer actually writes - they drifted apart before, hiding 3 of the 8
sheet types from every screen."""
from __future__ import annotations

import openpyxl

from mrg2opus.excel_io.writer import write_opus_workbook_multi
from mrg2opus.schema.opus_rows import (
    CmdtNoteRow, FreetimeRow, OpusRowSet, RatesRow, RouteNoteRow, VerticalRatesRow,
)
from mrg2opus.ui.sheets import OUTPUT_SHEET_FIELDS, apply_skips, output_sheets


def _rates_row():
    return RatesRow(
        commodity_group_code="G0001", commodity_group_description="FAK",
        origin_code="CNSHA", origin_description="SHANGHAI",
        destination_code="BEANR", destination_description="ANTWERP",
        prefix="D", cgo_type="DR", cur_20="USD", rate_20=100,
    )


def _full_row_set():
    return OpusRowSet(
        rates=[_rates_row()],
        cmdt_notes=[CmdtNoteRow(code="APP", charge_seq=1)],
        route_notes=[RouteNoteRow(code="APP", charge_seq=1, contents="x")],
        vertical_rates=[VerticalRatesRow(per="D2", rate=100)],
        freetime=[FreetimeRow()],
    )


class _OverridingParser:
    SHEET_NAME_OVERRIDES = {"route_notes": "ROUTE NOTE"}
    SCOPED_SHEET_NAME_OVERRIDES = {"AEW": {"cmdt_notes": "SRCHG"}}


def test_every_output_field_is_listed():
    """Regression: the wizard hardcoded 5 of these, so ROUTE NOTE,
    VERTICAL RATES and FREETIME were invisible and unskippable."""
    assert set(OUTPUT_SHEET_FIELDS) >= {"route_notes", "vertical_rates", "freetime"}
    fields = {s.field for s in output_sheets({"": _full_row_set()})}
    assert fields == {"rates", "cmdt_notes", "route_notes", "vertical_rates", "freetime"}


def test_listed_names_match_the_written_workbook(tmp_path):
    row_sets = {"": _full_row_set()}
    out = tmp_path / "o.xlsx"
    write_opus_workbook_multi(row_sets, out)
    assert [s.name for s in output_sheets(row_sets)] == openpyxl.load_workbook(out).sheetnames


def test_listed_names_honour_sheet_name_overrides(tmp_path):
    """A lane can rename its sheets per scope; the UI must show the name
    the user will actually see, not the default."""
    row_sets = {"AEW": _full_row_set()}
    cls = _OverridingParser
    out = tmp_path / "o.xlsx"
    write_opus_workbook_multi(
        row_sets, out,
        sheet_name_overrides=cls.SHEET_NAME_OVERRIDES,
        scoped_sheet_name_overrides=cls.SCOPED_SHEET_NAME_OVERRIDES,
    )
    names = [s.name for s in output_sheets(row_sets, cls)]
    assert names == openpyxl.load_workbook(out).sheetnames
    assert "SRCHG" in names and "ROUTE NOTE-AEW" in names


def test_empty_sheets_are_not_listed():
    """The writer skips empty sheets, so listing them would promise a
    sheet the workbook won't contain."""
    assert [s.field for s in output_sheets({"": OpusRowSet(rates=[_rates_row()])})] == ["rates"]


def test_skipping_a_previously_unskippable_sheet_works():
    row_sets = {"": _full_row_set()}
    kept = apply_skips(row_sets, {"VERTICAL RATES": True, "FREETIME": True})
    assert [s.name for s in output_sheets(kept)] == ["RATES", "CMDT NOTE", "RN"]


def test_skip_keys_resolve_per_scope_with_overrides():
    """Skips are keyed by the real sheet name, so a renamed sheet must be
    skippable under the name the UI showed."""
    kept = apply_skips({"AEW": _full_row_set()}, {"SRCHG": True}, _OverridingParser)
    assert "SRCHG" not in [s.name for s in output_sheets(kept, _OverridingParser)]


def test_no_skips_returns_the_same_object():
    row_sets = {"": _full_row_set()}
    assert apply_skips(row_sets, {}) is row_sets
    assert apply_skips(row_sets, {"RATES": False}) is row_sets
