from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.tad_oew_omw import TADOewOmwParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

RAW_PATH = REFERENCE_DIR / "1_MRGs" / "25_TAD FILING OEW OMW" / "AE WB Sept MRG Dated 26th Aug (OEWOMW).xlsx"
GROUND_TRUTH = {
    "OEW": REFERENCE_DIR / "2_OPUS" / "25_TAD FILING OEW OMW" / "OEW POLLY.xlsx",
    "OMW": REFERENCE_DIR / "2_OPUS" / "25_TAD FILING OEW OMW" / "OMW.xlsx",
}

pytestmark = pytest.mark.skipif(
    not RAW_PATH.exists() or any(not p.exists() for p in GROUND_TRUTH.values()),
    reason="reference/ ground-truth files not present in this checkout",
)

# The real ground truth's own cmdt_seq (16 for OEW, 18 for OMW) is a
# leftover from the TAD VBA tool's own internal running counter across
# whatever scopes the filer happened to process that session - not
# reproducible from the raw MRG alone, same category of gap as every
# other lane's cmdt_seq (see RATES_IGNORE_FIELDS_BY_LANE). Unlike every
# other lane, route_seq here IS directly comparable: TAD's real filing is
# one raw row -> one RATES row in raw order, so a fresh single-file parse
# reproduces the exact literal ground-truth numbers, not just the shape.
RATES_IGNORE_FIELDS = {"type", "cmdt_seq", "commodity_note"}

# RFA effective/expiry override confirmed against ground truth (see
# MappingProfile.rfa_effective_date/rfa_expiry_date) - every child CMDT
# NOTE row's Application Effective/Expires is the RFA window, not the
# weekly rate validity window the parent (APP) row uses.
RFA_EFFECTIVE = date(2026, 8, 19)
RFA_EXPIRY = date(2026, 12, 31)


def _read_tad_route_note_sheet(wb, sheet_name: str) -> list[dict]:
    """Every real TAD FILING 'ROUTE NOTE' ground truth checked (23/25/27 -
    AEW/AMW, OEW/OMW, WMW/WEW, all 5 files) has ONE stray value in cell A1
    only (the group's own header_seq, e.g. 16) sitting before the true
    header row, shifting every real column one to the right versus the
    RN_HEADER/RN_ROW_FIELDS schema (confirmed clean, no such shift, against
    LAWC's own 'RN' sheet). Every data row's column A is blank - this is a
    peculiarity of whatever template these specific reference files were
    filed from, not a general TAD convention our own writer should
    reproduce (RN_HEADER's un-shifted layout is the correct one to write)."""
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[1 : 1 + len(cols.RN_ROW_FIELDS)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.RN_ROW_FIELDS, values)))
    return rows


def _run_tad() -> dict:
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    profile = MappingProfile(rfa_effective_date=RFA_EFFECTIVE, rfa_expiry_date=RFA_EXPIRY)
    return TADOewOmwParser().run_multi(wb, profile)


@pytest.mark.parametrize("scope", ["OEW", "OMW"])
def test_tad_rates_matches_ground_truth(scope):
    row_set = _run_tad()[scope]
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize("scope", ["OEW", "OMW"])
def test_tad_route_seq_is_one_continuous_counter(scope):
    """Every raw row maps 1:1 to a RATES row in raw order - confirmed
    against ground truth, route_seq 1..N with no resets (single commodity
    group in both real files seen so far)."""
    row_set = _run_tad()[scope]
    assert [r.route_seq for r in row_set.rates] == list(range(1, len(row_set.rates) + 1))


@pytest.mark.parametrize("scope", ["OEW", "OMW"])
def test_tad_cmdt_note_matches_ground_truth(scope):
    row_set = _run_tad()[scope]
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "CMDT NOTE")

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


@pytest.mark.parametrize("scope", ["OEW", "OMW"])
def test_tad_route_note_matches_ground_truth(scope):
    """Covers both route-note triggers: OEW's populated T/S Port column
    (NLRTM) and OMW's 4 hardcoded special-node commodity-group-name tags
    (Alexandria x2, Haydarpasa, Marport) - see SPECIAL_NODE_ROUTE_NOTES."""
    row_set = _run_tad()[scope]
    generated = [n.model_dump() for n in row_set.route_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = _read_tad_route_note_sheet(ref_wb, "ROUTE NOTE")

    assert len(generated) == len(expected)
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in ("route_seq", "contents", "charge_seq", "code", "application_effective", "application_expires", "application"):
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_tad_caf_gets_its_own_cmdt_note_child_row():
    """Regression guard: TAD's 'Include Surcharge' column is an authoritative
    structured list, not free text - it must NOT be filtered through
    INDIVIDUAL_CHARGE_CODES (a whitelist built for other lanes' ambiguous
    'Includes X/Y/Z' text parsing, which doesn't include CAF even though
    CAF is confirmed - ground truth - to always get its own child row here)."""
    row_set = _run_tad()["OEW"]
    codes = {n.code for n in row_set.cmdt_notes}
    assert "CAF" in codes


def test_tad_excluded_charge_codes_drops_it_end_to_end():
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    profile = MappingProfile(rfa_effective_date=RFA_EFFECTIVE, rfa_expiry_date=RFA_EXPIRY, excluded_charge_codes=["CAF"])
    row_set = TADOewOmwParser().run_multi(wb, profile)["OEW"]

    codes = {n.code for n in row_set.cmdt_notes}
    assert "CAF" not in codes
    assert "CSS" in codes  # untouched - only CAF was excluded
    assert "CAF" not in row_set.cmdt_notes[0].contents
