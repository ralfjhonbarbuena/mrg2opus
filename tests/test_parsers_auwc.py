from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.auwc import AUWCParser, DEFAULT_MAIN_DESCRIPTION, DEFAULT_NOR_DESCRIPTION, DEFAULT_RF_DESCRIPTION
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

# 2 weekly pairs, both real ground truth (tracker status "Completed").
PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "33_AUS NEA to AUWC FAK" / "ONE AU MRG 2026_0815 to 2026_0831 - ex NEA to AUWC (07 August 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "33_AUS NEA to AUWC FAK" / "AUWC 15 TO 31.xlsx",
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "34_AUS NEA to AUWC FAK" / "ONE AU MRG 2026_0901 to 2026_0914 - ex NEA to AUWC (25 August 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "34_AUS NEA to AUWC FAK" / "AUWC 1 TO 14.xlsx",
    ),
]

pytestmark = pytest.mark.skipif(
    any(not p.exists() for pair in PAIRS for p in pair),
    reason="reference/ ground-truth files not present in this checkout",
)

# Known, deliberate deviations from ground truth (same categories already
# documented for every other lane, see RATES_IGNORE_FIELDS_BY_LANE in
# audit/compare.py):
#   - type: forced to "C" on every row, a user-directed business rule.
#   - commodity_group_code/commodity_group_description/cmdt_seq: OPUS's
#     own global running sequence (G0004/G0005/G0006 in this ground
#     truth) is not reproducible from the raw MRG alone.
#   - route_seq: single running counter per commodity group - verified
#     correct in shape (see test_auwc_route_seq_resets_per_commodity_group
#     below) but not pinned to the exact literal ground-truth numbers.
#   - commodity_note: exact text verified separately via the CMDT NOTE
#     sheet comparison below.
RATES_IGNORE_FIELDS = {
    "type", "commodity_group_code", "commodity_group_description", "cmdt_seq", "commodity_note", "route_seq",
}


def _run_auwc(raw_path: Path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    return AUWCParser().run(wb, MappingProfile())


@pytest.mark.parametrize("raw_path,opus_path", PAIRS)
def test_auwc_rates_matches_ground_truth(raw_path, opus_path):
    row_set = _run_auwc(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize("raw_path,opus_path", PAIRS)
def test_auwc_cmdt_note_matches_ground_truth(raw_path, opus_path):
    """Covers the lane-specific POR scoping (ISL only for Taiwan-origin
    shipments; EFS filed as two separate rows scoped to Korea and Hong
    Kong respectively) - note this lane stamps the scope on POR, not POL
    like AUEC - see auwc.py's INCLUDED_CHARGES module comment."""
    row_set = _run_auwc(raw_path)
    generated = [r.model_dump() for r in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SRCHG")

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            if field_name in ("application_effective", "application_expires") and g.get("code") != "APP":
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_auwc_route_seq_resets_per_commodity_group():
    """Confirmed against ground truth: Route Seq. is a single running
    counter per commodity group - the main group's DG rows continue the
    same counter as their DR parent (same group), but the separate RF and
    NOR groups each restart at 1."""
    raw_path, _ = PAIRS[0]
    row_set = _run_auwc(raw_path)

    for description in (DEFAULT_MAIN_DESCRIPTION, DEFAULT_RF_DESCRIPTION, DEFAULT_NOR_DESCRIPTION):
        group_rows = [r for r in row_set.rates if r.commodity_group_description == description]
        assert group_rows, f"no rows found for {description}"
        assert [r.route_seq for r in group_rows] == list(range(1, len(group_rows) + 1))


def test_auwc_three_commodity_groups_are_distinct():
    """Unlike AUEC (RF/RAD folded into the main group), AUWC's RF and NOR
    rows each get their OWN separate commodity group - confirmed 3
    distinct groups in ground truth."""
    raw_path, _ = PAIRS[0]
    row_set = _run_auwc(raw_path)
    descriptions = {r.commodity_group_description for r in row_set.rates}
    assert descriptions == {DEFAULT_MAIN_DESCRIPTION, DEFAULT_RF_DESCRIPTION, DEFAULT_NOR_DESCRIPTION}


def test_auwc_excluded_charge_codes_drops_efs_end_to_end():
    """MappingProfile.excluded_charge_codes wired through: excluding EFS
    drops BOTH its POR-scoped rows (Korea and Hong Kong), not just one."""
    raw_path, _ = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUWCParser().run(wb, MappingProfile(excluded_charge_codes=["EFS"]))

    codes = {n.code for n in row_set.cmdt_notes}
    assert "EFS" not in codes
    assert "ISL" in codes  # untouched - only EFS was excluded


def test_auwc_skip_dg_generation_suppresses_dg_rows_for_main_group_only():
    """skip_dg_generation is keyed per commodity group's default
    description - suppressing the main group's DG rows must not affect
    the separate RF/NOR groups (which never had DG rows to begin with)."""
    raw_path, _ = PAIRS[0]
    default_row_set = _run_auwc(raw_path)
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUWCParser().run(wb, MappingProfile(skip_dg_generation={DEFAULT_MAIN_DESCRIPTION: True}))

    default_main_cgo = {r.cgo_type for r in default_row_set.rates if r.commodity_group_description == DEFAULT_MAIN_DESCRIPTION}
    assert "DG" in default_main_cgo

    main_cgo = {r.cgo_type for r in row_set.rates if r.commodity_group_description == DEFAULT_MAIN_DESCRIPTION}
    assert "DG" not in main_cgo
    assert "DR" in main_cgo

    assert len(row_set.rates) < len(default_row_set.rates)
