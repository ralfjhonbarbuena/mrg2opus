from __future__ import annotations

from dataclasses import dataclass
from datetime import date

import openpyxl

from mrg2opus.parsers.common.tad_snapshots import find_snapshot_sheets, merge_dated_snapshots


@dataclass
class _Row:
    validity_start: date | None
    validity_end: date | None
    tag: str


def test_single_occurrence_is_a_no_op():
    rows = [_Row(date(2026, 9, 1), date(2026, 9, 15), "a")]
    assert merge_dated_snapshots([rows]) == rows


def test_no_occurrences_returns_empty():
    assert merge_dated_snapshots([]) == []
    assert merge_dated_snapshots([[]]) == []


def test_earlier_occurrence_truncated_to_day_before_later_starts():
    """Confirmed against reference/23_TAD FILING AEW AMW's own ground
    truth: a rate correction issued mid-period truncates the earlier
    filing's validity_end to end the day before the later one starts."""
    week1 = [_Row(date(2026, 9, 1), date(2026, 9, 15), "week1")]
    week2 = [_Row(date(2026, 9, 7), date(2026, 9, 15), "week2")]

    merged = merge_dated_snapshots([week1, week2])

    assert len(merged) == 2
    by_tag = {r.tag: r for r in merged}
    assert by_tag["week1"].validity_end == date(2026, 9, 6)
    assert by_tag["week2"].validity_end == date(2026, 9, 15)


def test_occurrence_order_in_the_input_list_does_not_matter():
    """Sorted internally by each occurrence's own validity_start, not by
    upload/list order."""
    week2 = [_Row(date(2026, 9, 7), date(2026, 9, 15), "week2")]
    week1 = [_Row(date(2026, 9, 1), date(2026, 9, 15), "week1")]

    merged = merge_dated_snapshots([week2, week1])

    by_tag = {r.tag: r for r in merged}
    assert by_tag["week1"].validity_end == date(2026, 9, 6)
    assert by_tag["week2"].validity_end == date(2026, 9, 15)


def test_three_occurrences_each_truncated_to_the_next():
    w1 = [_Row(date(2026, 9, 1), date(2026, 9, 30), "w1")]
    w2 = [_Row(date(2026, 9, 10), date(2026, 9, 30), "w2")]
    w3 = [_Row(date(2026, 9, 20), date(2026, 9, 30), "w3")]

    merged = merge_dated_snapshots([w1, w2, w3])

    by_tag = {r.tag: r for r in merged}
    assert by_tag["w1"].validity_end == date(2026, 9, 9)
    assert by_tag["w2"].validity_end == date(2026, 9, 19)
    assert by_tag["w3"].validity_end == date(2026, 9, 30)


def test_does_not_extend_a_row_whose_own_end_is_already_earlier():
    """Truncation only ever shortens - a row already ending before the
    next occurrence's start keeps its own (earlier) end date."""
    week1 = [_Row(date(2026, 9, 1), date(2026, 9, 3), "short")]
    week2 = [_Row(date(2026, 9, 7), date(2026, 9, 15), "week2")]

    merged = merge_dated_snapshots([week1, week2])

    by_tag = {r.tag: r for r in merged}
    assert by_tag["short"].validity_end == date(2026, 9, 3)


def test_find_snapshot_sheets_matches_base_and_numbered_occurrences():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("AEW")
    wb.create_sheet("AEW (2)")
    wb.create_sheet("AEW ARBS")  # must NOT match - different sheet entirely

    assert find_snapshot_sheets(wb, "AEW") == ["AEW", "AEW (2)"]


def test_find_snapshot_sheets_returns_empty_when_absent():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("AMW")

    assert find_snapshot_sheets(wb, "AEW") == []
