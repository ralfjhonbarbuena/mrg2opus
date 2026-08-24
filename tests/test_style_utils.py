from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color

from mrg2opus.excel_io.style_utils import is_blacked_out, is_excluded, is_struck_through

# A real (non-default) theme, byte-for-byte the same shape as one found in
# a real user MRG file that triggered this: dk1/dk2 = black, lt1/lt2 =
# white - confirmed via that file's own xl/theme/theme1.xml, not a guess.
# OOXML SpreadsheetML swaps theme color index 0<->1 and 2<->3 relative to
# this declaration order (index 0=lt1, 1=dk1, 2=lt2, 3=dk2) - a documented
# quirk, verified against the real file: a cell with theme index 1 there
# renders visibly black, matching dk1, not lt1 (which the unswapped
# reading would incorrectly resolve to).
_THEME_XML = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Sheets">
<a:themeElements><a:clrScheme name="Sheets">
<a:dk1><a:srgbClr val="000000"/></a:dk1>
<a:lt1><a:srgbClr val="FFFFFF"/></a:lt1>
<a:dk2><a:srgbClr val="000000"/></a:dk2>
<a:lt2><a:srgbClr val="FFFFFF"/></a:lt2>
<a:accent1><a:srgbClr val="4472C4"/></a:accent1>
<a:accent2><a:srgbClr val="ED7D31"/></a:accent2>
<a:accent3><a:srgbClr val="A5A5A5"/></a:accent3>
<a:accent4><a:srgbClr val="FFC000"/></a:accent4>
<a:accent5><a:srgbClr val="5B9BD5"/></a:accent5>
<a:accent6><a:srgbClr val="70AD47"/></a:accent6>
<a:hlink><a:srgbClr val="0563C1"/></a:hlink>
<a:folHlink><a:srgbClr val="0563C1"/></a:folHlink>
</a:clrScheme></a:themeElements></a:theme>"""


def test_strikethrough_cell_is_detected():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "cancelled rate"
    ws["A1"].font = Font(strike=True)
    ws["A2"] = "normal rate"

    assert is_struck_through(ws["A1"])
    assert not is_struck_through(ws["A2"])
    assert is_excluded(ws["A1"])
    assert not is_excluded(ws["A2"])


def test_blacked_out_fill_is_detected():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hidden"
    ws["A1"].fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    ws["A2"] = "visible"
    ws["A2"].fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    assert is_blacked_out(ws["A1"])
    assert not is_blacked_out(ws["A2"])
    assert is_excluded(ws["A1"])


def test_theme_referenced_black_fill_is_detected():
    """A trader can black out a cell via Excel's theme color swatch (Fill
    Color -> a theme black square) instead of a literal RGB color picker -
    openpyxl represents this as a `theme` color reference, not `rgb`, and
    the cell's true color can only be found by resolving that index
    against the WORKBOOK's own theme (themes vary per file, never assume
    a fixed palette). Real-world case: a lane's raw sheet zeroed out one
    origin's rates across every destination via exactly this styling, and
    the un-fixed code silently filed real $0.00 rates for all of them."""
    wb = Workbook()
    wb.loaded_theme = _THEME_XML
    ws = wb.active
    ws["A1"] = "blacked via theme"
    ws["A1"].fill = PatternFill(fgColor=Color(theme=1, type="theme"), fill_type="solid")
    ws["A2"] = "theme white, not black"
    ws["A2"].fill = PatternFill(fgColor=Color(theme=0, type="theme"), fill_type="solid")

    assert is_blacked_out(ws["A1"])
    assert not is_blacked_out(ws["A2"])
    assert is_excluded(ws["A1"])
