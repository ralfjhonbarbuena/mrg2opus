from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.laec import LAECLuxParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

_PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "49_LAEC LUX" / "LAEC Pricing Guideline - (20260815-20260831-IN) (FAK) for via LUX Service.xlsx",
        REFERENCE_DIR / "2_OPUS" / "49_LAEC LUX" / "LAEC Pricing Guideline - (20260815-20260831-IN) (FAK) for via LUX Service_opus.xlsx",
        date(2026, 5, 14),
        date(2026, 9, 30),
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "50_LAEC LUX" / "LAEC Pricing Guideline - (20260901-20260914-IN) (FAK) for via LUX Service.xlsx",
        REFERENCE_DIR / "2_OPUS" / "50_LAEC LUX" / "LAEC Pricing Guideline - (20260901-20260914-IN) (FAK) for via LUX Service_opus.xlsx",
        date(2026, 8, 23),
        date(2026, 12, 31),
    ),
]

pytestmark = pytest.mark.skipif(
    any(not raw.exists() or not opus.exists() for raw, opus, _, _ in _PAIRS),
    reason="reference/ ground-truth files not present in this checkout",
)

# Known, verified gaps - same categories already accepted for the main LAEC
# lane (see test_parsers_laec.py's own RATES_IGNORE_FIELDS comment):
#   - cmdt_seq/route_seq: externally-assigned running numbers.
#   - type: forced to "C" on every generated row; ground truth leaves blank.
#   - commodity_group_description/commodity_group_code: user-customizable,
#     this real reference file's own values ("G0007"/"FAK -  for via LUX
#     service only") are the parser's default but not enforced by this diff.
RATES_IGNORE_FIELDS = {"cmdt_seq", "route_seq", "type", "commodity_group_description", "commodity_group_code"}


def _run(raw_path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    return LAECLuxParser().run(wb, MappingProfile())


@pytest.mark.parametrize("raw_path,opus_path,rfa_eff,rfa_exp", _PAIRS)
def test_laec_lux_rates_matches_ground_truth(raw_path, opus_path, rfa_eff, rfa_exp):
    row_set = _run(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


def test_laec_lux_rates_no_arbs_or_special_note():
    """Reduced output scope confirmed against ground truth: RATES + SUR
    (CMDT NOTE) + FREETIME only - no ORIGIN ARBS, no SPECIAL NOTE."""
    row_set = _run(_PAIRS[0][0])
    assert row_set.arbs == []
    assert row_set.special_notes == []


@pytest.mark.parametrize("raw_path,opus_path,rfa_eff,rfa_exp", _PAIRS)
def test_laec_lux_cmdt_note_matches_ground_truth(raw_path, opus_path, rfa_eff, rfa_exp):
    """Single commodity group (G0007), single CMDT NOTE block. The child
    rows' own Application Effective/Expires dates are an externally-
    assigned RFA window, not derivable from the raw MRG (confirmed
    different between both real samples) - supplied here the same way the
    main LAEC lane already supports via MappingProfile.rfa_effective_date/
    rfa_expiry_date."""
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    profile = MappingProfile(rfa_effective_date=rfa_eff, rfa_expiry_date=rfa_exp)
    row_set = LAECLuxParser().run(wb, profile)

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SUR")

    assert len(row_set.cmdt_notes) == len(expected) == 6  # 1 APP parent + 5 charge-code children

    ignore_fields = {"header_seq", "note_seq", "pol"}
    for i, (g, e) in enumerate(zip(row_set.cmdt_notes, expected)):
        gd, ed = g.model_dump(), e
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore_fields:
                continue
            gv, ev = _normalize(gd.get(field_name)), _normalize(ed.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def _find_freetime_sheet(wb):
    for name in wb.sheetnames:
        if "FREETIME" in name.upper():
            return wb[name]
    raise KeyError(f"no FREETIME sheet in {wb.sheetnames}")


def _read_freetime_rows(path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_freetime_sheet(wb)
    rows = []
    for r in range(3, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, 47)]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(cols.FREETIME_ROW_FIELDS, values)))
    return rows


def _norm(v):
    v = _normalize(v)
    return None if v == "" else v


@pytest.mark.parametrize("raw_path,opus_path,rfa_eff,rfa_exp", _PAIRS)
def test_laec_lux_freetime_matches_ground_truth(raw_path, opus_path, rfa_eff, rfa_exp):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = LAECLuxParser().run(wb, MappingProfile())
    generated = [r.model_dump() for r in row_set.freetime]
    expected = _read_freetime_rows(opus_path)

    assert len(generated) == len(expected) == 8
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.FREETIME_ROW_FIELDS:
            gv, ev = _norm(g.get(field_name)), _norm(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"
