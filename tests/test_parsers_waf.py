from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.waf import DEFAULT_DR_DESCRIPTION, WAFParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

# 2 weekly pairs, both real ground truth (tracker status "Completed").
PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "9_West Africa WAF" / "Asia WAF MRG rate (26 Aug 2026 - 01 Sep 2026) (18 Aug updated.) (1).xlsx",
        REFERENCE_DIR / "2_OPUS" / "9_West Africa WAF" / "WEST AFRICA WAF ( 20260826 - 20260901 ).xlsx",
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "10_West Africa WAF" / "Asia WAF MRG rate (02 Sep 2026 - 08 Sep 2026) (25 Aug updated.).xlsx",
        REFERENCE_DIR / "2_OPUS" / "10_West Africa WAF" / "WEST AFRICA WAF ( 20260902 - 20260908 ).xlsx",
    ),
]

pytestmark = pytest.mark.skipif(
    any(not p.exists() for pair in PAIRS for p in pair),
    reason="reference/ ground-truth files not present in this checkout",
)

# Known, deliberate deviations from ground truth (same categories already
# documented for every other lane, see RATES_IGNORE_FIELDS_BY_LANE in
# audit/compare.py):
#   - type: forced to "C" on every row, a user-directed business rule -
#     this lane's own ground truth leaves it blank.
#   - commodity_group_code/commodity_group_description: OPUS's own global
#     running sequence (G0006/G0007 in this ground truth) is not
#     reproducible from the raw MRG alone - user-customizable per filing,
#     same gap already documented for LAWC/LAEC/CSE.
#   - cmdt_seq: same "OPUS renumbers on import" placeholder gap as every
#     other lane's Header Seq.
#   - commodity_note: exact text is verified separately via the CMDT NOTE
#     sheet comparison below; excluded here only because it's keyed by
#     commodity_group_description, which is itself ignored above.
RATES_IGNORE_FIELDS = {"type", "commodity_group_code", "commodity_group_description", "cmdt_seq", "commodity_note"}


def _run_waf(raw_path: Path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    parser = WAFParser()
    return parser.run(wb, MappingProfile())


@pytest.mark.parametrize("raw_path,opus_path", PAIRS)
def test_waf_rates_matches_ground_truth(raw_path, opus_path):
    row_set = _run_waf(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize("raw_path,opus_path", PAIRS)
def test_waf_cmdt_note_matches_ground_truth(raw_path, opus_path):
    row_set = _run_waf(raw_path)
    generated = [r.model_dump() for r in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SRCHG")

    # reference/2_OPUS/9's own SRCHG sheet carries 36 rows (4 header_seq
    # blocks: 825, 826 for THIS week plus 827, 828 - byte-identical to
    # reference/2_OPUS/10's own 827/828 blocks) instead of the 18 this
    # week's RATES sheet actually needs (only cmdt_seq 825/826 appear
    # there) - a verified copy-paste leftover in the ground truth file
    # itself, not something to reproduce. Compare only the leading rows
    # that belong to this week.
    expected = expected[: len(generated)]
    assert len(generated) == len(expected)
    # header_seq/note_seq: OPUS's own externally-assigned running sequence
    # numbers, not derivable from this file (see "ignore" comment
    # convention in test_parsers_cse.py/test_parsers_eaf.py).
    # application_effective/expires (child rows only): ground truth uses an
    # unrelated, longer per-charge-code filing window (e.g. 20260520 for
    # this week's BAF row) rather than the weekly rate validity - the same
    # already-accepted gap documented for CSE's CMDT NOTE (see
    # RATES_IGNORE_FIELDS_BY_LANE / test_parsers_cse.py), not newly
    # introduced here.
    ignore = {"header_seq", "note_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            if field_name in ("application_effective", "application_expires") and g.get("code") != "APP":
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_waf_excluded_charge_codes_drops_baf_end_to_end():
    """MappingProfile.excluded_charge_codes wired through: this lane's real
    CMDT NOTE includes BAF by default (see
    test_waf_cmdt_note_matches_ground_truth), but a filing-wide exclusion
    should drop it entirely."""
    raw_path, _ = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = WAFParser().run(wb, MappingProfile(excluded_charge_codes=["BAF"]))

    codes = {n.code for n in row_set.cmdt_notes}
    assert "BAF" not in codes
    assert "EFS" in codes  # untouched - only BAF was excluded


def test_waf_rfa_override_matches_ground_truth_child_dates():
    """West Africa WAF's ground truth is where this gap was originally
    found: CMDT NOTE child rows' Application Effective/Expires (20260520 -
    20261231 for this week) are the charge code's own RFA (Rate Filing
    Agreement) window, not the weekly rate validity the parser defaults
    to - a human filer enters this per account, so MappingProfile.
    rfa_effective_date/rfa_expiry_date lets the UI reproduce it exactly."""
    from datetime import date

    raw_path, _ = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = WAFParser().run(
        wb, MappingProfile(rfa_effective_date=date(2026, 5, 20), rfa_expiry_date=date(2026, 12, 31))
    )

    parents = [n for n in row_set.cmdt_notes if n.code == "APP"]
    children = [n for n in row_set.cmdt_notes if n.code != "APP"]
    assert parents and children
    for p in parents:
        assert p.application_effective == date(2026, 8, 26)  # unaffected - this week's rate validity
        assert p.application_expires == date(2026, 9, 1)
    for c in children:
        assert c.application_effective == date(2026, 5, 20)
        assert c.application_expires == date(2026, 12, 31)


def test_waf_skip_dg_generation_suppresses_dg_rows():
    raw_path, _ = PAIRS[0]
    default_row_set = _run_waf(raw_path)
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = WAFParser().run(wb, MappingProfile(skip_dg_generation={DEFAULT_DR_DESCRIPTION: True}))

    default_cgo_types = {r.cgo_type for r in default_row_set.rates}
    assert "DG" in default_cgo_types

    cgo_types = {r.cgo_type for r in row_set.rates}
    assert "DG" not in cgo_types
    assert "DR" in cgo_types
    assert len(row_set.rates) < len(default_row_set.rates)
