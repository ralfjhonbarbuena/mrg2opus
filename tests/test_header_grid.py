from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

from mrg2opus.parsers.common.header_grid import flatten_pod_header

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"
CSE_RAW_PATH = (
    REFERENCE_DIR / "1_MRGs" / "1_CSE FAK, CSE FAK FOR VELAG AND VEPBL"
    / "CSE Pricing Guideline (15-21  AUG 2026 ) FAK.xlsx"
)


@pytest.mark.skipif(not CSE_RAW_PATH.exists(), reason="reference/ ground-truth files not present in this checkout")
def test_flatten_pod_header_against_real_cse_sample():
    wb = openpyxl.load_workbook(CSE_RAW_PATH, data_only=False)
    ws = wb["CSE"]

    columns = flatten_pod_header(ws, pod_label_row=7, container_label_row=9, min_col=3, max_col=11)

    labels = [(c.pod_label, c.container_label) for c in columns]
    assert labels == [
        ("Manzanillo, PA", "D2"), ("Manzanillo, PA", "D4"), ("Manzanillo, PA", "D5"),
        ("Colon Free Zone\n (Door via Manzanillo, PA)", "D2"),
        ("Colon Free Zone\n (Door via Manzanillo, PA)", "D4"),
        ("Colon Free Zone\n (Door via Manzanillo, PA)", "D5"),
        ("Caucedo", "D2"), ("Caucedo", "D4"), ("Caucedo", "D5"),
    ]


def _build_two_pod_sheet() -> "openpyxl.worksheet.worksheet.Worksheet":
    """3 columns per POD (D2/D4/D5), 2 PODs - mirrors CSE/LAEC/LAWC's shape,
    minus merged cells (flatten_pod_header's merge-range lookup is already
    covered by the real CSE sample test above; this isolates the exclusion
    check)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Manzanillo")
    ws.cell(row=1, column=4, value="Colon")
    for col, label in [(1, "D2"), (2, "D4"), (3, "D5"), (4, "D2"), (5, "D4"), (6, "D5")]:
        ws.cell(row=2, column=col, value=label)
    return ws


def test_flatten_pod_header_drops_struck_through_pod():
    ws = _build_two_pod_sheet()
    ws.cell(row=1, column=1).font = Font(strike=True)  # "Manzanillo" withdrawn

    columns = flatten_pod_header(ws, pod_label_row=1, container_label_row=2, min_col=1, max_col=6)

    pods = {c.pod_label for c in columns}
    assert pods == {"Colon"}
    assert len(columns) == 3  # only Colon's D2/D4/D5


def test_flatten_pod_header_drops_blacked_out_pod():
    ws = _build_two_pod_sheet()
    ws.cell(row=1, column=4).fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")

    columns = flatten_pod_header(ws, pod_label_row=1, container_label_row=2, min_col=1, max_col=6)

    pods = {c.pod_label for c in columns}
    assert pods == {"Manzanillo"}
    assert len(columns) == 3
