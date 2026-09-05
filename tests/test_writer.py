"""excel_io.writer output sheet naming + RN sheet - see
project-opus-note-sheet-taxonomy memory for why these real filing names
(RATES, ORIGIN ARBS, CMDT NOTE, SPECIAL NOTE, RN) differ from
schema.opus_columns's SHEET_NAME_* constants (those match the older
bundled-sample fixtures instead, see that module's own docstring)."""
from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl

from mrg2opus.excel_io.writer import write_opus_workbook_multi
from mrg2opus.schema.opus_rows import (
    ArbsRow,
    CmdtNoteRow,
    FreetimeRow,
    OpusRowSet,
    RatesPortPortRow,
    RatesRow,
    RouteNoteRow,
    SpecialNoteRow,
    VerticalRatesRow,
)

_RATES_KWARGS = dict(
    commodity_group_code="G0001", commodity_group_description="FAK",
    origin_code="CNSHA", origin_description="Shanghai",
    destination_code="USLAX", destination_description="Los Angeles",
    prefix="D", cgo_type="DR",
)


def _write_and_load(row_sets: dict[str, OpusRowSet]) -> openpyxl.Workbook:
    with tempfile.TemporaryDirectory() as tmp_dir:
        out_path = Path(tmp_dir) / "out.xlsx"
        write_opus_workbook_multi(row_sets, out_path)
        return openpyxl.load_workbook(out_path)


def test_writer_uses_real_filing_sheet_names():
    row_set = OpusRowSet(
        rates=[RatesRow(**_RATES_KWARGS)],
        rates_port_port=[RatesPortPortRow(**_RATES_KWARGS)],
        arbs=[ArbsRow(point="CNSHA")],
        cmdt_notes=[CmdtNoteRow(contents="note")],
        special_notes=[SpecialNoteRow(contents="note")],
        route_notes=[RouteNoteRow(contents="Rates are applicable for Vessel Service Lane: MX2")],
    )
    wb = _write_and_load({"": row_set})
    assert set(wb.sheetnames) == {"RATES", "RATES PORT-PORT", "ORIGIN ARBS", "CMDT NOTE", "SPECIAL NOTE", "RN"}


def test_writer_suffix_tags_real_filing_sheet_names():
    row_set = OpusRowSet(rates=[RatesRow(**_RATES_KWARGS)], route_notes=[RouteNoteRow(contents="note")])
    wb = _write_and_load({"TZDAR": row_set})
    assert set(wb.sheetnames) == {"RATES-TZDAR", "RN-TZDAR"}


def test_writer_omits_route_note_sheet_when_empty():
    row_set = OpusRowSet(rates=[RatesRow(**_RATES_KWARGS)])
    wb = _write_and_load({"": row_set})
    assert "RN" not in wb.sheetnames


def test_vertical_rates_written_as_one_sheet():
    """Real filings keep every commodity group on a single VERTICAL RATES
    sheet (confirmed: reference/2_OPUS/27's WEW sheet is 5,209 rows across
    4 groups), so this is never split - however many rows it runs to."""
    rows = [VerticalRatesRow(per="D2", rate=100) for _ in range(5)]
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "o.xlsx"
        write_opus_workbook_multi({"": OpusRowSet(vertical_rates=rows)}, out)
        wb = openpyxl.load_workbook(out)
        assert [s for s in wb.sheetnames if "VERTICAL" in s] == ["VERTICAL RATES"]
        assert wb["VERTICAL RATES"].max_row == 7  # 2 header rows + 5


def test_freetime_sheet_blanks_the_columns_opus_assigns_itself():
    """The reference FREETIME sheets are 46 columns wide because they're
    DOWNLOADS of live filings; an upload must not send OPUS's own
    identifiers back to it, so "RFA No.", "Status" and everything from
    column AO ("DAR No.") rightwards come out EMPTY (user direction,
    2026-08-31). The header keeps all 46 columns and every other value
    stays under its own heading - deleting the columns outright shifted
    the header away from the data."""
    row = FreetimeRow(seq="1", rfa_no="SINN02654A", status="Approved", tariff="CTIC", dar_no="X")
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "o.xlsx"
        write_opus_workbook_multi({"": OpusRowSet(freetime=[row])}, out)
        ws = openpyxl.load_workbook(out)["FREETIME"]

    assert ws.max_column == 46
    header = [ws.cell(row=1, column=c).value for c in range(1, 47)]
    assert header[1] == "RFA No." and header[2] == "Status"
    # "Customer" spans the last two columns (AS1:AT1), so only the first
    # of the pair holds the label - merging keeps the top-left cell.
    assert header[40] == "DAR No." and header[44] == "Customer" and header[45] is None

    values = {header[c - 1]: ws.cell(row=3, column=c).value for c in range(1, 47)}
    assert values["RFA No."] is None
    assert values["Status"] is None
    assert values["DAR No."] is None
    # Everything else still lines up with its own column.
    assert values["Seq."] == 1  # FreetimeRow coerces seq to int
    assert values["Tariff"] == "CTIC"
    assert ws.cell(row=3, column=4).value == "CTIC"  # Tariff is still column D
