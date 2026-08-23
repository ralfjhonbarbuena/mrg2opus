from __future__ import annotations

import openpyxl

from mrg2opus.audit.compare import (
    arbs_row_key,
    diff_by_key,
    find_sheet,
    rates_row_key,
    read_arbs_sheet,
    read_rates_sheet,
)
from mrg2opus.schema import opus_columns as cols


def _rates_sheet_wb(rows: list[list]) -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cols.SHEET_NAME_RATES
    # Rows start at row 3 (rows 1-2 are the 2-row header) - matches every
    # lane's writer output and golden.py's existing read_rates_sheet.
    for offset, row in enumerate(rows):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=3 + offset, column=col_idx, value=value)
    return wb


def test_find_sheet_exact_match():
    wb = openpyxl.Workbook()
    wb.active.title = "OPUS RATES"
    assert find_sheet(wb, "OPUS RATES") == "OPUS RATES"


def test_find_sheet_loose_whitespace_hyphen_match():
    wb = openpyxl.Workbook()
    wb.active.title = "OPUS RATES PORT - PORT"  # SAF's spacing quirk
    assert find_sheet(wb, "OPUS RATES PORT-PORT") == "OPUS RATES PORT - PORT"


def test_find_sheet_raises_when_missing():
    wb = openpyxl.Workbook()
    wb.active.title = "Something Else"
    try:
        find_sheet(wb, "OPUS RATES")
        assert False, "expected KeyError"
    except KeyError:
        pass


def test_read_rates_sheet_skips_blank_rows_and_reads_by_field_order():
    row = [None] * len(cols.RATES_ROW_FIELDS)
    row[cols.RATES_ROW_FIELDS.index("origin_code")] = "CNSHA"
    row[cols.RATES_ROW_FIELDS.index("destination_code")] = "USLAX"
    wb = _rates_sheet_wb([row, [None] * len(cols.RATES_ROW_FIELDS)])
    rows = read_rates_sheet(wb, cols.SHEET_NAME_RATES)
    assert len(rows) == 1
    assert rows[0]["origin_code"] == "CNSHA"
    assert rows[0]["destination_code"] == "USLAX"


def test_arbs_row_key():
    assert arbs_row_key({"point": "CNSHA", "over": "20MT", "per": "MT", "seq": 1}) == ("CNSHA", "20MT", "MT")


def test_diff_by_key_matched_missing_extra_and_field_mismatch():
    generated = [
        {"k": "A", "v": 1},
        {"k": "B", "v": 2},
        {"k": "C", "v": 99},
    ]
    expected = [
        {"k": "A", "v": 1},
        {"k": "C", "v": 3},
        {"k": "D", "v": 4},
    ]
    result = diff_by_key(generated, expected, key_fn=lambda r: (r["k"],), fields=["v"])
    assert result.matched == 1
    assert result.missing == {("D",)}
    assert result.extra == {("B",)}
    assert result.field_mismatches == [(("C",), "v", 99, 3)]


def test_diff_by_key_respects_ignore_fields():
    generated = [{"k": "A", "v": 1, "note": "x"}]
    expected = [{"k": "A", "v": 1, "note": "y"}]
    result = diff_by_key(
        generated, expected, key_fn=lambda r: (r["k"],), fields=["v", "note"], ignore_fields={"note"}
    )
    assert result.matched == 1
    assert result.field_mismatches == []


from mrg2opus.audit.compare import diff_cmdt_blocks, reconstruct_blocks


def _cmdt_row(header_seq=None, note_seq=None, contents=None, charge_seq=None, code=None, amount=None) -> dict:
    row = dict.fromkeys(cols.CMDT_NOTE_ROW_FIELDS)
    row.update(header_seq=header_seq, note_seq=note_seq, contents=contents, charge_seq=charge_seq, code=code, amount=amount)
    return row


def test_reconstruct_blocks_groups_children_under_parent():
    rows = [
        _cmdt_row(header_seq=1, note_seq=1, contents="Block A", charge_seq=1, code="PSS"),
        _cmdt_row(charge_seq=2, code="OBS"),
        _cmdt_row(header_seq=2, note_seq=2, contents="Block B", charge_seq=1, code="EFS"),
    ]
    blocks = reconstruct_blocks(rows)
    assert [b.key for b in blocks] == ["Block A", "Block B"]
    assert len(blocks[0].children) == 1
    assert blocks[0].children[0]["code"] == "OBS"
    assert len(blocks[1].children) == 0


def test_diff_cmdt_blocks_matched_identical():
    generated = [
        _cmdt_row(header_seq=1, note_seq=1, contents="Block A", charge_seq=1, code="PSS"),
        _cmdt_row(charge_seq=2, code="OBS"),
    ]
    expected = [
        _cmdt_row(header_seq=1, note_seq=1, contents="Block A", charge_seq=1, code="PSS"),
        _cmdt_row(charge_seq=2, code="OBS"),
    ]
    result = diff_cmdt_blocks(generated, expected, cols.CMDT_NOTE_ROW_FIELDS)
    assert result.missing_blocks == []
    assert result.extra_blocks == []
    assert result.field_mismatches == []


def test_diff_cmdt_blocks_missing_and_extra():
    generated = [_cmdt_row(header_seq=1, note_seq=1, contents="Only Generated", charge_seq=1, code="PSS")]
    expected = [_cmdt_row(header_seq=1, note_seq=1, contents="Only Reference", charge_seq=1, code="PSS")]
    result = diff_cmdt_blocks(generated, expected, cols.CMDT_NOTE_ROW_FIELDS)
    assert result.missing_blocks == ["Only Reference"]
    assert result.extra_blocks == ["Only Generated"]


def test_diff_cmdt_blocks_field_mismatch_within_matched_block():
    generated = [_cmdt_row(header_seq=1, note_seq=1, contents="Block A", charge_seq=1, code="PSS", amount=100)]
    expected = [_cmdt_row(header_seq=1, note_seq=1, contents="Block A", charge_seq=1, code="PSS", amount=200)]
    result = diff_cmdt_blocks(generated, expected, cols.CMDT_NOTE_ROW_FIELDS)
    assert result.missing_blocks == []
    assert result.extra_blocks == []
    assert len(result.field_mismatches) == 1
    key, idx, field_name, gv, ev = result.field_mismatches[0]
    assert (key, idx, field_name, gv, ev) == ("Block A", 0, "amount", 100, 200)
