"""The written headers must match the user's own OPUS HEADERS.xlsx, which
is the authority on what OPUS expects: two header rows with the group
label MERGED across its span rather than repeated (user direction,
2026-09-05, "maintain the format of the headers just like in the file").
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.excel_io.writer import write_opus_workbook_multi
from mrg2opus.schema import opus_columns as cols
from mrg2opus.schema.opus_rows import (
    CmdtNoteRow,
    FreetimeRow,
    OpusRowSet,
    RatesPortPortRow,
    RatesRow,
    SpecialNoteRow,
    VerticalRatesRow,
)

REFERENCE = Path(__file__).resolve().parents[1] / "reference" / "OPUS HEADERS.xlsx"

pytestmark = pytest.mark.skipif(not REFERENCE.exists(), reason="OPUS HEADERS.xlsx not in this checkout")


def _written() -> openpyxl.Workbook:
    row_set = OpusRowSet(
        rates_port_port=[
            RatesPortPortRow(
                commodity_group_code="G0001", commodity_group_description="d",
                origin_code="CNSHA", origin_description="SHANGHAI",
                destination_code="AEJEA", destination_description="JEBEL ALI",
                prefix="D", cgo_type="DR", rate_20=100,
            )
        ],
        rates=[
            RatesRow(
                commodity_group_code="G0001", commodity_group_description="d",
                origin_code="CNSHA", origin_description="SHANGHAI",
                destination_code="AEJEA", destination_description="JEBEL ALI",
                prefix="D", cgo_type="DR", rate_20=100,
            )
        ],
        vertical_rates=[VerticalRatesRow(per="D2", rate=100)],
        cmdt_notes=[CmdtNoteRow(contents="note", charge_seq=1)],
        special_notes=[SpecialNoteRow(contents="note", charge_seq=1)],
        freetime=[FreetimeRow(seq="1")],
    )
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "o.xlsx"
        write_opus_workbook_multi({"": row_set}, out)
        return openpyxl.load_workbook(out)


@pytest.mark.parametrize("sheet", ["RATES", "VERTICAL RATES", "FREETIME"])
def test_two_row_headers_match_the_reference_exactly(sheet):
    """Values in both rows AND the merged ranges."""
    ours = _written()[sheet]
    ref = openpyxl.load_workbook(REFERENCE)[sheet]

    assert ours.max_column == ref.max_column
    for row in (1, 2):
        assert [ours.cell(row=row, column=c).value for c in range(1, ref.max_column + 1)] == [
            ref.cell(row=row, column=c).value for c in range(1, ref.max_column + 1)
        ], f"{sheet} header row {row} differs"
    assert sorted(str(m) for m in ours.merged_cells.ranges) == sorted(
        str(m) for m in ref.merged_cells.ranges
    ), f"{sheet} merged ranges differ"


@pytest.mark.parametrize(
    "sheet,ref_sheet",
    [("CMDT NOTE", "CMDT NOTE"), ("SPECIAL NOTE", "SPECIAL NOTE"), ("RATES", "RATES")],
)
def test_single_row_headers_match_the_reference(sheet, ref_sheet):
    ours = _written()[sheet]
    ref = openpyxl.load_workbook(REFERENCE)[ref_sheet]
    assert ours.max_column == ref.max_column


def test_note_sheets_carry_the_nine_trailing_columns():
    """CMDT NOTE and SPECIAL NOTE are 39 wide in the reference, not 30 -
    the same trailing block RN has, minus RN's own Premium."""
    assert len(cols.CMDT_NOTE_HEADER) == len(cols.CMDT_NOTE_ROW_FIELDS) == 39
    assert len(cols.SPECIAL_NOTE_HEADER) == len(cols.SPECIAL_NOTE_ROW_FIELDS) == 39
    assert cols.CMDT_NOTE_ROW_FIELDS[-9:] == [
        "receiving_term", "delivery_term", "weight_gte_mt", "weight_lt_mt",
        "direct_call", "bar_type", "s_i", "mty_pickup_cy", "mty_return_cy",
    ]
    assert cols.RN_ROW_FIELDS[-1] == "premium"


def _style_signature(cell):
    border = cell.border
    return (
        cell.value,
        cell.fill.start_color.rgb if cell.fill and cell.fill.patternType else None,
        cell.font.color.rgb if cell.font.color and isinstance(cell.font.color.rgb, str) else None,
        cell.font.size, cell.font.name,
        cell.alignment.horizontal, cell.alignment.vertical, cell.alignment.wrap_text,
        border.left.style, border.right.style, border.top.style, border.bottom.style,
    )


@pytest.mark.parametrize(
    "sheet,ref_sheet",
    [
        ("RATES", "RATES"),
        ("RATES PORT-PORT", "RATES"),
        ("VERTICAL RATES", "VERTICAL RATES"),
        ("CMDT NOTE", "CMDT NOTE"),
        ("SPECIAL NOTE", "SPECIAL NOTE"),
        ("FREETIME", "FREETIME"),
    ],
)
def test_header_styling_matches_the_reference(sheet, ref_sheet):
    """Values alone aren't enough: an unstyled MERGED header renders
    left-aligned on white and reads as a broken merge (user-reported,
    2026-09-05, "the merging is incorrect"). Fill, font, alignment and
    borders have to match too.

    Cells hidden INSIDE a merge are skipped - Excel never draws them, and
    openpyxl leaves them at the workbook's default font, which differs
    between the file we write and one saved by a newer Excel.
    """
    from openpyxl.cell.cell import MergedCell

    ours = _written()[sheet]
    ref = openpyxl.load_workbook(REFERENCE)[ref_sheet]

    for row in range(1, ref.max_row + 1):
        for col in range(1, ref.max_column + 1):
            cell = ours.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            assert _style_signature(cell) == _style_signature(ref.cell(row=row, column=col)), (
                f"{sheet} r{row}c{col} differs"
            )
        assert ours.row_dimensions[row].height == ref.row_dimensions[row].height


def test_header_column_widths_match_the_reference():
    ours = _written()["RATES"]
    ref = openpyxl.load_workbook(REFERENCE)["RATES"]
    assert {k: round(v.width, 1) for k, v in ours.column_dimensions.items() if v.width} == {
        k: round(v.width, 1) for k, v in ref.column_dimensions.items() if v.width
    }
