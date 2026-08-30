from __future__ import annotations

import openpyxl
import pytest

from mrg2opus.excel_io.merge import DuplicateSheetError, merge_workbooks


def _wb_with_sheet(name: str, values: dict[tuple[int, int], object]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(name)
    for (row, col), value in values.items():
        ws.cell(row=row, column=col, value=value)
    return wb


def test_single_workbook_passthrough():
    wb = _wb_with_sheet("A", {(1, 1): "x"})
    assert merge_workbooks([wb]) is wb


def test_merges_sheets_from_multiple_workbooks():
    wb1 = _wb_with_sheet("CSE", {(1, 1): "main"})
    wb2 = _wb_with_sheet("CSE VE", {(1, 1): "ve"})

    merged = merge_workbooks([wb1, wb2])

    assert set(merged.sheetnames) == {"CSE", "CSE VE"}
    assert merged["CSE"].cell(row=1, column=1).value == "main"
    assert merged["CSE VE"].cell(row=1, column=1).value == "ve"


def test_preserves_merged_cell_ranges():
    wb1 = openpyxl.Workbook()
    wb1.remove(wb1.active)
    ws1 = wb1.create_sheet("Sheet1")
    ws1["A1"] = "spans"
    ws1.merge_cells("A1:C1")
    wb2 = _wb_with_sheet("Sheet2", {(1, 1): "y"})

    merged = merge_workbooks([wb1, wb2])

    assert "A1:C1" in [str(r) for r in merged["Sheet1"].merged_cells.ranges]


def test_preserves_font_and_fill_for_exclusion_detection():
    wb1 = openpyxl.Workbook()
    wb1.remove(wb1.active)
    ws1 = wb1.create_sheet("Sheet1")
    from openpyxl.styles import Font

    cell = ws1.cell(row=1, column=1, value=100)
    cell.font = Font(strike=True)
    wb2 = _wb_with_sheet("Sheet2", {(1, 1): "y"})

    merged = merge_workbooks([wb1, wb2])

    assert merged["Sheet1"].cell(row=1, column=1).font.strike is True


def test_duplicate_sheet_name_raises():
    wb1 = _wb_with_sheet("CSE", {(1, 1): "main"})
    wb2 = _wb_with_sheet("CSE", {(1, 1): "other"})

    with pytest.raises(DuplicateSheetError):
        merge_workbooks([wb1, wb2])


def test_duplicate_sheet_name_still_raises_when_names_given_but_not_venezuela():
    """names alone don't disable the safety check - only a name that
    actually matches the VELAG/VEPBL pattern does."""
    wb1 = _wb_with_sheet("CSE", {(1, 1): "main"})
    wb2 = _wb_with_sheet("CSE", {(1, 1): "other"})

    with pytest.raises(DuplicateSheetError):
        merge_workbooks([wb1, wb2], ["main.xlsx", "accidentally_reuploaded.xlsx"])


def test_cse_venezuela_supplement_renames_colliding_cse_sheet():
    """Real-world case (reference/1_MRGs/1_CSE FAK.../): CSE's second file
    is named "...for VELAG and VEPBL" and its own "CSE" sheet is genuinely
    different (Venezuela) data the parser expects under "CSE VE" - not an
    accidental duplicate upload."""
    wb1 = _wb_with_sheet("CSE", {(1, 1): "main"})
    wb2 = _wb_with_sheet("CSE", {(1, 1): "venezuela"})

    merged = merge_workbooks(
        [wb1, wb2], ["CSE Pricing Guideline FAK.xlsx", "CSE Pricing Guideline FAK for VELAG and VEPBL.xlsx"]
    )

    assert set(merged.sheetnames) == {"CSE", "CSE VE"}
    assert merged["CSE"].cell(row=1, column=1).value == "main"
    assert merged["CSE VE"].cell(row=1, column=1).value == "venezuela"


def test_cse_venezuela_supplement_drops_duplicate_support_sheets():
    """The same VELAG/VEPBL file re-bundles lane-wide support sheets
    ("DG surcharges", "Yangtze ARB Add-on", "Free Time") that aren't
    Venezuela-specific - the main file's copy wins, the duplicate is
    dropped rather than raising or silently overwriting."""
    wb1 = _wb_with_sheet("CSE", {(1, 1): "main"})
    wb1.create_sheet("DG surcharges")
    wb1["DG surcharges"].cell(row=1, column=1, value="main dg")
    wb2 = _wb_with_sheet("CSE", {(1, 1): "venezuela"})
    wb2.create_sheet("DG surcharges")
    wb2["DG surcharges"].cell(row=1, column=1, value="ve dg")

    merged = merge_workbooks([wb1, wb2], ["main.xlsx", "for VELAG and VEPBL.xlsx"])

    assert set(merged.sheetnames) == {"CSE", "CSE VE", "DG surcharges"}
    assert merged["DG surcharges"].cell(row=1, column=1).value == "main dg"


def test_venezuela_marker_matches_case_insensitively():
    wb1 = _wb_with_sheet("CSE", {(1, 1): "main"})
    wb2 = _wb_with_sheet("CSE", {(1, 1): "venezuela"})

    merged = merge_workbooks([wb1, wb2], ["main.xlsx", "for vepbl.xlsx"])

    assert "CSE VE" in merged.sheetnames


def test_tad_snapshot_sheet_collision_renames_instead_of_raising():
    """Real-world case (reference/1_MRGs/23_TAD FILING AEW AMW/): the team
    can issue a rate correction mid-period as a second raw file with the
    SAME sheet names - not an accidental duplicate upload, no filename
    marker needed (unlike CSE's Venezuela case) since any collision on a
    known TAD sheet name IS the legitimate multi-snapshot pattern."""
    wb1 = _wb_with_sheet("AEW", {(1, 1): "week1"})
    wb2 = _wb_with_sheet("AEW", {(1, 1): "week2"})

    merged = merge_workbooks([wb1, wb2])

    assert set(merged.sheetnames) == {"AEW", "AEW (2)"}
    assert merged["AEW"].cell(row=1, column=1).value == "week1"
    assert merged["AEW (2)"].cell(row=1, column=1).value == "week2"


def test_tad_snapshot_sheet_collision_numbers_a_third_occurrence():
    wb1 = _wb_with_sheet("OEW", {(1, 1): "a"})
    wb2 = _wb_with_sheet("OEW", {(1, 1): "b"})
    wb3 = _wb_with_sheet("OEW", {(1, 1): "c"})

    merged = merge_workbooks([wb1, wb2, wb3])

    assert set(merged.sheetnames) == {"OEW", "OEW (2)", "OEW (3)"}
    assert merged["OEW (3)"].cell(row=1, column=1).value == "c"


def test_non_tad_sheet_name_collision_still_raises():
    """A collision on a name outside the known TAD sheet set is treated
    exactly as before - no blanket "just rename everything" behavior."""
    wb1 = _wb_with_sheet("Random Sheet", {(1, 1): "a"})
    wb2 = _wb_with_sheet("Random Sheet", {(1, 1): "b"})

    with pytest.raises(DuplicateSheetError):
        merge_workbooks([wb1, wb2])
