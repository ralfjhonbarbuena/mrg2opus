from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.aubp import AUBPParser, DEFAULT_MAIN_DESCRIPTION, DEFAULT_RAD_DESCRIPTION, DEFAULT_RF_DESCRIPTION
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

# 2 weekly pairs, both real ground truth (tracker status "Completed").
PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "29_AUS SEA to AUBP FAK" / "ONE SEA to AU MRG 20260901 to 20260914 (24 AUG 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "29_AUS SEA to AUBP FAK" / "ONE SEA to AU MRG 20260901 to 20260914 (24 AUG 2026)_OPUS.xlsx",
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "30_AUS SEA to AUBP FAK" / "ONE SEA to AU MRG 20260815 to 20260831 (07 AUG 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "30_AUS SEA to AUBP FAK" / "ONE SEA to AU MRG 20260815 to 20260831 (07 AUG 2026)_OPUS.xlsx",
    ),
]

pytestmark = pytest.mark.skipif(
    any(not p.exists() for pair in PAIRS for p in pair),
    reason="reference/ ground-truth files not present in this checkout",
)

# Known, deliberate deviations from ground truth (same categories already
# documented for every other lane, see RATES_IGNORE_FIELDS_BY_LANE in
# audit/compare.py).
RATES_IGNORE_FIELDS = {
    "type", "commodity_group_code", "commodity_group_description", "cmdt_seq", "commodity_note", "route_seq",
}


def _run(raw_path: Path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    return AUBPParser().run(wb, MappingProfile())


def test_aubp_rates_matches_ground_truth_week2():
    """Week 2 is the "clean" reference week - full exact match. Week 1 has
    a known, documented gap (see below) around the Yangon/Thilawa merge."""
    raw_path, opus_path = PAIRS[1]
    row_set = _run(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


def test_aubp_rates_matches_ground_truth_week1_except_known_thilawa_gap():
    """Week 1's own ground truth drops Thilawa entirely instead of merging
    it with Yangon (both weeks otherwise end up with the same row count -
    week 1 by omission, week 2 by merge) - a real, confirmed divergence
    between the two reference weeks, not a parsing gap. This parser always
    merges (matches week 2's design, see aubp.py module docstring), so
    week 1 is verified as "matches ground truth except exactly the
    Yangon/Thilawa keys this documented difference explains"."""
    raw_path, opus_path = PAIRS[0]
    row_set = _run(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)

    unexplained_missing = [k for k in result.missing if k[0] != "MMRGN"]
    unexplained_extra = [k for k in result.extra if k[0] != "MMRGN;MMTLA"]
    assert not unexplained_missing, f"missing rows beyond the known Thilawa gap: {unexplained_missing[:5]}"
    assert not unexplained_extra, f"extra rows beyond the known Thilawa gap: {unexplained_extra[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"
    # Sanity: the gap is exactly the Dry group's 3 destination-blocks x 2
    # cgo_types (DR, DG) - Yangon/Thilawa never have Reefer/RAD data.
    assert len(result.missing) == 6
    assert len(result.extra) == 6


def test_aubp_cmdt_note_matches_ground_truth_week1_positionally():
    """Week 1's own charge_seq order (OBS, EFS, PSS, then the origin-scoped
    THL/ISL/DOC rows) matches this parser's fixed emission order exactly -
    verified positionally. POL/POR: week 1's ground truth scoped via POL
    (country code); this parser always uses POR (matching week 2's own
    choice for the identical business rule - see module docstring), so POL
    and POR are both excluded from this comparison and verified via week 2
    instead."""
    raw_path, opus_path = PAIRS[0]
    row_set = _run(raw_path)
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SUR")

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq", "pol", "por"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            if field_name in ("application_effective", "application_expires") and g.get("code") != "APP":
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_aubp_cmdt_note_matches_ground_truth_week2_by_content():
    """Week 2 both scopes via POR (matching this parser) AND orders PSS
    last instead of third (a genuine, confirmed per-filing difference, not
    a parsing gap - same category as AUEC's own TIER1/FAK order
    difference) - verified by content (code/por set), not position."""
    raw_path, opus_path = PAIRS[1]
    row_set = _run(raw_path)
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SUR")

    ignore = {"header_seq", "note_seq", "charge_seq", "application_effective", "application_expires"}

    def key(row):
        return tuple(_normalize(row.get(f)) for f in cols.CMDT_NOTE_ROW_FIELDS if f not in ignore)

    assert len(generated) == len(expected)
    assert {key(g) for g in generated} == {key(e) for e in expected}


def test_aubp_commodity_groups_are_separate_not_folded():
    """Confirmed via ground truth's own (cmdt_seq, code, description)
    tuple counts: G0001/G0002/G0003 are 3 distinct groups (matching AUWC's
    shape), not AUEC's "RAD folds into main" shape."""
    row_set = _run(PAIRS[1][0])
    descriptions = {r.commodity_group_description for r in row_set.rates}
    assert descriptions == {DEFAULT_MAIN_DESCRIPTION, DEFAULT_RF_DESCRIPTION, DEFAULT_RAD_DESCRIPTION}


def test_aubp_reefer_and_rad_40_maps_to_40hc_not_40():
    """For both Reefer and RAD, the raw sheet's own '40'' column feeds the
    OPUS 40HC slot, not 40' (which stays blank) - confirmed against every
    RF/RAD row in ground truth."""
    row_set = _run(PAIRS[1][0])
    for description in (DEFAULT_RF_DESCRIPTION, DEFAULT_RAD_DESCRIPTION):
        group_rows = [r for r in row_set.rates if r.commodity_group_description == description]
        assert group_rows
        for row in group_rows:
            assert row.rate_40 is None
            assert row.rate_40hc is not None


def test_aubp_route_seq_is_one_continuous_counter_per_group():
    row_set = _run(PAIRS[1][0])
    for description in (DEFAULT_MAIN_DESCRIPTION, DEFAULT_RF_DESCRIPTION, DEFAULT_RAD_DESCRIPTION):
        group_rows = [r for r in row_set.rates if r.commodity_group_description == description]
        assert [r.route_seq for r in group_rows] == list(range(1, len(group_rows) + 1))


def test_aubp_yangon_thilawa_merge():
    row_set = _run(PAIRS[1][0])
    merged = [r for r in row_set.rates if r.origin_code == "MMRGN;MMTLA"]
    assert merged
    for row in merged:
        assert row.route_note == "(FRT collect only)"
    assert not any(r.origin_code in ("MMRGN", "MMTLA") for r in row_set.rates if r is not merged)


def test_aubp_batam_term_override():
    row_set = _run(PAIRS[1][0])
    batam_rows = [r for r in row_set.rates if r.origin_code == "IDBTM"]
    assert batam_rows
    assert all(r.origin_term == "Door" for r in batam_rows)


def test_aubp_destination_codes_are_independently_sorted_semicolon_lists():
    row_set = _run(PAIRS[1][0])
    dry_rows = [r for r in row_set.rates if r.commodity_group_description == DEFAULT_MAIN_DESCRIPTION and r.cgo_type == "DR"]
    combined = [r for r in dry_rows if ";" in r.destination_code]
    assert combined
    for row in combined:
        assert row.destination_code == "AUADL;AUBNE;AUMEL"
        assert row.destination_description == "ADELAIDE, SA;BRISBANE, QLD;MELBOURNE, VIC"


def test_aubp_excluded_charge_codes_drops_scoped_rows_too():
    raw_path, _ = PAIRS[1]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUBPParser().run(wb, MappingProfile(excluded_charge_codes=["THL"]))
    codes = [n.code for n in row_set.cmdt_notes]
    assert "THL" not in codes
    assert "ISL" in codes  # untouched


def test_aubp_skip_dg_generation():
    raw_path, _ = PAIRS[1]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = AUBPParser().run(wb, MappingProfile(skip_dg_generation={DEFAULT_MAIN_DESCRIPTION: True}))
    main_cgo = {r.cgo_type for r in row_set.rates if r.commodity_group_description == DEFAULT_MAIN_DESCRIPTION}
    assert "DG" not in main_cgo
    assert "DR" in main_cgo
