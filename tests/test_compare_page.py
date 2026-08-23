from __future__ import annotations

import openpyxl

from mrg2opus.schema import opus_columns as cols
from mrg2opus.schema.opus_rows import OpusRowSet, RatesPortPortRow, RatesRow
from mrg2opus.ui.compare_page import _run_comparison


def _rates_row(**overrides) -> RatesRow:
    base = dict(
        commodity_group_code="G0001", commodity_group_description="FAK",
        origin_code="CNSHA", origin_description="Shanghai",
        destination_code="USLAX", destination_description="Los Angeles",
        prefix="D", cgo_type="DR", rate_20=1000,
    )
    base.update(overrides)
    return RatesRow(**base)


def _reference_workbook_with_rates_sheet(rows: list[list]) -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cols.SHEET_NAME_RATES
    for offset, row in enumerate(rows):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=3 + offset, column=col_idx, value=value)
    return wb


def test_run_comparison_matches_identical_rates_row():
    generated_row = _rates_row()
    row_sets = {"": OpusRowSet(rates=[generated_row])}

    ref_row = [None] * len(cols.RATES_ROW_FIELDS)
    for field_name, value in generated_row.model_dump().items():
        ref_row[cols.RATES_ROW_FIELDS.index(field_name)] = value
    ref_wb = _reference_workbook_with_rates_sheet([ref_row])

    results = _run_comparison(row_sets, ref_wb, "Both", "SAF")
    rates_result = next(r for r in results if r["sheet_type"] == "RATES")
    assert rates_result["found_in_reference"] is True
    assert rates_result["matched"] == 1
    assert rates_result["missing"] == []
    assert rates_result["extra"] == []


def test_run_comparison_reports_missing_sheet_as_all_extra():
    generated_row = _rates_row()
    row_sets = {"": OpusRowSet(rates=[generated_row])}
    ref_wb = openpyxl.Workbook()
    ref_wb.active.title = "Unrelated Sheet"

    results = _run_comparison(row_sets, ref_wb, "Grouped (RATES)", "SAF")
    rates_result = next(r for r in results if r["sheet_type"] == "RATES")
    assert rates_result["found_in_reference"] is False
    assert len(rates_result["extra"]) == 1


def test_run_comparison_respects_rates_mode_grouped_only():
    generated_row = _rates_row()
    row_set = OpusRowSet(rates=[generated_row], rates_port_port=[RatesPortPortRow(**generated_row.model_dump())])
    row_sets = {"": row_set}
    ref_wb = openpyxl.Workbook()
    ref_wb.active.title = "Unrelated Sheet"

    results = _run_comparison(row_sets, ref_wb, "Grouped (RATES)", "SAF")
    sheet_types = {r["sheet_type"] for r in results}
    assert sheet_types == {"RATES"}


def test_run_comparison_both_mode_includes_grouped_and_exploded():
    generated_row = _rates_row()
    row_set = OpusRowSet(rates=[generated_row], rates_port_port=[RatesPortPortRow(**generated_row.model_dump())])
    row_sets = {"": row_set}
    ref_wb = openpyxl.Workbook()
    ref_wb.active.title = "Unrelated Sheet"

    results = _run_comparison(row_sets, ref_wb, "Both", "SAF")
    sheet_types = {r["sheet_type"] for r in results}
    assert sheet_types == {"RATES", "RATES PORT-PORT"}
