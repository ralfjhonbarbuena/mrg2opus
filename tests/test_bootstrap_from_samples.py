"""The Location Bank miner - specifically the two ways it mined garbage
before, both caught by dry-running it against the real filings."""
from __future__ import annotations

import openpyxl
import pytest

from mrg2opus.location_bank.bootstrap_from_samples import _has_rates_layout, _scan_workbook


def _sheet(wb, title, header_row2, data_rows):
    ws = wb.create_sheet(title)
    ws.append([None] * 20)                      # row 1: group header
    ws.append(header_row2)                      # row 2: field header
    for r in data_rows:
        ws.append(r)
    return ws


RATES_HEADER = [None, None, "Code", "Description", "Code", "Description", None,
                "Code", "Description", "Term", "Transmode", "Code", " Code",
                "Code", "Description", "Term", "Transmode", "Prefix", "CGO TYPE", "Cur"]


def _row(origin_code, origin_desc, dest_code, dest_desc):
    r = [None] * 20
    r[7], r[8], r[13], r[14] = origin_code, origin_desc, dest_code, dest_desc
    return r


def test_mines_codes_and_names_from_a_rates_sheet(tmp_path):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "RATES", RATES_HEADER, [_row("CNSHA", "SHANGHAI", "BEANR", "ANTWERP")])
    p = tmp_path / "f.xlsx"; wb.save(p)

    clean, _ = _scan_workbook(p)
    assert clean == {"CNSHA": "SHANGHAI", "BEANR": "ANTWERP"}


def test_skips_a_shifted_sheet_that_only_looks_like_rates(tmp_path):
    """Real filings carry a vertical-rates sheet named "V RATES" whose
    columns sit one to the left. Selecting sheets by name mined its
    DESCRIPTIONS as codes - "ALEXANDRIA, EGYPT" went in as a port code."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    shifted = [None] * 20
    shifted[6], shifted[7], shifted[12], shifted[13] = "Code", "Description", "Code", "Description"
    ws = _sheet(wb, "V RATES", shifted, [])
    assert not _has_rates_layout(ws)

    data = [None] * 20
    data[6], data[7], data[12], data[13] = "LKCMB", "COLOMBO", "EGALY", "ALEXANDRIA, EGYPT"
    ws.append(data)
    p = tmp_path / "f.xlsx"; wb.save(p)

    clean, grouped = _scan_workbook(p)
    assert clean == {} and grouped == {}


def test_skips_names_carrying_the_comma_to_double_space_artifact(tmp_path):
    """Some exports replace every comma with two spaces. Mining those
    overwrote 16 correct names ("DALIAN, LIAONING" -> "DALIAN  LIAONING")."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "RATES", RATES_HEADER, [
        _row("CNDLC", "DALIAN  LIAONING", "BEANR", "ANTWERP"),
        _row("CNSHA", "SHANGHAI, SHANGHAI", "DEHAM", "HAMBURG"),
    ])
    p = tmp_path / "f.xlsx"; wb.save(p)

    clean, _ = _scan_workbook(p)
    assert "CNDLC" not in clean, "artifact name must not be mined"
    assert clean["CNSHA"] == "SHANGHAI, SHANGHAI", "a real comma'd name is kept"


def test_multi_port_rows_are_held_back_for_elimination(tmp_path):
    """Codes and names are each sorted independently, so position N of one
    doesn't pair with position N of the other - these can't be mined
    directly, only deduced once all but one member is known."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "RATES", RATES_HEADER, [_row("CNSHA;CNNGB", "NINGBO;SHANGHAI", "BEANR", "ANTWERP")])
    p = tmp_path / "f.xlsx"; wb.save(p)

    clean, grouped = _scan_workbook(p)
    assert "CNSHA" not in clean and "CNNGB" not in clean
    assert grouped == {"NINGBO; SHANGHAI": {"CNSHA", "CNNGB"}}
