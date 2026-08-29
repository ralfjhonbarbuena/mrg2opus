from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.nzj import DEFAULT_DESCRIPTION, NZJParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

# 2 weekly pairs, both real ground truth (tracker status "Completed"). The
# ground-truth workbooks' own sheet names drift between weeks (week 1:
# 'rates'/'SUR', week 2: 'RATES'/'SURCHARGE') - find_sheet()'s
# whitespace/case-insensitive matching covers 'rates' vs 'RATES', but
# 'SUR' vs 'SURCHARGE' is a genuinely different name, so each pair carries
# its own CMDT-note sheet name explicitly.
PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "45_NZJ NEA to NZ FAK" / "ONE NZ MRG 20260815 to 20260831 - ex NEA (07 August 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "45_NZJ NEA to NZ FAK" / "ONE NZ MRG 20260815 to 20260831 - ex NEA (07 August 2026)_opus.xlsx",
        "SUR",
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "46_NZJ NEA to NZ FAK" / "ONE NZ MRG 20260901 to 20260914 - ex NEA (20 Aug 2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "46_NZJ NEA to NZ FAK" / "ONE NZ MRG 20260901 to 20260914 - ex NEA (20 Aug 2026)_opus.xlsx",
        "SURCHARGE",
    ),
]

pytestmark = pytest.mark.skipif(
    any(not p.exists() for pair in PAIRS for p in (pair[0], pair[1])),
    reason="reference/ ground-truth files not present in this checkout",
)

# Known, deliberate deviations from ground truth (same categories already
# documented for every other lane, see RATES_IGNORE_FIELDS_BY_LANE in
# audit/compare.py):
#   - type/commodity_group_code/commodity_group_description/cmdt_seq/
#     commodity_note: OPUS's own user-customizable group identity and
#     running sequence, not reproducible from the raw MRG alone.
#   - route_seq: single running counter per commodity group - verified
#     correct in shape (see test_nzj_route_seq_is_one_continuous_counter
#     below), not pinned to the literal ground-truth numbers.
RATES_IGNORE_FIELDS = {
    "type", "commodity_group_code", "commodity_group_description", "cmdt_seq", "commodity_note", "route_seq",
}


def _run(raw_path: Path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    return NZJParser().run(wb, MappingProfile())


@pytest.mark.parametrize("raw_path,opus_path,cmdt_sheet", PAIRS)
def test_nzj_rates_matches_ground_truth(raw_path, opus_path, cmdt_sheet):
    row_set = _run(raw_path)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


def test_nzj_cmdt_note_matches_ground_truth_week1():
    """Week 1's ground truth confirms this parser's child-row order (EFS,
    OBS, ISL) and the ISL->Taiwan POL scope exactly. charge_seq is ignored:
    ground truth's own numbering (15-18) is an OPUS-assigned global running
    sequence across the whole filing, not derivable from the raw MRG -
    confirmed non-derivable by week 2 starting its own block at 1 instead
    (see week2 test below), a genuine cross-week inconsistency."""
    raw_path, opus_path, cmdt_sheet = PAIRS[0]
    row_set = _run(raw_path)
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, cmdt_sheet)

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq", "charge_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_nzj_cmdt_note_matches_ground_truth_week2_by_content():
    """Week 2 diverges from week 1 only in the parent row's own "inclusive
    of X and Y and Z" text word order (EFS/ISL/OBS vs. EFS/OBS/ISL) - a
    human filing inconsistency (see module docstring's "known, accepted
    gap"), not a parsing gap. Every other field, including the ISL POL
    scope, matches exactly; the charge-code SET mentioned in the text is
    verified separately, order-independent."""
    raw_path, opus_path, cmdt_sheet = PAIRS[1]
    row_set = _run(raw_path)
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, cmdt_sheet)

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq", "charge_seq", "contents"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"

    gen_codes = set(re.findall(r"\(([A-Z]+)\)", generated[0]["contents"] or ""))
    exp_codes = set(re.findall(r"\(([A-Z]+)\)", expected[0]["contents"] or ""))
    assert gen_codes == exp_codes == {"EFS", "OBS", "ISL"}


def test_nzj_route_seq_is_one_continuous_counter_per_group():
    """Confirmed against ground truth: Route Seq. is a single running
    counter across the whole commodity group, in generation order
    (Auckland-DR, combined-DR(LNT), RF, RAD, Auckland-DG, combined-DG(LNT))
    - not resetting per destination block or container type."""
    raw_path, _, _ = PAIRS[0]
    row_set = _run(raw_path)
    group_rows = [r for r in row_set.rates if r.commodity_group_description == DEFAULT_DESCRIPTION]
    assert [r.route_seq for r in group_rows] == list(range(1, len(group_rows) + 1))


def test_nzj_isl_scoped_to_taiwan():
    raw_path, _, _ = PAIRS[0]
    row_set = _run(raw_path)
    isl_rows = [n for n in row_set.cmdt_notes if n.code == "ISL"]
    assert len(isl_rows) == 1
    assert isl_rows[0].pol == "TW"
    other_rows = [n for n in row_set.cmdt_notes if n.code != "ISL"]
    assert all(n.pol is None for n in other_rows)


def test_nzj_every_origin_feeds_both_dry_destination_blocks():
    """Every origin's D/DR rate is filed twice: once under Auckland alone
    (NZAKL), once under the other 3 ports combined
    (NZLYT;NZNPE;NZTRG) - confirmed via both reference weeks' identical
    49-origin-row counts on each side (some rows share one origin_code,
    e.g. Keelung's two "direct loading"/"via Kaohsiung" rows both resolve
    to TWKEL, so unique codes alone undercounts - compare row counts)."""
    raw_path, _, _ = PAIRS[0]
    row_set = _run(raw_path)
    dr_rows = [r for r in row_set.rates if r.cgo_type == "DR"]
    akl_rows = [r for r in dr_rows if r.destination_code == "NZAKL"]
    lnt_rows = [r for r in dr_rows if r.destination_code == "NZLYT;NZNPE;NZTRG"]
    assert [r.origin_code for r in akl_rows] == [r.origin_code for r in lnt_rows]
    assert len(akl_rows) == 49


def test_nzj_rad_and_reefer_never_get_dg_duplicate():
    raw_path, _, _ = PAIRS[0]
    row_set = _run(raw_path)
    combined_dest = "NZAKL;NZLYT;NZNPE;NZTRG"
    assert not [r for r in row_set.rates if r.destination_code == combined_dest and r.cgo_type == "DG"]
    assert {r.cgo_type for r in row_set.rates if r.destination_code == combined_dest} == {"RF", "DR"}


def test_nzj_plain_via_clause_without_parens_resolves_o_via():
    """'Keelung \\n    via Kaohsiung' - a plain trailing via-clause with no
    parens and no 'by rail' - sets O.Via to Kaohsiung's code but leaves
    origin_transmode blank, same convention as the parenthetical
    '(via CNYTN by rail)' form used elsewhere in this same raw sheet."""
    raw_path, _, _ = PAIRS[0]
    row_set = _run(raw_path)
    via_rows = [r for r in row_set.rates if r.origin_code == "TWKEL" and r.o_via_code is not None]
    assert via_rows
    assert all(r.o_via_code == "TWKHH" for r in via_rows)
    assert all(r.origin_transmode is None for r in via_rows)


def test_nzj_rail_via_clause_sets_transmode():
    raw_path, _, _ = PAIRS[0]
    row_set = _run(raw_path)
    rail_rows = [r for r in row_set.rates if r.origin_code == "CNCTU" and r.o_via_code == "CNYTN"]
    assert rail_rows
    assert all(r.origin_transmode == "Rail" for r in rail_rows)
    assert all(r.origin_term == "CY" for r in rail_rows)


def test_nzj_excluded_charge_codes_drops_isl():
    raw_path, _, _ = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = NZJParser().run(wb, MappingProfile(excluded_charge_codes=["ISL"]))
    codes = [n.code for n in row_set.cmdt_notes]
    assert "ISL" not in codes
    assert "EFS" in codes
    assert "OBS" in codes


def test_nzj_skip_dg_generation():
    raw_path, _, _ = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = NZJParser().run(wb, MappingProfile(skip_dg_generation={DEFAULT_DESCRIPTION: True}))
    assert "DG" not in {r.cgo_type for r in row_set.rates}
    assert "DR" in {r.cgo_type for r in row_set.rates}
