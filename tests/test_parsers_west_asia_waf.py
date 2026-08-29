from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_cmdt_note_sheet, read_rates_sheet
from mrg2opus.parsers.west_asia_waf import WestAsiaWAFParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

PAIRS = [
    (
        REFERENCE_DIR / "1_MRGs" / "11_West Asia to West Africa" / "WEST Asia WAF MRG Rate (AIM)  (1 - 14 Aug  2026).xlsx",
        REFERENCE_DIR / "2_OPUS" / "11_West Asia to West Africa" / "West Asia to West Africa ( 20260801 - 20260814 ).xlsx",
        date(2026, 4, 29), date(2026, 12, 31),
    ),
    (
        REFERENCE_DIR / "1_MRGs" / "12_West Asia to West Africa" / "West Asia - WAF_SAF_EAF_MZ  MRG (15 - 31 Aug 2026) (1).xlsx",
        REFERENCE_DIR / "2_OPUS" / "12_West Asia to West Africa" / "West Asia to West Africa ( 20260815 - 20260831 ).xlsx",
        date(2026, 5, 15), date(2026, 12, 31),
    ),
]

pytestmark = pytest.mark.skipif(
    any(not p.exists() for pair in PAIRS for p in (pair[0], pair[1])),
    reason="reference/ ground-truth files not present in this checkout",
)

# type/commodity_group_code/commodity_group_description/cmdt_seq/
# commodity_note/route_seq: same accepted-gap categories as every other
# lane (see RATES_IGNORE_FIELDS_BY_LANE in audit/compare.py) - OPUS's own
# running sequence and user-customizable group identity aren't derivable
# from the raw MRG alone.
RATES_IGNORE_FIELDS = {"type", "commodity_group_code", "commodity_group_description", "cmdt_seq", "commodity_note", "route_seq"}


def _run(raw_path: Path, rfa_eff: date, rfa_exp: date):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    return WestAsiaWAFParser().run(wb, MappingProfile(rfa_effective_date=rfa_eff, rfa_expiry_date=rfa_exp))


@pytest.mark.parametrize("raw_path,opus_path,rfa_eff,rfa_exp", PAIRS)
def test_west_asia_waf_rates_matches_ground_truth(raw_path, opus_path, rfa_eff, rfa_exp):
    row_set = _run(raw_path, rfa_eff, rfa_exp)
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, "RATES")

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


def test_west_asia_waf_cmdt_note_matches_ground_truth_week1():
    """Week 1's ground truth confirms this parser's fully-alphabetized
    child-row order exactly. Week 2 diverges (see module docstring's
    "known, accepted gap") and is checked separately, by content only."""
    raw_path, opus_path, rfa_eff, rfa_exp = PAIRS[0]
    row_set = _run(raw_path, rfa_eff, rfa_exp)
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SRCHG")

    assert len(generated) == len(expected)
    ignore = {"header_seq", "note_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_west_asia_waf_cmdt_note_matches_ground_truth_week2_by_content():
    raw_path, opus_path, rfa_eff, rfa_exp = PAIRS[1]
    row_set = _run(raw_path, rfa_eff, rfa_exp)
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(opus_path, data_only=True, read_only=True)
    expected = read_cmdt_note_sheet(ref_wb, "SRCHG")

    ignore = {"header_seq", "note_seq", "charge_seq"}

    def key(row):
        return tuple(_normalize(row.get(f)) for f in cols.CMDT_NOTE_ROW_FIELDS if f not in ignore)

    assert len(generated) == len(expected)
    assert {key(g) for g in generated} == {key(e) for e in expected}


def test_west_asia_waf_thl_origin_scoping():
    """THL is confirmed origin-scoped (LKCMB/PKKHI/BDCGP) via 3 extra
    child rows, on top of appearing in the main "inclusive of" sentence -
    not a separate blank-origin child row of its own."""
    raw_path, _, rfa_eff, rfa_exp = PAIRS[0]
    row_set = _run(raw_path, rfa_eff, rfa_exp)
    thl_rows = [n for n in row_set.cmdt_notes if n.code == "THL"]
    assert {n.pol for n in thl_rows} == {"LKCMB", "PKKHI", "BDCGP"}
    assert "TERMINAL HANDLING CHARGE" in row_set.cmdt_notes[0].contents


def test_west_asia_waf_dg_duplicate_at_same_rate():
    raw_path, _, rfa_eff, rfa_exp = PAIRS[0]
    row_set = _run(raw_path, rfa_eff, rfa_exp)
    dr_by_key = {(r.origin_code, r.destination_code): r for r in row_set.rates if r.cgo_type == "DR"}
    dg_by_key = {(r.origin_code, r.destination_code): r for r in row_set.rates if r.cgo_type == "DG"}
    assert dr_by_key.keys() == dg_by_key.keys()
    for key, dr_row in dr_by_key.items():
        dg_row = dg_by_key[key]
        assert dr_row.rate_20 == dg_row.rate_20
        assert dr_row.rate_40 == dg_row.rate_40
        assert dr_row.rate_40hc == dg_row.rate_40hc


def test_west_asia_waf_skip_dg_generation():
    raw_path, _, rfa_eff, rfa_exp = PAIRS[0]
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    row_set = WestAsiaWAFParser().run(
        wb, MappingProfile(rfa_effective_date=rfa_eff, rfa_expiry_date=rfa_exp, skip_dg_generation={"West Asia West Africa MRG": True})
    )
    assert "DG" not in {r.cgo_type for r in row_set.rates}
    assert "DR" in {r.cgo_type for r in row_set.rates}
