from __future__ import annotations

import functools
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_arbs_sheet, read_rates_sheet
from mrg2opus.excel_io.merge import merge_workbooks
from mrg2opus.parsers.tad_aew_amw import TADAewAmwParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"
RAW_DIR = REFERENCE_DIR / "1_MRGs" / "23_TAD FILING AEW AMW"
OPUS_DIR = REFERENCE_DIR / "2_OPUS" / "23_TAD FILING AEW AMW"

RAW_PATH_WEEK1 = RAW_DIR / "Sep MRG dated 20th Aug (AEWAMW).xlsx"
RAW_PATH_WEEK2 = RAW_DIR / "Sep MRG dated 27th Aug (AEWAMW).xlsx"
RAW_PATH_JP = RAW_DIR / "AE WB Sep MRG Dated 20 Aug (AEWAMW ex.JP).xlsx"
RAW_PATHS = [RAW_PATH_WEEK1, RAW_PATH_WEEK2, RAW_PATH_JP]
GROUND_TRUTH = {
    "AEW": OPUS_DIR / "AEW POLLY.xlsx",
    "AMW": OPUS_DIR / "AMW POLLY.xlsx",
    "JAPAN": OPUS_DIR / "JAPAN POLLY.xlsx",
}

pytestmark = pytest.mark.skipif(
    any(not p.exists() for p in RAW_PATHS) or any(not p.exists() for p in GROUND_TRUTH.values()),
    reason="reference/ ground-truth files not present in this checkout",
)

RATES_IGNORE_FIELDS = {"type", "cmdt_seq", "commodity_note"}

# Confirmed against ground truth: this filing round used a rate correction
# issued mid-period (see parsers/common/tad_snapshots.py) AND a genuinely
# different per-scope RFA effective date - AEW's own SRCHG children all
# use 2026-08-30, AMW's own CMDT NOTE children all use 2026-08-19, both
# expiring 2026-12-31. MappingProfile.rfa_effective_date is one filing-wide
# setting, so AEW and AMW are run with their own separate profile.
RFA_EXPIRY = date(2026, 12, 31)
AEW_RFA_EFFECTIVE = date(2026, 8, 30)
AMW_RFA_EFFECTIVE = date(2026, 8, 19)

# The Japan scopes drift the same way again: Japan-AEW's own SRCHG children
# use the plain rate-validity window (2026-09-01 to -30, i.e. no RFA
# override at all) while Japan-AMW's use 2026-08-30 to 2026-12-31. The DG
# duplicate toggle also differs - ON for the main scopes, OFF for Japan -
# and it too is filing-wide, so each Japan scope gets its own profile.
JP_VALIDITY_START = date(2026, 9, 1)
JP_VALIDITY_END = date(2026, 9, 30)
JP_AEW_RFA = (JP_VALIDITY_START, JP_VALIDITY_END)
JP_AMW_RFA = (date(2026, 8, 30), RFA_EXPIRY)


def _normalize_export_punctuation_rows(rows: list[dict]) -> list[dict]:
    """Same comma->space export artifact already confirmed for TAD-WMW-WEW
    and LAWC TIER 1 - normalize the ground-truth side (and generated,
    idempotently) so the comparison isn't tripped up by it."""
    out = []
    for r in rows:
        r = dict(r)
        for field_name in ("origin_description", "destination_description"):
            value = r.get(field_name)
            if isinstance(value, str):
                r[field_name] = value.replace(",", " ")
        out.append(r)
    return out


def _read_shifted_sheet(wb, sheet_name: str, fields: list[str]) -> list[dict]:
    """AEW's own SRCHG sheet AND AMW's own CMDT NOTE/ROUTE NOTE sheets all
    carry the same stray-value-in-column-A shift already confirmed for
    TAD-WMW-WEW's ground truth - see that test module's identical helper."""
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[1 : 1 + len(fields)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(fields, values)))
    return rows


def _cmdt_groups(notes: list[dict]) -> dict[frozenset, dict]:
    """Same rationale as TAD-WMW-WEW's own helper: AEW has 3 CMDT NOTE
    blocks (one per validity/charge-code combo from the 2-file merge, plus
    a 3rd for a code-set variant), AMW has its own set - ground truth's own
    header_seq numbering order doesn't match this parser's build order, so
    match by the block's own code set instead of position."""
    groups: dict[frozenset, dict] = {}
    header_seq = None
    children: list[dict] = []
    parent: dict | None = None
    for n in notes:
        if n.get("code") == "APP":
            if parent is not None:
                groups[frozenset(c["code"] for c in children)] = {
                    "parent": parent, "children": children, "header_seq": header_seq,
                }
            parent = n
            header_seq = n.get("header_seq")
            children = []
        else:
            children.append(n)
    if parent is not None:
        groups[frozenset(c["code"] for c in children)] = {
            "parent": parent, "children": children, "header_seq": header_seq,
        }
    return groups


def _assert_cmdt_groups_match(generated_groups: dict, expected_groups: dict) -> None:
    assert set(generated_groups) == set(expected_groups), (
        f"code-set groups differ: only-generated={set(generated_groups) - set(expected_groups)}, "
        f"only-expected={set(expected_groups) - set(generated_groups)}"
    )
    ignore = {"header_seq", "note_seq"}
    for codes, g_group in generated_groups.items():
        e_group = expected_groups[codes]
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g_group["parent"].get(field_name)), _normalize(e_group["parent"].get(field_name))
            assert gv == ev, f"group {sorted(codes)} parent {field_name}: {gv!r} != {ev!r}"
        assert len(g_group["children"]) == len(e_group["children"])
        for gc, ec in zip(g_group["children"], e_group["children"]):
            for field_name in cols.CMDT_NOTE_ROW_FIELDS:
                if field_name in ignore:
                    continue
                gv, ev = _normalize(gc.get(field_name)), _normalize(ec.get(field_name))
                assert gv == ev, f"group {sorted(codes)} child {gc.get('code')} {field_name}: {gv!r} != {ev!r}"


def _route_notes_by_rates_identity(route_notes: list[dict], rates: list[dict]) -> dict[tuple, dict]:
    """Same join-back-to-RATES-identity approach as TAD-WMW-WEW's test -
    route note header_seq/route_seq numbering isn't derivable, but the
    RATES row it decorates (by origin/dest/cargo) is."""
    by_seq = {(_normalize(r.get("cmdt_seq")), _normalize(r.get("route_seq"))): rates_row_key(r) for r in rates}
    out: dict[tuple, dict] = {}
    for n in route_notes:
        key = by_seq.get((_normalize(n.get("header_seq")), _normalize(n.get("route_seq"))))
        if key is not None:
            out[key] = n
    return out


@functools.lru_cache(maxsize=1)
def _load_merged_workbook():
    """All three raw files are wide (100 columns x ~900 rows x 3 sheets
    each) - merge_workbooks() touches every cell in that range, so
    reloading and re-merging per test made this module alone take 8+
    minutes. Cached once; every test here only reads the result, never
    mutates it."""
    wbs = [openpyxl.load_workbook(p, data_only=True) for p in RAW_PATHS]
    return merge_workbooks(wbs, names=[p.name for p in RAW_PATHS])


@functools.lru_cache(maxsize=None)
def _run_tad(rfa_effective: date) -> dict:
    """The main AEW/AMW scopes' settings: both toggles on, matching this
    filing round's own real settings."""
    profile = MappingProfile(
        rfa_effective_date=rfa_effective, rfa_expiry_date=RFA_EXPIRY,
        include_tad_d7=True, generate_tad_dg_duplicate=True,
    )
    return TADAewAmwParser().run_multi(_load_merged_workbook(), profile)


@functools.lru_cache(maxsize=None)
def _run_tad_jp(rfa: tuple[date, date]) -> dict:
    """The Japan scopes' settings: DG duplication off (and D7 never
    applies there at all) - see the module docstring's inconsistency note."""
    profile = MappingProfile(rfa_effective_date=rfa[0], rfa_expiry_date=rfa[1])
    return TADAewAmwParser().run_multi(_load_merged_workbook(), profile)


@pytest.mark.parametrize("scope,rfa_effective", [("AEW", AEW_RFA_EFFECTIVE), ("AMW", AMW_RFA_EFFECTIVE)])
def test_tad_aew_amw_rates_matches_ground_truth(scope, rfa_effective):
    row_set = _run_tad(rfa_effective)[scope]
    generated = _normalize_export_punctuation_rows([r.model_dump() for r in row_set.rates])

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = _normalize_export_punctuation_rows(read_rates_sheet(ref_wb, "RATES"))

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


def test_tad_aew_amw_route_seq_is_one_continuous_counter_per_group():
    row_set = _run_tad(AEW_RFA_EFFECTIVE)["AEW"]
    by_cmdt: dict[int, list[int]] = {}
    for r in row_set.rates:
        by_cmdt.setdefault(r.cmdt_seq, []).append(r.route_seq)
    for cmdt_seq, seqs in by_cmdt.items():
        assert seqs == list(range(1, len(seqs) + 1)), f"cmdt_seq {cmdt_seq}: {seqs}"


def test_tad_aew_amw_cmdt_note_matches_ground_truth_aew():
    """AEW's own CMDT-NOTE-equivalent sheet is literally named "SRCHG"."""
    row_set = _run_tad(AEW_RFA_EFFECTIVE)["AEW"]
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["AEW"], data_only=True, read_only=True)
    expected = _read_shifted_sheet(ref_wb, "SRCHG", cols.CMDT_NOTE_ROW_FIELDS)

    _assert_cmdt_groups_match(_cmdt_groups(generated), _cmdt_groups(expected))


def test_tad_aew_amw_cmdt_note_matches_ground_truth_amw():
    row_set = _run_tad(AMW_RFA_EFFECTIVE)["AMW"]
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["AMW"], data_only=True, read_only=True)
    expected = _read_shifted_sheet(ref_wb, "CMDT NOTE", cols.CMDT_NOTE_ROW_FIELDS)

    _assert_cmdt_groups_match(_cmdt_groups(generated), _cmdt_groups(expected))


def test_tad_aew_amw_no_route_notes_for_aew():
    """AEW's own ground truth has zero T/S Port, Service Lane, or
    special-node rows this period - confirmed no ROUTE NOTE sheet at all."""
    row_set = _run_tad(AEW_RFA_EFFECTIVE)["AEW"]
    assert row_set.route_notes == []


def test_tad_aew_amw_route_note_matches_ground_truth_amw():
    """AMW covers all 3 real triggers: 2 special-node commodity-group-name
    tags (Alexandria x2 reused from TAD-OEW-OMW's own table), Haydarpasa,
    Marport, and a populated Service Lane (FE3)."""
    row_set = _run_tad(AMW_RFA_EFFECTIVE)["AMW"]
    generated_rates = [r.model_dump() for r in row_set.rates]
    generated = _route_notes_by_rates_identity([n.model_dump() for n in row_set.route_notes], generated_rates)

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["AMW"], data_only=True, read_only=True)
    expected_rates = read_rates_sheet(ref_wb, "RATES")
    expected_notes = _read_shifted_sheet(ref_wb, "ROUTE NOTE", cols.RN_ROW_FIELDS)
    expected = _route_notes_by_rates_identity(expected_notes, expected_rates)

    assert set(generated) == set(expected), (
        f"only-generated={set(generated) - set(expected)}, only-expected={set(expected) - set(generated)}"
    )
    for key, g in generated.items():
        e = expected[key]
        for field_name in ("contents", "charge_seq", "code", "application_effective", "application_expires", "application"):
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"{key} {field_name}: {gv!r} != {ev!r}"


def test_tad_aew_amw_d7_addon_is_opt_in():
    profile = MappingProfile(rfa_effective_date=AEW_RFA_EFFECTIVE, rfa_expiry_date=RFA_EXPIRY)
    row_set = TADAewAmwParser().run_multi(_load_merged_workbook(), profile)["AEW"]
    assert all(r.rate_45 is None for r in row_set.rates)


def test_tad_aew_amw_dg_duplicate_is_opt_in():
    profile = MappingProfile(rfa_effective_date=AEW_RFA_EFFECTIVE, rfa_expiry_date=RFA_EXPIRY)
    row_set = TADAewAmwParser().run_multi(_load_merged_workbook(), profile)["AEW"]
    assert all(r.cgo_type != "DG" for r in row_set.rates)


def test_tad_aew_amw_multi_snapshot_merge_truncates_earlier_validity():
    """The 20th-Aug file's own declared validity (2026-09-01 to -15) must
    be truncated to end the day before the 27th-Aug file's own validity
    starts (2026-09-07), not used as filed - confirmed against ground
    truth's own 2 CMDT groups per route."""
    row_set = _run_tad(AEW_RFA_EFFECTIVE)["AEW"]
    windows = {(n.application_effective, n.application_expires) for n in row_set.cmdt_notes if n.code == "APP"}
    assert (date(2026, 9, 1), date(2026, 9, 6)) in windows
    assert (date(2026, 9, 7), date(2026, 9, 15)) in windows


def _arbs_key(row: dict) -> tuple:
    return (row.get("point"), row.get("over"), row.get("via"), row.get("per"))


@pytest.mark.parametrize("scope,sheet", [("AEW", "AEW ARBS"), ("AMW", "ORIGIN ARBS")])
def test_tad_aew_amw_arbs_matches_ground_truth(scope, sheet):
    """AEW's own ARBS sheet is scope-prefixed ("AEW ARBS"), AMW's is bare
    ("ORIGIN ARBS", no scope tag at all) - another real naming drift on
    top of the CMDT NOTE one (see SCOPED_SHEET_NAME_OVERRIDES)."""
    row_set = _run_tad(AEW_RFA_EFFECTIVE if scope == "AEW" else AMW_RFA_EFFECTIVE)[scope]
    generated = [r.model_dump() for r in row_set.arbs]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = read_arbs_sheet(ref_wb, sheet)

    gk = {_arbs_key(r): r for r in generated}
    ek = {_arbs_key(r): r for r in expected}
    assert set(gk) == set(ek), f"missing={set(ek) - set(gk)}, extra={set(gk) - set(ek)}"
    for key, g in gk.items():
        e = ek[key]
        for field_name in ("description", "trans_mode", "term", "cur", "proposal", "eff_date", "exp_date", "cgo_type"):
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"{key} {field_name}: {gv!r} != {ev!r}"


def test_tad_aew_amw_arbs_drops_self_transshipment_rows():
    """Confirmed against ground truth: a raw row whose "Over" equals its
    own origin Point (CNXMN transshipping via CNXMN) never survives into
    the output - not even as a same-code no-op."""
    row_set = _run_tad(AEW_RFA_EFFECTIVE)["AEW"]
    assert not any(a.point == a.over for a in row_set.arbs)


def test_tad_aew_amw_arbs_keeps_zero_rates():
    """A literal $0 ARBS rate is real data (confirmed: VNBHA's own raw
    row), not a "blank means no value" case - must not be dropped."""
    row_set = _run_tad(AEW_RFA_EFFECTIVE)["AEW"]
    assert any(a.point == "VNBHA" and a.proposal == 0 for a in row_set.arbs)


# --- Japan-origin scope -----------------------------------------------
#
# JAPAN POLLY.xlsx is the team's VBA tool caught mid-pipeline, not a
# finished OPUS filing (see the parser's module docstring): Descriptions
# blank, Commodity Note still the raw charge-code CSV, CMDT/Route Seq
# blank. This parser deliberately produces the COMPLETE output instead, so
# only the columns ground truth actually derived are compared.
JP_RATES_IGNORE_FIELDS = RATES_IGNORE_FIELDS | {"route_seq", "origin_description", "destination_description"}


@pytest.mark.parametrize(
    "scope,sheet,rfa",
    [("JAPAN AEW", "AEW RATES", JP_AEW_RFA), ("JAPAN AMW", "AMW RATES", JP_AMW_RFA)],
)
def test_tad_aew_amw_japan_rates_matches_ground_truth(scope, sheet, rfa):
    row_set = _run_tad_jp(rfa)[scope]
    generated = [r.model_dump() for r in row_set.rates]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["JAPAN"], data_only=True, read_only=True)
    expected = read_rates_sheet(ref_wb, sheet)

    result = diff_by_key(
        generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=JP_RATES_IGNORE_FIELDS
    )
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize(
    "scope,sheet,rfa",
    [("JAPAN AEW", "AEW SRCHG", JP_AEW_RFA), ("JAPAN AMW", "AMW SRCHG", JP_AMW_RFA)],
)
def test_tad_aew_amw_japan_cmdt_note_matches_ground_truth(scope, sheet, rfa):
    """One block each, 4 child codes (CAF/CSS/EFS/MBS - notably shorter
    than the main scopes' list, straight off the raw rows' own column)."""
    row_set = _run_tad_jp(rfa)[scope]
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["JAPAN"], data_only=True, read_only=True)
    expected = _read_shifted_sheet(ref_wb, sheet, cols.CMDT_NOTE_ROW_FIELDS)

    assert len(generated) == len(expected) == 5
    ignore = {"header_seq", "note_seq"}
    for i, (g, e) in enumerate(zip(generated, expected)):
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            # This one sheet's Contents carries literal "_x000D_" escapes
            # from however it was written; the text is otherwise identical.
            if isinstance(gv, str) and isinstance(ev, str):
                ev = ev.replace("_x000D_", "")
            assert gv == ev, f"row {i} {field_name}: {gv!r} != {ev!r}"


def test_tad_aew_amw_japan_commodity_group_differs_per_subscope():
    """Both Japan sub-scopes share the description "FAK - JAPAN" but carry
    genuinely different default codes - confirmed, not a typo."""
    aew = _run_tad_jp(JP_AEW_RFA)["JAPAN AEW"]
    amw = _run_tad_jp(JP_AMW_RFA)["JAPAN AMW"]
    assert {r.commodity_group_code for r in aew.rates} == {"G0011"}
    assert {r.commodity_group_code for r in amw.rates} == {"G0001"}
    assert {r.commodity_group_description for r in aew.rates + amw.rates} == {"FAK - JAPAN"}


def test_tad_aew_amw_japan_never_generates_d7():
    """The OFT 45 add-on is AEW/AMW-only - even with the toggle explicitly
    on, the Japan scopes must not produce a rate_45."""
    profile = MappingProfile(
        rfa_effective_date=JP_VALIDITY_START, rfa_expiry_date=JP_VALIDITY_END, include_tad_d7=True
    )
    row_sets = TADAewAmwParser().run_multi(_load_merged_workbook(), profile)
    for scope in ("JAPAN AEW", "JAPAN AMW"):
        assert all(r.rate_45 is None for r in row_sets[scope].rates), scope


def test_tad_aew_amw_japan_aew_has_no_route_notes():
    """No T/S Port, Service Lane, or special-node tag on any AEW-Japan raw
    row - confirmed against JAPAN POLLY's own ROUTE sheet, which carries
    AMW rows only."""
    assert _run_tad_jp(JP_AEW_RFA)["JAPAN AEW"].route_notes == []


def test_tad_aew_amw_japan_amw_route_notes_match_ground_truth():
    """JAPAN POLLY's RATES sheet leaves CMDT/Route Seq blank, so its ROUTE
    rows can't be joined back to a RATES identity the way the main scopes'
    can - compared as a content multiset instead, the strongest check the
    ground truth actually supports."""
    generated = [n.model_dump() for n in _run_tad_jp(JP_AMW_RFA)["JAPAN AMW"].route_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["JAPAN"], data_only=True, read_only=True)
    expected = _read_shifted_sheet(ref_wb, "ROUTE", cols.RN_ROW_FIELDS)

    def summary(rows):
        return Counter(
            (
                r.get("contents"),
                r.get("code"),
                _normalize(r.get("charge_seq")),
                r.get("application"),
                _normalize(r.get("application_effective")),
                _normalize(r.get("application_expires")),
            )
            for r in rows
        )

    assert summary(generated) == summary(expected)


@pytest.mark.parametrize(
    "scope,sheet,rfa",
    [("JAPAN AEW", "ORIGIN ARBS AEW", JP_AEW_RFA), ("JAPAN AMW", "ORIGIN ARBS AMW", JP_AMW_RFA)],
)
def test_tad_aew_amw_japan_arbs_matches_ground_truth(scope, sheet, rfa):
    """Compared as a multiset: unlike the main sheets, a Japan ARBS route
    can carry both a Dry General and a Dry Dangerous raw row, so
    (point, over, via, per) alone isn't a unique key - cgo_type is what
    separates them."""
    generated = [r.model_dump() for r in _run_tad_jp(rfa)[scope].arbs]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH["JAPAN"], data_only=True, read_only=True)
    expected = read_arbs_sheet(ref_wb, sheet)

    def summary(rows):
        return Counter(
            (
                r.get("point"), r.get("over"), r.get("via"), r.get("per"), r.get("cgo_type"),
                r.get("description"), r.get("trans_mode"), r.get("term"), r.get("cur"),
                _normalize(r.get("proposal")), _normalize(r.get("eff_date")), _normalize(r.get("exp_date")),
            )
            for r in rows
        )

    assert summary(generated) == summary(expected)


def test_tad_aew_amw_japan_arbs_stamps_cargo_type_on_dry_rows():
    """The two ARBS sheets in this same filing use different CGO Type
    conventions (see _ARBS_CGO_TYPE_MAIN vs _ARBS_CGO_TYPE_JP): the Japan
    one stamps DR/DG on dry rows where the main one leaves them blank."""
    jp_arbs = _run_tad_jp(JP_AEW_RFA)["JAPAN AEW"].arbs
    main_arbs = _run_tad(AEW_RFA_EFFECTIVE)["AEW"].arbs
    assert {a.cgo_type for a in jp_arbs if a.per.startswith("D")} == {"DR", "DG"}
    assert {a.cgo_type for a in main_arbs if a.per.startswith("D")} == {None}


def test_tad_aew_amw_japan_arbs_duplicates_the_copies_in_the_main_files():
    """AEW/AMW POLLY.xlsx each bundle a copy of the Japan ARBS ("AEW ARBS
    JP" / "ORIGIN ARBS JP") that is byte-identical to JAPAN POLLY's own -
    this parser emits it once, under the Japan scopes, rather than twice."""
    ref_jp = openpyxl.load_workbook(GROUND_TRUTH["JAPAN"], data_only=True, read_only=True)
    ref_aew = openpyxl.load_workbook(GROUND_TRUTH["AEW"], data_only=True, read_only=True)
    assert read_arbs_sheet(ref_jp, "ORIGIN ARBS AEW") == read_arbs_sheet(ref_aew, "AEW ARBS JP")
