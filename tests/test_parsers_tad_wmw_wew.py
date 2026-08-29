from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from mrg2opus.audit.compare import _normalize, diff_by_key, rates_row_key, read_rates_sheet
from mrg2opus.parsers.tad_wmw_wew import TADWmwWewParser
from mrg2opus.presets.models import MappingProfile
from mrg2opus.schema import opus_columns as cols

REFERENCE_DIR = Path(__file__).resolve().parents[1] / "reference"

RAW_PATH = (
    REFERENCE_DIR / "1_MRGs" / "27_TAD FILING WMW WEW"
    / "WEW-AET WB 1H September MRG 2026 as of 19th August 2026 (WEW and WMW ).xlsx"
)
GROUND_TRUTH = {
    "WEW": REFERENCE_DIR / "2_OPUS" / "27_TAD FILING WMW WEW" / "WEW POLLY.xlsx",
    "WMW": REFERENCE_DIR / "2_OPUS" / "27_TAD FILING WMW WEW" / "WMW POLLY.xlsx",
}

pytestmark = pytest.mark.skipif(
    not RAW_PATH.exists() or any(not p.exists() for p in GROUND_TRUTH.values()),
    reason="reference/ ground-truth files not present in this checkout",
)

RATES_IGNORE_FIELDS = {"type", "cmdt_seq", "commodity_note"}

RFA_EFFECTIVE = date(2026, 8, 19)
RFA_EXPIRY = date(2026, 12, 31)


def _normalize_export_punctuation_rows(rows: list[dict]) -> list[dict]:
    """WEW POLLY.xlsx's own RATES export has a comma->space artifact in
    destination_description (e.g. raw "GRANGEMOUTH, FALKIRK" -> ground
    truth "GRANGEMOUTH  FALKIRK", the comma itself replaced by a space
    while the pre-existing space after it survives) - same category of
    downstream export-tool quirk as LAWC TIER 1's own comma artifact (see
    test_parsers_lawc.py), confirmed across 203 WEW rows and absent from
    WMW's own ground truth entirely. Normalize the ground-truth side only;
    this parser's raw-text-copied descriptions are already correct."""
    out = []
    for r in rows:
        r = dict(r)
        for field_name in ("origin_description", "destination_description"):
            value = r.get(field_name)
            if isinstance(value, str):
                r[field_name] = value.replace(",", " ")
        out.append(r)
    return out


def _read_shifted_sheet(wb, sheet_name: str, fields: list[str]) -> list[dict]:
    """Unlike TAD-OEW-OMW's ground truth, WMW/WEW's own CMDT NOTE *and*
    ROUTE NOTE sheets both carry a stray value in column A that shifts
    every real column one to the right (confirmed against both WEW
    POLLY.xlsx and WMW POLLY.xlsx - OEW/OMW's CMDT NOTE sheet has no such
    shift, so this quirk is per-filing, not a general TAD convention; our
    own writer's un-shifted layout is still the correct one to produce)."""
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[1 : 1 + len(fields)]]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(fields, values)))
    return rows


def _cmdt_groups(notes: list[dict]) -> dict[frozenset, dict]:
    """Group a flat CMDT NOTE row list into {frozenset(child codes): {parent,
    children, header_seq}} blocks (a new block starts at each 'APP' row).
    Keying by the child-code set (rather than raw position/header_seq)
    sidesteps the ordering problem below: WMW/WEW has multiple CMDT NOTE
    groups per scope (unlike OEW/OMW's single group each), and ground
    truth's own header_seq numbering (a leftover VBA running counter, not
    derivable from the raw MRG - see TAD-OEW-OMW's own comment) assigns
    them in a DIFFERENT order than this parser's first-seen-in-raw-data
    order (confirmed: WEW's 3 groups have identical code sets both sides,
    just numbered 219/220/221 in ground truth vs 1/2/3 here in reverse)."""
    groups: dict[frozenset, dict] = {}
    header_seq = None
    children: list[dict] = []
    parent: dict | None = None
    for n in notes:
        if n.get("code") == "APP":
            if parent is not None:
                groups[frozenset(c["code"] for c in children)] = {
                    "parent": parent, "children": children, "header_seq": header_seq,
                }
            parent = n
            header_seq = n.get("header_seq")
            children = []
        else:
            children.append(n)
    if parent is not None:
        groups[frozenset(c["code"] for c in children)] = {
            "parent": parent, "children": children, "header_seq": header_seq,
        }
    return groups


def _assert_cmdt_groups_match(generated_groups: dict, expected_groups: dict) -> None:
    assert set(generated_groups) == set(expected_groups), (
        f"code-set groups differ: only-generated={set(generated_groups) - set(expected_groups)}, "
        f"only-expected={set(expected_groups) - set(generated_groups)}"
    )
    ignore = {"header_seq", "note_seq"}
    for codes, g_group in generated_groups.items():
        e_group = expected_groups[codes]
        for field_name in cols.CMDT_NOTE_ROW_FIELDS:
            if field_name in ignore:
                continue
            gv, ev = _normalize(g_group["parent"].get(field_name)), _normalize(e_group["parent"].get(field_name))
            assert gv == ev, f"group {sorted(codes)} parent {field_name}: {gv!r} != {ev!r}"
        assert len(g_group["children"]) == len(e_group["children"])
        for gc, ec in zip(g_group["children"], e_group["children"]):
            for field_name in cols.CMDT_NOTE_ROW_FIELDS:
                if field_name in ignore:
                    continue
                gv, ev = _normalize(gc.get(field_name)), _normalize(ec.get(field_name))
                assert gv == ev, f"group {sorted(codes)} child {gc.get('code')} {field_name}: {gv!r} != {ev!r}"


def _route_notes_by_rates_identity(route_notes: list[dict], rates: list[dict]) -> dict[tuple, dict]:
    """Route notes carry no origin/destination of their own - only a
    (header_seq, route_seq) pair, and header_seq numbering is the same
    non-derivable VBA-counter gap as CMDT NOTE's (see _cmdt_groups). Join
    each route note back to its RATES row by that pair to recover a
    business-meaningful key (origin/dest/cargo), which IS directly
    comparable across generated and ground truth regardless of numbering."""
    by_seq = {(_normalize(r.get("cmdt_seq")), _normalize(r.get("route_seq"))): rates_row_key(r) for r in rates}
    out: dict[tuple, dict] = {}
    for n in route_notes:
        key = by_seq.get((_normalize(n.get("header_seq")), _normalize(n.get("route_seq"))))
        if key is not None:
            out[key] = n
    return out


def _run_tad() -> dict:
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    profile = MappingProfile(rfa_effective_date=RFA_EFFECTIVE, rfa_expiry_date=RFA_EXPIRY)
    return TADWmwWewParser().run_multi(wb, profile)


@pytest.mark.parametrize("scope", ["WEW", "WMW"])
def test_tad_wmw_wew_rates_matches_ground_truth(scope):
    row_set = _run_tad()[scope]
    generated = _normalize_export_punctuation_rows([r.model_dump() for r in row_set.rates])

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = _normalize_export_punctuation_rows(read_rates_sheet(ref_wb, "RATES"))

    result = diff_by_key(generated, expected, key_fn=rates_row_key, fields=cols.RATES_ROW_FIELDS, ignore_fields=RATES_IGNORE_FIELDS)
    assert not result.missing, f"missing {len(result.missing)} expected rows, e.g. {list(result.missing)[:5]}"
    assert not result.extra, f"{len(result.extra)} unexpected generated rows, e.g. {list(result.extra)[:5]}"
    assert not result.field_mismatches, f"{len(result.field_mismatches)} field mismatches, e.g. {result.field_mismatches[:10]}"


@pytest.mark.parametrize("scope", ["WEW", "WMW"])
def test_tad_wmw_wew_route_seq_is_one_continuous_counter_per_group(scope):
    row_set = _run_tad()[scope]
    by_cmdt: dict[int, list[int]] = {}
    for r in row_set.rates:
        by_cmdt.setdefault(r.cmdt_seq, []).append(r.route_seq)
    for cmdt_seq, seqs in by_cmdt.items():
        assert seqs == list(range(1, len(seqs) + 1)), f"cmdt_seq {cmdt_seq}: {seqs}"


@pytest.mark.parametrize("scope", ["WEW", "WMW"])
def test_tad_wmw_wew_cmdt_note_matches_ground_truth(scope):
    row_set = _run_tad()[scope]
    generated = [n.model_dump() for n in row_set.cmdt_notes]

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected = _read_shifted_sheet(ref_wb, "CMDT NOTE", cols.CMDT_NOTE_ROW_FIELDS)

    _assert_cmdt_groups_match(_cmdt_groups(generated), _cmdt_groups(expected))


@pytest.mark.parametrize("scope", ["WEW", "WMW"])
def test_tad_wmw_wew_route_note_matches_ground_truth(scope):
    """WMW covers the combined-note case: a raw row with BOTH T/S Port and
    Service Lane populated (INCOK->EGALY, route_seq 24) joins both lines
    into one cell with ' | ', never splitting into two note_seq rows."""
    row_set = _run_tad()[scope]
    generated_rates = [r.model_dump() for r in row_set.rates]
    generated = _route_notes_by_rates_identity([n.model_dump() for n in row_set.route_notes], generated_rates)

    ref_wb = openpyxl.load_workbook(GROUND_TRUTH[scope], data_only=True, read_only=True)
    expected_rates = read_rates_sheet(ref_wb, "RATES")
    expected_notes = _read_shifted_sheet(ref_wb, "ROUTE NOTE", cols.RN_ROW_FIELDS)
    expected = _route_notes_by_rates_identity(expected_notes, expected_rates)

    assert set(generated) == set(expected), (
        f"only-generated={set(generated) - set(expected)}, only-expected={set(expected) - set(generated)}"
    )
    for key, g in generated.items():
        e = expected[key]
        for field_name in ("contents", "charge_seq", "code", "application_effective", "application_expires", "application"):
            gv, ev = _normalize(g.get(field_name)), _normalize(e.get(field_name))
            assert gv == ev, f"{key} {field_name}: {gv!r} != {ev!r}"


def test_tad_wmw_wew_combined_ts_port_and_service_lane_join_with_pipe():
    row_set = _run_tad()["WMW"]
    contents = {n.contents for n in row_set.route_notes}
    assert "Rates are Subject to Transhipment Port: INMUN | Rates are applicable for Vessel Service Lane: IOM" in contents


def test_tad_wmw_wew_caf_gets_its_own_cmdt_note_child_row():
    row_set = _run_tad()["WEW"]
    codes = {n.code for n in row_set.cmdt_notes}
    assert "CAF" in codes


def test_tad_wmw_wew_excluded_charge_codes_drops_it_end_to_end():
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    profile = MappingProfile(rfa_effective_date=RFA_EFFECTIVE, rfa_expiry_date=RFA_EXPIRY, excluded_charge_codes=["CAF"])
    row_set = TADWmwWewParser().run_multi(wb, profile)["WEW"]

    codes = {n.code for n in row_set.cmdt_notes}
    assert "CAF" not in codes
    assert "CSS" in codes
    assert "CAF" not in row_set.cmdt_notes[0].contents
